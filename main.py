import base64
import calendar
import hashlib
import hmac
import html
import json
import mimetypes
import os
import secrets
import shutil
import sqlite3
import subprocess
import uuid
import zipfile
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path, PurePosixPath
import re
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.parse import quote, urlencode, urlsplit

from fastapi import Depends, FastAPI, File, Form, HTTPException, Query, Request, UploadFile, status as http_status
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, UnidentifiedImageError


BASE_DIR = Path(__file__).resolve().parent
CONFIG_DIR = BASE_DIR / "config"
TEMPLATES_DIR = BASE_DIR / "templates"
STATIC_DIR = BASE_DIR / "static"
DEFAULT_DATA_DIR = BASE_DIR / "data"
DEFAULT_UPLOAD_DIR = BASE_DIR / "uploads"
ENV_FILE = BASE_DIR / ".env"
SESSION_COOKIE_NAME = "admin_session"
DEFAULT_SESSION_SECRET = "store-request-tool-local-dev-session-secret"
PLACEHOLDER_SESSION_SECRET = "change-this-to-a-random-long-secret"
DEFAULT_SESSION_MAX_AGE_HOURS = 12

DEFAULT_STORES = ["南京门东店", "南昌万寿宫店", "山城巷店", "东郊记忆店", "蟠龙天地店", "秀水街店", "湾里店", "下浩里店", "烟台山店"]
DEFAULT_REQUEST_TYPES = ["建单需求", "审单需求", "商品异常", "缺货需求", "新品需求", "系统问题", "其他"]
DEFAULT_URGENCY_LEVELS = ["普通", "加急", "当天必须处理"]
DEFAULT_STATUSES = ["待处理", "处理中", "待门店补充", "已完成", "已驳回"]
DEFAULT_BRANDS: List[str] = []
DEFAULT_HANDLERS: List[str] = []
TASK_STATUSES = ["待处理", "处理中", "已完成"]
EMPLOYEE_STATUSES = ["在职", "离职", "停用"]
EMPLOYEE_ROLE_GROUPS = [
    ("management", "店长 / 经理"),
    ("staff", "店员"),
    ("part_time", "兼职"),
    ("regional", "区域经理"),
    ("unset", "未设置角色"),
    ("other", "其他角色"),
]
SHIFT_DATA_SCOPES = {
    "current": "启用 + 停用",
    "active": "启用班次",
    "inactive": "停用班次",
    "archive": "归档班次",
    "trash": "回收站班次",
    "all": "全部",
}
DUE_STATUS_OPTIONS = ["已超时", "今日到期", "未到期", "未设置", "超时完成", "按时完成"]
LEGACY_ADMIN_REDIRECTS = {
    "/admin/personnel": "/admin/employees",
    "/admin/staff": "/admin/employees",
    "/admin/employee": "/admin/employees",
    "/admin/schedule": "/admin/schedules",
    "/admin/store-schedule": "/admin/schedules",
    "/admin/shift-type": "/admin/shift-types",
    "/admin/archive-list": "/admin/archive",
    "/admin/recycle": "/admin/trash",
    "/admin/trashes": "/admin/trash",
    "/admin/home": "/admin/dashboard",
    "/admin/index": "/admin/dashboard",
    "/admin/tickets": "/admin",
    "/admin/orders": "/admin",
}
LEGACY_PUBLIC_REDIRECTS = {
    "/ticket": "/query",
    "/tickets": "/query",
    "/new": "/submit",
    "/create": "/submit",
    "/form": "/submit",
}
CANONICAL_ACCESS_PATHS = {
    "home": "/",
    "submit": "/submit",
    "query": "/query",
    "store_schedule": "/schedule",
    "admin_login": "/admin/login",
    "dashboard": "/admin/dashboard",
    "tickets": "/admin",
    "my_work": "/admin/my-work",
    "archive": "/admin/archive",
    "trash": "/admin/trash",
    "cleanup": "/admin/cleanup",
    "employees": "/admin/employees",
    "shift_types": "/admin/shift-types",
    "schedules": "/admin/schedules",
    "settings": "/admin/settings",
    "account": "/admin/account",
    "system": "/admin/system",
    "embedded_pages": "/admin/embedded-pages",
    "route_health": "/admin/route-health",
    "version": "/__version",
    "healthz": "/healthz",
}
REQUIRED_RUNTIME_ROUTES = [
    "/",
    "/submit",
    "/query",
    "/query/ticket/{ticket_id}",
    "/query/ticket/{ticket_id}/supplement",
    "/schedule",
    "/admin/login",
    "/admin/dashboard",
    "/admin",
    "/admin/my-work",
    "/admin/archive",
    "/admin/trash",
    "/admin/cleanup",
    "/admin/employees",
    "/admin/shift-types",
    "/admin/schedules",
    "/admin/settings",
    "/admin/account",
    "/admin/system",
    "/admin/embedded-pages",
    "/admin/route-health",
    "/admin/tickets/create",
    "/admin/tickets/bulk-archive",
    "/admin/tickets/bulk-delete",
    "/__version",
    "/healthz",
]
REQUIRED_RUNTIME_ROUTE_LABELS = {
    "/": "首页跳转",
    "/submit": "门店提交工单",
    "/query": "门店查询工单",
    "/query/ticket/{ticket_id}": "门店查看工单详情",
    "/query/ticket/{ticket_id}/supplement": "门店补充资料",
    "/schedule": "门店查看排班",
    "/admin/login": "后台登录",
    "/admin/dashboard": "业务总览",
    "/admin": "工单管理",
    "/admin/my-work": "我的待办",
    "/admin/archive": "归档工单",
    "/admin/trash": "回收站",
    "/admin/cleanup": "测试数据清理",
    "/admin/employees": "员工管理",
    "/admin/shift-types": "班次设置",
    "/admin/schedules": "门店排班",
    "/admin/settings": "配置管理",
    "/admin/account": "账号设置",
    "/admin/system": "系统设置",
    "/admin/embedded-pages": "嵌入页面管理",
    "/admin/route-health": "路由体检",
    "/admin/tickets/create": "后台新建工单",
    "/admin/tickets/bulk-archive": "批量归档工单",
    "/admin/tickets/bulk-delete": "批量移入回收站",
    "/__version": "当前运行版本",
    "/healthz": "健康检查",
}
BULK_SELECTION_REQUIRED_MESSAGE = "请选择要操作的工单"
ROUTE_HEALTH_ATTRIBUTE_RE = re.compile(r'\b(href|action|formaction)\s*=\s*(["\'])(.*?)\2', re.IGNORECASE | re.DOTALL)
ROUTE_HEALTH_FORM_METHOD_RE = re.compile(r'\bmethod\s*=\s*(["\']?)(get|post)\1', re.IGNORECASE)
ROUTE_HEALTH_JINJA_EXPRESSION_RE = re.compile(r"{{.*?}}", re.DOTALL)
APP_STARTED_AT = datetime.now().isoformat(timespec="seconds")
DEFAULT_SYSTEM = {
    "app_name": "门店需求工单系统",
    "port": 8701,
    "max_image_mb": 10,
    "allowed_image_extensions": ["jpg", "jpeg", "png", "webp"],
    "default_status": "待处理",
    "excel_filename_prefix": "门店需求工单",
    "page_size": 50,
    "max_image_count": 5,
    "max_total_upload_mb": 30,
    "allowed_file_extensions": ["pdf", "doc", "docx", "xls", "xlsx", "csv", "txt", "zip", "rar"],
    "max_file_mb": 20,
    "max_file_count": 5,
    "max_total_file_upload_mb": 50,
    "max_embedded_html_mb": 20,
    "max_embedded_zip_mb": 100,
    "store_query_default_days": 30,
    "store_query_page_size": 20,
    "supplement_status_after_store_update": "待处理",
    "max_bulk_schedule_count": 200,
}
COMPLETED_STATUS = "已完成"
BLOCKED_FILE_EXTENSIONS = {"exe", "bat", "cmd", "js", "py", "sh", "php", "jar", "msi"}
EMBEDDED_BLOCKED_EXTENSIONS = {"exe", "bat", "cmd", "sh", "py", "php", "jar", "msi"}
EMBEDDED_ALLOWED_EXTENSIONS = {
    "html",
    "css",
    "js",
    "png",
    "jpg",
    "jpeg",
    "webp",
    "gif",
    "svg",
    "json",
    "csv",
    "txt",
    "pdf",
    "xlsx",
    "docx",
}
IMAGE_FORMAT_EXTENSIONS = {
    "JPEG": {"jpg", "jpeg"},
    "PNG": {"png"},
    "WEBP": {"webp"},
}


@dataclass(frozen=True)
class AppConfig:
    stores: List[str]
    request_types: List[str]
    urgency_levels: List[str]
    statuses: List[str]
    brands: List[str]
    handlers: List[str]
    app_name: str
    port: int
    max_image_mb: int
    allowed_image_extensions: List[str]
    default_status: str
    excel_filename_prefix: str
    page_size: int
    max_image_count: int
    max_total_upload_mb: int
    allowed_file_extensions: List[str]
    max_file_mb: int
    max_file_count: int
    max_total_file_upload_mb: int
    max_embedded_html_mb: int
    max_embedded_zip_mb: int
    store_query_default_days: int
    store_query_page_size: int
    supplement_status_after_store_update: str
    max_bulk_schedule_count: int

    @property
    def max_image_bytes(self) -> int:
        return self.max_image_mb * 1024 * 1024

    @property
    def max_total_upload_bytes(self) -> int:
        return self.max_total_upload_mb * 1024 * 1024

    @property
    def max_file_bytes(self) -> int:
        return self.max_file_mb * 1024 * 1024

    @property
    def max_total_file_upload_bytes(self) -> int:
        return self.max_total_file_upload_mb * 1024 * 1024

    @property
    def max_embedded_html_bytes(self) -> int:
        return self.max_embedded_html_mb * 1024 * 1024

    @property
    def max_embedded_zip_bytes(self) -> int:
        return self.max_embedded_zip_mb * 1024 * 1024


@dataclass(frozen=True)
class PreparedFile:
    original_filename: str
    file_ext: str
    file_size: int
    content: bytes


@dataclass(frozen=True)
class PreparedEmbeddedUpload:
    storage_type: str
    entry_file: str
    file_size: int
    content: bytes
    source_entry_file: str = "index.html"

EXCEL_HEADERS = [
    "工单号",
    "提交时间",
    "门店",
    "提报人",
    "需求类型",
    "紧急程度",
    "品牌",
    "商品名称",
    "规格条码",
    "数量",
    "问题说明",
    "图片路径",
    "文件名称",
    "文件路径",
    "期望完成时间",
    "当前状态",
    "处理人",
    "处理备注",
    "完成时间",
    "最后更新时间",
    "处理时长小时",
    "是否超时",
    "时效状态",
]


def load_env_file() -> None:
    if not ENV_FILE.exists():
        return
    for raw_line in ENV_FILE.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        if not key or key in os.environ:
            continue
        os.environ[key] = value.strip().strip("\"'")


def parse_admin_users(raw_users: str) -> List[Tuple[str, str]]:
    users: List[Tuple[str, str]] = []
    for raw_item in raw_users.split(","):
        username, separator, password = raw_item.partition(":")
        if not separator:
            continue
        username = username.strip()
        password = password.strip()
        if username and password:
            users.append((username, password))
    return users


def get_admin_credentials() -> List[Tuple[str, str]]:
    admin_users = os.environ.get("ADMIN_USERS")
    if admin_users is not None:
        users = parse_admin_users(admin_users)
        if users:
            return users
        raise HTTPException(status_code=503, detail="Admin credentials are not configured.")

    username = os.environ.get("ADMIN_USERNAME", "").strip()
    password = os.environ.get("ADMIN_PASSWORD", "")
    if not username or not password:
        raise HTTPException(status_code=503, detail="Admin credentials are not configured.")
    return [(username, password)]


def admin_usernames_for_handlers() -> List[str]:
    admin_users = os.environ.get("ADMIN_USERS")
    if admin_users is not None:
        return unique_clean_values(username for username, _password in parse_admin_users(admin_users))

    username = os.environ.get("ADMIN_USERNAME", "").strip()
    password = os.environ.get("ADMIN_PASSWORD", "")
    if username and password:
        return [username]
    return []


def authenticate_admin(username: str, password: str) -> Optional[str]:
    input_username = username.strip()
    for expected_username, expected_password in get_admin_credentials():
        username_ok = secrets.compare_digest(input_username, expected_username)
        password_ok = secrets.compare_digest(password, expected_password)
        if username_ok and password_ok:
            return expected_username
    return None


def admin_username_exists(username: str) -> bool:
    input_username = username.strip()
    for expected_username, _expected_password in get_admin_credentials():
        if secrets.compare_digest(input_username, expected_username):
            return True
    return False


def get_session_secret() -> str:
    return os.environ.get("SESSION_SECRET", DEFAULT_SESSION_SECRET)


def get_app_env() -> str:
    return os.environ.get("APP_ENV", "development").strip().lower() or "development"


def get_session_max_age_seconds() -> int:
    raw_value = os.environ.get("SESSION_MAX_AGE_HOURS", str(DEFAULT_SESSION_MAX_AGE_HOURS)).strip()
    try:
        hours = float(raw_value)
    except ValueError:
        hours = DEFAULT_SESSION_MAX_AGE_HOURS
    if hours <= 0:
        hours = DEFAULT_SESSION_MAX_AGE_HOURS
    return int(hours * 3600)


def session_cookie_secure() -> bool:
    return os.environ.get("SESSION_COOKIE_SECURE", "false").strip().lower() in {"1", "true", "yes", "on"}


def warn_if_insecure_production_session() -> None:
    if get_app_env() == "production" and get_session_secret() in {DEFAULT_SESSION_SECRET, PLACEHOLDER_SESSION_SECRET}:
        print(
            "WARNING: APP_ENV=production but SESSION_SECRET is still the default or example value. "
            "Set a random long SESSION_SECRET before exposing the admin system.",
            flush=True,
        )


def base64url_encode(raw: bytes) -> str:
    return base64.urlsafe_b64encode(raw).decode("ascii").rstrip("=")


def base64url_decode(raw: str) -> bytes:
    padding = "=" * (-len(raw) % 4)
    return base64.urlsafe_b64decode((raw + padding).encode("ascii"))


def sign_session_payload(payload: str) -> str:
    signature = hmac.new(get_session_secret().encode("utf-8"), payload.encode("ascii"), hashlib.sha256).digest()
    return base64url_encode(signature)


def create_admin_session(
    username: str,
    issued_at: Optional[int] = None,
    max_age_seconds: Optional[int] = None,
) -> str:
    issued_at_value = int(issued_at if issued_at is not None else datetime.now().timestamp())
    max_age_value = int(max_age_seconds if max_age_seconds is not None else get_session_max_age_seconds())
    payload = base64url_encode(
        json.dumps(
            {
                "username": username,
                "issued_at": issued_at_value,
                "expires_at": issued_at_value + max_age_value,
                "csrf_token": secrets.token_urlsafe(32),
            },
            ensure_ascii=False,
            separators=(",", ":"),
        ).encode("utf-8")
    )
    return f"{payload}.{sign_session_payload(payload)}"


def read_admin_session_data(token: str) -> Optional[Dict[str, Any]]:
    if not token or "." not in token:
        return None
    payload, signature = token.rsplit(".", 1)
    expected_signature = sign_session_payload(payload)
    if not secrets.compare_digest(signature, expected_signature):
        return None
    try:
        data = json.loads(base64url_decode(payload).decode("utf-8"))
    except (ValueError, json.JSONDecodeError):
        return None
    username = str(data.get("username") or "").strip()
    expires_at = data.get("expires_at")
    if not isinstance(expires_at, (int, float)) or datetime.now().timestamp() >= float(expires_at):
        return None
    csrf_token = str(data.get("csrf_token") or "").strip()
    if username and csrf_token and admin_username_exists(username):
        return data
    return None


def read_admin_session(token: str) -> Optional[str]:
    data = read_admin_session_data(token)
    if not data:
        return None
    return str(data.get("username") or "").strip() or None


def current_admin_username(request: Request) -> Optional[str]:
    return read_admin_session(request.cookies.get(SESSION_COOKIE_NAME, ""))


def current_csrf_token(request: Request) -> str:
    data = read_admin_session_data(request.cookies.get(SESSION_COOKIE_NAME, ""))
    if not data:
        return ""
    return str(data.get("csrf_token") or "")


def require_admin_csrf(request: Request, csrf_token: str) -> None:
    expected_token = current_csrf_token(request)
    if not expected_token or not secrets.compare_digest(csrf_token or "", expected_token):
        raise HTTPException(status_code=http_status.HTTP_403_FORBIDDEN, detail="CSRF token invalid.")


def safe_admin_return_url(value: str) -> str:
    value = (value or "").strip()
    if not value:
        return "/admin"
    parsed = urlsplit(value)
    if parsed.scheme or parsed.netloc or not value.startswith("/admin") or value.startswith("//"):
        return "/admin"
    return value


def safe_query_return_url(value: str) -> str:
    value = (value or "").strip()
    if not value:
        return "/query"
    parsed = urlsplit(value)
    if parsed.scheme or parsed.netloc or value.startswith("//"):
        return "/query"
    if parsed.path != "/query" and not parsed.path.startswith("/query/"):
        return "/query"
    return value


def request_path_with_query(request: Request) -> str:
    return request.url.path + (f"?{request.url.query}" if request.url.query else "")


def shift_types_redirect_url(
    request: Request,
    store_names: Optional[List[str]] = None,
    status_filter: str = "",
    global_scope: str = "",
    data_scope: str = "",
    **flags: str,
) -> str:
    if store_names is not None or status_filter or global_scope or data_scope:
        query_items: List[Tuple[str, str]] = []
        query_items.extend(("store_names", store) for store in unique_clean_values(store_names or []))
        if status_filter and status_filter != "all":
            query_items.append(("status", status_filter))
        if global_scope and global_scope != "all":
            query_items.append(("global_scope", global_scope))
        if data_scope and data_scope != "current":
            query_items.append(("data_scope", data_scope))
    else:
        query_items = [
            (key, value)
            for key, value in request.query_params.multi_items()
            if key not in {"error", "success", "updated"} and value
        ]
    query_items.extend((key, value) for key, value in flags.items() if value)
    return "/admin/shift-types" + (f"?{urlencode(query_items)}" if query_items else "")


def build_ticket_detail_url(ticket_id: int, return_url: str = "", **flags: str) -> str:
    params = {key: value for key, value in flags.items() if value}
    safe_return_url = safe_admin_return_url(return_url)
    if safe_return_url != "/admin":
        params["return_url"] = safe_return_url
    query = urlencode(params)
    return f"/admin/ticket/{ticket_id}" + (f"?{query}" if query else "")


def build_store_query_url(store_name: str) -> str:
    clean_store_name = store_name.strip()
    if not clean_store_name:
        return "/query"
    return "/query?" + urlencode({"store_name": clean_store_name})


def store_query_list_return_url(store_name: str, return_url: str = "") -> str:
    if not return_url.strip():
        return build_store_query_url(store_name)
    safe_return_url = safe_query_return_url(return_url)
    if urlsplit(safe_return_url).path == "/query":
        return safe_return_url
    return build_store_query_url(store_name)


def build_store_ticket_detail_url(ticket_id: int, store_name: str, return_url: str = "") -> str:
    params = {"store_name": store_name.strip()}
    safe_return_url = safe_query_return_url(return_url)
    if safe_return_url != "/query":
        params["return_url"] = safe_return_url
    return f"/query/ticket/{ticket_id}?" + urlencode(params)


def build_store_ticket_supplement_url(ticket_id: int, store_name: str, return_url: str = "") -> str:
    params = {"store_name": store_name.strip()}
    safe_return_url = safe_query_return_url(return_url)
    if safe_return_url != "/query":
        params["return_url"] = safe_return_url
    return f"/query/ticket/{ticket_id}/supplement?" + urlencode(params)


def login_redirect_location(request: Request) -> str:
    next_url = safe_admin_return_url(request_path_with_query(request))
    return "/admin/login?" + urlencode({"next": next_url})


def should_redirect_to_login(request: Request) -> bool:
    path = request.url.path
    if path.startswith("/admin/uploads") or path.startswith("/admin/files") or path.startswith("/admin/export"):
        return False
    return (
        path == "/admin"
        or path == "/admin/dashboard"
        or path == "/admin/settings"
        or path == "/admin/account"
        or path == "/admin/system"
        or path == "/admin/route-health"
        or path == "/admin/my-work"
        or path == "/admin/schedules"
        or path == "/admin/employees"
        or path == "/admin/shift-types"
        or path == "/admin/archive"
        or path == "/admin/trash"
        or path == "/admin/cleanup"
        or path.startswith("/admin/schedules/")
        or path.startswith("/admin/employees/")
        or path.startswith("/admin/shift-types/")
        or path.startswith("/admin/ticket")
        or path.startswith("/admin/tickets")
        or path.startswith("/admin/embedded-pages")
        or path.startswith("/admin/embed")
        or path in LEGACY_ADMIN_REDIRECTS
    )


def require_admin(request: Request) -> str:
    username = current_admin_username(request)
    if username:
        return username
    if should_redirect_to_login(request):
        raise HTTPException(status_code=303, detail="Login required.", headers={"Location": login_redirect_location(request)})
    raise HTTPException(
        status_code=http_status.HTTP_401_UNAUTHORIZED,
        detail="Login required.",
    )


def route_pattern(path: str) -> re.Pattern[str]:
    escaped = re.escape(path)
    escaped = re.sub(r"\\\{[^{}:]+:path\\\}", ".+", escaped)
    escaped = re.sub(r"\\\{[^{}]+\\\}", "[^/]+", escaped)
    return re.compile(f"^{escaped}$")


def route_exists(route_pairs: Iterable[Tuple[str, str]], method: str, path: str) -> bool:
    for route_method, route_path in route_pairs:
        if route_method == method and route_pattern(route_path).match(path):
            return True
    return False


def normalize_admin_route_path(raw_value: str) -> Optional[str]:
    value = html.unescape(raw_value.strip())
    if not value.startswith("/admin"):
        return None
    value = ROUTE_HEALTH_JINJA_EXPRESSION_RE.sub("1", value)
    return urlsplit(value).path


def infer_template_route_method(template_text: str, attr: str, attr_start: int) -> str:
    if attr == "href":
        return "GET"
    if attr == "formaction":
        return "POST"
    form_start = template_text.rfind("<form", 0, attr_start)
    form_tag_end = template_text.find(">", form_start)
    form_tag = template_text[form_start:form_tag_end] if form_start >= 0 and form_tag_end >= form_start else ""
    method_match = ROUTE_HEALTH_FORM_METHOD_RE.search(form_tag)
    return method_match.group(2).upper() if method_match else "GET"


def registered_route_pairs(app: FastAPI) -> List[Tuple[str, str]]:
    pairs: List[Tuple[str, str]] = []
    for route in app.routes:
        path = getattr(route, "path", "")
        methods = getattr(route, "methods", None) or []
        for method in sorted(methods):
            if method != "HEAD":
                pairs.append((method, path))
    return sorted(set(pairs), key=lambda item: (item[1], item[0]))


def scan_template_admin_routes(app: FastAPI) -> List[Dict[str, object]]:
    route_pairs = registered_route_pairs(app)
    results: List[Dict[str, object]] = []
    for template_path in sorted(TEMPLATES_DIR.glob("*.html")):
        template_text = template_path.read_text(encoding="utf-8")
        for match in ROUTE_HEALTH_ATTRIBUTE_RE.finditer(template_text):
            attr = match.group(1).lower()
            path = normalize_admin_route_path(match.group(3))
            if not path:
                continue
            method = infer_template_route_method(template_text, attr, match.start())
            exists = route_exists(route_pairs, method, path)
            results.append(
                {
                    "template": template_path.name,
                    "attr": attr,
                    "raw_path": match.group(3),
                    "path": path,
                    "method": method,
                    "exists": exists,
                    "status": "OK" if exists else "MISSING",
                }
            )
    return results


def required_missing_routes(app: FastAPI) -> List[str]:
    paths = {getattr(route, "path", "") for route in app.routes}
    return [path for path in REQUIRED_RUNTIME_ROUTES if path not in paths]


def required_route_items(app: FastAPI) -> List[Dict[str, object]]:
    paths = {getattr(route, "path", "") for route in app.routes}
    return [
        {
            "label": REQUIRED_RUNTIME_ROUTE_LABELS.get(path, path),
            "path": path,
            "exists": path in paths,
            "status": "OK" if path in paths else "MISSING",
        }
        for path in REQUIRED_RUNTIME_ROUTES
    ]


def public_access_urls(port: Optional[int] = None) -> Dict[str, str]:
    active_port = int(port or load_app_config().port)
    base_url = f"http://127.0.0.1:{active_port}"
    return {key: base_url + path for key, path in CANONICAL_ACCESS_PATHS.items()}


def recommended_access_url_entries(port: Optional[int] = None) -> List[Dict[str, str]]:
    labels = {
        "submit": "门店提交",
        "query": "门店查询",
        "admin_login": "后台登录",
        "dashboard": "业务总览",
        "tickets": "工单管理",
        "schedules": "门店排班",
        "employees": "员工管理",
        "shift_types": "班次设置",
        "embedded_pages": "嵌入页面管理",
        "route_health": "路由体检",
        "version": "当前版本",
        "healthz": "健康检查",
    }
    urls = public_access_urls(port)
    return [{"key": key, "label": label, "url": urls[key]} for key, label in labels.items()]


def read_git_head_short() -> str:
    git_dir = BASE_DIR / ".git"
    head_file = git_dir / "HEAD"
    if not head_file.is_file():
        return "unknown"
    try:
        head = head_file.read_text(encoding="utf-8").strip()
    except OSError:
        return "unknown"
    if head.startswith("ref:"):
        ref_path = git_dir / head.removeprefix("ref:").strip()
        try:
            commit = ref_path.read_text(encoding="utf-8").strip()
        except OSError:
            commit = ""
    else:
        commit = head
    return commit[:7] if re.fullmatch(r"[0-9a-fA-F]{7,40}", commit or "") else "unknown"


def current_git_commit() -> str:
    git_candidates = [
        "git",
        r"C:\Program Files\Git\cmd\git.exe",
        r"C:\Program Files\Git\bin\git.exe",
    ]
    for git_exe in git_candidates:
        try:
            result = subprocess.run(
                [git_exe, "rev-parse", "--short", "HEAD"],
                cwd=BASE_DIR,
                capture_output=True,
                text=True,
                timeout=3,
                check=False,
            )
        except (OSError, subprocess.TimeoutExpired):
            continue
        commit = result.stdout.strip()
        if result.returncode == 0 and commit:
            return commit
    return read_git_head_short()


def current_asset_version() -> str:
    commit = current_git_commit()
    if commit != "unknown":
        return commit
    return re.sub(r"\D+", "", APP_STARTED_AT) or "dev"


def get_db_path() -> Path:
    return Path(os.environ.get("STORE_REQUEST_DB_PATH", DEFAULT_DATA_DIR / "tickets.db"))


def get_upload_dir() -> Path:
    return Path(os.environ.get("STORE_REQUEST_UPLOAD_DIR", DEFAULT_UPLOAD_DIR))


def get_embedded_pages_dir() -> Path:
    return get_db_path().parent / "embedded_pages"


def get_config_dir() -> Path:
    return Path(os.environ.get("STORE_REQUEST_CONFIG_DIR", CONFIG_DIR))


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def ensure_directories() -> None:
    get_db_path().parent.mkdir(parents=True, exist_ok=True)
    get_upload_dir().mkdir(parents=True, exist_ok=True)
    get_config_dir().mkdir(parents=True, exist_ok=True)
    TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
    STATIC_DIR.mkdir(parents=True, exist_ok=True)


def get_connection() -> sqlite3.Connection:
    connection = sqlite3.connect(get_db_path())
    connection.row_factory = sqlite3.Row
    return connection


def table_exists(connection: sqlite3.Connection, table_name: str) -> bool:
    row = connection.execute(
        "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = ?",
        (table_name,),
    ).fetchone()
    return row is not None


def column_exists(connection: sqlite3.Connection, table_name: str, column_name: str) -> bool:
    if not table_exists(connection, table_name):
        return False
    rows = connection.execute(f"PRAGMA table_info({table_name})").fetchall()
    return any(str(row["name"]) == column_name for row in rows)


def add_column_if_missing(
    connection: sqlite3.Connection,
    table_name: str,
    column_name: str,
    column_definition: str,
) -> None:
    if not column_exists(connection, table_name, column_name):
        connection.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_definition}")


def add_soft_delete_columns(connection: sqlite3.Connection, table_name: str) -> None:
    add_column_if_missing(connection, table_name, "deleted_at", "TEXT")
    add_column_if_missing(connection, table_name, "deleted_by", "TEXT")
    add_column_if_missing(connection, table_name, "delete_reason", "TEXT")


def add_archive_columns(connection: sqlite3.Connection, table_name: str) -> None:
    add_column_if_missing(connection, table_name, "archived_at", "TEXT")
    add_column_if_missing(connection, table_name, "archived_by", "TEXT")
    add_column_if_missing(connection, table_name, "archive_reason", "TEXT")


MULTI_VALUE_SPLIT_RE = re.compile(r"[,\uFF0C\u3001;\uFF1B]+")
EMBEDDED_PAGE_KEY_RE = re.compile(r"^[a-z0-9]+(?:-[a-z0-9]+)*$")


def unique_clean_values(values: Iterable[object]) -> List[str]:
    cleaned: List[str] = []
    seen: set[str] = set()
    for value in values:
        text = str(value or "").strip()
        if not text or text in seen:
            continue
        cleaned.append(text)
        seen.add(text)
    return cleaned


def split_multi_value_text(value: object) -> List[str]:
    return unique_clean_values(MULTI_VALUE_SPLIT_RE.split(str(value or "")))


def join_display_values(values: Iterable[object]) -> str:
    return "、".join(unique_clean_values(values))


def is_valid_embedded_page_key(page_key: str) -> bool:
    return bool(EMBEDDED_PAGE_KEY_RE.fullmatch(page_key.strip()))


def embedded_page_version(updated_at: object) -> str:
    version = re.sub(r"\D+", "", str(updated_at or ""))
    return version or "0"


def normalize_store_names(raw_store_names: Optional[List[str]], legacy_store_name: str = "") -> List[str]:
    candidates: List[object] = []
    for item in raw_store_names or []:
        candidates.extend(split_multi_value_text(item))
    if not candidates and legacy_store_name.strip():
        candidates.extend(split_multi_value_text(legacy_store_name))
    return unique_clean_values(candidates)


def normalize_employee_scope(value: str) -> str:
    clean_value = value.strip()
    return clean_value if clean_value in {"all", "primary", "support"} else "all"


def normalize_schedule_view(value: str) -> str:
    clean_value = value.strip()
    return clean_value if clean_value in {"calendar", "employee", "table", "store-summary"} else "calendar"


def normalize_employee_record_scope(value: str) -> str:
    clean_value = value.strip()
    return clean_value if clean_value in {"active", "archive", "trash"} else "active"


def normalize_schedule_store_filter(
    raw_store_names: Optional[List[object]],
    legacy_store_name: str,
    config: AppConfig,
    default_to_first: bool = True,
) -> Tuple[List[str], bool, bool]:
    explicit_multi = raw_store_names is not None
    raw_values = clean_multi_value_list(raw_store_names)
    if any(value in {"__all__", "全部门店", "all"} for value in raw_values):
        return [], True, False
    valid_stores = set(config.stores)
    selected_stores: List[str] = []
    has_invalid = False
    for store_name in raw_values:
        if store_name in valid_stores and store_name not in selected_stores:
            selected_stores.append(store_name)
        elif store_name:
            has_invalid = True
    if selected_stores:
        return selected_stores, False, has_invalid
    if explicit_multi:
        return [], True, has_invalid

    legacy_store = legacy_store_name.strip()
    if legacy_store:
        if legacy_store in valid_stores:
            return [legacy_store], False, False
        return [], False, True
    if default_to_first and config.stores:
        return [config.stores[0]], False, False
    return [], True, False


def normalize_employee_status_filters(values: Optional[List[object]], legacy_status: str = "") -> Tuple[List[str], bool]:
    raw_values = clean_multi_value_list(values)
    if not raw_values and legacy_status.strip():
        raw_values = [legacy_status.strip()]
    clean_statuses: List[str] = []
    has_invalid = False
    for status in raw_values:
        if status in EMPLOYEE_STATUSES and status not in clean_statuses:
            clean_statuses.append(status)
        elif status:
            has_invalid = True
    return clean_statuses, has_invalid


def normalize_employee_scope_filter(values: Optional[List[object]], legacy_scope: str = "all") -> str:
    scopes = [normalize_employee_scope(value) for value in clean_multi_value_list(values)]
    scopes = [scope for scope in scopes if scope in {"primary", "support", "all"}]
    if not scopes:
        return normalize_employee_scope(legacy_scope)
    if "all" in scopes or {"primary", "support"}.issubset(set(scopes)):
        return "all"
    return scopes[0]


def parse_optional_int(value: object) -> Optional[int]:
    clean_value = str(value or "").strip()
    if not clean_value:
        return None
    try:
        return int(clean_value)
    except (TypeError, ValueError):
        return None


def parse_optional_positive_int(value: object) -> Optional[int]:
    parsed = parse_optional_int(value)
    return parsed if parsed and parsed > 0 else None


def parse_optional_int_list(values: Optional[List[object]]) -> Tuple[List[int], bool]:
    parsed_values: List[int] = []
    has_invalid = False
    for value in values or []:
        clean_value = str(value or "").strip()
        if not clean_value:
            continue
        parsed = parse_optional_positive_int(clean_value)
        if parsed is None:
            has_invalid = True
            continue
        if parsed not in parsed_values:
            parsed_values.append(parsed)
    return parsed_values, has_invalid


def clean_multi_value_list(values: Optional[List[object]]) -> List[str]:
    clean_values: List[str] = []
    for value in values or []:
        for item in split_multi_value_text(value):
            clean_item = str(item or "").strip()
            if clean_item and clean_item not in clean_values:
                clean_values.append(clean_item)
    return clean_values


def parse_shift_filter_values(values: Optional[List[object]], legacy_value: object = "") -> Tuple[List[int], bool, bool]:
    raw_values = list(values or [])
    if not raw_values and str(legacy_value or "").strip():
        raw_values = [legacy_value]
    parsed_values: List[int] = []
    include_custom = False
    has_invalid = False
    for value in raw_values:
        clean_value = str(value or "").strip()
        if not clean_value:
            continue
        if clean_value == "custom":
            include_custom = True
            continue
        parsed = parse_optional_positive_int(clean_value)
        if parsed is None:
            has_invalid = True
            continue
        if parsed not in parsed_values:
            parsed_values.append(parsed)
    return parsed_values, include_custom, has_invalid


def combine_error_messages(*messages: str) -> str:
    return "；".join(message.strip() for message in messages if message and message.strip())


def normalize_brand_names(
    raw_brands: Optional[List[str]],
    brand_extra: str = "",
    legacy_brand: str = "",
) -> List[str]:
    candidates: List[object] = []
    for item in raw_brands or []:
        candidates.extend(split_multi_value_text(item))
    candidates.extend(split_multi_value_text(brand_extra))
    if not candidates and legacy_brand.strip():
        candidates.extend(split_multi_value_text(legacy_brand))
    return unique_clean_values(candidates)


def backfill_ticket_relations(connection: sqlite3.Connection) -> None:
    timestamp = now_text()
    rows = connection.execute("SELECT id, store_name, brand, created_at FROM tickets").fetchall()
    for row in rows:
        ticket_id = int(row["id"])
        created_at = str(row["created_at"] or timestamp)
        for store_name in split_multi_value_text(row["store_name"]):
            connection.execute(
                """
                INSERT OR IGNORE INTO ticket_stores (ticket_id, store_name, created_at)
                VALUES (?, ?, ?)
                """,
                (ticket_id, store_name, created_at),
            )
        for brand in split_multi_value_text(row["brand"]):
            connection.execute(
                """
                INSERT OR IGNORE INTO ticket_brands (ticket_id, brand, created_at)
                VALUES (?, ?, ?)
                """,
                (ticket_id, brand, created_at),
            )


def backfill_employee_store_map(connection: sqlite3.Connection) -> None:
    timestamp = now_text()
    rows = connection.execute("SELECT id, store_name, primary_store_name, created_at FROM employees").fetchall()
    for row in rows:
        employee_id = int(row["id"])
        created_at = str(row["created_at"] or timestamp)
        existing_store_names = split_multi_value_text(row["store_name"])
        primary_store = str(row["primary_store_name"] or "").strip() or (existing_store_names[0] if existing_store_names else "")
        if primary_store and not str(row["primary_store_name"] or "").strip():
            connection.execute(
                "UPDATE employees SET primary_store_name = ?, store_name = ? WHERE id = ?",
                (primary_store, primary_store, employee_id),
            )
        for store_name in unique_clean_values([primary_store, *existing_store_names]):
            connection.execute(
                """
                INSERT OR IGNORE INTO employee_store_map (employee_id, store_name, created_at)
                VALUES (?, ?, ?)
                """,
                (employee_id, store_name, created_at),
            )


DEFAULT_SHIFT_TYPES = [
    ("早班", "09:30", "17:30", 8.0, "#2563eb"),
    ("晚班", "14:00", "22:00", 8.0, "#7c3aed"),
    ("全天", "09:30", "22:00", 12.5, "#059669"),
    ("休息", "", "", 0.0, "#64748b"),
]


def ensure_default_shift_types(connection: sqlite3.Connection) -> None:
    row = connection.execute("SELECT COUNT(*) AS total FROM shift_types").fetchone()
    if int(row["total"] or 0) > 0:
        return
    timestamp = now_text()
    for shift_name, start_time, end_time, duration_hours, color in DEFAULT_SHIFT_TYPES:
        connection.execute(
            """
            INSERT INTO shift_types (
                shift_name, start_time, end_time, duration_hours,
                color, is_active, store_name, is_global,
                business_start_time, business_end_time, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, 1, '', 1, '', '', ?, ?)
            """,
            (shift_name, start_time, end_time, duration_hours, color, timestamp, timestamp),
        )


def shift_types_has_unique_name_constraint(connection: sqlite3.Connection) -> bool:
    if not table_exists(connection, "shift_types"):
        return False
    for index_row in connection.execute("PRAGMA index_list(shift_types)").fetchall():
        if int(index_row["unique"] or 0) != 1:
            continue
        index_name = str(index_row["name"] or "")
        columns = [str(row["name"] or "") for row in connection.execute(f"PRAGMA index_info({index_name})").fetchall()]
        if columns == ["shift_name"]:
            return True
    return False


def recreate_shift_types_without_unique_name(connection: sqlite3.Connection) -> None:
    if not shift_types_has_unique_name_constraint(connection):
        return
    backup_table = f"shift_types_unique_backup_{datetime.now().strftime('%Y%m%d%H%M%S%f')}"
    source_columns = {
        str(row["name"])
        for row in connection.execute("PRAGMA table_info(shift_types)").fetchall()
    }

    def source_expr(column_name: str, fallback: str) -> str:
        return column_name if column_name in source_columns else fallback

    connection.execute(f"ALTER TABLE shift_types RENAME TO {backup_table}")
    connection.execute(
        """
        CREATE TABLE shift_types (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            shift_name TEXT NOT NULL,
            store_name TEXT,
            is_global INTEGER DEFAULT 0,
            business_start_time TEXT,
            business_end_time TEXT,
            start_time TEXT,
            end_time TEXT,
            duration_hours REAL DEFAULT 0,
            color TEXT,
            is_active INTEGER NOT NULL DEFAULT 1,
            archived_at TEXT,
            archived_by TEXT,
            archive_reason TEXT,
            deleted_at TEXT,
            deleted_by TEXT,
            delete_reason TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    connection.execute(
        f"""
        INSERT INTO shift_types (
            id, shift_name, store_name, is_global, business_start_time, business_end_time,
            start_time, end_time, duration_hours, color, is_active,
            archived_at, archived_by, archive_reason, deleted_at, deleted_by, delete_reason,
            created_at, updated_at
        )
        SELECT
            {source_expr('id', 'NULL')},
            {source_expr('shift_name', "''")},
            {source_expr('store_name', "''")},
            {source_expr('is_global', '1')},
            {source_expr('business_start_time', "''")},
            {source_expr('business_end_time', "''")},
            {source_expr('start_time', "''")},
            {source_expr('end_time', "''")},
            {source_expr('duration_hours', '0')},
            {source_expr('color', "''")},
            {source_expr('is_active', '1')},
            {source_expr('archived_at', 'NULL')},
            {source_expr('archived_by', 'NULL')},
            {source_expr('archive_reason', 'NULL')},
            {source_expr('deleted_at', 'NULL')},
            {source_expr('deleted_by', 'NULL')},
            {source_expr('delete_reason', 'NULL')},
            {source_expr('created_at', "datetime('now')")},
            {source_expr('updated_at', "datetime('now')")}
        FROM {backup_table}
        ORDER BY id
        """
    )
    connection.execute(f"DROP TABLE {backup_table}")


def ensure_schedule_schema_migrations(connection: sqlite3.Connection) -> None:
    add_column_if_missing(connection, "shift_types", "store_name", "TEXT")
    add_column_if_missing(connection, "shift_types", "is_global", "INTEGER DEFAULT 0")
    add_column_if_missing(connection, "shift_types", "business_start_time", "TEXT")
    add_column_if_missing(connection, "shift_types", "business_end_time", "TEXT")
    recreate_shift_types_without_unique_name(connection)
    add_archive_columns(connection, "shift_types")
    add_soft_delete_columns(connection, "shift_types")
    connection.execute(
        """
        UPDATE shift_types
        SET is_global = 1, store_name = COALESCE(store_name, '')
        WHERE COALESCE(store_name, '') = ''
        """
    )
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS store_business_hours (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            store_name TEXT NOT NULL UNIQUE,
            business_start_time TEXT,
            business_end_time TEXT,
            created_at TEXT,
            updated_at TEXT
        )
        """
    )
    add_column_if_missing(connection, "store_schedules", "is_custom_time", "INTEGER DEFAULT 0")
    add_column_if_missing(connection, "store_schedules", "custom_start_time", "TEXT")
    add_column_if_missing(connection, "store_schedules", "custom_end_time", "TEXT")
    add_column_if_missing(connection, "store_schedules", "custom_duration_hours", "REAL")
    add_column_if_missing(connection, "store_schedules", "custom_label", "TEXT")


def init_db() -> None:
    ensure_directories()
    with get_connection() as connection:
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS tickets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ticket_no TEXT NOT NULL UNIQUE,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                store_name TEXT NOT NULL,
                submitter TEXT NOT NULL,
                request_type TEXT NOT NULL,
                urgency TEXT NOT NULL,
                brand TEXT,
                product_name TEXT,
                sku_barcode TEXT,
                quantity INTEGER,
                description TEXT NOT NULL,
                expected_finish_date TEXT,
                status TEXT NOT NULL,
                assigned_to TEXT,
                handler_note TEXT,
                closed_at TEXT
            )
            """
        )
        add_column_if_missing(connection, "tickets", "assigned_to", "TEXT")
        add_column_if_missing(connection, "tickets", "closed_at", "TEXT")
        add_soft_delete_columns(connection, "tickets")
        add_archive_columns(connection, "tickets")
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS ticket_stores (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ticket_id INTEGER NOT NULL,
                store_name TEXT NOT NULL,
                created_at TEXT NOT NULL,
                UNIQUE(ticket_id, store_name),
                FOREIGN KEY (ticket_id) REFERENCES tickets(id) ON DELETE CASCADE
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS ticket_brands (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ticket_id INTEGER NOT NULL,
                brand TEXT NOT NULL,
                created_at TEXT NOT NULL,
                UNIQUE(ticket_id, brand),
                FOREIGN KEY (ticket_id) REFERENCES tickets(id) ON DELETE CASCADE
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS ticket_images (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ticket_id INTEGER NOT NULL,
                image_path TEXT NOT NULL,
                uploaded_at TEXT NOT NULL,
                source TEXT,
                uploaded_by TEXT,
                supplement_id INTEGER,
                FOREIGN KEY (ticket_id) REFERENCES tickets(id) ON DELETE CASCADE
            )
            """
        )
        add_column_if_missing(connection, "ticket_images", "source", "TEXT")
        add_column_if_missing(connection, "ticket_images", "uploaded_by", "TEXT")
        add_column_if_missing(connection, "ticket_images", "supplement_id", "INTEGER")
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS ticket_files (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ticket_id INTEGER NOT NULL,
                original_filename TEXT NOT NULL,
                stored_filename TEXT NOT NULL,
                file_path TEXT NOT NULL,
                file_ext TEXT NOT NULL,
                file_size INTEGER NOT NULL,
                uploaded_at TEXT NOT NULL,
                source TEXT,
                uploaded_by TEXT,
                supplement_id INTEGER,
                FOREIGN KEY (ticket_id) REFERENCES tickets(id) ON DELETE CASCADE
            )
            """
        )
        add_column_if_missing(connection, "ticket_files", "source", "TEXT")
        add_column_if_missing(connection, "ticket_files", "uploaded_by", "TEXT")
        add_column_if_missing(connection, "ticket_files", "supplement_id", "INTEGER")
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS ticket_supplements (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ticket_id INTEGER NOT NULL,
                store_name TEXT NOT NULL,
                submitter TEXT NOT NULL,
                note TEXT,
                image_count INTEGER DEFAULT 0,
                file_count INTEGER DEFAULT 0,
                created_at TEXT NOT NULL,
                FOREIGN KEY (ticket_id) REFERENCES tickets(id) ON DELETE CASCADE
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS ticket_participants (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ticket_id INTEGER NOT NULL,
                participant_type TEXT NOT NULL,
                participant_name TEXT NOT NULL,
                role TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (ticket_id) REFERENCES tickets(id) ON DELETE CASCADE
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS ticket_comments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ticket_id INTEGER NOT NULL,
                author_type TEXT NOT NULL,
                author_name TEXT NOT NULL,
                content TEXT NOT NULL,
                visibility TEXT NOT NULL,
                created_at TEXT NOT NULL,
                FOREIGN KEY (ticket_id) REFERENCES tickets(id) ON DELETE CASCADE
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS ticket_tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ticket_id INTEGER NOT NULL,
                title TEXT NOT NULL,
                assignee TEXT,
                status TEXT NOT NULL,
                due_date TEXT,
                completed_at TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY (ticket_id) REFERENCES tickets(id) ON DELETE CASCADE
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS ticket_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ticket_id INTEGER NOT NULL,
                action TEXT NOT NULL,
                old_status TEXT,
                new_status TEXT,
                old_assigned_to TEXT,
                new_assigned_to TEXT,
                note TEXT,
                operator TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (ticket_id) REFERENCES tickets(id) ON DELETE CASCADE
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS notification_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                event_type TEXT NOT NULL,
                ticket_id INTEGER,
                ticket_no TEXT,
                store_name TEXT,
                title TEXT NOT NULL,
                content TEXT,
                severity TEXT NOT NULL DEFAULT 'info',
                created_by TEXT,
                created_at TEXT NOT NULL
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS notification_reads (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                event_id INTEGER NOT NULL,
                username TEXT NOT NULL,
                read_at TEXT NOT NULL,
                UNIQUE(event_id, username),
                FOREIGN KEY (event_id) REFERENCES notification_events(id) ON DELETE CASCADE
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS embedded_pages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                page_key TEXT NOT NULL UNIQUE,
                title TEXT NOT NULL,
                nav_label TEXT NOT NULL,
                filename TEXT NOT NULL,
                storage_type TEXT NOT NULL DEFAULT 'html',
                entry_file TEXT NOT NULL DEFAULT 'index.html',
                file_size INTEGER NOT NULL DEFAULT 0,
                enabled INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                updated_by TEXT
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS employees (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_name TEXT NOT NULL,
                store_name TEXT NOT NULL,
                primary_store_name TEXT,
                role TEXT,
                phone TEXT,
                status TEXT NOT NULL DEFAULT '在职',
                archived_at TEXT,
                archived_by TEXT,
                archive_reason TEXT,
                deleted_at TEXT,
                deleted_by TEXT,
                delete_reason TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS employee_store_map (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL,
                store_name TEXT NOT NULL,
                created_at TEXT NOT NULL,
                UNIQUE(employee_id, store_name),
                FOREIGN KEY (employee_id) REFERENCES employees(id)
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS shift_types (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                shift_name TEXT NOT NULL,
                store_name TEXT,
                is_global INTEGER DEFAULT 0,
                business_start_time TEXT,
                business_end_time TEXT,
                start_time TEXT,
                end_time TEXT,
                duration_hours REAL DEFAULT 0,
                color TEXT,
                is_active INTEGER NOT NULL DEFAULT 1,
                archived_at TEXT,
                archived_by TEXT,
                archive_reason TEXT,
                deleted_at TEXT,
                deleted_by TEXT,
                delete_reason TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS store_schedules (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                store_name TEXT NOT NULL,
                employee_id INTEGER NOT NULL,
                schedule_date TEXT NOT NULL,
                shift_type_id INTEGER NOT NULL DEFAULT 0,
                note TEXT,
                is_custom_time INTEGER DEFAULT 0,
                custom_start_time TEXT,
                custom_end_time TEXT,
                custom_duration_hours REAL,
                custom_label TEXT,
                created_by TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(employee_id, schedule_date),
                FOREIGN KEY (employee_id) REFERENCES employees(id),
                FOREIGN KEY (shift_type_id) REFERENCES shift_types(id)
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS schedule_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                schedule_id INTEGER,
                action TEXT NOT NULL,
                old_value TEXT,
                new_value TEXT,
                operator TEXT,
                created_at TEXT NOT NULL
            )
            """
        )
        connection.execute("CREATE INDEX IF NOT EXISTS idx_tickets_created_at ON tickets(created_at)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_tickets_status ON tickets(status)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_tickets_assigned_to ON tickets(assigned_to)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_tickets_deleted_at ON tickets(deleted_at)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_tickets_archived_at ON tickets(archived_at)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_ticket_stores_store_name ON ticket_stores(store_name)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_ticket_stores_ticket_id ON ticket_stores(ticket_id)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_ticket_brands_brand ON ticket_brands(brand)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_ticket_brands_ticket_id ON ticket_brands(ticket_id)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_ticket_images_ticket_id ON ticket_images(ticket_id)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_ticket_files_ticket_id ON ticket_files(ticket_id)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_ticket_logs_ticket_id ON ticket_logs(ticket_id)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_ticket_supplements_ticket_id ON ticket_supplements(ticket_id)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_ticket_participants_ticket_id ON ticket_participants(ticket_id)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_ticket_comments_ticket_id ON ticket_comments(ticket_id)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_ticket_comments_visibility ON ticket_comments(visibility)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_ticket_tasks_ticket_id ON ticket_tasks(ticket_id)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_ticket_tasks_status ON ticket_tasks(status)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_notification_events_created_at ON notification_events(created_at)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_notification_events_ticket_id ON notification_events(ticket_id)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_notification_reads_username ON notification_reads(username)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_embedded_pages_enabled ON embedded_pages(enabled)")
        add_column_if_missing(connection, "employees", "primary_store_name", "TEXT")
        add_archive_columns(connection, "employees")
        add_soft_delete_columns(connection, "employees")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_employees_store_name ON employees(store_name)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_employees_primary_store_name ON employees(primary_store_name)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_employees_deleted_at ON employees(deleted_at)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_employees_archived_at ON employees(archived_at)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_employee_store_map_employee_id ON employee_store_map(employee_id)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_employee_store_map_store_name ON employee_store_map(store_name)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_store_schedules_store_name ON store_schedules(store_name)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_store_schedules_schedule_date ON store_schedules(schedule_date)")
        ensure_schedule_schema_migrations(connection)
        connection.execute("CREATE INDEX IF NOT EXISTS idx_shift_types_store_name ON shift_types(store_name)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_shift_types_is_global ON shift_types(is_global)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_shift_types_archived_at ON shift_types(archived_at)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_shift_types_deleted_at ON shift_types(deleted_at)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_store_business_hours_store_name ON store_business_hours(store_name)")
        add_column_if_missing(connection, "embedded_pages", "storage_type", "TEXT NOT NULL DEFAULT 'html'")
        add_column_if_missing(connection, "embedded_pages", "entry_file", "TEXT NOT NULL DEFAULT 'index.html'")
        add_column_if_missing(connection, "embedded_pages", "file_size", "INTEGER NOT NULL DEFAULT 0")
        for table_name in (
            "ticket_supplements",
            "ticket_participants",
            "ticket_comments",
            "ticket_tasks",
        ):
            add_soft_delete_columns(connection, table_name)
        add_soft_delete_columns(connection, "embedded_pages")
        backfill_ticket_relations(connection)
        backfill_employee_store_map(connection)
        ensure_default_shift_types(connection)


def clean_string_list(value: object, default: List[str], allow_empty: bool = False) -> List[str]:
    if not isinstance(value, list):
        return list(default)
    clean_values = [str(item).strip() for item in value if str(item).strip()]
    if clean_values or allow_empty:
        return clean_values
    return list(default)


def load_json_file(filename: str, default: object) -> object:
    try:
        with (get_config_dir() / filename).open("r", encoding="utf-8") as file:
            return json.load(file)
    except (OSError, json.JSONDecodeError):
        return default


def load_list_config(filename: str, default: List[str], allow_empty: bool = False) -> List[str]:
    return clean_string_list(load_json_file(filename, default), default, allow_empty=allow_empty)


def load_handlers() -> List[str]:
    admin_handlers = admin_usernames_for_handlers()
    handlers_path = get_config_dir() / "handlers.json"
    if handlers_path.exists():
        configured_handlers = load_list_config("handlers.json", [], allow_empty=True)
        return unique_clean_values([*configured_handlers, *admin_handlers])
    return unique_clean_values(admin_handlers)


def positive_int_config(system: Dict[str, object], key: str) -> int:
    value = system.get(key)
    default_value = int(DEFAULT_SYSTEM[key])
    if isinstance(value, bool):
        return default_value
    if isinstance(value, int) and value > 0:
        return value
    return default_value


def normalized_extensions(value: object, default: List[str], blocked: Optional[set[str]] = None) -> List[str]:
    extensions = clean_string_list(value, default)
    blocked = blocked or set()
    normalized: List[str] = []
    for extension in extensions:
        clean_extension = extension.lower().lstrip(".")
        if clean_extension and clean_extension not in blocked and clean_extension not in normalized:
            normalized.append(clean_extension)
    if normalized:
        return normalized
    return [extension for extension in default if extension not in blocked]


def load_system_config(statuses: List[str]) -> Dict[str, object]:
    raw_system = load_json_file("system.json", DEFAULT_SYSTEM)
    system = dict(DEFAULT_SYSTEM)
    if isinstance(raw_system, dict):
        system.update(raw_system)

    for key in (
        "max_image_mb",
        "page_size",
        "max_image_count",
        "max_total_upload_mb",
        "max_file_mb",
        "max_file_count",
        "max_total_file_upload_mb",
        "max_embedded_html_mb",
        "max_embedded_zip_mb",
        "store_query_default_days",
        "store_query_page_size",
        "max_bulk_schedule_count",
    ):
        system[key] = positive_int_config(system, key)

    system["allowed_image_extensions"] = normalized_extensions(
        system.get("allowed_image_extensions"),
        list(DEFAULT_SYSTEM["allowed_image_extensions"]),
    )
    system["allowed_file_extensions"] = normalized_extensions(
        system.get("allowed_file_extensions"),
        list(DEFAULT_SYSTEM["allowed_file_extensions"]),
        BLOCKED_FILE_EXTENSIONS,
    )

    default_status = str(system.get("default_status", "")).strip()
    system["default_status"] = default_status if default_status in statuses else statuses[0]

    supplement_status = str(system.get("supplement_status_after_store_update", "")).strip()
    system["supplement_status_after_store_update"] = supplement_status if supplement_status in statuses else "待处理"

    for key in ("app_name", "excel_filename_prefix"):
        value = str(system.get(key, "")).strip()
        if value:
            system[key] = value
        else:
            system[key] = DEFAULT_SYSTEM[key]

    port = system.get("port")
    if not isinstance(port, int) or port <= 0:
        system["port"] = DEFAULT_SYSTEM["port"]

    return system


def load_app_config() -> AppConfig:
    statuses = load_list_config("statuses.json", DEFAULT_STATUSES)
    system = load_system_config(statuses)
    return AppConfig(
        stores=load_list_config("stores.json", DEFAULT_STORES),
        request_types=load_list_config("request_types.json", DEFAULT_REQUEST_TYPES),
        urgency_levels=load_list_config("urgency_levels.json", DEFAULT_URGENCY_LEVELS),
        statuses=statuses,
        brands=load_list_config("brands.json", DEFAULT_BRANDS, allow_empty=True),
        handlers=load_handlers(),
        app_name=str(system["app_name"]),
        port=int(system["port"]),
        max_image_mb=int(system["max_image_mb"]),
        allowed_image_extensions=list(system["allowed_image_extensions"]),
        default_status=str(system["default_status"]),
        excel_filename_prefix=str(system["excel_filename_prefix"]),
        page_size=int(system["page_size"]),
        max_image_count=int(system["max_image_count"]),
        max_total_upload_mb=int(system["max_total_upload_mb"]),
        allowed_file_extensions=list(system["allowed_file_extensions"]),
        max_file_mb=int(system["max_file_mb"]),
        max_file_count=int(system["max_file_count"]),
        max_total_file_upload_mb=int(system["max_total_file_upload_mb"]),
        max_embedded_html_mb=int(system["max_embedded_html_mb"]),
        max_embedded_zip_mb=int(system["max_embedded_zip_mb"]),
        store_query_default_days=int(system["store_query_default_days"]),
        store_query_page_size=int(system["store_query_page_size"]),
        supplement_status_after_store_update=str(system["supplement_status_after_store_update"]),
        max_bulk_schedule_count=int(system["max_bulk_schedule_count"]),
    )


def load_stores() -> List[str]:
    return load_app_config().stores


REQUEST_RULE_FIELD_LABELS = {
    "brand": "品牌",
    "product_name": "商品名称",
    "sku_barcode": "规格条码",
    "quantity": "数量",
    "description": "问题说明",
    "expected_finish_date": "期望完成时间",
}


def load_request_type_rules() -> Dict[str, Dict[str, object]]:
    raw_rules = load_json_file("request_type_rules.json", {})
    if not isinstance(raw_rules, dict):
        return {}
    rules: Dict[str, Dict[str, object]] = {}
    for request_type, raw_rule in raw_rules.items():
        if not isinstance(raw_rule, dict):
            continue
        required_fields = [
            str(field).strip()
            for field in raw_rule.get("required_fields", [])
            if str(field).strip() in REQUEST_RULE_FIELD_LABELS
        ]
        rules[str(request_type)] = {
            "required_fields": required_fields,
            "require_image": bool(raw_rule.get("require_image", False)),
            "require_file": bool(raw_rule.get("require_file", False)),
            "require_any_attachment": bool(raw_rule.get("require_any_attachment", False)),
            "description_hint": str(raw_rule.get("description_hint") or "").strip(),
        }
    return rules


def compact_text(value: Optional[str], max_len: int = 36) -> str:
    text = (value or "").strip()
    if len(text) <= max_len:
        return text
    return text[:max_len] + "..."


def image_filename(image_path: str) -> str:
    return Path(str(image_path).replace("\\", "/")).name


def protected_upload_url(image_path: str) -> str:
    return f"/admin/uploads/{quote(image_filename(image_path))}"


def protected_file_url(file_id: object) -> str:
    return f"/admin/files/{file_id}"


def file_size_label(size: object) -> str:
    try:
        size_value = int(size)
    except (TypeError, ValueError):
        return ""
    if size_value >= 1024 * 1024:
        return f"{size_value / (1024 * 1024):.1f}MB"
    if size_value >= 1024:
        return f"{size_value / 1024:.1f}KB"
    return f"{size_value}B"


def safe_uploaded_name(filename: str) -> str:
    return Path(str(filename or "").replace("\\", "/")).name.strip()


def resolve_upload_path(filename: str) -> Optional[Path]:
    normalized = filename.replace("\\", "/")
    if not normalized or "/" in normalized or normalized in {".", ".."}:
        return None
    upload_root = get_upload_dir().resolve()
    target = (upload_root / normalized).resolve()
    try:
        target.relative_to(upload_root)
    except ValueError:
        return None
    return target


def resolve_upload_file(filename: str) -> Optional[Path]:
    target = resolve_upload_path(filename)
    if target is None:
        return None
    if not target.is_file():
        return None
    return target


def active_ticket_image_exists(filename: str) -> bool:
    clean_filename = safe_uploaded_name(filename)
    if not clean_filename:
        return False
    with get_connection() as connection:
        row = connection.execute(
            """
            SELECT 1
            FROM ticket_images
            JOIN tickets ON tickets.id = ticket_images.ticket_id
            WHERE tickets.deleted_at IS NULL
              AND (ticket_images.image_path = ? OR ticket_images.image_path = ?)
            """,
            (clean_filename, f"uploads/{clean_filename}"),
        ).fetchone()
    return bool(row)


def generate_ticket_no(connection: sqlite3.Connection) -> str:
    today = datetime.now().strftime("%Y%m%d")
    prefix = f"REQ-{today}-"
    row = connection.execute(
        "SELECT MAX(CAST(SUBSTR(ticket_no, 14) AS INTEGER)) AS max_no FROM tickets WHERE ticket_no LIKE ?",
        (prefix + "%",),
    ).fetchone()
    next_no = (row["max_no"] or 0) + 1
    return f"{prefix}{next_no:04d}"


def parse_datetime_text(value: object) -> Optional[datetime]:
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def processing_hours(ticket: Dict[str, object]) -> object:
    created_at = parse_datetime_text(ticket.get("created_at"))
    if not created_at:
        return ""
    finished_at = parse_datetime_text(ticket.get("closed_at")) or datetime.now()
    seconds = max((finished_at - created_at).total_seconds(), 0)
    return round(seconds / 3600, 2)


def overdue_text(ticket: Dict[str, object]) -> str:
    expected = parse_datetime_text(ticket.get("expected_finish_date"))
    if not expected:
        return ""
    closed_at = parse_datetime_text(ticket.get("closed_at"))
    compare_date = closed_at.date() if closed_at else datetime.now().date()
    return "是" if compare_date > expected.date() else "否"


def due_status_label(ticket: Dict[str, object]) -> str:
    expected = parse_datetime_text(ticket.get("expected_finish_date"))
    if not expected:
        return "未设置"
    expected_date = expected.date()
    status = str(ticket.get("status") or "")
    closed_at = parse_datetime_text(ticket.get("closed_at"))
    if status == COMPLETED_STATUS:
        closed_date = (closed_at or parse_datetime_text(ticket.get("updated_at")) or datetime.now()).date()
        return "超时完成" if closed_date > expected_date else "按时完成"
    today = datetime.now().date()
    if today > expected_date:
        return "已超时"
    if today == expected_date:
        return "今日到期"
    return "未到期"


def due_status_class(label: str) -> str:
    mapping = {
        "已超时": "overdue",
        "今日到期": "due-today",
        "未到期": "not-due",
        "未设置": "unset",
        "超时完成": "completed-late",
        "按时完成": "completed-on-time",
    }
    return mapping.get(label, "unset")


def annotate_ticket_runtime(ticket: Dict[str, object]) -> Dict[str, object]:
    ticket.setdefault("store_names", split_multi_value_text(ticket.get("store_name")))
    ticket.setdefault("brand_names", split_multi_value_text(ticket.get("brand")))
    label = due_status_label(ticket)
    ticket["due_status"] = label
    ticket["due_status_class"] = due_status_class(label)
    ticket["description_summary"] = compact_text(str(ticket.get("description") or ""))
    ticket["is_archived"] = bool(ticket.get("archived_at"))
    ticket["is_deleted"] = bool(ticket.get("deleted_at"))
    return ticket


def relation_map_for_tickets(ticket_ids: List[int], table_name: str, value_column: str) -> Dict[int, List[str]]:
    if not ticket_ids:
        return {}
    placeholders = ",".join("?" for _ in ticket_ids)
    mapping: Dict[int, List[str]] = {ticket_id: [] for ticket_id in ticket_ids}
    with get_connection() as connection:
        rows = connection.execute(
            f"""
            SELECT ticket_id, {value_column} AS value
            FROM {table_name}
            WHERE ticket_id IN ({placeholders})
            ORDER BY id
            """,
            ticket_ids,
        ).fetchall()
    for row in rows:
        ticket_id = int(row["ticket_id"])
        value = str(row["value"] or "").strip()
        if value:
            mapping.setdefault(ticket_id, []).append(value)
    return mapping


def attach_ticket_relations(tickets: List[Dict[str, object]]) -> List[Dict[str, object]]:
    ticket_ids = [int(ticket["id"]) for ticket in tickets if ticket.get("id") is not None]
    store_map = relation_map_for_tickets(ticket_ids, "ticket_stores", "store_name")
    brand_map = relation_map_for_tickets(ticket_ids, "ticket_brands", "brand")
    for ticket in tickets:
        ticket_id = int(ticket["id"])
        ticket["store_names"] = unique_clean_values(store_map.get(ticket_id, [])) or split_multi_value_text(ticket.get("store_name"))
        ticket["brand_names"] = unique_clean_values(brand_map.get(ticket_id, [])) or split_multi_value_text(ticket.get("brand"))
        annotate_ticket_runtime(ticket)
    return tickets


def build_query_params(filters: Dict[str, str], sort: str, page: Optional[int] = None) -> str:
    params = {
        key: value
        for key, value in filters.items()
        if not key.startswith("__") and str(value or "").strip()
    }
    if sort:
        params["sort"] = sort
    if page is not None:
        params["page"] = str(page)
    return urlencode(params)


def validate_submission(
    store_names: List[str],
    submitter: str,
    request_type: str,
    urgency: str,
    quantity: str,
    description: str,
    stores: Iterable[str],
    request_types: Iterable[str],
    urgency_levels: Iterable[str],
) -> Optional[str]:
    valid_stores = set(stores)
    if not store_names:
        return "请至少选择一个门店。"
    if any(store_name not in valid_stores for store_name in store_names):
        return "请选择有效门店。"
    if not submitter.strip():
        return "请填写提报人。"
    if request_type not in request_types:
        return "请选择有效需求类型。"
    if urgency not in urgency_levels:
        return "请选择有效紧急程度。"
    if not description.strip():
        return "请填写问题说明。"
    if quantity.strip() and not quantity.strip().isdigit():
        return "数量只能填写数字。"
    return None


def validate_request_type_rule(
    request_type: str,
    values: Dict[str, object],
    prepared_images: List[Tuple[str, bytes]],
    prepared_files: List[PreparedFile],
) -> Optional[str]:
    rule = load_request_type_rules().get(request_type)
    if not rule:
        return None
    for field in rule.get("required_fields", []):
        if not str(values.get(str(field), "")).strip():
            if str(field) == "brand":
                return f"{request_type}必须至少选择或填写一个品牌。"
            return f"{request_type}必须填写{REQUEST_RULE_FIELD_LABELS.get(str(field), str(field))}。"
    if rule.get("require_image") and not prepared_images:
        return f"{request_type}必须上传至少一张图片。"
    if rule.get("require_file") and not prepared_files:
        return f"{request_type}必须上传至少一个文件附件。"
    if rule.get("require_any_attachment") and not prepared_images and not prepared_files:
        return f"{request_type}必须上传图片或文件附件。"
    return None


def build_ticket_submission_values(
    store_names: Optional[List[str]],
    brands: Optional[List[str]],
    brand_extra: str,
    store_name: str,
    submitter: str,
    request_type: str,
    urgency: str,
    brand: str,
    product_name: str,
    sku_barcode: str,
    quantity: str,
    description: str,
    expected_finish_date: str,
) -> Tuple[List[str], List[str], str, str, Dict[str, object]]:
    normalized_stores = normalize_store_names(store_names, store_name)
    normalized_brands = normalize_brand_names(brands, brand_extra, brand)
    store_display = join_display_values(normalized_stores)
    brand_display = join_display_values(normalized_brands)
    form_values = {
        "store_name": store_name,
        "store_names": normalized_stores,
        "submitter": submitter,
        "request_type": request_type,
        "urgency": urgency,
        "brand": brand,
        "brands": normalized_brands,
        "brand_extra": brand_extra,
        "product_name": product_name,
        "sku_barcode": sku_barcode,
        "quantity": quantity,
        "description": description,
        "expected_finish_date": expected_finish_date,
    }
    return normalized_stores, normalized_brands, store_display, brand_display, form_values


async def create_ticket_from_submission(
    *,
    store_names: Optional[List[str]],
    brands: Optional[List[str]],
    brand_extra: str,
    store_name: str,
    submitter: str,
    request_type: str,
    urgency: str,
    brand: str,
    product_name: str,
    sku_barcode: str,
    quantity: str,
    description: str,
    expected_finish_date: str,
    images: Optional[List[UploadFile]],
    files: Optional[List[UploadFile]],
    config: AppConfig,
) -> Dict[str, object]:
    normalized_stores, normalized_brands, store_display, brand_display, form_values = build_ticket_submission_values(
        store_names,
        brands,
        brand_extra,
        store_name,
        submitter,
        request_type,
        urgency,
        brand,
        product_name,
        sku_barcode,
        quantity,
        description,
        expected_finish_date,
    )
    error = validate_submission(
        normalized_stores,
        submitter,
        request_type,
        urgency,
        quantity,
        description,
        config.stores,
        config.request_types,
        config.urgency_levels,
    )
    if error:
        return {"ok": False, "status_code": 400, "error": error, "values": form_values}

    try:
        prepared_images = await prepare_images(images, config)
        prepared_files = await prepare_files(files, config)
    except ValueError as exc:
        return {"ok": False, "status_code": 400, "error": str(exc), "values": form_values}

    rule_values = dict(form_values)
    rule_values["brand"] = brand_display
    rule_error = validate_request_type_rule(request_type, rule_values, prepared_images, prepared_files)
    if rule_error:
        return {"ok": False, "status_code": 400, "error": rule_error, "values": form_values}

    timestamp = now_text()
    quantity_value = int(quantity.strip()) if quantity.strip() else None
    try:
        ticket_id, ticket_no = create_ticket_with_images(
            {
                "created_at": timestamp,
                "store_name": store_display,
                "store_names": normalized_stores,
                "submitter": submitter.strip(),
                "request_type": request_type,
                "urgency": urgency,
                "brand": brand_display,
                "brands": normalized_brands,
                "product_name": product_name.strip(),
                "sku_barcode": sku_barcode.strip(),
                "quantity": quantity_value,
                "description": description.strip(),
                "expected_finish_date": expected_finish_date.strip(),
            },
            prepared_images,
            config,
            prepared_files,
        )
        try:
            created_ticket = fetch_ticket(ticket_id)
            if created_ticket:
                create_new_ticket_notification(created_ticket)
        except Exception:
            pass
    except RuntimeError as exc:
        return {"ok": False, "status_code": 500, "error": str(exc), "values": form_values}
    except OSError:
        return {"ok": False, "status_code": 500, "error": "附件保存失败，请稍后重试。", "values": form_values}

    return {
        "ok": True,
        "status_code": 200,
        "ticket_id": ticket_id,
        "ticket_no": ticket_no,
        "values": form_values,
    }


async def prepare_images(images: Optional[List[UploadFile]], config: AppConfig) -> List[Tuple[str, bytes]]:
    raw_images: List[Tuple[str, bytes]] = []
    for image in images or []:
        if not image or not image.filename:
            continue
        original_name = safe_uploaded_name(image.filename)
        content = await image.read()
        if not content:
            continue
        raw_images.append((original_name, content))

    if len(raw_images) > config.max_image_count:
        raise ValueError(f"最多上传 {config.max_image_count} 张图片。")

    total_bytes = sum(len(content) for _, content in raw_images)
    if total_bytes > config.max_total_upload_bytes:
        raise ValueError(f"图片总大小不能超过 {config.max_total_upload_mb}MB。")

    prepared_images: List[Tuple[str, bytes]] = []
    allowed_extensions = set(config.allowed_image_extensions)
    for original_name, content in raw_images:
        extension = Path(original_name).suffix.lower().lstrip(".")
        if extension not in allowed_extensions:
            allowed_text = "、".join(config.allowed_image_extensions)
            raise ValueError(f"图片仅支持 {allowed_text} 格式。")
        if len(content) > config.max_image_bytes:
            raise ValueError(f"单张图片不能超过 {config.max_image_mb}MB。")

        try:
            with Image.open(BytesIO(content)) as image_file:
                image_format = str(image_file.format or "").upper()
                image_file.verify()
        except (UnidentifiedImageError, OSError, ValueError):
            raise ValueError("图片文件无法识别，请重新上传。") from None

        valid_extensions = IMAGE_FORMAT_EXTENSIONS.get(image_format, {image_format.lower()})
        if extension not in valid_extensions:
            raise ValueError("图片后缀与实际格式不匹配，请重新上传。")

        prepared_images.append((extension, content))
    return prepared_images


async def prepare_files(files: Optional[List[UploadFile]], config: AppConfig) -> List[PreparedFile]:
    raw_files: List[PreparedFile] = []
    for upload in files or []:
        if not upload or not upload.filename:
            continue
        original_name = safe_uploaded_name(upload.filename)
        content = await upload.read()
        if not original_name or not content:
            continue
        extension = Path(original_name).suffix.lower().lstrip(".")
        raw_files.append(
            PreparedFile(
                original_filename=original_name,
                file_ext=extension,
                file_size=len(content),
                content=content,
            )
        )

    if len(raw_files) > config.max_file_count:
        raise ValueError(f"最多上传 {config.max_file_count} 个文件。")

    total_bytes = sum(file.file_size for file in raw_files)
    if total_bytes > config.max_total_file_upload_bytes:
        raise ValueError(f"文件总大小不能超过 {config.max_total_file_upload_mb}MB。")

    allowed_extensions = set(config.allowed_file_extensions)
    for file in raw_files:
        if not file.file_ext or file.file_ext in BLOCKED_FILE_EXTENSIONS or file.file_ext not in allowed_extensions:
            allowed_text = "、".join(config.allowed_file_extensions)
            raise ValueError(f"文件仅支持 {allowed_text} 格式。")
        if file.file_size > config.max_file_bytes:
            raise ValueError(f"单个文件不能超过 {config.max_file_mb}MB。")
    return raw_files


def save_images(ticket_no: str, prepared_images: List[Tuple[str, bytes]]) -> Tuple[List[str], List[Path]]:
    image_paths: List[str] = []
    saved_files: List[Path] = []
    upload_dir = get_upload_dir()
    upload_dir.mkdir(parents=True, exist_ok=True)
    try:
        for index, (extension, content) in enumerate(prepared_images, start=1):
            filename = f"{ticket_no}_{index}_{uuid.uuid4().hex[:8]}.{extension}"
            target_path = upload_dir / filename
            target_path.write_bytes(content)
            saved_files.append(target_path)
            image_paths.append(f"uploads/{filename}")
        return image_paths, saved_files
    except Exception:
        cleanup_saved_files(saved_files)
        raise


def save_files(ticket_no: str, prepared_files: List[PreparedFile]) -> Tuple[List[Dict[str, object]], List[Path]]:
    file_records: List[Dict[str, object]] = []
    saved_files: List[Path] = []
    upload_dir = get_upload_dir()
    upload_dir.mkdir(parents=True, exist_ok=True)
    try:
        for prepared_file in prepared_files:
            stored_filename = f"FILE-{ticket_no}-{uuid.uuid4().hex}.{prepared_file.file_ext}"
            target_path = upload_dir / stored_filename
            target_path.write_bytes(prepared_file.content)
            saved_files.append(target_path)
            file_records.append(
                {
                    "original_filename": prepared_file.original_filename,
                    "stored_filename": stored_filename,
                    "file_path": f"uploads/{stored_filename}",
                    "file_ext": prepared_file.file_ext,
                    "file_size": prepared_file.file_size,
                }
            )
        return file_records, saved_files
    except Exception:
        cleanup_saved_files(saved_files)
        raise


def cleanup_saved_files(paths: List[Path]) -> None:
    for path in paths:
        try:
            path.unlink(missing_ok=True)
        except OSError:
            pass


def is_ticket_no_conflict(error: sqlite3.IntegrityError) -> bool:
    return "ticket_no" in str(error).lower()


def create_ticket_with_images(
    ticket_data: Dict[str, object],
    prepared_images: List[Tuple[str, bytes]],
    config: AppConfig,
    prepared_files: Optional[List[PreparedFile]] = None,
) -> Tuple[int, str]:
    prepared_files = prepared_files or []
    for _attempt in range(3):
        saved_files: List[Path] = []
        timestamp = now_text()
        try:
            with get_connection() as connection:
                ticket_no = generate_ticket_no(connection)
                try:
                    cursor = connection.execute(
                        """
                        INSERT INTO tickets (
                            ticket_no, created_at, updated_at, store_name, submitter,
                            request_type, urgency, brand, product_name, sku_barcode,
                            quantity, description, expected_finish_date, status,
                            assigned_to, handler_note, closed_at
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            ticket_no,
                            timestamp,
                            timestamp,
                            ticket_data["store_name"],
                            ticket_data["submitter"],
                            ticket_data["request_type"],
                            ticket_data["urgency"],
                            ticket_data["brand"],
                            ticket_data["product_name"],
                            ticket_data["sku_barcode"],
                            ticket_data["quantity"],
                            ticket_data["description"],
                            ticket_data["expected_finish_date"],
                            config.default_status,
                            "",
                            "",
                            None,
                        ),
                    )
                except sqlite3.IntegrityError as exc:
                    if is_ticket_no_conflict(exc):
                        continue
                    raise

                ticket_id = int(cursor.lastrowid)
                for store_name in ticket_data.get("store_names", []):
                    connection.execute(
                        """
                        INSERT OR IGNORE INTO ticket_stores (ticket_id, store_name, created_at)
                        VALUES (?, ?, ?)
                        """,
                        (ticket_id, store_name, timestamp),
                    )
                for brand in ticket_data.get("brands", []):
                    connection.execute(
                        """
                        INSERT OR IGNORE INTO ticket_brands (ticket_id, brand, created_at)
                        VALUES (?, ?, ?)
                        """,
                        (ticket_id, brand, timestamp),
                    )
                image_paths, saved_image_files = save_images(ticket_no, prepared_images)
                saved_files.extend(saved_image_files)
                for image_path in image_paths:
                    connection.execute(
                        """
                        INSERT INTO ticket_images (ticket_id, image_path, uploaded_at, source, uploaded_by, supplement_id)
                        VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        (ticket_id, image_path, timestamp, "store_initial", ticket_data["submitter"], None),
                    )
                file_records, saved_attachment_files = save_files(ticket_no, prepared_files)
                saved_files.extend(saved_attachment_files)
                for file_record in file_records:
                    connection.execute(
                        """
                        INSERT INTO ticket_files (
                            ticket_id, original_filename, stored_filename, file_path,
                            file_ext, file_size, uploaded_at, source, uploaded_by, supplement_id
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            ticket_id,
                            file_record["original_filename"],
                            file_record["stored_filename"],
                            file_record["file_path"],
                            file_record["file_ext"],
                            file_record["file_size"],
                            timestamp,
                            "store_initial",
                            ticket_data["submitter"],
                            None,
                        ),
                    )
            return ticket_id, ticket_no
        except sqlite3.IntegrityError as exc:
            cleanup_saved_files(saved_files)
            if is_ticket_no_conflict(exc):
                continue
            raise
        except Exception:
            cleanup_saved_files(saved_files)
            raise
    raise RuntimeError("工单号生成冲突，请稍后重试。")


def build_ticket_where(filters: Dict[str, str]) -> Tuple[str, List[str]]:
    include_deleted = str(filters.get("__include_deleted") or "").strip() == "1"
    deleted_only = str(filters.get("__deleted_only") or "").strip() == "1"
    ticket_scope = str(filters.get("__ticket_scope") or "").strip()
    clauses: List[str] = []
    params: List[str] = []
    if deleted_only or ticket_scope == "deleted":
        clauses.append("tickets.deleted_at IS NOT NULL")
    elif ticket_scope == "archive":
        clauses.append("tickets.deleted_at IS NULL")
        clauses.append("tickets.archived_at IS NOT NULL")
    elif ticket_scope == "store":
        clauses.append("tickets.deleted_at IS NULL")
    elif not include_deleted:
        clauses.append("tickets.deleted_at IS NULL")
        clauses.append("tickets.archived_at IS NULL")

    exact_fields = {
        "ticket_no": "ticket_no",
        "submitter": "submitter",
        "request_type": "request_type",
        "urgency": "urgency",
        "status": "status",
        "assigned_to": "assigned_to",
    }
    for filter_key, column_name in exact_fields.items():
        value = filters.get(filter_key, "").strip()
        if value:
            clauses.append(f"{column_name} = ?")
            params.append(value)

    store_name = filters.get("store_name", "").strip()
    if store_name:
        clauses.append(
            """
            EXISTS (
                SELECT 1
                FROM ticket_stores
                WHERE ticket_stores.ticket_id = tickets.id
                  AND ticket_stores.store_name = ?
            )
            """
        )
        params.append(store_name)

    date_start = filters.get("date_start", "").strip()
    if date_start:
        clauses.append("DATE(created_at) >= DATE(?)")
        params.append(date_start)

    date_end = filters.get("date_end", "").strip()
    if date_end:
        clauses.append("DATE(created_at) <= DATE(?)")
        params.append(date_end)

    keyword = filters.get("keyword", "").strip()
    if keyword:
        like_value = f"%{keyword}%"
        clauses.append(
            """
            (
                ticket_no LIKE ?
                OR store_name LIKE ?
                OR submitter LIKE ?
                OR brand LIKE ?
                OR product_name LIKE ?
                OR sku_barcode LIKE ?
                OR description LIKE ?
                OR assigned_to LIKE ?
                OR handler_note LIKE ?
            )
            """
        )
        params.extend([like_value] * 9)

    where_sql = " WHERE " + " AND ".join(clauses) if clauses else ""
    return where_sql, params


def build_case_order(values: List[str]) -> str:
    cases = []
    for index, value in enumerate(values):
        escaped_value = value.replace("'", "''")
        cases.append(f"WHEN '{escaped_value}' THEN {index}")
    return " ".join(cases)


def build_order_sql(sort: str, config: AppConfig) -> str:
    if sort == "urgency":
        urgency_cases = build_case_order(config.urgency_levels)
        return """
        ORDER BY
            CASE urgency
                {urgency_cases}
                ELSE {fallback_index}
            END,
            created_at DESC,
            tickets.id DESC
        """.format(urgency_cases=urgency_cases, fallback_index=len(config.urgency_levels))
    if sort == "status":
        status_cases = build_case_order(config.statuses)
        return """
        ORDER BY
            CASE status
                {status_cases}
                ELSE {fallback_index}
            END,
            created_at DESC,
            tickets.id DESC
        """.format(status_cases=status_cases, fallback_index=len(config.statuses))
    return "ORDER BY created_at DESC, tickets.id DESC"


def count_tickets(filters: Dict[str, str]) -> int:
    if filters.get("due_status", "").strip():
        return len(fetch_tickets(filters, "newest", load_app_config()))
    where_sql, params = build_ticket_where(filters)
    with get_connection() as connection:
        row = connection.execute(f"SELECT COUNT(*) AS total FROM tickets {where_sql}", params).fetchone()
    return int(row["total"] or 0)


def fetch_ticket_summary(filters: Dict[str, str]) -> Dict[str, int]:
    if filters.get("due_status", "").strip():
        tickets = fetch_tickets(filters, "newest", load_app_config())
        return {
            "total_count": len(tickets),
            "pending_count": sum(1 for ticket in tickets if ticket.get("status") == "待处理"),
            "processing_count": sum(1 for ticket in tickets if ticket.get("status") == "处理中"),
            "today_urgent_count": sum(1 for ticket in tickets if ticket.get("urgency") == "当天必须处理"),
            "completed_count": sum(1 for ticket in tickets if ticket.get("status") == COMPLETED_STATUS),
        }
    where_sql, params = build_ticket_where(filters)
    with get_connection() as connection:
        row = connection.execute(
            f"""
            SELECT
                COUNT(*) AS total_count,
                SUM(CASE WHEN status = '待处理' THEN 1 ELSE 0 END) AS pending_count,
                SUM(CASE WHEN status = '处理中' THEN 1 ELSE 0 END) AS processing_count,
                SUM(CASE WHEN urgency = '当天必须处理' THEN 1 ELSE 0 END) AS today_urgent_count,
                SUM(CASE WHEN status = '已完成' THEN 1 ELSE 0 END) AS completed_count
            FROM tickets
            {where_sql}
            """,
            params,
        ).fetchone()
    return {
        "total_count": int(row["total_count"] or 0),
        "pending_count": int(row["pending_count"] or 0),
        "processing_count": int(row["processing_count"] or 0),
        "today_urgent_count": int(row["today_urgent_count"] or 0),
        "completed_count": int(row["completed_count"] or 0),
    }


def fetch_tickets(
    filters: Dict[str, str],
    sort: str,
    config: AppConfig,
    limit: Optional[int] = None,
    offset: int = 0,
) -> List[Dict[str, object]]:
    where_sql, params = build_ticket_where(filters)
    order_sql = build_order_sql(sort, config)
    limit_sql = ""
    query_params: List[object] = list(params)
    if limit is not None:
        limit_sql = "LIMIT ? OFFSET ?"
        query_params.extend([limit, offset])
    with get_connection() as connection:
        rows = connection.execute(
            f"""
            SELECT
                tickets.id, ticket_no, created_at, updated_at, store_name, submitter,
                request_type, urgency, brand, product_name, sku_barcode,
                quantity, description, expected_finish_date, status, assigned_to,
                handler_note, closed_at, deleted_at, deleted_by, delete_reason,
                archived_at, archived_by, archive_reason,
                COALESCE(image_counts.image_count, 0) AS image_count,
                COALESCE(file_counts.file_count, 0) AS file_count
            FROM tickets
            LEFT JOIN (
                SELECT ticket_id, COUNT(*) AS image_count
                FROM ticket_images
                GROUP BY ticket_id
            ) AS image_counts ON image_counts.ticket_id = tickets.id
            LEFT JOIN (
                SELECT ticket_id, COUNT(*) AS file_count
                FROM ticket_files
                GROUP BY ticket_id
            ) AS file_counts ON file_counts.ticket_id = tickets.id
            {where_sql}
            {order_sql}
            {limit_sql}
            """,
            query_params,
        ).fetchall()
    tickets = attach_ticket_relations([dict(row) for row in rows])
    due_filter = filters.get("due_status", "").strip()
    if due_filter:
        tickets = [ticket for ticket in tickets if ticket.get("due_status") == due_filter]
    return tickets


def fetch_ticket_page(filters: Dict[str, str], sort: str, config: AppConfig, page: int) -> Tuple[List[Dict[str, object]], Dict[str, int]]:
    if filters.get("due_status", "").strip():
        all_tickets = fetch_tickets(filters, sort, config)
        total_count = len(all_tickets)
        total_pages = max((total_count + config.page_size - 1) // config.page_size, 1)
        current_page = min(max(page, 1), total_pages)
        offset = (current_page - 1) * config.page_size
        return all_tickets[offset : offset + config.page_size], {
            "current_page": current_page,
            "total_pages": total_pages,
            "total_count": total_count,
            "page_size": config.page_size,
            "has_prev": 1 if current_page > 1 else 0,
            "has_next": 1 if current_page < total_pages else 0,
            "prev_page": max(current_page - 1, 1),
            "next_page": min(current_page + 1, total_pages),
        }
    total_count = count_tickets(filters)
    total_pages = max((total_count + config.page_size - 1) // config.page_size, 1)
    current_page = min(max(page, 1), total_pages)
    offset = (current_page - 1) * config.page_size
    tickets = fetch_tickets(filters, sort, config, limit=config.page_size, offset=offset)
    return tickets, {
        "current_page": current_page,
        "total_pages": total_pages,
        "total_count": total_count,
        "page_size": config.page_size,
        "has_prev": 1 if current_page > 1 else 0,
        "has_next": 1 if current_page < total_pages else 0,
        "prev_page": max(current_page - 1, 1),
        "next_page": min(current_page + 1, total_pages),
    }


def fetch_ticket(ticket_id: int, include_deleted: bool = False) -> Optional[Dict[str, object]]:
    deleted_sql = "" if include_deleted else " AND deleted_at IS NULL"
    with get_connection() as connection:
        row = connection.execute(f"SELECT * FROM tickets WHERE id = ?{deleted_sql}", (ticket_id,)).fetchone()
    if not row:
        return None
    tickets = attach_ticket_relations([dict(row)])
    return tickets[0] if tickets else None


def fetch_store_ticket(ticket_id: int, store_name: str) -> Optional[Dict[str, object]]:
    clean_store_name = store_name.strip()
    if not clean_store_name:
        return None
    with get_connection() as connection:
        row = connection.execute(
            """
            SELECT tickets.*
            FROM tickets
            WHERE tickets.id = ?
              AND tickets.deleted_at IS NULL
              AND EXISTS (
                  SELECT 1
                  FROM ticket_stores
                  WHERE ticket_stores.ticket_id = tickets.id
                    AND ticket_stores.store_name = ?
              )
            """,
            (ticket_id, clean_store_name),
        ).fetchone()
    if not row:
        return None
    tickets = attach_ticket_relations([dict(row)])
    ticket = tickets[0]
    ticket["query_store_name"] = clean_store_name
    return ticket


def fetch_store_ticket_supplements(ticket_id: int) -> List[Dict[str, object]]:
    with get_connection() as connection:
        rows = connection.execute(
            """
            SELECT id, ticket_id, store_name, submitter, note, image_count, file_count, created_at
            FROM ticket_supplements
            WHERE ticket_id = ?
              AND deleted_at IS NULL
            ORDER BY created_at DESC, id DESC
            """,
            (ticket_id,),
        ).fetchall()
    return [dict(row) for row in rows]


def fetch_store_visible_logs(ticket_id: int) -> List[Dict[str, object]]:
    with get_connection() as connection:
        rows = connection.execute(
            """
            SELECT id, ticket_id, action, old_status, new_status, note, created_at
            FROM ticket_logs
            WHERE ticket_id = ?
              AND action IN ('update', '门店补充资料')
            ORDER BY created_at DESC, id DESC
            """,
            (ticket_id,),
        ).fetchall()
        supplement_count = int(
            connection.execute(
                "SELECT COUNT(*) AS total FROM ticket_supplements WHERE ticket_id = ? AND deleted_at IS NULL",
                (ticket_id,),
            ).fetchone()["total"]
            or 0
        )
    logs = [dict(row) for row in rows]
    if supplement_count == 0:
        for log in logs:
            if str(log.get("action") or "") == "门店补充资料":
                log["note"] = "相关门店补充记录已移入回收站"
    return logs


def fetch_ticket_attachment_counts(ticket_id: int) -> Dict[str, object]:
    with get_connection() as connection:
        image_row = connection.execute(
            "SELECT COUNT(*) AS total FROM ticket_images WHERE ticket_id = ?",
            (ticket_id,),
        ).fetchone()
        file_row = connection.execute(
            "SELECT COUNT(*) AS total FROM ticket_files WHERE ticket_id = ?",
            (ticket_id,),
        ).fetchone()
        supplement_row = connection.execute(
            """
            SELECT COUNT(*) AS total, MAX(created_at) AS latest_created_at
            FROM ticket_supplements
            WHERE ticket_id = ?
              AND deleted_at IS NULL
            """,
            (ticket_id,),
        ).fetchone()
    return {
        "image_count": int(image_row["total"] or 0),
        "file_count": int(file_row["total"] or 0),
        "supplement_count": int(supplement_row["total"] or 0),
        "latest_supplement_at": supplement_row["latest_created_at"] or "",
    }


def fetch_ticket_images(ticket_id: int) -> List[Dict[str, object]]:
    with get_connection() as connection:
        rows = connection.execute(
            """
            SELECT id, ticket_id, image_path, uploaded_at, source, uploaded_by, supplement_id
            FROM ticket_images
            WHERE ticket_id = ?
            ORDER BY id
            """,
            (ticket_id,),
        ).fetchall()
    images = [dict(row) for row in rows]
    for image in images:
        image["filename"] = image_filename(str(image.get("image_path") or ""))
        image["protected_url"] = protected_upload_url(str(image.get("image_path") or ""))
    return images


def fetch_ticket_files(ticket_id: int) -> List[Dict[str, object]]:
    with get_connection() as connection:
        rows = connection.execute(
            """
            SELECT id, ticket_id, original_filename, stored_filename, file_path,
                   file_ext, file_size, uploaded_at, source, uploaded_by, supplement_id
            FROM ticket_files
            WHERE ticket_id = ?
            ORDER BY id
            """,
            (ticket_id,),
        ).fetchall()
    files = [dict(row) for row in rows]
    for file in files:
        file["download_url"] = protected_file_url(file["id"])
        file["size_label"] = file_size_label(file.get("file_size"))
    return files


def fetch_ticket_file(file_id: int, ticket_id: Optional[int] = None) -> Optional[Dict[str, object]]:
    sql = """
        SELECT ticket_files.id, ticket_files.ticket_id, original_filename, stored_filename, file_path,
               file_ext, file_size, uploaded_at, source, uploaded_by, supplement_id
        FROM ticket_files
        JOIN tickets ON tickets.id = ticket_files.ticket_id
        WHERE ticket_files.id = ? AND tickets.deleted_at IS NULL
    """
    params: List[object] = [file_id]
    if ticket_id is not None:
        sql += " AND ticket_files.ticket_id = ?"
        params.append(ticket_id)
    with get_connection() as connection:
        row = connection.execute(sql, params).fetchone()
    return dict(row) if row else None


def fetch_ticket_image(image_id: int, ticket_id: int) -> Optional[Dict[str, object]]:
    with get_connection() as connection:
        row = connection.execute(
            """
            SELECT id, ticket_id, image_path, uploaded_at, source, uploaded_by, supplement_id
            FROM ticket_images
            WHERE id = ? AND ticket_id = ?
              AND EXISTS (
                  SELECT 1
                  FROM tickets
                  WHERE tickets.id = ticket_images.ticket_id
                    AND tickets.deleted_at IS NULL
              )
            """,
            (image_id, ticket_id),
        ).fetchone()
    return dict(row) if row else None


def fetch_ticket_logs(ticket_id: int) -> List[Dict[str, object]]:
    with get_connection() as connection:
        rows = connection.execute(
            """
            SELECT id, ticket_id, action, old_status, new_status, old_assigned_to,
                   new_assigned_to, note, operator, created_at
            FROM ticket_logs
            WHERE ticket_id = ?
            ORDER BY created_at DESC, id DESC
            """,
            (ticket_id,),
        ).fetchall()
        active_counts = {
            "comments": int(
                connection.execute(
                    "SELECT COUNT(*) AS total FROM ticket_comments WHERE ticket_id = ? AND deleted_at IS NULL",
                    (ticket_id,),
                ).fetchone()["total"]
                or 0
            ),
            "tasks": int(
                connection.execute(
                    "SELECT COUNT(*) AS total FROM ticket_tasks WHERE ticket_id = ? AND deleted_at IS NULL",
                    (ticket_id,),
                ).fetchone()["total"]
                or 0
            ),
            "participants": int(
                connection.execute(
                    "SELECT COUNT(*) AS total FROM ticket_participants WHERE ticket_id = ? AND deleted_at IS NULL",
                    (ticket_id,),
                ).fetchone()["total"]
                or 0
            ),
            "supplements": int(
                connection.execute(
                    "SELECT COUNT(*) AS total FROM ticket_supplements WHERE ticket_id = ? AND deleted_at IS NULL",
                    (ticket_id,),
                ).fetchone()["total"]
                or 0
            ),
        }
    logs = [dict(row) for row in rows]
    for log in logs:
        action = str(log.get("action") or "")
        if action in {"新增评论", "门店评论"} and active_counts["comments"] == 0:
            log["note"] = "相关沟通记录已移入回收站"
        elif action in {"新增子任务", "更新子任务"} and active_counts["tasks"] == 0:
            log["note"] = "相关子任务已移入回收站"
        elif action == "新增协作人" and active_counts["participants"] == 0:
            log["note"] = "相关协作人已移入回收站"
        elif action == "门店补充资料" and active_counts["supplements"] == 0:
            log["note"] = "相关门店补充记录已移入回收站"
    return logs


def fetch_ticket_participants(ticket_id: int) -> List[Dict[str, object]]:
    with get_connection() as connection:
        rows = connection.execute(
            """
            SELECT id, ticket_id, participant_type, participant_name, role, created_at
            FROM ticket_participants
            WHERE ticket_id = ?
              AND deleted_at IS NULL
            ORDER BY id ASC
            """,
            (ticket_id,),
        ).fetchall()
    return [dict(row) for row in rows]


def fetch_ticket_comments(ticket_id: int, public_only: bool = False) -> List[Dict[str, object]]:
    clauses = ["ticket_id = ?", "deleted_at IS NULL"]
    params: List[object] = [ticket_id]
    if public_only:
        clauses.append("visibility = ?")
        params.append("public")
    where_sql = " AND ".join(clauses)
    with get_connection() as connection:
        rows = connection.execute(
            f"""
            SELECT id, ticket_id, author_type, author_name, content, visibility, created_at
            FROM ticket_comments
            WHERE {where_sql}
            ORDER BY created_at ASC, id ASC
            """,
            params,
        ).fetchall()
    return [dict(row) for row in rows]


def fetch_ticket_tasks(ticket_id: int) -> List[Dict[str, object]]:
    with get_connection() as connection:
        rows = connection.execute(
            """
            SELECT id, ticket_id, title, assignee, status, due_date,
                   completed_at, created_at, updated_at
            FROM ticket_tasks
            WHERE ticket_id = ?
              AND deleted_at IS NULL
            ORDER BY
                CASE status
                    WHEN '待处理' THEN 0
                    WHEN '处理中' THEN 1
                    WHEN '已完成' THEN 2
                    ELSE 3
                END,
                due_date IS NULL,
                due_date,
                id DESC
            """,
            (ticket_id,),
        ).fetchall()
    return [dict(row) for row in rows]


def detail_handler_options(config: AppConfig, ticket: Dict[str, object], admin_user: str) -> List[str]:
    options: List[str] = []
    for value in (str(ticket.get("assigned_to") or ""), admin_user, *config.handlers):
        clean_value = str(value or "").strip()
        if clean_value and clean_value not in options:
            options.append(clean_value)
    return options


def should_prompt_close_ticket(ticket: Dict[str, object], tasks: List[Dict[str, object]]) -> bool:
    if str(ticket.get("status") or "") == COMPLETED_STATUS:
        return False
    return bool(tasks) and all(str(task.get("status") or "") == COMPLETED_STATUS for task in tasks)


def fetch_my_work(username: str, config: AppConfig) -> Dict[str, object]:
    clean_username = username.strip()
    assigned_tickets = fetch_tickets(
        {"assigned_to": clean_username},
        "newest",
        config,
        limit=50,
        offset=0,
    )
    for ticket in assigned_tickets:
        ticket["detail_url"] = build_ticket_detail_url(int(ticket["id"]), "/admin/my-work")

    with get_connection() as connection:
        task_rows = connection.execute(
            """
            SELECT
                ticket_tasks.id AS task_id,
                ticket_tasks.ticket_id,
                ticket_tasks.title,
                ticket_tasks.assignee,
                ticket_tasks.status AS task_status,
                ticket_tasks.due_date,
                ticket_tasks.completed_at,
                ticket_tasks.created_at AS task_created_at,
                ticket_tasks.updated_at AS task_updated_at,
                tickets.ticket_no,
                tickets.store_name,
                tickets.submitter,
                tickets.request_type,
                tickets.urgency,
                tickets.status AS ticket_status,
                tickets.assigned_to,
                tickets.description
            FROM ticket_tasks
            JOIN tickets ON tickets.id = ticket_tasks.ticket_id
            WHERE ticket_tasks.deleted_at IS NULL
              AND ticket_tasks.status != ?
              AND TRIM(COALESCE(ticket_tasks.assignee, '')) = ?
              AND tickets.deleted_at IS NULL
              AND tickets.archived_at IS NULL
            ORDER BY
                ticket_tasks.due_date IS NULL,
                ticket_tasks.due_date,
                ticket_tasks.updated_at DESC,
                ticket_tasks.id DESC
            LIMIT 100
            """,
            (COMPLETED_STATUS, clean_username),
        ).fetchall()
        reply_rows = connection.execute(
            """
            SELECT *
            FROM (
                SELECT
                    '门店沟通' AS source_type,
                    ticket_comments.id AS source_id,
                    ticket_comments.created_at,
                    ticket_comments.author_name AS author_name,
                    ticket_comments.content AS content,
                    tickets.id AS ticket_id,
                    tickets.ticket_no,
                    tickets.store_name,
                    tickets.status,
                    tickets.request_type
                FROM ticket_comments
                JOIN tickets ON tickets.id = ticket_comments.ticket_id
                WHERE ticket_comments.deleted_at IS NULL
                  AND ticket_comments.author_type = 'store'
                  AND tickets.assigned_to = ?
                  AND tickets.deleted_at IS NULL
                  AND tickets.archived_at IS NULL
                UNION ALL
                SELECT
                    '补充资料' AS source_type,
                    ticket_supplements.id AS source_id,
                    ticket_supplements.created_at,
                    ticket_supplements.submitter AS author_name,
                    ticket_supplements.note AS content,
                    tickets.id AS ticket_id,
                    tickets.ticket_no,
                    tickets.store_name,
                    tickets.status,
                    tickets.request_type
                FROM ticket_supplements
                JOIN tickets ON tickets.id = ticket_supplements.ticket_id
                WHERE ticket_supplements.deleted_at IS NULL
                  AND tickets.assigned_to = ?
                  AND tickets.deleted_at IS NULL
                  AND tickets.archived_at IS NULL
            )
            ORDER BY created_at DESC, source_id DESC
            LIMIT 20
            """,
            (clean_username, clean_username),
        ).fetchall()

    tasks = [dict(row) for row in task_rows]
    for task in tasks:
        task["detail_url"] = build_ticket_detail_url(int(task["ticket_id"]), "/admin/my-work")
        task["description_summary"] = compact_text(str(task.get("description") or ""))
    reply_items = [dict(row) for row in reply_rows]
    for item in reply_items:
        item["detail_url"] = build_ticket_detail_url(int(item["ticket_id"]), "/admin/my-work") + "#comments"
        item["content_summary"] = compact_text(str(item.get("content") or ""))
    return {
        "assigned_tickets": assigned_tickets,
        "tasks": tasks,
        "reply_items": reply_items,
        "assigned_count": len(assigned_tickets),
        "task_count": len(tasks),
        "reply_count": len(reply_items),
        "total_count": len(assigned_tickets) + len(tasks) + len(reply_items),
    }


def parse_duration_hours(value: str) -> float:
    clean_value = str(value or "").strip()
    if not clean_value:
        return 0.0
    try:
        return round(float(clean_value), 2)
    except ValueError as exc:
        raise ValueError("工时必须是数字。") from exc


def normalize_month(value: str) -> str:
    clean_value = str(value or "").strip()
    if not clean_value:
        return datetime.now().strftime("%Y-%m")
    try:
        return datetime.strptime(clean_value, "%Y-%m").strftime("%Y-%m")
    except ValueError as exc:
        raise ValueError("月份格式不正确。") from exc


def normalize_schedule_date(value: str) -> str:
    clean_value = str(value or "").strip()
    try:
        return datetime.strptime(clean_value, "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError as exc:
        raise ValueError("排班日期格式不正确。") from exc


def weekday_label(date_text: str) -> str:
    labels = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
    parsed = datetime.strptime(date_text, "%Y-%m-%d")
    return labels[parsed.weekday()]


def month_date_items(month: str) -> List[Dict[str, object]]:
    normalized = normalize_month(month)
    year, month_number = [int(part) for part in normalized.split("-")]
    last_day = calendar.monthrange(year, month_number)[1]
    today = datetime.now().strftime("%Y-%m-%d")
    days: List[Dict[str, object]] = []
    for day in range(1, last_day + 1):
        date_text = f"{normalized}-{day:02d}"
        weekday = weekday_label(date_text)
        days.append(
            {
                "date": date_text,
                "day": day,
                "weekday": weekday,
                "is_weekend": weekday in {"周六", "周日"},
                "is_today": date_text == today,
            }
        )
    return days


def calendar_week_rows(days: List[Dict[str, object]]) -> List[List[Optional[Dict[str, object]]]]:
    if not days:
        return []
    first_date = datetime.strptime(str(days[0]["date"]), "%Y-%m-%d")
    cells: List[Optional[Dict[str, object]]] = [None] * first_date.weekday()
    cells.extend(days)
    while len(cells) % 7:
        cells.append(None)
    return [cells[index : index + 7] for index in range(0, len(cells), 7)]


def annotate_schedule_employee(employee: Dict[str, object], selected_store: str) -> Dict[str, object]:
    store_names = list(employee.get("store_names") or [])
    primary_store = str(employee.get("primary_store_name") or employee.get("store_name") or (store_names[0] if store_names else "")).strip()
    annotated = dict(employee)
    annotated["primary_store"] = primary_store
    annotated["is_current_store_employee"] = bool(not selected_store or selected_store in store_names)
    annotated["is_support_store_employee"] = bool(selected_store and selected_store in store_names and primary_store != selected_store)
    annotated["is_primary_store_employee"] = bool(selected_store and primary_store == selected_store)
    annotated["is_cross_store_visible"] = annotated["is_support_store_employee"]
    return annotated


def annotate_schedule_employee_for_stores(employee: Dict[str, object], selected_stores: List[str]) -> Dict[str, object]:
    if len(selected_stores) == 1:
        return annotate_schedule_employee(employee, selected_stores[0])
    store_names = list(employee.get("store_names") or [])
    selected_store_set = set(selected_stores)
    primary_store = str(employee.get("primary_store_name") or employee.get("store_name") or (store_names[0] if store_names else "")).strip()
    matching_stores = [store_name for store_name in store_names if not selected_store_set or store_name in selected_store_set]
    annotated = dict(employee)
    annotated["primary_store"] = primary_store
    annotated["is_current_store_employee"] = bool(matching_stores or not selected_store_set)
    annotated["is_primary_store_employee"] = bool(not selected_store_set or primary_store in selected_store_set)
    annotated["is_support_store_employee"] = bool(selected_store_set and matching_stores and primary_store not in selected_store_set)
    annotated["is_cross_store_visible"] = annotated["is_support_store_employee"]
    return annotated


def build_schedule_employee_summaries(
    employees: List[Dict[str, object]],
    rows: List[Dict[str, object]],
    days: List[Dict[str, object]],
) -> List[Dict[str, object]]:
    rows_by_employee: Dict[int, List[Dict[str, object]]] = {}
    for row in rows:
        rows_by_employee.setdefault(int(row["employee_id"]), []).append(row)
    summaries: List[Dict[str, object]] = []
    for employee in employees:
        employee_id = int(employee["id"])
        employee_rows = rows_by_employee.get(employee_id, [])
        scheduled_dates = {str(row.get("schedule_date") or "") for row in employee_rows}
        timeline = []
        for day in days:
            matching = next((row for row in employee_rows if str(row.get("schedule_date") or "") == str(day["date"])), None)
            timeline.append({"day": day, "schedule": matching})
        summaries.append(
            {
                "employee": employee,
                "rows": employee_rows,
                "timeline": timeline,
                "total_hours": round(sum(float(row.get("duration_hours") or 0) for row in employee_rows), 2),
                "schedule_day_count": len(scheduled_dates),
                "rest_day_count": sum(1 for row in employee_rows if str(row.get("shift_name") or "") == "休息"),
                "unscheduled_day_count": max(len(days) - len(scheduled_dates), 0),
            }
        )
    return summaries


def build_schedule_calendar_summary(days: List[Dict[str, object]], rows: List[Dict[str, object]]) -> List[Dict[str, object]]:
    rows_by_date: Dict[str, List[Dict[str, object]]] = {}
    for row in rows:
        rows_by_date.setdefault(str(row["schedule_date"]), []).append(row)
    summary: List[Dict[str, object]] = []
    for day in days:
        date_text = str(day["date"])
        day_rows = rows_by_date.get(date_text, [])
        shift_counts: Dict[str, int] = {}
        store_counts: Dict[str, int] = {}
        for row in day_rows:
            shift_name = str(row.get("shift_name") or "未设置班次")
            store_name = str(row.get("store_name") or "未设置门店")
            shift_counts[shift_name] = shift_counts.get(shift_name, 0) + 1
            store_counts[store_name] = store_counts.get(store_name, 0) + 1
        summary.append(
            {
                "day": day,
                "rows": day_rows,
                "schedule_count": len(day_rows),
                "shift_counts": shift_counts,
                "store_counts": store_counts,
            }
        )
    return summary


def build_schedule_dashboard(rows: List[Dict[str, object]], month: str) -> Dict[str, object]:
    normalized_month = normalize_month(month)
    year, month_number = [int(part) for part in normalized_month.split("-")]
    last_day = calendar.monthrange(year, month_number)[1]
    today = datetime.now().date()
    month_start = date(year, month_number, 1)
    month_end = date(year, month_number, last_day)
    if month_end < today:
        cutoff_date = month_end
    elif month_start > today:
        cutoff_date = None
    else:
        cutoff_date = today

    total_hours = round(sum(float(row.get("duration_hours") or 0) for row in rows), 2)
    cutoff_hours = round(
        sum(
            float(row.get("duration_hours") or 0)
            for row in rows
            if cutoff_date is not None and str(row.get("schedule_date") or "") <= cutoff_date.strftime("%Y-%m-%d")
        ),
        2,
    )
    store_hours: Dict[str, float] = {}
    shift_distribution: Dict[str, int] = {}
    for row in rows:
        store_name = str(row.get("store_name") or "未设置门店")
        shift_name = str(row.get("shift_name") or "未设置班次")
        if int(row.get("is_custom_time") or 0) == 1:
            shift_name = "自定义"
        store_hours[store_name] = round(store_hours.get(store_name, 0.0) + float(row.get("duration_hours") or 0), 2)
        shift_distribution[shift_name] = shift_distribution.get(shift_name, 0) + 1
    max_store_hours = max(store_hours.values(), default=0)
    store_rank = [
        {
            "store_name": store_name,
            "hours": hours,
            "percent": round((hours / max_store_hours * 100), 2) if max_store_hours else 0,
        }
        for store_name, hours in sorted(store_hours.items(), key=lambda item: item[1], reverse=True)
    ]
    total_shift_count = max(sum(shift_distribution.values()), 1)
    shift_distribution_items = [
        {
            "shift_name": shift_name,
            "count": count,
            "percent": round(count / total_shift_count * 100, 2),
        }
        for shift_name, count in sorted(shift_distribution.items(), key=lambda item: item[1], reverse=True)
    ]
    return {
        "total_hours": total_hours,
        "scheduled_until_today_hours": cutoff_hours,
        "person_time_count": len(rows),
        "employee_count": len({int(row["employee_id"]) for row in rows}),
        "rest_shift_count": sum(1 for row in rows if "休息" in str(row.get("shift_name") or "")),
        "custom_shift_count": sum(1 for row in rows if int(row.get("is_custom_time") or 0) == 1),
        "store_rank": store_rank,
        "shift_distribution": shift_distribution_items,
    }


def normalize_employee_store_data(
    primary_store_name: str,
    raw_store_names: Optional[List[str]],
    legacy_store_name: str,
    config: AppConfig,
) -> Tuple[str, List[str]]:
    store_names = normalize_store_names(raw_store_names, "")
    primary_store = primary_store_name.strip() or legacy_store_name.strip() or (store_names[0] if store_names else "")
    if not primary_store:
        raise ValueError("请选择主门店。")
    valid_stores = set(config.stores)
    if primary_store not in valid_stores or any(store_name not in valid_stores for store_name in store_names):
        raise ValueError("请选择有效门店。")
    clean_store_names = unique_clean_values([primary_store, *store_names])
    return primary_store, clean_store_names


def employee_store_map_for_ids(connection: sqlite3.Connection, employee_ids: Iterable[int]) -> Dict[int, List[str]]:
    ids = [int(employee_id) for employee_id in employee_ids]
    if not ids:
        return {}
    placeholders = ",".join("?" for _ in ids)
    rows = connection.execute(
        f"""
        SELECT employee_id, store_name
        FROM employee_store_map
        WHERE employee_id IN ({placeholders})
        ORDER BY id
        """,
        ids,
    ).fetchall()
    store_map: Dict[int, List[str]] = {}
    for row in rows:
        employee_id = int(row["employee_id"])
        store_map.setdefault(employee_id, [])
        store_name = str(row["store_name"] or "").strip()
        if store_name and store_name not in store_map[employee_id]:
            store_map[employee_id].append(store_name)
    return store_map


def build_employee_schedule_url(employee: Dict[str, object]) -> str:
    employee_id = str(employee.get("id") or "").strip()
    primary_store = str(employee.get("primary_store_name") or employee.get("store_name") or "").strip()
    store_names = list(employee.get("store_names") or [])
    target_store = primary_store or (str(store_names[0]).strip() if store_names else "")
    query_items: List[Tuple[str, str]] = []
    if target_store:
        query_items.append(("store_names", target_store))
    if employee_id:
        query_items.append(("employee_ids", employee_id))
    query_items.append(("view_mode", "employee"))
    return "/admin/schedules?" + urlencode(query_items) + "#employee-schedule-view"


def employee_role_group_key(role: object) -> str:
    role_text = str(role or "").strip()
    if not role_text:
        return "unset"
    if "区域" in role_text and ("经理" in role_text or "督导" in role_text):
        return "regional"
    if "兼职" in role_text:
        return "part_time"
    if "店员" in role_text:
        return "staff"
    if "店长" in role_text or "经理" in role_text:
        return "management"
    return "other"


def employee_sort_key(employee: Dict[str, object]) -> Tuple[int, str, str]:
    status_rank = 0 if str(employee.get("status") or "") == "在职" else 1
    primary_store = str(employee.get("primary_store_name") or employee.get("store_name") or "")
    employee_name = str(employee.get("employee_name") or "")
    return (status_rank, primary_store, employee_name)


def group_employees_by_role(employees: Iterable[Dict[str, object]]) -> List[Dict[str, object]]:
    grouped: Dict[str, List[Dict[str, object]]] = {key: [] for key, _label in EMPLOYEE_ROLE_GROUPS}
    for employee in employees:
        grouped.setdefault(employee_role_group_key(employee.get("role")), []).append(employee)
    result: List[Dict[str, object]] = []
    for key, label in EMPLOYEE_ROLE_GROUPS:
        group_employees = sorted(grouped.get(key, []), key=employee_sort_key)
        if group_employees:
            result.append({"key": key, "label": label, "count": len(group_employees), "employees": group_employees})
    return result


def attach_employee_store_bindings(
    connection: sqlite3.Connection,
    rows: Iterable[sqlite3.Row],
) -> List[Dict[str, object]]:
    employees = [dict(row) for row in rows]
    store_map = employee_store_map_for_ids(connection, [int(employee["id"]) for employee in employees])
    for employee in employees:
        employee_id = int(employee["id"])
        store_names = store_map.get(employee_id, []) or split_multi_value_text(employee.get("store_name"))
        primary_store = str(employee.get("primary_store_name") or employee.get("store_name") or (store_names[0] if store_names else "")).strip()
        effective_store_names = unique_clean_values([primary_store, *store_names])
        employee["primary_store_name"] = primary_store
        employee["store_name"] = primary_store
        employee["store_names"] = effective_store_names
        employee["store_names_text"] = join_display_values(effective_store_names)
        employee["secondary_store_names"] = [store_name for store_name in effective_store_names if store_name != primary_store]
        employee["secondary_store_names_text"] = join_display_values(employee["secondary_store_names"])
        employee["schedule_url"] = build_employee_schedule_url(employee)
    return employees


def replace_employee_store_bindings(
    connection: sqlite3.Connection,
    employee_id: int,
    store_names: Iterable[str],
    timestamp: str,
) -> None:
    connection.execute("DELETE FROM employee_store_map WHERE employee_id = ?", (employee_id,))
    for store_name in unique_clean_values(store_names):
        connection.execute(
            """
            INSERT OR IGNORE INTO employee_store_map (employee_id, store_name, created_at)
            VALUES (?, ?, ?)
            """,
            (employee_id, store_name, timestamp),
        )


def employee_belongs_to_store(employee: Dict[str, object], store_name: str) -> bool:
    clean_store = store_name.strip()
    return bool(clean_store and clean_store in list(employee.get("store_names") or []))


def employee_record_scope_clause(scope: str) -> str:
    normalized_scope = normalize_employee_record_scope(scope)
    if normalized_scope == "archive":
        return "employees.deleted_at IS NULL AND employees.archived_at IS NOT NULL"
    if normalized_scope == "trash":
        return "employees.deleted_at IS NOT NULL"
    return "employees.deleted_at IS NULL AND employees.archived_at IS NULL"


def normalize_shift_data_scope(data_scope: str) -> str:
    clean_scope = str(data_scope or "").strip()
    return clean_scope if clean_scope in SHIFT_DATA_SCOPES else "current"


def shift_type_scope_clause(data_scope: str) -> str:
    normalized_scope = normalize_shift_data_scope(data_scope)
    if normalized_scope == "archive":
        return "deleted_at IS NULL AND archived_at IS NOT NULL"
    if normalized_scope == "trash":
        return "deleted_at IS NOT NULL"
    if normalized_scope == "all":
        return "1 = 1"
    return "deleted_at IS NULL AND archived_at IS NULL"


def fetch_employees(store_name: str = "", status: str = "", scope: str = "active") -> List[Dict[str, object]]:
    clauses: List[str] = []
    params: List[object] = []
    join_sql = ""
    if store_name.strip():
        join_sql = "JOIN employee_store_map AS employee_store_map_filter ON employee_store_map_filter.employee_id = employees.id"
        clauses.append("employee_store_map_filter.store_name = ?")
        params.append(store_name.strip())
    if status.strip():
        clauses.append("employees.status = ?")
        params.append(status.strip())
    clauses.append(employee_record_scope_clause(scope))
    where_sql = "WHERE " + " AND ".join(clauses) if clauses else ""
    with get_connection() as connection:
        rows = connection.execute(
            f"""
            SELECT
                employees.id, employees.employee_name, employees.store_name, employees.primary_store_name,
                employees.role, employees.phone, employees.status,
                employees.archived_at, employees.archived_by, employees.archive_reason,
                employees.deleted_at, employees.deleted_by, employees.delete_reason,
                employees.created_at, employees.updated_at
            FROM employees
            {join_sql}
            {where_sql}
            ORDER BY COALESCE(NULLIF(employees.primary_store_name, ''), employees.store_name), employees.status, employees.employee_name, employees.id
            """,
            params,
        ).fetchall()
        return attach_employee_store_bindings(connection, rows)


def fetch_employee(employee_id: int) -> Optional[Dict[str, object]]:
    with get_connection() as connection:
        row = connection.execute("SELECT * FROM employees WHERE id = ?", (employee_id,)).fetchone()
        if not row:
            return None
        return attach_employee_store_bindings(connection, [row])[0]


def fetch_schedulable_employees(
    store_name: str,
    employee_status: str = "",
    employee_scope: str = "all",
    selected_employee_ids: Optional[List[int]] = None,
) -> List[Dict[str, object]]:
    statuses = [employee_status.strip()] if employee_status.strip() else ["在职"]
    return fetch_schedulable_employees_for_stores(
        [store_name.strip()] if store_name.strip() else [],
        statuses,
        employee_scope,
        selected_employee_ids,
    )


def fetch_schedulable_employees_for_stores(
    store_names: List[str],
    employee_statuses: Optional[List[str]] = None,
    employee_scope: str = "all",
    selected_employee_ids: Optional[List[int]] = None,
) -> List[Dict[str, object]]:
    clean_stores = unique_clean_values(store_names)
    clean_statuses = unique_clean_values(employee_statuses or ["在职"])
    employees = fetch_employees("", "", scope="active")
    normalized_scope = normalize_employee_scope(employee_scope)
    selected_ids = {int(employee_id) for employee_id in selected_employee_ids or [] if int(employee_id) > 0}
    selected_store_set = set(clean_stores)
    result: List[Dict[str, object]] = []
    for employee in employees:
        if clean_statuses and str(employee.get("status") or "") not in clean_statuses:
            continue
        employee_stores = set(employee.get("store_names") or [])
        if selected_store_set and not (employee_stores & selected_store_set):
            continue
        annotated = annotate_schedule_employee_for_stores(employee, clean_stores)
        if normalized_scope == "primary" and not annotated.get("is_primary_store_employee"):
            continue
        if normalized_scope == "support" and not annotated.get("is_support_store_employee"):
            continue
        if selected_ids and int(annotated["id"]) not in selected_ids:
            continue
        result.append(annotated)
    return result


def create_employee(
    employee_name: str,
    store_name: str,
    role: str,
    phone: str,
    status: str,
    config: AppConfig,
    store_names: Optional[List[str]] = None,
    primary_store_name: str = "",
) -> int:
    clean_name = employee_name.strip()
    clean_status = status.strip() or "在职"
    if not clean_name:
        raise ValueError("请填写员工姓名。")
    primary_store, clean_stores = normalize_employee_store_data(primary_store_name, store_names, store_name, config)
    if clean_status not in EMPLOYEE_STATUSES:
        raise ValueError("员工状态不正确。")
    timestamp = now_text()
    with get_connection() as connection:
        cursor = connection.execute(
            """
            INSERT INTO employees (employee_name, store_name, primary_store_name, role, phone, status, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (clean_name, primary_store, primary_store, role.strip(), phone.strip(), clean_status, timestamp, timestamp),
        )
        employee_id = int(cursor.lastrowid)
        replace_employee_store_bindings(connection, employee_id, clean_stores, timestamp)
        return employee_id


def update_employee(
    employee_id: int,
    employee_name: str,
    store_name: str,
    role: str,
    phone: str,
    status: str,
    config: AppConfig,
    store_names: Optional[List[str]] = None,
    primary_store_name: str = "",
) -> None:
    existing_employee = fetch_employee(employee_id)
    if not existing_employee:
        raise HTTPException(status_code=404, detail="员工不存在")
    clean_name = employee_name.strip()
    clean_status = status.strip() or "在职"
    if not clean_name:
        raise ValueError("请填写员工姓名。")
    legacy_primary_store = store_name or str(existing_employee.get("primary_store_name") or existing_employee.get("store_name") or "")
    primary_store, clean_stores = normalize_employee_store_data(primary_store_name, store_names, legacy_primary_store, config)
    if clean_status not in EMPLOYEE_STATUSES:
        raise ValueError("员工状态不正确。")
    timestamp = now_text()
    with get_connection() as connection:
        connection.execute(
            """
            UPDATE employees
            SET employee_name = ?, store_name = ?, primary_store_name = ?, role = ?, phone = ?, status = ?, updated_at = ?
            WHERE id = ?
            """,
            (clean_name, primary_store, primary_store, role.strip(), phone.strip(), clean_status, timestamp, employee_id),
        )
        replace_employee_store_bindings(connection, employee_id, clean_stores, timestamp)


def disable_employee(employee_id: int) -> None:
    if not fetch_employee(employee_id):
        raise HTTPException(status_code=404, detail="员工不存在")
    with get_connection() as connection:
        connection.execute(
            "UPDATE employees SET status = ?, updated_at = ? WHERE id = ?",
            ("离职", now_text(), employee_id),
        )


def archive_employee(employee_id: int, operator: str, archive_reason: str = "") -> None:
    if not fetch_employee(employee_id):
        raise HTTPException(status_code=404, detail="员工不存在")
    timestamp = now_text()
    with get_connection() as connection:
        connection.execute(
            """
            UPDATE employees
            SET archived_at = ?, archived_by = ?, archive_reason = ?, updated_at = ?
            WHERE id = ? AND deleted_at IS NULL AND archived_at IS NULL
            """,
            (timestamp, operator, archive_reason.strip() or "后台归档员工", timestamp, employee_id),
        )


def unarchive_employee(employee_id: int) -> None:
    if not fetch_employee(employee_id):
        raise HTTPException(status_code=404, detail="员工不存在")
    with get_connection() as connection:
        connection.execute(
            """
            UPDATE employees
            SET archived_at = NULL, archived_by = NULL, archive_reason = NULL, updated_at = ?
            WHERE id = ? AND deleted_at IS NULL
            """,
            (now_text(), employee_id),
        )


def soft_delete_employee(employee_id: int, operator: str, delete_reason: str = "") -> None:
    if not fetch_employee(employee_id):
        raise HTTPException(status_code=404, detail="员工不存在")
    timestamp = now_text()
    with get_connection() as connection:
        connection.execute(
            """
            UPDATE employees
            SET deleted_at = ?, deleted_by = ?, delete_reason = ?, updated_at = ?
            WHERE id = ? AND deleted_at IS NULL
            """,
            (timestamp, operator, delete_reason.strip() or "后台移入回收站", timestamp, employee_id),
        )


def restore_employee(employee_id: int) -> None:
    if not fetch_employee(employee_id):
        raise HTTPException(status_code=404, detail="员工不存在")
    with get_connection() as connection:
        connection.execute(
            """
            UPDATE employees
            SET deleted_at = NULL, deleted_by = NULL, delete_reason = NULL, updated_at = ?
            WHERE id = ?
            """,
            (now_text(), employee_id),
        )


def hard_delete_employee(employee_id: int) -> None:
    if not fetch_employee(employee_id):
        raise HTTPException(status_code=404, detail="员工不存在")
    with get_connection() as connection:
        schedule_count = connection.execute(
            "SELECT COUNT(*) AS total FROM store_schedules WHERE employee_id = ?",
            (employee_id,),
        ).fetchone()["total"]
        if int(schedule_count or 0) > 0:
            raise ValueError("该员工已有排班记录，建议归档而不是永久删除。")
        connection.execute("DELETE FROM employee_store_map WHERE employee_id = ?", (employee_id,))
        connection.execute("DELETE FROM employees WHERE id = ?", (employee_id,))


def fetch_shift_types(
    include_inactive: bool = True,
    active_only: bool = False,
    store_names: Optional[List[str]] = None,
    global_scope: str = "all",
    data_scope: str = "current",
) -> List[Dict[str, object]]:
    clauses: List[str] = []
    params: List[object] = []
    normalized_global_scope = global_scope if global_scope in {"all", "global", "store"} else "all"
    normalized_data_scope = normalize_shift_data_scope(data_scope)
    clauses.append(shift_type_scope_clause(normalized_data_scope))
    if normalized_data_scope == "active":
        clauses.append("is_active = 1")
    elif normalized_data_scope == "inactive":
        clauses.append("is_active = 0")
    if active_only or not include_inactive:
        clauses.append("is_active = 1")
    clean_store_names = unique_clean_values(store_names or [])
    if normalized_global_scope == "global":
        clauses.append("COALESCE(is_global, 0) = 1")
    elif normalized_global_scope == "store":
        clauses.append("COALESCE(is_global, 0) = 0")
    if clean_store_names:
        placeholders = ",".join("?" for _ in clean_store_names)
        if normalized_global_scope == "store":
            clauses.append(f"COALESCE(store_name, '') IN ({placeholders})")
            params.extend(clean_store_names)
        elif normalized_global_scope == "global":
            pass
        else:
            clauses.append(f"(COALESCE(is_global, 0) = 1 OR COALESCE(store_name, '') IN ({placeholders}))")
            params.extend(clean_store_names)
    with get_connection() as connection:
        where_sql = "WHERE " + " AND ".join(clauses) if clauses else ""
        rows = connection.execute(
            f"""
            SELECT
                id, shift_name, store_name, is_global, business_start_time, business_end_time,
                start_time, end_time, duration_hours, color, is_active,
                archived_at, archived_by, archive_reason, deleted_at, deleted_by, delete_reason,
                created_at, updated_at
            FROM shift_types
            {where_sql}
            ORDER BY
                CASE WHEN deleted_at IS NOT NULL THEN 2 WHEN archived_at IS NOT NULL THEN 1 ELSE 0 END,
                is_active DESC, COALESCE(is_global, 0) DESC, COALESCE(store_name, ''), id
            """,
            params,
        ).fetchall()
    result = [dict(row) for row in rows]
    for row in result:
        row["store_label"] = "通用班次" if int(row.get("is_global") or 0) == 1 else (row.get("store_name") or "未绑定门店")
    return result


def fetch_shift_type(shift_type_id: int) -> Optional[Dict[str, object]]:
    with get_connection() as connection:
        row = connection.execute("SELECT * FROM shift_types WHERE id = ?", (shift_type_id,)).fetchone()
    return dict(row) if row else None


def validate_shift_scope(store_name: str, is_global: bool, config: AppConfig) -> Tuple[str, int]:
    if is_global:
        return "", 1
    clean_store = store_name.strip()
    if not clean_store:
        raise ValueError("请选择班次所属门店，或勾选通用班次。")
    if clean_store not in config.stores:
        raise ValueError("请选择有效门店。")
    return clean_store, 0


def is_global_from_shift_scope(shift_scope: str, legacy_is_global: str = "") -> bool:
    normalized_scope = shift_scope.strip()
    if normalized_scope == "global":
        return True
    if normalized_scope == "store":
        return False
    return legacy_is_global in {"1", "true", "on", "yes"}


def shift_scope_duplicate_exists(
    connection: sqlite3.Connection,
    shift_name: str,
    store_name: str,
    is_global: int,
    exclude_shift_type_id: int = 0,
) -> bool:
    row = connection.execute(
        """
        SELECT id FROM shift_types
        WHERE shift_name = ?
          AND COALESCE(store_name, '') = ?
          AND COALESCE(is_global, 0) = ?
          AND id != ?
        LIMIT 1
        """,
        (shift_name, store_name, is_global, exclude_shift_type_id),
    ).fetchone()
    return row is not None


def create_shift_type(
    shift_name: str,
    start_time: str,
    end_time: str,
    duration_hours: str,
    color: str,
    store_name: str = "",
    is_global: bool = False,
    config: Optional[AppConfig] = None,
) -> int:
    clean_name = shift_name.strip()
    if not clean_name:
        raise ValueError("请填写班次名称。")
    clean_store, normalized_global = validate_shift_scope(store_name, is_global, config or load_app_config())
    timestamp = now_text()
    with get_connection() as connection:
        if shift_scope_duplicate_exists(connection, clean_name, clean_store, normalized_global):
            raise ValueError("同一门店或通用范围内班次名称不能重复。")
        cursor = connection.execute(
            """
            INSERT INTO shift_types (
                shift_name, store_name, is_global, business_start_time, business_end_time,
                start_time, end_time, duration_hours, color,
                is_active, created_at, updated_at
            )
            VALUES (?, ?, ?, '', '', ?, ?, ?, ?, 1, ?, ?)
            """,
            (
                clean_name,
                clean_store,
                normalized_global,
                start_time.strip(),
                end_time.strip(),
                parse_duration_hours(duration_hours),
                color.strip(),
                timestamp,
                timestamp,
            ),
        )
        return int(cursor.lastrowid)


def update_shift_type(
    shift_type_id: int,
    shift_name: str,
    start_time: str,
    end_time: str,
    duration_hours: str,
    color: str,
    is_active: bool,
    store_name: str = "",
    is_global: bool = False,
    config: Optional[AppConfig] = None,
) -> None:
    existing_shift = fetch_shift_type(shift_type_id)
    if not existing_shift:
        raise HTTPException(status_code=404, detail="班次不存在")
    clean_name = shift_name.strip()
    if not clean_name:
        raise ValueError("请填写班次名称。")
    clean_store, normalized_global = validate_shift_scope(
        store_name or str(existing_shift.get("store_name") or ""),
        is_global,
        config or load_app_config(),
    )
    with get_connection() as connection:
        if shift_scope_duplicate_exists(connection, clean_name, clean_store, normalized_global, exclude_shift_type_id=shift_type_id):
            raise ValueError("同一门店或通用范围内班次名称不能重复。")
        connection.execute(
            """
            UPDATE shift_types
            SET shift_name = ?, store_name = ?, is_global = ?,
                start_time = ?, end_time = ?, duration_hours = ?,
                color = ?, is_active = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                clean_name,
                clean_store,
                normalized_global,
                start_time.strip(),
                end_time.strip(),
                parse_duration_hours(duration_hours),
                color.strip(),
                1 if is_active else 0,
                now_text(),
                shift_type_id,
            ),
        )


def disable_shift_type(shift_type_id: int) -> None:
    if not fetch_shift_type(shift_type_id):
        raise HTTPException(status_code=404, detail="班次不存在")
    with get_connection() as connection:
        connection.execute("UPDATE shift_types SET is_active = 0, updated_at = ? WHERE id = ?", (now_text(), shift_type_id))


def archive_shift_type(shift_type_id: int, operator: str, archive_reason: str = "") -> None:
    if not fetch_shift_type(shift_type_id):
        raise HTTPException(status_code=404, detail="班次不存在")
    timestamp = now_text()
    with get_connection() as connection:
        connection.execute(
            """
            UPDATE shift_types
            SET archived_at = ?, archived_by = ?, archive_reason = ?,
                deleted_at = NULL, deleted_by = NULL, delete_reason = NULL,
                updated_at = ?
            WHERE id = ? AND deleted_at IS NULL
            """,
            (timestamp, operator, archive_reason.strip() or "后台归档班次", timestamp, shift_type_id),
        )


def soft_delete_shift_type(shift_type_id: int, operator: str, delete_reason: str = "") -> None:
    if not fetch_shift_type(shift_type_id):
        raise HTTPException(status_code=404, detail="班次不存在")
    timestamp = now_text()
    with get_connection() as connection:
        connection.execute(
            """
            UPDATE shift_types
            SET deleted_at = ?, deleted_by = ?, delete_reason = ?, updated_at = ?
            WHERE id = ? AND deleted_at IS NULL
            """,
            (timestamp, operator, delete_reason.strip() or "后台移入班次回收站", timestamp, shift_type_id),
        )


def restore_shift_type(shift_type_id: int) -> None:
    if not fetch_shift_type(shift_type_id):
        raise HTTPException(status_code=404, detail="班次不存在")
    with get_connection() as connection:
        connection.execute(
            """
            UPDATE shift_types
            SET archived_at = NULL, archived_by = NULL, archive_reason = NULL,
                deleted_at = NULL, deleted_by = NULL, delete_reason = NULL,
                updated_at = ?
            WHERE id = ?
            """,
            (now_text(), shift_type_id),
        )


def hard_delete_shift_type(shift_type_id: int) -> None:
    if not fetch_shift_type(shift_type_id):
        raise HTTPException(status_code=404, detail="班次不存在")
    with get_connection() as connection:
        schedule_count = connection.execute(
            "SELECT COUNT(*) AS total FROM store_schedules WHERE shift_type_id = ?",
            (shift_type_id,),
        ).fetchone()["total"]
        if int(schedule_count or 0) > 0:
            raise ValueError("该班次已有排班记录，建议停用或归档，不允许永久删除。")
        connection.execute("DELETE FROM shift_types WHERE id = ?", (shift_type_id,))


def fetch_store_business_hours() -> Dict[str, Dict[str, object]]:
    with get_connection() as connection:
        if not table_exists(connection, "store_business_hours"):
            return {}
        rows = connection.execute("SELECT * FROM store_business_hours ORDER BY store_name").fetchall()
    return {str(row["store_name"]): dict(row) for row in rows}


def upsert_store_business_hours(store_name: str, business_start_time: str, business_end_time: str, config: AppConfig) -> None:
    clean_store = store_name.strip()
    if clean_store not in config.stores:
        raise ValueError("请选择有效门店。")
    timestamp = now_text()
    with get_connection() as connection:
        existing = connection.execute("SELECT id FROM store_business_hours WHERE store_name = ?", (clean_store,)).fetchone()
        if existing:
            connection.execute(
                """
                UPDATE store_business_hours
                SET business_start_time = ?, business_end_time = ?, updated_at = ?
                WHERE store_name = ?
                """,
                (business_start_time.strip(), business_end_time.strip(), timestamp, clean_store),
            )
        else:
            connection.execute(
                """
                INSERT INTO store_business_hours (store_name, business_start_time, business_end_time, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?)
                """,
                (clean_store, business_start_time.strip(), business_end_time.strip(), timestamp, timestamp),
            )


def schedule_redirect_url(
    store_name: str,
    month: str,
    saved: int = 0,
    deleted: int = 0,
    error: str = "",
    saved_count: int = 0,
    created_count: int = 0,
    updated_count: int = 0,
    skipped_count: int = 0,
    anchor: str = "schedule-view",
) -> str:
    try:
        normalized_month = normalize_month(month)
    except ValueError:
        normalized_month = datetime.now().strftime("%Y-%m")
    params = {"store_name": store_name.strip(), "month": normalized_month}
    if saved:
        params["saved"] = str(saved)
    if saved_count:
        params["saved_count"] = str(saved_count)
    if created_count:
        params["created_count"] = str(created_count)
    if updated_count:
        params["updated_count"] = str(updated_count)
    if skipped_count:
        params["skipped_count"] = str(skipped_count)
    if deleted:
        params["deleted"] = str(deleted)
    if error:
        params["error"] = error
    hash_fragment = f"#{anchor.strip()}" if anchor.strip() else ""
    return "/admin/schedules?" + urlencode({key: value for key, value in params.items() if value}) + hash_fragment


def schedule_filters(
    store_name: str,
    month: str,
    employee_status: str = "",
    shift_type_id: int = 0,
    store_names: Optional[List[str]] = None,
    employee_statuses: Optional[List[str]] = None,
    shift_type_ids: Optional[List[int]] = None,
    include_custom_shift: bool = False,
    is_all_stores: bool = False,
) -> Dict[str, object]:
    selected_store_names = unique_clean_values(store_names or ([store_name.strip()] if store_name.strip() and not is_all_stores else []))
    selected_employee_statuses = unique_clean_values(employee_statuses or ([employee_status.strip()] if employee_status.strip() else []))
    selected_shift_type_ids = [int(value) for value in shift_type_ids or ([] if shift_type_id <= 0 else [shift_type_id]) if int(value) > 0]
    return {
        "store_name": selected_store_names[0] if len(selected_store_names) == 1 else "",
        "store_names": selected_store_names,
        "store_names_text": join_display_values(selected_store_names) if selected_store_names else "全部门店",
        "is_all_stores": bool(is_all_stores or not selected_store_names),
        "month": normalize_month(month),
        "employee_status": selected_employee_statuses[0] if len(selected_employee_statuses) == 1 else "",
        "employee_statuses": selected_employee_statuses,
        "shift_type_id": str(selected_shift_type_ids[0]) if len(selected_shift_type_ids) == 1 and not include_custom_shift else "",
        "shift_type_ids": [str(value) for value in selected_shift_type_ids],
        "include_custom_shift": include_custom_shift,
    }


def fetch_schedule_rows(
    store_name: str,
    month: str,
    employee_status: str = "",
    shift_type_id: int = 0,
    store_names: Optional[List[str]] = None,
    employee_statuses: Optional[List[str]] = None,
    shift_type_ids: Optional[List[int]] = None,
    include_custom_shift: bool = False,
    is_all_stores: bool = False,
) -> List[Dict[str, object]]:
    filters = schedule_filters(
        store_name,
        month,
        employee_status,
        shift_type_id,
        store_names=store_names,
        employee_statuses=employee_statuses,
        shift_type_ids=shift_type_ids,
        include_custom_shift=include_custom_shift,
        is_all_stores=is_all_stores,
    )
    year, month_number = [int(part) for part in filters["month"].split("-")]
    last_day = calendar.monthrange(year, month_number)[1]
    clauses = ["schedules.schedule_date >= ?", "schedules.schedule_date <= ?"]
    params: List[object] = [f"{filters['month']}-01", f"{filters['month']}-{last_day:02d}"]
    selected_store_names = list(filters["store_names"])
    if selected_store_names:
        placeholders = ",".join("?" for _ in selected_store_names)
        clauses.append(f"schedules.store_name IN ({placeholders})")
        params.extend(selected_store_names)
    selected_employee_statuses = list(filters["employee_statuses"])
    if selected_employee_statuses:
        placeholders = ",".join("?" for _ in selected_employee_statuses)
        clauses.append(f"employees.status IN ({placeholders})")
        params.extend(selected_employee_statuses)
    selected_shift_type_ids = [int(value) for value in filters["shift_type_ids"]]
    include_custom = bool(filters["include_custom_shift"])
    if selected_shift_type_ids and include_custom:
        placeholders = ",".join("?" for _ in selected_shift_type_ids)
        clauses.append(f"(schedules.shift_type_id IN ({placeholders}) OR COALESCE(schedules.is_custom_time, 0) = 1)")
        params.extend(selected_shift_type_ids)
    elif selected_shift_type_ids:
        placeholders = ",".join("?" for _ in selected_shift_type_ids)
        clauses.append(f"COALESCE(schedules.is_custom_time, 0) = 0 AND schedules.shift_type_id IN ({placeholders})")
        params.extend(selected_shift_type_ids)
    elif include_custom:
        clauses.append("COALESCE(schedules.is_custom_time, 0) = 1")
    where_sql = " AND ".join(clauses)
    with get_connection() as connection:
        rows = connection.execute(
            f"""
            SELECT
                schedules.id, schedules.store_name, schedules.employee_id, schedules.schedule_date,
                schedules.shift_type_id, schedules.note, schedules.created_by,
                COALESCE(schedules.is_custom_time, 0) AS is_custom_time,
                schedules.custom_start_time, schedules.custom_end_time,
                schedules.custom_duration_hours, schedules.custom_label,
                employees.employee_name, employees.role, employees.status AS employee_status,
                shift_types.shift_name, shift_types.start_time, shift_types.end_time,
                shift_types.duration_hours, shift_types.color, shift_types.is_active,
                shift_types.store_name AS shift_store_name, shift_types.is_global AS shift_is_global
            FROM store_schedules AS schedules
            JOIN employees ON employees.id = schedules.employee_id
            LEFT JOIN shift_types ON shift_types.id = schedules.shift_type_id
            WHERE {where_sql}
            ORDER BY schedules.schedule_date, schedules.store_name, employees.employee_name
            """,
            params,
        ).fetchall()
    result = [dict(row) for row in rows]
    for row in result:
        is_custom = int(row.get("is_custom_time") or 0) == 1
        if is_custom:
            row["shift_name"] = str(row.get("custom_label") or "").strip() or "自定义"
            row["start_time"] = row.get("custom_start_time") or ""
            row["end_time"] = row.get("custom_end_time") or ""
            row["duration_hours"] = float(row.get("custom_duration_hours") or 0)
            row["color"] = row.get("color") or "#f97316"
        else:
            row["shift_name"] = row.get("shift_name") or "未设置班次"
            row["duration_hours"] = float(row.get("duration_hours") or 0)
        row["weekday"] = weekday_label(str(row["schedule_date"]))
    return result


def fetch_schedule_context(
    store_name: str,
    month: str,
    employee_status: str,
    config: AppConfig,
    view: str = "calendar",
    show_cross_store: bool = False,
    scope: str = "store",
    employee_scope: str = "all",
    selected_employee_ids: Optional[List[int]] = None,
    shift_type_id: int = 0,
    store_names: Optional[List[str]] = None,
    employee_statuses: Optional[List[str]] = None,
    shift_type_ids: Optional[List[int]] = None,
    include_custom_shift: bool = False,
    is_all_stores: bool = False,
) -> Dict[str, object]:
    normalized_view = normalize_schedule_view(view)
    selected_store_names = unique_clean_values(store_names or ([store_name.strip()] if store_name.strip() and not is_all_stores else []))
    normalized_scope = "all" if is_all_stores or scope == "all" or (normalized_view == "store-summary" and not selected_store_names) else "store"
    if normalized_scope == "all":
        selected_store_names = []
    filters = schedule_filters(
        "" if normalized_scope == "all" else store_name,
        month,
        employee_status,
        shift_type_id,
        store_names=selected_store_names,
        employee_statuses=employee_statuses,
        shift_type_ids=shift_type_ids,
        include_custom_shift=include_custom_shift,
        is_all_stores=normalized_scope == "all",
    )
    days = month_date_items(filters["month"])
    normalized_employee_scope = normalize_employee_scope(employee_scope)
    selected_ids = [int(employee_id) for employee_id in selected_employee_ids or [] if int(employee_id) > 0]
    selected_store_names = list(filters["store_names"])
    is_all = bool(filters["is_all_stores"])
    employee_status_filters = list(filters["employee_statuses"]) or ["在职"]
    employee_options = fetch_schedulable_employees_for_stores(
        selected_store_names,
        employee_status_filters,
        normalized_employee_scope,
    )
    employees = fetch_schedulable_employees_for_stores(
        selected_store_names,
        employee_status_filters,
        normalized_employee_scope,
        selected_employee_ids=selected_ids,
    )
    rows = fetch_schedule_rows(
        filters["store_name"],
        filters["month"],
        filters["employee_status"],
        shift_type_id,
        store_names=selected_store_names,
        employee_statuses=list(filters["employee_statuses"]),
        shift_type_ids=[int(value) for value in filters["shift_type_ids"]],
        include_custom_shift=bool(filters["include_custom_shift"]),
        is_all_stores=is_all,
    )
    schedule_map: Dict[str, Dict[int, Dict[str, object]]] = {}
    daily_store_counts: Dict[str, Dict[str, int]] = {}
    for row in rows:
        schedule_map.setdefault(str(row["schedule_date"]), {})[int(row["employee_id"])] = row
        daily_store_counts.setdefault(str(row["schedule_date"]), {}).setdefault(str(row["store_name"]), 0)
        daily_store_counts[str(row["schedule_date"])][str(row["store_name"])] += 1
    dashboard = build_schedule_dashboard(rows, str(filters["month"]))
    total_hours = round(sum(float(row.get("duration_hours") or 0) for row in rows), 2)
    scheduled_dates = {str(row["schedule_date"]) for row in rows}
    stats = {
        "scheduled_day_count": len(scheduled_dates),
        "employee_count": len(employees),
        "total_hours": total_hours,
        "unscheduled_day_count": max(len(days) - len(scheduled_dates), 0),
        "rest_shift_count": sum(1 for row in rows if str(row.get("shift_name") or "") == "休息"),
    }
    selected_id_strings = [str(employee_id) for employee_id in selected_ids]
    query_items: List[Tuple[str, str]] = [("month", str(filters["month"])), ("employee_scope", normalized_employee_scope), ("scope", normalized_scope)]
    if selected_store_names:
        query_items.extend(("store_names", store) for store in selected_store_names)
    else:
        query_items.append(("store_names", "__all__"))
    query_items.extend(("employee_statuses", status) for status in filters["employee_statuses"])
    query_items.extend(("shift_type_ids", shift_id) for shift_id in filters["shift_type_ids"])
    if filters["include_custom_shift"]:
        query_items.append(("shift_type_ids", "custom"))
    query_items.extend(("employee_ids", employee_id) for employee_id in selected_id_strings)
    if show_cross_store:
        query_items.append(("show_cross_store", "1"))
    view_anchors = {
        "calendar": "calendar-summary",
        "employee": "employee-schedule-view",
        "table": "schedule-table-view",
        "store-summary": "store-summary-view",
    }
    view_urls = {
        view_name: "/admin/schedules?"
        + urlencode([*query_items, ("view_mode", view_name)])
        + f"#{view_anchors[view_name]}"
        for view_name in ("calendar", "employee", "table", "store-summary")
    }
    toggle_cross_store_url = "/admin/schedules?" + urlencode(
        [*query_items, ("view_mode", normalized_view), ("show_cross_store", "" if show_cross_store else "1")]
    )
    global_view_url = "/admin/schedules?" + urlencode(
        [("store_names", "__all__"), ("month", str(filters["month"])), ("scope", "all"), ("view_mode", "store-summary")]
    ) + "#store-summary-view"
    calendar_summary = build_schedule_calendar_summary(days, rows)
    export_query = [("month", filters["month"])]
    export_query.extend(("store_names", store) for store in selected_store_names)
    export_query.extend(("employee_statuses", status) for status in filters["employee_statuses"])
    export_query.extend(("shift_type_ids", shift_id) for shift_id in filters["shift_type_ids"])
    if filters["include_custom_shift"]:
        export_query.append(("shift_type_ids", "custom"))
    is_single_store = len(selected_store_names) == 1 and not is_all
    bulk_shift_types = fetch_shift_types(active_only=True, store_names=selected_store_names if is_single_store else [], global_scope="all" if is_single_store else "global")
    all_shift_types = fetch_shift_types(store_names=selected_store_names or None)
    return {
        "filters": filters,
        "stores": config.stores,
        "selected_store_names": selected_store_names,
        "selected_store_count": len(selected_store_names) if selected_store_names else len(config.stores),
        "is_all_stores": is_all,
        "is_single_store": is_single_store,
        "is_multi_store": len(selected_store_names) > 1,
        "days": days,
        "employees": employees,
        "bulk_employees": employee_options,
        "employee_options": employee_options,
        "selected_employee_ids": selected_id_strings,
        "employee_scope": normalized_employee_scope,
        "shift_types": bulk_shift_types,
        "all_shift_types": all_shift_types,
        "rows": rows,
        "schedule_map": schedule_map,
        "daily_store_counts": daily_store_counts,
        "calendar_summary": calendar_summary,
        "store_summary_calendar": calendar_summary,
        "stats": stats,
        "dashboard": dashboard,
        "calendar_weeks": calendar_week_rows(days),
        "employee_summaries": build_schedule_employee_summaries(employees, rows, days),
        "view": normalized_view,
        "show_cross_store": show_cross_store,
        "scope": normalized_scope,
        "form_store": filters["store_name"] if is_single_store else "",
        "view_urls": view_urls,
        "toggle_cross_store_url": toggle_cross_store_url,
        "global_view_url": global_view_url,
        "max_bulk_schedule_count": config.max_bulk_schedule_count,
        "export_url": "/admin/schedules/export?" + urlencode(export_query),
    }


def insert_schedule_log(
    connection: sqlite3.Connection,
    schedule_id: Optional[int],
    action: str,
    old_value: object,
    new_value: object,
    operator: str,
    created_at: str,
) -> None:
    connection.execute(
        """
        INSERT INTO schedule_logs (schedule_id, action, old_value, new_value, operator, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (
            schedule_id,
            action,
            json.dumps(old_value, ensure_ascii=False) if old_value not in (None, "") else "",
            json.dumps(new_value, ensure_ascii=False) if new_value not in (None, "") else "",
            operator,
            created_at,
        ),
    )


def validate_shift_available_for_store(shift_type: Dict[str, object], store_name: str) -> None:
    if int(shift_type.get("is_active") or 0) != 1:
        raise ValueError("班次必须启用。")
    if int(shift_type.get("is_global") or 0) == 1:
        return
    if str(shift_type.get("store_name") or "").strip() != store_name.strip():
        raise ValueError("班次不属于当前门店。")


def upsert_schedule(
    store_name: str,
    employee_id: int,
    schedule_date: str,
    shift_type_id: int,
    note: str,
    operator: str,
    custom_schedule: Optional[Dict[str, object]] = None,
) -> int:
    clean_store = store_name.strip()
    clean_date = normalize_schedule_date(schedule_date)
    employee = fetch_employee(employee_id)
    if not employee:
        raise ValueError("员工不存在。")
    if not employee_belongs_to_store(employee, clean_store):
        raise ValueError("员工必须绑定该门店。")
    if employee.get("deleted_at"):
        raise ValueError("员工已在回收站，不能排班。")
    if employee.get("archived_at"):
        raise ValueError("员工已归档，不能排班。")
    if str(employee.get("status") or "") != "在职":
        raise ValueError("员工必须是在职状态。")
    is_custom = bool(custom_schedule)
    if is_custom:
        clean_shift_type_id = 0
    else:
        shift_type = fetch_shift_type(shift_type_id)
        if not shift_type:
            raise ValueError("班次不存在。")
        validate_shift_available_for_store(shift_type, clean_store)
        clean_shift_type_id = shift_type_id
    timestamp = now_text()
    clean_note = note.strip()
    new_value = {
        "store_name": clean_store,
        "employee_id": employee_id,
        "schedule_date": clean_date,
        "shift_type_id": clean_shift_type_id,
        "note": clean_note,
        "custom_schedule": custom_schedule or {},
    }
    custom_payload = custom_schedule or {
        "is_custom_time": 0,
        "custom_start_time": "",
        "custom_end_time": "",
        "custom_duration_hours": None,
        "custom_label": "",
    }
    with get_connection() as connection:
        existing = connection.execute(
            "SELECT * FROM store_schedules WHERE employee_id = ? AND schedule_date = ?",
            (employee_id, clean_date),
        ).fetchone()
        if existing:
            old_value = dict(existing)
            schedule_id = int(existing["id"])
            connection.execute(
                """
                UPDATE store_schedules
                SET store_name = ?, shift_type_id = ?, note = ?,
                    is_custom_time = ?, custom_start_time = ?, custom_end_time = ?,
                    custom_duration_hours = ?, custom_label = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    clean_store,
                    clean_shift_type_id,
                    clean_note,
                    int(custom_payload.get("is_custom_time") or 0),
                    str(custom_payload.get("custom_start_time") or ""),
                    str(custom_payload.get("custom_end_time") or ""),
                    custom_payload.get("custom_duration_hours"),
                    str(custom_payload.get("custom_label") or ""),
                    timestamp,
                    schedule_id,
                ),
            )
            insert_schedule_log(connection, schedule_id, "更新排班", old_value, new_value, operator, timestamp)
            return schedule_id
        cursor = connection.execute(
            """
            INSERT INTO store_schedules (
                store_name, employee_id, schedule_date, shift_type_id,
                note, is_custom_time, custom_start_time, custom_end_time,
                custom_duration_hours, custom_label, created_by, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                clean_store,
                employee_id,
                clean_date,
                clean_shift_type_id,
                clean_note,
                int(custom_payload.get("is_custom_time") or 0),
                str(custom_payload.get("custom_start_time") or ""),
                str(custom_payload.get("custom_end_time") or ""),
                custom_payload.get("custom_duration_hours"),
                str(custom_payload.get("custom_label") or ""),
                operator,
                timestamp,
                timestamp,
            ),
        )
        schedule_id = int(cursor.lastrowid)
        insert_schedule_log(connection, schedule_id, "新增排班", "", new_value, operator, timestamp)
        return schedule_id


def existing_schedule_for(employee_id: int, schedule_date: str) -> Optional[Dict[str, object]]:
    with get_connection() as connection:
        row = connection.execute(
            "SELECT * FROM store_schedules WHERE employee_id = ? AND schedule_date = ?",
            (employee_id, schedule_date),
        ).fetchone()
    return dict(row) if row else None


def normalize_schedule_employee_ids(employee_ids: Optional[List[object]], employee_id: object = 0) -> List[int]:
    raw_ids = list(employee_ids or [])
    if not raw_ids and employee_id:
        raw_ids = [employee_id]
    clean_ids: List[int] = []
    for raw_id in raw_ids:
        try:
            clean_id = int(raw_id)
        except (TypeError, ValueError) as exc:
            raise ValueError("员工不存在。") from exc
        if clean_id > 0 and clean_id not in clean_ids:
            clean_ids.append(clean_id)
    if not clean_ids:
        raise ValueError("请选择至少 1 名员工。")
    return clean_ids


def normalize_schedule_dates(schedule_dates: Optional[List[str]], schedule_date: str = "") -> List[str]:
    raw_dates = list(schedule_dates or [])
    if not raw_dates and schedule_date.strip():
        raw_dates = [schedule_date]
    clean_dates: List[str] = []
    for raw_date in raw_dates:
        clean_date = normalize_schedule_date(str(raw_date or ""))
        if clean_date not in clean_dates:
            clean_dates.append(clean_date)
    if not clean_dates:
        raise ValueError("请选择至少 1 个日期。")
    return clean_dates


def time_to_minutes(value: str) -> int:
    clean_value = str(value or "").strip()
    try:
        parsed_time = datetime.strptime(clean_value, "%H:%M")
    except ValueError as exc:
        raise ValueError("自定义时间格式不正确。") from exc
    return parsed_time.hour * 60 + parsed_time.minute


def calculate_duration_from_times(start_time: str, end_time: str) -> float:
    start_minutes = time_to_minutes(start_time)
    end_minutes = time_to_minutes(end_time)
    if end_minutes <= start_minutes:
        end_minutes += 24 * 60
    return round((end_minutes - start_minutes) / 60, 2)


def normalize_custom_schedule_payload(
    schedule_mode: str,
    shift_type_id: int,
    custom_label: str,
    custom_start_time: str,
    custom_end_time: str,
    custom_duration_hours: str,
) -> Tuple[int, Optional[Dict[str, object]]]:
    is_custom = schedule_mode.strip() == "custom"
    if not is_custom:
        if shift_type_id <= 0:
            raise ValueError("请选择班次。")
        return shift_type_id, None
    start_time = custom_start_time.strip()
    end_time = custom_end_time.strip()
    if not start_time or not end_time:
        raise ValueError("自定义时间必须填写开始时间和结束时间。")
    duration = parse_duration_hours(custom_duration_hours)
    if duration <= 0:
        duration = calculate_duration_from_times(start_time, end_time)
    if duration <= 0:
        raise ValueError("自定义工时必须大于 0。")
    return 0, {
        "is_custom_time": 1,
        "custom_label": custom_label.strip() or "自定义",
        "custom_start_time": start_time,
        "custom_end_time": end_time,
        "custom_duration_hours": duration,
    }


def bulk_upsert_schedules(
    store_name: str,
    employee_ids: List[int],
    schedule_dates: List[str],
    shift_type_id: int,
    note: str,
    overwrite_existing: bool,
    operator: str,
    max_bulk_count: int,
    custom_schedule: Optional[Dict[str, object]] = None,
) -> Dict[str, int]:
    clean_store = store_name.strip()
    if not clean_store:
        raise ValueError("请选择具体门店后进行排班操作。请选择单个具体门店后进行排班操作。")
    total_count = len(employee_ids) * len(schedule_dates)
    if total_count <= 0:
        raise ValueError("请选择至少 1 名员工和 1 个日期。")
    if total_count > max_bulk_count:
        raise ValueError(f"一次最多批量生成 {max_bulk_count} 条排班，请减少员工或日期数量。")
    if not custom_schedule:
        if shift_type_id <= 0:
            raise ValueError("请选择班次。")
        shift_type = fetch_shift_type(shift_type_id)
        if not shift_type:
            raise ValueError("班次不存在。")
        validate_shift_available_for_store(shift_type, clean_store)
    for selected_employee_id in employee_ids:
        employee = fetch_employee(selected_employee_id)
        if not employee:
            raise ValueError("员工不存在。")
        if not employee_belongs_to_store(employee, clean_store):
            raise ValueError("员工必须绑定该门店。")
        if employee.get("deleted_at"):
            raise ValueError("员工已在回收站，不能排班。")
        if employee.get("archived_at"):
            raise ValueError("员工已归档，不能排班。")
        if str(employee.get("status") or "") != "在职":
            raise ValueError("员工必须是在职状态。")

    created_count = 0
    updated_count = 0
    skipped_count = 0
    for selected_employee_id in employee_ids:
        for selected_date in schedule_dates:
            existing = existing_schedule_for(selected_employee_id, selected_date)
            if existing and not overwrite_existing:
                skipped_count += 1
                continue
            upsert_schedule(clean_store, selected_employee_id, selected_date, shift_type_id, note, operator, custom_schedule=custom_schedule)
            if existing:
                updated_count += 1
            else:
                created_count += 1

    return {
        "created_count": created_count,
        "updated_count": updated_count,
        "skipped_count": skipped_count,
        "saved_count": created_count + updated_count,
        "total_count": total_count,
    }


def delete_schedule(schedule_id: int, operator: str) -> Dict[str, object]:
    timestamp = now_text()
    with get_connection() as connection:
        row = connection.execute("SELECT * FROM store_schedules WHERE id = ?", (schedule_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="排班不存在")
        old_value = dict(row)
        connection.execute("DELETE FROM store_schedules WHERE id = ?", (schedule_id,))
        insert_schedule_log(connection, schedule_id, "删除排班", old_value, "", operator, timestamp)
    return old_value


def copy_previous_day_schedules(store_name: str, target_date: str, operator: str) -> Dict[str, object]:
    clean_store = store_name.strip()
    clean_target_date = normalize_schedule_date(target_date)
    if not clean_store:
        raise ValueError("请选择有效门店。")
    source_date = (datetime.strptime(clean_target_date, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
    with get_connection() as connection:
        rows = connection.execute(
            """
            SELECT
                employee_id, shift_type_id, note,
                COALESCE(is_custom_time, 0) AS is_custom_time,
                custom_start_time, custom_end_time, custom_duration_hours, custom_label
            FROM store_schedules
            WHERE store_name = ? AND schedule_date = ?
            ORDER BY employee_id
            """,
            (clean_store, source_date),
        ).fetchall()
    copied_count = 0
    for row in rows:
        upsert_schedule(
            clean_store,
            int(row["employee_id"]),
            clean_target_date,
            int(row["shift_type_id"]),
            str(row["note"] or ""),
            operator,
            custom_schedule={
                "is_custom_time": 1,
                "custom_start_time": str(row["custom_start_time"] or ""),
                "custom_end_time": str(row["custom_end_time"] or ""),
                "custom_duration_hours": float(row["custom_duration_hours"] or 0),
                "custom_label": str(row["custom_label"] or "自定义"),
            }
            if int(row["is_custom_time"] or 0) == 1
            else None,
        )
        copied_count += 1
    return {"source_date": source_date, "target_date": clean_target_date, "copied_count": copied_count}


def clear_employee_month_schedules(store_name: str, employee_id: int, month: str, operator: str) -> int:
    clean_store = store_name.strip()
    clean_month = normalize_month(month)
    employee = fetch_employee(employee_id)
    if not employee:
        raise ValueError("员工不存在。")
    year, month_number = [int(part) for part in clean_month.split("-")]
    last_day = calendar.monthrange(year, month_number)[1]
    clauses = ["employee_id = ?", "schedule_date >= ?", "schedule_date <= ?"]
    params: List[object] = [employee_id, f"{clean_month}-01", f"{clean_month}-{last_day:02d}"]
    if clean_store:
        clauses.append("store_name = ?")
        params.append(clean_store)
    with get_connection() as connection:
        rows = connection.execute(
            f"SELECT id FROM store_schedules WHERE {' AND '.join(clauses)} ORDER BY schedule_date",
            params,
        ).fetchall()
    for row in rows:
        delete_schedule(int(row["id"]), operator)
    return len(rows)


def build_schedule_excel(rows: List[Dict[str, object]]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "门店排班"
    headers = ["门店", "员工", "日期", "星期", "班次", "开始时间", "结束时间", "工时", "是否自定义", "备注"]
    sheet.append(headers)
    header_fill = PatternFill(fill_type="solid", fgColor="E8EEF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        sheet.append(
            [
                row.get("store_name"),
                row.get("employee_name"),
                row.get("schedule_date"),
                row.get("weekday"),
                row.get("shift_name"),
                row.get("start_time") or "",
                row.get("end_time") or "",
                row.get("duration_hours") or 0,
                "是" if int(row.get("is_custom_time") or 0) == 1 else "否",
                row.get("note") or "",
            ]
        )
    widths = [18, 16, 14, 10, 14, 12, 12, 10, 12, 28]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def embedded_html_filename(page_key: str) -> str:
    return f"{page_key}/index.html"


def legacy_embedded_html_filename(page_key: str) -> str:
    return f"{page_key}.html"


def embedded_page_dir(page_key: str) -> Optional[Path]:
    clean_key = page_key.strip()
    if not is_valid_embedded_page_key(clean_key):
        return None
    return get_embedded_pages_dir() / clean_key


def embedded_html_path(filename: str) -> Optional[Path]:
    clean_filename = safe_uploaded_name(filename)
    if clean_filename != filename or not clean_filename.lower().endswith(".html"):
        return None
    base_dir = get_embedded_pages_dir()
    target = base_dir / clean_filename
    try:
        target.resolve().relative_to(base_dir.resolve())
    except (OSError, ValueError):
        return None
    return target


def legacy_embedded_html_path(page_key: str, filename: object = "") -> Optional[Path]:
    clean_key = page_key.strip()
    if not is_valid_embedded_page_key(clean_key):
        return None
    candidate = str(filename or legacy_embedded_html_filename(clean_key)).strip()
    if "/" in candidate.replace("\\", "/"):
        candidate = legacy_embedded_html_filename(clean_key)
    return embedded_html_path(candidate)


def embedded_file_size_label(size: object) -> str:
    try:
        value = int(size or 0)
    except (TypeError, ValueError):
        value = 0
    if value < 1024:
        return f"{value} B"
    if value < 1024 * 1024:
        return f"{value / 1024:.1f} KB"
    return f"{value / (1024 * 1024):.1f} MB"


def normalize_embedded_storage_type(value: object) -> str:
    storage_type = str(value or "html").strip().lower()
    return "zip" if storage_type == "zip" else "html"


def embedded_storage_label(storage_type: object) -> str:
    return "ZIP" if normalize_embedded_storage_type(storage_type) == "zip" else "HTML"


def safe_embedded_resource_path(resource_path: object) -> Optional[str]:
    clean_path = str(resource_path or "").strip()
    if not clean_path:
        return None
    if "\\" in clean_path or clean_path.startswith("/"):
        return None
    parts = clean_path.split("/")
    if any(part in {"", ".", ".."} for part in parts):
        return None
    posix_path = PurePosixPath(clean_path)
    if posix_path.is_absolute():
        return None
    return "/".join(parts)


def resolve_embedded_resource(page: Dict[str, object], resource_path: object = "index.html") -> Optional[Path]:
    page_key = str(page.get("page_key") or "").strip()
    entry_file = safe_embedded_resource_path(page.get("entry_file") or "index.html") or "index.html"
    clean_resource = safe_embedded_resource_path(resource_path or entry_file)
    if not clean_resource:
        return None

    page_dir = embedded_page_dir(page_key)
    if page_dir and page_dir.is_dir():
        root = page_dir.resolve()
        target = (root / clean_resource).resolve()
        try:
            target.relative_to(root)
        except ValueError:
            return None
        return target if target.is_file() else None

    if clean_resource == entry_file:
        legacy_target = legacy_embedded_html_path(page_key, page.get("filename"))
        if legacy_target and legacy_target.is_file():
            return legacy_target
    return None


def embedded_page_file_exists(page: Dict[str, object]) -> bool:
    return resolve_embedded_resource(page, page.get("entry_file") or "index.html") is not None


def embedded_media_type(path: Path) -> str:
    extension = path.suffix.lower().lstrip(".")
    overrides = {
        "html": "text/html",
        "css": "text/css",
        "js": "application/javascript",
        "svg": "image/svg+xml",
        "json": "application/json",
        "csv": "text/csv",
        "txt": "text/plain",
        "pdf": "application/pdf",
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    }
    if extension in overrides:
        return overrides[extension]
    guessed, _encoding = mimetypes.guess_type(path.name)
    return guessed or "application/octet-stream"


def enrich_embedded_page(page: Dict[str, object]) -> Dict[str, object]:
    page["is_enabled"] = int(page.get("enabled") or 0) == 1
    page["version"] = embedded_page_version(page.get("updated_at"))
    page["storage_type"] = normalize_embedded_storage_type(page.get("storage_type"))
    page["storage_label"] = embedded_storage_label(page.get("storage_type"))
    page["entry_file"] = str(page.get("entry_file") or "index.html")
    page["file_size"] = int(page.get("file_size") or 0)
    page["file_size_label"] = embedded_file_size_label(page.get("file_size"))
    page["file_exists"] = embedded_page_file_exists(page)
    page["file_exists_label"] = "存在" if page["file_exists"] else "缺失"
    return page


def fetch_embedded_pages(enabled_only: bool = False, deleted_only: bool = False) -> List[Dict[str, object]]:
    clauses: List[str] = []
    params: List[object] = []
    if deleted_only:
        clauses.append("deleted_at IS NOT NULL")
    else:
        clauses.append("deleted_at IS NULL")
    if enabled_only:
        clauses.append("enabled = 1")
    where_sql = "WHERE " + " AND ".join(clauses) if clauses else ""
    with get_connection() as connection:
        rows = connection.execute(
            f"""
            SELECT id, page_key, title, nav_label, filename, storage_type,
                   entry_file, file_size, enabled, created_at, updated_at, updated_by,
                   deleted_at, deleted_by, delete_reason
            FROM embedded_pages
            {where_sql}
            ORDER BY nav_label ASC, id ASC
            """,
            params,
        ).fetchall()
    pages = [dict(row) for row in rows]
    for page in pages:
        enrich_embedded_page(page)
    return pages


def fetch_enabled_embedded_pages() -> List[Dict[str, object]]:
    return fetch_embedded_pages(enabled_only=True)


def fetch_embedded_page(page_key: str, enabled_only: bool = False, include_deleted: bool = False) -> Optional[Dict[str, object]]:
    clean_key = page_key.strip()
    if not is_valid_embedded_page_key(clean_key):
        return None
    clauses = ["page_key = ?"]
    params: List[object] = [clean_key]
    if not include_deleted:
        clauses.append("deleted_at IS NULL")
    if enabled_only:
        clauses.append("enabled = 1")
    with get_connection() as connection:
        row = connection.execute(
            f"""
            SELECT id, page_key, title, nav_label, filename, storage_type,
                   entry_file, file_size, enabled, created_at, updated_at, updated_by,
                   deleted_at, deleted_by, delete_reason
            FROM embedded_pages
            WHERE {" AND ".join(clauses)}
            """,
            params,
        ).fetchone()
    if not row:
        return None
    return enrich_embedded_page(dict(row))


def validate_embedded_zip_member(filename: str) -> Optional[str]:
    raw_name = str(filename or "").strip()
    if not raw_name:
        raise ValueError("ZIP 文件路径不安全：存在空文件名。")
    if "\\" in raw_name or raw_name.startswith("/"):
        raise ValueError("ZIP 文件路径不安全：不能包含绝对路径、反斜杠或路径穿越。")
    is_directory = raw_name.endswith("/")
    clean_name = raw_name[:-1] if is_directory else raw_name
    if not clean_name:
        raise ValueError("ZIP 文件路径不安全：存在空文件名。")
    parts = clean_name.split("/")
    if any(part in {"", ".", ".."} for part in parts):
        raise ValueError("ZIP 文件路径不安全：不能包含绝对路径、反斜杠或路径穿越。")
    if PurePosixPath(clean_name).is_absolute():
        raise ValueError("ZIP 文件路径不安全：不能包含绝对路径、反斜杠或路径穿越。")
    if is_directory:
        return None
    leaf = parts[-1]
    extension = leaf.rsplit(".", 1)[-1].lower() if "." in leaf else ""
    if extension in EMBEDDED_BLOCKED_EXTENSIONS:
        raise ValueError(f"ZIP 内不允许包含 .{extension} 文件。")
    if extension not in EMBEDDED_ALLOWED_EXTENSIONS:
        raise ValueError(f"ZIP 内文件类型 .{extension or '无后缀'} 不受支持。")
    return "/".join(parts)


def select_embedded_zip_entry_file(safe_files: List[str]) -> str:
    root_html_files = [
        file_path
        for file_path in safe_files
        if "/" not in file_path and file_path.lower().endswith(".html")
    ]
    index_files = [file_path for file_path in root_html_files if file_path.lower() == "index.html"]
    if index_files:
        return index_files[0]
    if len(root_html_files) == 1:
        return root_html_files[0]
    if len(root_html_files) > 1:
        raise ValueError("ZIP 根目录包含多个 HTML 文件，请将入口文件命名为 index.html。")
    raise ValueError("ZIP 根目录必须包含 index.html 或一个 HTML 文件。")


def validate_embedded_zip_content(content: bytes) -> str:
    try:
        with zipfile.ZipFile(BytesIO(content)) as archive:
            safe_files: List[str] = []
            for info in archive.infolist():
                member_path = validate_embedded_zip_member(info.filename)
                if member_path:
                    safe_files.append(member_path)
            return select_embedded_zip_entry_file(safe_files)
    except zipfile.BadZipFile as exc:
        raise ValueError("ZIP 文件无法读取，请重新打包后上传。") from exc


async def prepare_embedded_upload_file(html_file: Optional[UploadFile], config: AppConfig) -> PreparedEmbeddedUpload:
    if not html_file or not html_file.filename:
        raise ValueError("请选择要上传的 HTML 或 ZIP 文件。")
    filename = safe_uploaded_name(html_file.filename)
    lower_filename = filename.lower()
    content = await html_file.read()
    if lower_filename.endswith(".html"):
        if len(content) > config.max_embedded_html_bytes:
            raise ValueError(f"HTML 文件不能超过 {config.max_embedded_html_mb}MB。")
        if not content.strip():
            raise ValueError("HTML 文件内容不能为空。")
        return PreparedEmbeddedUpload("html", "index.html", len(content), content)
    if lower_filename.endswith(".zip"):
        if len(content) > config.max_embedded_zip_bytes:
            raise ValueError(f"ZIP 文件不能超过 {config.max_embedded_zip_mb}MB。")
        if not content:
            raise ValueError("ZIP 文件内容不能为空。")
        source_entry_file = validate_embedded_zip_content(content)
        return PreparedEmbeddedUpload("zip", "index.html", len(content), content, source_entry_file)
    raise ValueError("只允许上传 .html 或 .zip 文件。")


def validate_embedded_page_form(page_key: str, title: str, nav_label: str) -> Tuple[str, str, str]:
    clean_key = page_key.strip()
    clean_title = title.strip()
    clean_nav_label = nav_label.strip()
    if not is_valid_embedded_page_key(clean_key):
        raise ValueError("page_key 只能使用小写字母、数字和短横线。")
    if not clean_title:
        raise ValueError("请填写页面标题。")
    if not clean_nav_label:
        raise ValueError("请填写导航名称。")
    return clean_key, clean_title, clean_nav_label


def write_embedded_upload_to_temp_dir(target_dir: Path, upload: PreparedEmbeddedUpload) -> None:
    target_dir.mkdir(parents=True, exist_ok=True)
    if upload.storage_type == "html":
        (target_dir / "index.html").write_bytes(upload.content)
        return
    with zipfile.ZipFile(BytesIO(upload.content)) as archive:
        for info in archive.infolist():
            member_path = validate_embedded_zip_member(info.filename)
            if not member_path:
                continue
            output_path = upload.entry_file if member_path == upload.source_entry_file else member_path
            target = (target_dir / output_path).resolve()
            try:
                target.relative_to(target_dir.resolve())
            except ValueError as exc:
                raise ValueError("ZIP 文件路径不安全：不能包含绝对路径、反斜杠或路径穿越。") from exc
            target.parent.mkdir(parents=True, exist_ok=True)
            target.write_bytes(archive.read(info))


def replace_embedded_page_directory(page_key: str, upload: PreparedEmbeddedUpload) -> Path:
    page_dir = embedded_page_dir(page_key)
    if not page_dir:
        raise ValueError("嵌入页面目录不正确。")
    base_dir = get_embedded_pages_dir()
    base_dir.mkdir(parents=True, exist_ok=True)
    token = uuid.uuid4().hex
    temp_dir = base_dir / f".{page_key}.tmp-{token}"
    backup_dir = base_dir / f".{page_key}.bak-{token}"
    backup_created = False
    try:
        write_embedded_upload_to_temp_dir(temp_dir, upload)
        if page_dir.exists():
            page_dir.rename(backup_dir)
            backup_created = True
        temp_dir.rename(page_dir)
        if backup_created and backup_dir.exists():
            shutil.rmtree(backup_dir)
        return page_dir / upload.entry_file
    except Exception:
        if temp_dir.exists():
            shutil.rmtree(temp_dir, ignore_errors=True)
        if backup_created and backup_dir.exists() and not page_dir.exists():
            backup_dir.rename(page_dir)
        raise


def create_embedded_page(
    page_key: str,
    title: str,
    nav_label: str,
    enabled: bool,
    upload: PreparedEmbeddedUpload,
    updated_by: str,
) -> int:
    clean_key, clean_title, clean_nav_label = validate_embedded_page_form(page_key, title, nav_label)
    filename = embedded_html_filename(clean_key)
    timestamp = now_text()
    with get_connection() as connection:
        existing = connection.execute("SELECT id FROM embedded_pages WHERE page_key = ?", (clean_key,)).fetchone()
        if existing:
            raise ValueError("page_key 已存在，请使用替换文件。")
    target = replace_embedded_page_directory(clean_key, upload)
    try:
        with get_connection() as connection:
            cursor = connection.execute(
                """
                INSERT INTO embedded_pages (
                    page_key, title, nav_label, filename, storage_type, entry_file,
                    file_size, enabled, created_at, updated_at, updated_by
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    clean_key,
                    clean_title,
                    clean_nav_label,
                    filename,
                    upload.storage_type,
                    upload.entry_file,
                    upload.file_size,
                    1 if enabled else 0,
                    timestamp,
                    timestamp,
                    updated_by,
                ),
            )
            return int(cursor.lastrowid)
    except Exception:
        try:
            if target.exists():
                shutil.rmtree(target.parent, ignore_errors=True)
        except OSError:
            pass
        raise


def replace_embedded_page_file(page_key: str, upload: PreparedEmbeddedUpload, updated_by: str) -> None:
    page = fetch_embedded_page(page_key)
    if not page:
        raise HTTPException(status_code=404, detail="嵌入页面不存在")
    clean_key = str(page.get("page_key") or page_key).strip()
    replace_embedded_page_directory(clean_key, upload)
    timestamp = now_text()
    with get_connection() as connection:
        connection.execute(
            """
            UPDATE embedded_pages
            SET filename = ?, storage_type = ?, entry_file = ?, file_size = ?,
                updated_at = ?, updated_by = ?
            WHERE page_key = ?
            """,
            (
                embedded_html_filename(clean_key),
                upload.storage_type,
                upload.entry_file,
                upload.file_size,
                timestamp,
                updated_by,
                clean_key,
            ),
        )


def set_embedded_page_enabled(page_key: str, enabled: bool, updated_by: str) -> None:
    page = fetch_embedded_page(page_key)
    if not page:
        raise HTTPException(status_code=404, detail="嵌入页面不存在")
    timestamp = now_text()
    with get_connection() as connection:
        connection.execute(
            """
            UPDATE embedded_pages
            SET enabled = ?, updated_at = ?, updated_by = ?
            WHERE page_key = ?
            """,
            (1 if enabled else 0, timestamp, updated_by, page_key.strip()),
        )


def soft_delete_embedded_page(page_key: str, deleted_by: str, delete_reason: str = "") -> None:
    page = fetch_embedded_page(page_key)
    if not page:
        raise HTTPException(status_code=404, detail="嵌入页面不存在")
    timestamp = now_text()
    with get_connection() as connection:
        connection.execute(
            """
            UPDATE embedded_pages
            SET deleted_at = ?, deleted_by = ?, delete_reason = ?,
                enabled = 0, updated_at = ?, updated_by = ?
            WHERE page_key = ? AND deleted_at IS NULL
            """,
            (timestamp, deleted_by, delete_reason.strip(), timestamp, deleted_by, page_key.strip()),
        )


def restore_embedded_page(page_key: str, updated_by: str) -> None:
    page = fetch_embedded_page(page_key, include_deleted=True)
    if not page or not page.get("deleted_at"):
        raise HTTPException(status_code=404, detail="回收站中没有该嵌入页面")
    timestamp = now_text()
    with get_connection() as connection:
        connection.execute(
            """
            UPDATE embedded_pages
            SET deleted_at = NULL, deleted_by = NULL, delete_reason = NULL,
                enabled = 1, updated_at = ?, updated_by = ?
            WHERE page_key = ?
            """,
            (timestamp, updated_by, page_key.strip()),
        )


def hard_delete_embedded_page(page_key: str) -> None:
    page = fetch_embedded_page(page_key, include_deleted=True)
    if not page or not page.get("deleted_at"):
        raise HTTPException(status_code=404, detail="只能永久删除回收站中的嵌入页面")
    page_dir = embedded_page_dir(page_key)
    legacy_file = legacy_embedded_html_path(page_key, page.get("filename"))
    with get_connection() as connection:
        connection.execute("DELETE FROM embedded_pages WHERE page_key = ?", (page_key.strip(),))
    if page_dir and page_dir.exists():
        shutil.rmtree(page_dir, ignore_errors=True)
    if legacy_file and legacy_file.exists():
        try:
            legacy_file.unlink(missing_ok=True)
        except OSError:
            pass


def insert_ticket_log(
    connection: sqlite3.Connection,
    ticket_id: int,
    action: str,
    note: str,
    operator: str,
    created_at: str,
    old_status: Optional[str] = None,
    new_status: Optional[str] = None,
    old_assigned_to: Optional[str] = None,
    new_assigned_to: Optional[str] = None,
) -> None:
    connection.execute(
        """
        INSERT INTO ticket_logs (
            ticket_id, action, old_status, new_status, old_assigned_to,
            new_assigned_to, note, operator, created_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            ticket_id,
            action,
            old_status,
            new_status,
            old_assigned_to,
            new_assigned_to,
            note,
            operator,
            created_at,
        ),
    )


def normalize_ticket_ids(ticket_ids: Iterable[object]) -> List[int]:
    normalized: List[int] = []
    seen: set[int] = set()
    for raw_id in ticket_ids:
        try:
            ticket_id = int(raw_id)
        except (TypeError, ValueError):
            continue
        if ticket_id <= 0 or ticket_id in seen:
            continue
        normalized.append(ticket_id)
        seen.add(ticket_id)
    return normalized


def ticket_filters_from_params(
    store_name: str = "",
    request_type: str = "",
    urgency: str = "",
    status: str = "",
    assigned_to: str = "",
    date_start: str = "",
    date_end: str = "",
    keyword: str = "",
    due_status: str = "",
    scope: str = "active",
) -> Dict[str, str]:
    filters = {
        "store_name": store_name.strip(),
        "request_type": request_type.strip(),
        "urgency": urgency.strip(),
        "status": status.strip(),
        "assigned_to": assigned_to.strip(),
        "date_start": date_start.strip(),
        "date_end": date_end.strip(),
        "keyword": keyword.strip(),
        "due_status": due_status.strip(),
    }
    if scope:
        filters["__ticket_scope"] = scope
    return filters


def ticket_scope_for_source_view(source_view: str) -> str:
    if source_view == "archive":
        return "archive"
    if source_view == "trash":
        return "deleted"
    if source_view == "store":
        return "store"
    return "active"


def admin_path_for_source_view(source_view: str) -> str:
    if source_view == "archive":
        return "/admin/archive"
    if source_view == "trash":
        return "/admin/trash"
    return "/admin"


def admin_redirect_url(source_view: str, **params: object) -> str:
    path = admin_path_for_source_view(source_view)
    query = urlencode({key: str(value) for key, value in params.items() if value not in (None, "")})
    return path + (f"?{query}" if query else "")


def form_error_message(exc: Exception) -> str:
    if isinstance(exc, HTTPException):
        return str(exc.detail)
    return str(exc)


def admin_ticket_filters_from_form(
    store_name: str = Form(""),
    request_type: str = Form(""),
    urgency: str = Form(""),
    status: str = Form(""),
    assigned_to: str = Form(""),
    date_start: str = Form(""),
    date_end: str = Form(""),
    keyword: str = Form(""),
    due_status: str = Form(""),
    source_view: str = Form("active"),
) -> Dict[str, str]:
    return ticket_filters_from_params(
        store_name,
        request_type,
        urgency,
        status,
        assigned_to,
        date_start,
        date_end,
        keyword,
        due_status,
        scope=ticket_scope_for_source_view(source_view),
    )


def bulk_ticket_ids_from_scope(
    ticket_ids: Iterable[object],
    select_scope: str,
    filters: Dict[str, str],
    sort: str,
    config: AppConfig,
) -> List[int]:
    if select_scope == "filtered":
        return normalize_ticket_ids(ticket["id"] for ticket in fetch_tickets(filters, sort, config))
    return normalize_ticket_ids(ticket_ids)


def in_clause_for_ids(ticket_ids: List[int]) -> Tuple[str, List[int]]:
    placeholders = ",".join("?" for _ in ticket_ids)
    return placeholders, ticket_ids


def bulk_archive_tickets(ticket_ids: Iterable[object], operator: str, archive_reason: str = "") -> int:
    ids = normalize_ticket_ids(ticket_ids)
    if not ids:
        return 0
    reason = archive_reason.strip() or "批量归档"
    timestamp = now_text()
    placeholders, params = in_clause_for_ids(ids)
    with get_connection() as connection:
        rows = connection.execute(
            f"""
            SELECT id
            FROM tickets
            WHERE id IN ({placeholders})
              AND deleted_at IS NULL
              AND archived_at IS NULL
            ORDER BY id
            """,
            params,
        ).fetchall()
        target_ids = [int(row["id"]) for row in rows]
        for ticket_id in target_ids:
            connection.execute(
                """
                UPDATE tickets
                SET archived_at = ?, archived_by = ?, archive_reason = ?, updated_at = ?
                WHERE id = ?
                """,
                (timestamp, operator, reason, timestamp, ticket_id),
            )
            insert_ticket_log(connection, ticket_id, "归档工单", reason, operator, timestamp)
    return len(target_ids)


def bulk_unarchive_tickets(ticket_ids: Iterable[object], operator: str) -> int:
    ids = normalize_ticket_ids(ticket_ids)
    if not ids:
        return 0
    timestamp = now_text()
    placeholders, params = in_clause_for_ids(ids)
    with get_connection() as connection:
        rows = connection.execute(
            f"""
            SELECT id
            FROM tickets
            WHERE id IN ({placeholders})
              AND deleted_at IS NULL
              AND archived_at IS NOT NULL
            ORDER BY id
            """,
            params,
        ).fetchall()
        target_ids = [int(row["id"]) for row in rows]
        for ticket_id in target_ids:
            connection.execute(
                """
                UPDATE tickets
                SET archived_at = NULL, archived_by = NULL, archive_reason = NULL, updated_at = ?
                WHERE id = ?
                """,
                (timestamp, ticket_id),
            )
            insert_ticket_log(connection, ticket_id, "取消归档", "", operator, timestamp)
    return len(target_ids)


def accept_ticket(ticket_id: int, operator: str) -> None:
    clean_operator = operator.strip()
    timestamp = now_text()
    with get_connection() as connection:
        row = connection.execute(
            "SELECT * FROM tickets WHERE id = ? AND deleted_at IS NULL",
            (ticket_id,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="工单不存在")
        ticket = dict(row)
        old_assigned_to = str(ticket.get("assigned_to") or "").strip()
        if old_assigned_to and old_assigned_to != clean_operator:
            raise HTTPException(status_code=400, detail=f"工单已由 {old_assigned_to} 负责。")
        if old_assigned_to == clean_operator:
            return
        connection.execute(
            """
            UPDATE tickets
            SET assigned_to = ?, updated_at = ?
            WHERE id = ?
            """,
            (clean_operator, timestamp, ticket_id),
        )
        insert_ticket_log(
            connection,
            ticket_id,
            "接单",
            f"{clean_operator} 接单",
            clean_operator,
            timestamp,
            old_status=str(ticket.get("status") or ""),
            new_status=str(ticket.get("status") or ""),
            old_assigned_to=old_assigned_to,
            new_assigned_to=clean_operator,
        )


def bulk_soft_delete_tickets(ticket_ids: Iterable[object], operator: str, delete_reason: str = "") -> int:
    ids = normalize_ticket_ids(ticket_ids)
    if not ids:
        return 0
    reason = delete_reason.strip() or "批量移入回收站"
    timestamp = now_text()
    placeholders, params = in_clause_for_ids(ids)
    with get_connection() as connection:
        rows = connection.execute(
            f"""
            SELECT id
            FROM tickets
            WHERE id IN ({placeholders})
              AND deleted_at IS NULL
            ORDER BY id
            """,
            params,
        ).fetchall()
        target_ids = [int(row["id"]) for row in rows]
        for ticket_id in target_ids:
            connection.execute(
                """
                UPDATE tickets
                SET deleted_at = ?, deleted_by = ?, delete_reason = ?, updated_at = ?
                WHERE id = ?
                """,
                (timestamp, operator, reason, timestamp, ticket_id),
            )
            insert_ticket_log(connection, ticket_id, "移入回收站", reason, operator, timestamp)
    return len(target_ids)


def bulk_restore_tickets(ticket_ids: Iterable[object], operator: str) -> int:
    ids = normalize_ticket_ids(ticket_ids)
    if not ids:
        return 0
    timestamp = now_text()
    placeholders, params = in_clause_for_ids(ids)
    with get_connection() as connection:
        rows = connection.execute(
            f"""
            SELECT id
            FROM tickets
            WHERE id IN ({placeholders})
              AND deleted_at IS NOT NULL
            ORDER BY id
            """,
            params,
        ).fetchall()
        target_ids = [int(row["id"]) for row in rows]
        for ticket_id in target_ids:
            connection.execute(
                """
                UPDATE tickets
                SET deleted_at = NULL, deleted_by = NULL, delete_reason = NULL, updated_at = ?
                WHERE id = ?
                """,
                (timestamp, ticket_id),
            )
            insert_ticket_log(connection, ticket_id, "恢复工单", "", operator, timestamp)
    return len(target_ids)


def bulk_hard_delete_tickets(ticket_ids: Iterable[object], operator: str) -> int:
    ids = normalize_ticket_ids(ticket_ids)
    if not ids:
        return 0
    placeholders, params = in_clause_for_ids(ids)
    with get_connection() as connection:
        rows = connection.execute(
            f"""
            SELECT id
            FROM tickets
            WHERE id IN ({placeholders})
              AND deleted_at IS NOT NULL
            ORDER BY id
            """,
            params,
        ).fetchall()
    target_ids = [int(row["id"]) for row in rows]
    for ticket_id in target_ids:
        hard_delete_ticket(ticket_id, operator)
    return len(target_ids)


def soft_delete_ticket(ticket_id: int, deleted_by: str, delete_reason: str = "") -> None:
    ticket = fetch_ticket(ticket_id)
    if not ticket:
        raise HTTPException(status_code=404, detail="工单不存在或已删除")
    reason = delete_reason.strip()
    timestamp = now_text()
    with get_connection() as connection:
        connection.execute(
            """
            UPDATE tickets
            SET deleted_at = ?, deleted_by = ?, delete_reason = ?, updated_at = ?
            WHERE id = ? AND deleted_at IS NULL
            """,
            (timestamp, deleted_by, reason, timestamp, ticket_id),
        )
        insert_ticket_log(connection, ticket_id, "移入回收站", reason, deleted_by, timestamp)
    try:
        create_notification_event(
            event_type="ticket_deleted",
            ticket_id=ticket_id,
            ticket_no=str(ticket.get("ticket_no") or ""),
            store_name=str(ticket.get("store_name") or ""),
            title=f"工单已删除：{ticket.get('ticket_no')}",
            content=reason,
            severity="warning",
            created_by=deleted_by,
        )
    except Exception:
        pass


def restore_ticket(ticket_id: int, operator: str) -> None:
    ticket = fetch_ticket(ticket_id, include_deleted=True)
    if not ticket or not ticket.get("deleted_at"):
        raise HTTPException(status_code=404, detail="回收站中没有该工单")
    timestamp = now_text()
    with get_connection() as connection:
        connection.execute(
            """
            UPDATE tickets
            SET deleted_at = NULL, deleted_by = NULL, delete_reason = NULL, updated_at = ?
            WHERE id = ?
            """,
            (timestamp, ticket_id),
        )
        insert_ticket_log(connection, ticket_id, "恢复工单", "", operator, timestamp)


def archive_ticket(ticket_id: int, operator: str, archive_reason: str = "") -> None:
    archived_count = bulk_archive_tickets([ticket_id], operator, archive_reason or "后台手动归档")
    if archived_count == 0:
        raise HTTPException(status_code=404, detail="工单不存在、已删除或已归档")


def unarchive_ticket(ticket_id: int, operator: str) -> None:
    unarchived_count = bulk_unarchive_tickets([ticket_id], operator)
    if unarchived_count == 0:
        raise HTTPException(status_code=404, detail="工单不存在、已删除或未归档")


def hard_delete_ticket(ticket_id: int, operator: str) -> None:
    ticket = fetch_ticket(ticket_id, include_deleted=True)
    if not ticket or not ticket.get("deleted_at"):
        raise HTTPException(status_code=404, detail="只能永久删除回收站中的工单")
    with get_connection() as connection:
        image_rows = connection.execute("SELECT image_path FROM ticket_images WHERE ticket_id = ?", (ticket_id,)).fetchall()
        file_rows = connection.execute("SELECT stored_filename FROM ticket_files WHERE ticket_id = ?", (ticket_id,)).fetchall()
        for table_name in (
            "ticket_images",
            "ticket_files",
            "ticket_supplements",
            "ticket_comments",
            "ticket_tasks",
            "ticket_participants",
            "ticket_logs",
            "notification_reads",
        ):
            if table_name == "notification_reads":
                connection.execute(
                    "DELETE FROM notification_reads WHERE event_id IN (SELECT id FROM notification_events WHERE ticket_id = ?)",
                    (ticket_id,),
                )
            else:
                connection.execute(f"DELETE FROM {table_name} WHERE ticket_id = ?", (ticket_id,))
        connection.execute("DELETE FROM notification_events WHERE ticket_id = ?", (ticket_id,))
        connection.execute("DELETE FROM ticket_stores WHERE ticket_id = ?", (ticket_id,))
        connection.execute("DELETE FROM ticket_brands WHERE ticket_id = ?", (ticket_id,))
        connection.execute("DELETE FROM tickets WHERE id = ?", (ticket_id,))
    for row in image_rows:
        target = resolve_upload_path(image_filename(str(row["image_path"] or "")))
        if target:
            try:
                target.unlink(missing_ok=True)
            except OSError:
                pass
    for row in file_rows:
        target = resolve_upload_path(safe_uploaded_name(str(row["stored_filename"] or "")))
        if target:
            try:
                target.unlink(missing_ok=True)
            except OSError:
                pass


def soft_delete_ticket_child(
    table_name: str,
    item_id: int,
    ticket_id: int,
    deleted_by: str,
    action: str,
    note_column: str,
    fallback_note: str = "",
) -> None:
    allowed_tables = {
        "ticket_comments",
        "ticket_tasks",
        "ticket_participants",
        "ticket_supplements",
    }
    if table_name not in allowed_tables:
        raise ValueError("不支持的删除对象")
    ticket = fetch_ticket(ticket_id)
    if not ticket:
        raise HTTPException(status_code=404, detail="工单不存在")
    timestamp = now_text()
    with get_connection() as connection:
        row = connection.execute(
            f"SELECT * FROM {table_name} WHERE id = ? AND ticket_id = ? AND deleted_at IS NULL",
            (item_id, ticket_id),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="记录不存在或已删除")
        note = str(row[note_column] if note_column in row.keys() else fallback_note or "").strip()
        if table_name == "ticket_participants":
            role = str(row["role"] or "").strip()
            note = f"{row['participant_name']} {role}".strip()
        connection.execute(
            f"""
            UPDATE {table_name}
            SET deleted_at = ?, deleted_by = ?, delete_reason = ?
            WHERE id = ? AND ticket_id = ?
            """,
            (timestamp, deleted_by, note[:200], item_id, ticket_id),
        )
        insert_ticket_log(connection, ticket_id, action, f"记录 ID {item_id} 已移入回收站", deleted_by, timestamp)


def add_ticket_attachments(
    ticket_id: int,
    ticket_no: str,
    prepared_images: List[Tuple[str, bytes]],
    prepared_files: List[PreparedFile],
    operator: str,
) -> Tuple[int, int]:
    saved_files: List[Path] = []
    timestamp = now_text()
    image_count = 0
    file_count = 0
    try:
        image_paths, saved_image_files = save_images(ticket_no, prepared_images)
        saved_files.extend(saved_image_files)
        file_records, saved_attachment_files = save_files(ticket_no, prepared_files)
        saved_files.extend(saved_attachment_files)
        image_count = len(image_paths)
        file_count = len(file_records)
        note = f"新增图片 {image_count} 张，文件 {file_count} 个"
        with get_connection() as connection:
            ticket_exists = connection.execute("SELECT 1 FROM tickets WHERE id = ?", (ticket_id,)).fetchone()
            if not ticket_exists:
                raise HTTPException(status_code=404, detail="工单不存在")
            for image_path in image_paths:
                connection.execute(
                    """
                    INSERT INTO ticket_images (ticket_id, image_path, uploaded_at, source, uploaded_by, supplement_id)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (ticket_id, image_path, timestamp, "admin_supplement", operator, None),
                )
            for file_record in file_records:
                connection.execute(
                    """
                    INSERT INTO ticket_files (
                        ticket_id, original_filename, stored_filename, file_path,
                        file_ext, file_size, uploaded_at, source, uploaded_by, supplement_id
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        ticket_id,
                        file_record["original_filename"],
                        file_record["stored_filename"],
                        file_record["file_path"],
                        file_record["file_ext"],
                        file_record["file_size"],
                        timestamp,
                        "admin_supplement",
                        operator,
                        None,
                    ),
                )
            connection.execute("UPDATE tickets SET updated_at = ? WHERE id = ?", (timestamp, ticket_id))
            connection.execute(
                """
                INSERT INTO ticket_logs (
                    ticket_id, action, old_status, new_status, old_assigned_to,
                    new_assigned_to, note, operator, created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (ticket_id, "补充附件", None, None, None, None, note, operator, timestamp),
            )
    except Exception:
        cleanup_saved_files(saved_files)
        raise
    return image_count, file_count


def notification_severity_for_urgency(urgency: str) -> str:
    if urgency == "当天必须处理":
        return "urgent"
    if urgency == "加急":
        return "warning"
    return "info"


def create_notification_event(
    event_type: str,
    ticket_id: Optional[int],
    ticket_no: str,
    store_name: str,
    title: str,
    content: str,
    severity: str = "info",
    created_by: str = "",
) -> int:
    severity_value = severity if severity in {"info", "warning", "urgent"} else "info"
    with get_connection() as connection:
        cursor = connection.execute(
            """
            INSERT INTO notification_events (
                event_type, ticket_id, ticket_no, store_name, title, content,
                severity, created_by, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                event_type,
                ticket_id,
                ticket_no,
                store_name,
                title,
                content,
                severity_value,
                created_by,
                now_text(),
            ),
        )
        return int(cursor.lastrowid)


def create_new_ticket_notification(ticket: Dict[str, object]) -> None:
    create_notification_event(
        event_type="new_ticket",
        ticket_id=int(ticket["id"]),
        ticket_no=str(ticket.get("ticket_no") or ""),
        store_name=str(ticket.get("store_name") or ""),
        title="新工单",
        content=f"{ticket.get('store_name')} 提交了 {ticket.get('request_type')} 工单",
        severity=notification_severity_for_urgency(str(ticket.get("urgency") or "")),
        created_by=f"门店:{ticket.get('submitter') or ''}",
    )


def create_store_supplement_notification(ticket: Dict[str, object], submitter: str) -> None:
    assigned_to = str(ticket.get("assigned_to") or "").strip()
    handler_hint = f"，请处理人 {assigned_to} 及时回复" if assigned_to else ""
    create_notification_event(
        event_type="store_supplement",
        ticket_id=int(ticket["id"]),
        ticket_no=str(ticket.get("ticket_no") or ""),
        store_name=str(ticket.get("store_name") or ""),
        title="门店补充资料",
        content=f"{ticket.get('store_name')} 为工单 {ticket.get('ticket_no')} 补充了资料{handler_hint}",
        severity="warning",
        created_by=f"门店:{submitter}",
    )


def create_need_store_supplement_notification(ticket: Dict[str, object], operator: str) -> None:
    create_notification_event(
        event_type="need_store_supplement",
        ticket_id=int(ticket["id"]),
        ticket_no=str(ticket.get("ticket_no") or ""),
        store_name=str(ticket.get("store_name") or ""),
        title="待门店补充",
        content=f"工单 {ticket.get('ticket_no')} 已标记为待门店补充",
        severity="warning",
        created_by=operator,
    )


def visibility_label(visibility: str) -> str:
    return "内部备注" if visibility == "internal" else "门店可见"


def author_type_label(author_type: str) -> str:
    return "门店" if author_type == "store" else "总部"


def create_collaboration_notification(
    ticket: Dict[str, object],
    event_type: str,
    title: str,
    content: str,
    created_by: str,
    severity: str = "info",
) -> int:
    return create_notification_event(
        event_type=event_type,
        ticket_id=int(ticket["id"]),
        ticket_no=str(ticket.get("ticket_no") or ""),
        store_name=str(ticket.get("store_name") or ""),
        title=title,
        content=content,
        severity=severity,
        created_by=created_by,
    )


def create_ticket_participant(
    ticket: Dict[str, object],
    participant_type: str,
    participant_name: str,
    role: str,
    operator: str,
) -> int:
    clean_type = participant_type.strip() or "team"
    clean_name = participant_name.strip()
    clean_role = role.strip()
    if not clean_name:
        raise ValueError("请填写协作人名称。")
    ticket_id = int(ticket["id"])
    timestamp = now_text()
    note = f"{clean_name}（{clean_role or clean_type}）"
    with get_connection() as connection:
        cursor = connection.execute(
            """
            INSERT INTO ticket_participants (
                ticket_id, participant_type, participant_name, role, created_at
            )
            VALUES (?, ?, ?, ?, ?)
            """,
            (ticket_id, clean_type, clean_name, clean_role, timestamp),
        )
        connection.execute("UPDATE tickets SET updated_at = ? WHERE id = ?", (timestamp, ticket_id))
        insert_ticket_log(connection, ticket_id, "新增协作人", note, operator, timestamp)
        participant_id = int(cursor.lastrowid)
    create_collaboration_notification(
        ticket,
        "ticket_participant",
        "新增协作人",
        f"工单 {ticket.get('ticket_no')} 新增协作人：{note}",
        operator,
    )
    return participant_id


def create_ticket_comment(
    ticket: Dict[str, object],
    author_type: str,
    author_name: str,
    content: str,
    visibility: str,
    operator: str,
) -> int:
    clean_author_type = author_type.strip() or "admin"
    clean_author_name = author_name.strip() or author_type_label(clean_author_type)
    clean_content = content.strip()
    clean_visibility = visibility.strip() or "public"
    if clean_visibility not in {"public", "internal"}:
        raise ValueError("评论可见范围不正确。")
    if clean_author_type == "store":
        clean_visibility = "public"
    if not clean_content:
        raise ValueError("请填写沟通内容。")
    ticket_id = int(ticket["id"])
    timestamp = now_text()
    action = "门店评论" if clean_author_type == "store" else "新增评论"
    note = f"{visibility_label(clean_visibility)}：{clean_content}"
    with get_connection() as connection:
        cursor = connection.execute(
            """
            INSERT INTO ticket_comments (
                ticket_id, author_type, author_name, content, visibility, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (ticket_id, clean_author_type, clean_author_name, clean_content, clean_visibility, timestamp),
        )
        connection.execute("UPDATE tickets SET updated_at = ? WHERE id = ?", (timestamp, ticket_id))
        insert_ticket_log(connection, ticket_id, action, note, operator, timestamp)
        comment_id = int(cursor.lastrowid)
    if clean_author_type == "store":
        title = "门店新增沟通"
        created_by = f"门店:{clean_author_name}"
        severity = "warning"
    elif clean_visibility == "internal":
        title = "新增内部备注"
        created_by = operator
        severity = "info"
    else:
        title = "新增门店可见回复"
        created_by = operator
        severity = "info"
    create_collaboration_notification(
        ticket,
        "ticket_comment",
        title,
        f"{author_type_label(clean_author_type)}为工单 {ticket.get('ticket_no')} 新增沟通记录",
        created_by,
        severity=severity,
    )
    return comment_id


def create_ticket_task(
    ticket: Dict[str, object],
    title: str,
    assignee: str,
    status: str,
    due_date: str,
    operator: str,
) -> int:
    clean_title = title.strip()
    clean_assignee = assignee.strip()
    clean_status = status.strip() or "待处理"
    clean_due_date = due_date.strip()
    if not clean_title:
        raise ValueError("请填写子任务标题。")
    if clean_status not in TASK_STATUSES:
        raise ValueError("子任务状态不正确。")
    ticket_id = int(ticket["id"])
    timestamp = now_text()
    completed_at = timestamp if clean_status == "已完成" else None
    note = f"{clean_title} / {clean_assignee or '未指定'} / {clean_status}"
    with get_connection() as connection:
        cursor = connection.execute(
            """
            INSERT INTO ticket_tasks (
                ticket_id, title, assignee, status, due_date,
                completed_at, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (ticket_id, clean_title, clean_assignee, clean_status, clean_due_date, completed_at, timestamp, timestamp),
        )
        connection.execute("UPDATE tickets SET updated_at = ? WHERE id = ?", (timestamp, ticket_id))
        insert_ticket_log(connection, ticket_id, "新增子任务", note, operator, timestamp)
        task_id = int(cursor.lastrowid)
    create_collaboration_notification(
        ticket,
        "ticket_task",
        "新增子任务",
        f"工单 {ticket.get('ticket_no')} 新增子任务：{clean_title}",
        operator,
    )
    return task_id


def update_ticket_task(
    ticket: Dict[str, object],
    task_id: int,
    title: str,
    assignee: str,
    status: str,
    due_date: str,
    operator: str,
) -> None:
    clean_title = title.strip()
    clean_assignee = assignee.strip()
    clean_status = status.strip() or "待处理"
    clean_due_date = due_date.strip()
    if not clean_title:
        raise ValueError("请填写子任务标题。")
    if clean_status not in TASK_STATUSES:
        raise ValueError("子任务状态不正确。")
    ticket_id = int(ticket["id"])
    timestamp = now_text()
    with get_connection() as connection:
        old_row = connection.execute(
            "SELECT * FROM ticket_tasks WHERE id = ? AND ticket_id = ?",
            (task_id, ticket_id),
        ).fetchone()
        if not old_row:
            raise HTTPException(status_code=404, detail="子任务不存在")
        old_task = dict(old_row)
        completed_at = old_task.get("completed_at")
        if clean_status == "已完成" and not completed_at:
            completed_at = timestamp
        elif clean_status != "已完成":
            completed_at = None
        connection.execute(
            """
            UPDATE ticket_tasks
            SET title = ?, assignee = ?, status = ?, due_date = ?,
                completed_at = ?, updated_at = ?
            WHERE id = ? AND ticket_id = ?
            """,
            (clean_title, clean_assignee, clean_status, clean_due_date, completed_at, timestamp, task_id, ticket_id),
        )
        connection.execute("UPDATE tickets SET updated_at = ? WHERE id = ?", (timestamp, ticket_id))
        note = f"{old_task.get('status') or '未设置'} → {clean_status}：{clean_title}"
        insert_ticket_log(connection, ticket_id, "更新子任务", note, operator, timestamp)
    create_collaboration_notification(
        ticket,
        "ticket_task",
        "子任务已更新",
        f"工单 {ticket.get('ticket_no')} 子任务已更新：{clean_title}",
        operator,
    )


def notification_actions(
    notification: Dict[str, object],
    ticket: Optional[Dict[str, object]],
    detail_url: str,
) -> List[Dict[str, object]]:
    event_id = int(notification.get("id") or 0)
    actions: List[Dict[str, object]] = []
    if detail_url:
        actions.append({"label": "查看工单", "url": detail_url, "method": "get"})
    if ticket and detail_url:
        assigned_to = str(ticket.get("assigned_to") or "").strip()
        if not assigned_to:
            actions.append({"label": "接单", "url": f"/admin/ticket/{int(ticket['id'])}/accept", "method": "post"})
        actions.append({"label": "回复", "url": f"{detail_url}#comments", "method": "get"})
    if event_id:
        actions.append(
            {
                "label": "标记已读",
                "url": f"/admin/api/notifications/{event_id}/read",
                "method": "post",
                "disabled": bool(notification.get("is_read")),
            }
        )
    return actions


def fetch_notifications(
    username: str,
    after_id: int = 0,
    limit: int = 20,
    unread_only: bool = False,
) -> List[Dict[str, object]]:
    clean_username = username.strip()
    clean_limit = min(max(int(limit or 20), 1), 100)
    clauses: List[str] = []
    params: List[object] = [clean_username]
    if after_id > 0:
        clauses.append("events.id > ?")
        params.append(after_id)
    if unread_only:
        clauses.append("reads.id IS NULL")
    where_sql = "WHERE " + " AND ".join(clauses) if clauses else ""
    params.append(clean_limit)
    with get_connection() as connection:
        rows = connection.execute(
            f"""
            SELECT
                events.id, events.event_type, events.ticket_id, events.ticket_no,
                events.store_name, events.title, events.content, events.severity,
                events.created_by, events.created_at,
                CASE WHEN reads.id IS NULL THEN 0 ELSE 1 END AS is_read
            FROM notification_events AS events
            LEFT JOIN notification_reads AS reads
                ON reads.event_id = events.id AND reads.username = ?
            {where_sql}
            ORDER BY events.id DESC
            LIMIT ?
            """,
            params,
        ).fetchall()
    notifications = [dict(row) for row in rows]
    for notification in notifications:
        ticket_id = notification.get("ticket_id")
        notification["is_read"] = bool(notification.get("is_read"))
        notification["detail_url"] = ""
        ticket: Optional[Dict[str, object]] = None
        if ticket_id:
            ticket = fetch_ticket(int(ticket_id))
            if ticket:
                notification["detail_url"] = f"/admin/ticket/{ticket_id}"
            else:
                deleted_ticket = fetch_ticket(int(ticket_id), include_deleted=True)
                if deleted_ticket and deleted_ticket.get("deleted_at"):
                    ticket_no = str(deleted_ticket.get("ticket_no") or notification.get("ticket_no") or "")
                    notification["detail_url"] = "/admin/trash" + (f"?{urlencode({'keyword': ticket_no})}" if ticket_no else "")
                    notification["content"] = "工单已在回收站"
        notification["actions"] = notification_actions(notification, ticket, str(notification.get("detail_url") or ""))
    return notifications


def count_unread_notifications(username: str) -> int:
    with get_connection() as connection:
        row = connection.execute(
            """
            SELECT COUNT(*) AS unread_count
            FROM notification_events AS events
            LEFT JOIN notification_reads AS reads
                ON reads.event_id = events.id AND reads.username = ?
            WHERE reads.id IS NULL
            """,
            (username.strip(),),
        ).fetchone()
    return int(row["unread_count"] or 0)


def latest_notification_id() -> int:
    with get_connection() as connection:
        row = connection.execute("SELECT MAX(id) AS latest_id FROM notification_events").fetchone()
    return int(row["latest_id"] or 0)


def mark_notification_read(username: str, event_id: int) -> bool:
    timestamp = now_text()
    with get_connection() as connection:
        event = connection.execute("SELECT id FROM notification_events WHERE id = ?", (event_id,)).fetchone()
        if not event:
            return False
        connection.execute(
            """
            INSERT OR IGNORE INTO notification_reads (event_id, username, read_at)
            VALUES (?, ?, ?)
            """,
            (event_id, username.strip(), timestamp),
        )
    return True


def mark_all_notifications_read(username: str) -> None:
    timestamp = now_text()
    with get_connection() as connection:
        connection.execute(
            """
            INSERT OR IGNORE INTO notification_reads (event_id, username, read_at)
            SELECT id, ?, ? FROM notification_events
            """,
            (username.strip(), timestamp),
        )


def create_store_supplement(
    ticket: Dict[str, object],
    submitter: str,
    note: str,
    prepared_images: List[Tuple[str, bytes]],
    prepared_files: List[PreparedFile],
    config: AppConfig,
) -> int:
    ticket_id = int(ticket["id"])
    ticket_no = str(ticket["ticket_no"])
    store_name = str(ticket.get("query_store_name") or ticket.get("store_name") or "").strip()
    old_status = str(ticket.get("status") or "")
    old_assigned_to = str(ticket.get("assigned_to") or "")
    timestamp = now_text()
    saved_files: List[Path] = []
    try:
        image_paths, saved_image_files = save_images(ticket_no, prepared_images)
        saved_files.extend(saved_image_files)
        file_records, saved_attachment_files = save_files(ticket_no, prepared_files)
        saved_files.extend(saved_attachment_files)
        image_count = len(image_paths)
        file_count = len(file_records)
        new_status = (
            config.supplement_status_after_store_update
            if old_status == "待门店补充"
            else old_status
        )
        closed_at = ticket.get("closed_at")
        if new_status == COMPLETED_STATUS and not closed_at:
            closed_at = timestamp
        elif old_status == COMPLETED_STATUS and new_status != COMPLETED_STATUS:
            closed_at = None
        note_parts = []
        clean_note = note.strip()
        if clean_note:
            note_parts.append(clean_note)
        note_parts.append(f"新增图片 {image_count} 张，文件 {file_count} 个")
        log_note = "；".join(note_parts)

        with get_connection() as connection:
            current = connection.execute("SELECT * FROM tickets WHERE id = ?", (ticket_id,)).fetchone()
            if not current:
                raise HTTPException(status_code=404, detail="工单不存在")
            store_match = connection.execute(
                """
                SELECT 1
                FROM ticket_stores
                WHERE ticket_id = ? AND store_name = ?
                """,
                (ticket_id, store_name),
            ).fetchone()
            if not store_match:
                raise HTTPException(status_code=403, detail="门店不匹配")
            cursor = connection.execute(
                """
                INSERT INTO ticket_supplements (
                    ticket_id, store_name, submitter, note, image_count, file_count, created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (ticket_id, store_name, submitter, clean_note, image_count, file_count, timestamp),
            )
            supplement_id = int(cursor.lastrowid)
            for image_path in image_paths:
                connection.execute(
                    """
                    INSERT INTO ticket_images (ticket_id, image_path, uploaded_at, source, uploaded_by, supplement_id)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (ticket_id, image_path, timestamp, "store_supplement", submitter, supplement_id),
                )
            for file_record in file_records:
                connection.execute(
                    """
                    INSERT INTO ticket_files (
                        ticket_id, original_filename, stored_filename, file_path,
                        file_ext, file_size, uploaded_at, source, uploaded_by, supplement_id
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        ticket_id,
                        file_record["original_filename"],
                        file_record["stored_filename"],
                        file_record["file_path"],
                        file_record["file_ext"],
                        file_record["file_size"],
                        timestamp,
                        "store_supplement",
                        submitter,
                        supplement_id,
                    ),
                )
            connection.execute(
                "UPDATE tickets SET status = ?, closed_at = ?, updated_at = ? WHERE id = ?",
                (new_status, closed_at, timestamp, ticket_id),
            )
            connection.execute(
                """
                INSERT INTO ticket_logs (
                    ticket_id, action, old_status, new_status, old_assigned_to,
                    new_assigned_to, note, operator, created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    ticket_id,
                    "门店补充资料",
                    old_status,
                    new_status,
                    old_assigned_to,
                    old_assigned_to,
                    log_note,
                    f"门店:{submitter}",
                    timestamp,
                ),
            )
        return supplement_id
    except Exception:
        cleanup_saved_files(saved_files)
        raise


def build_public_query_params(filters: Dict[str, str], page: Optional[int] = None) -> str:
    params = {key: value for key, value in filters.items() if str(value or "").strip()}
    if page is not None:
        params["page"] = str(page)
    return urlencode(params)


def normalize_store_query_filters(raw_filters: Dict[str, str], config: AppConfig) -> Dict[str, str]:
    filters = {key: str(value or "").strip() for key, value in raw_filters.items()}
    store_name = filters.get("store_name", "")
    if store_name and store_name in config.stores:
        optional_keys = ["ticket_no", "submitter", "keyword", "status", "date_start", "date_end"]
        if not any(filters.get(key) for key in optional_keys):
            filters["date_start"] = (datetime.now().date() - timedelta(days=config.store_query_default_days)).isoformat()
    return filters


def fetch_store_query_page(
    filters: Dict[str, str],
    config: AppConfig,
    page: int,
) -> Tuple[List[Dict[str, object]], Dict[str, int]]:
    filters = dict(filters)
    filters["__ticket_scope"] = "store"
    page_size = config.store_query_page_size
    total_count = count_tickets(filters)
    total_pages = max((total_count + page_size - 1) // page_size, 1)
    current_page = min(max(page, 1), total_pages)
    offset = (current_page - 1) * page_size
    tickets = fetch_tickets(filters, "newest", config, limit=page_size, offset=offset)
    return tickets, {
        "current_page": current_page,
        "total_pages": total_pages,
        "total_count": total_count,
        "page_size": page_size,
        "has_prev": 1 if current_page > 1 else 0,
        "has_next": 1 if current_page < total_pages else 0,
        "prev_page": max(current_page - 1, 1),
        "next_page": min(current_page + 1, total_pages),
    }


TEST_DATA_KEYWORDS = ("codex", "smoke", "test", "测试")


def fetch_archived_tickets(filters: Optional[Dict[str, str]] = None, sort: str = "newest") -> List[Dict[str, object]]:
    scoped_filters = dict(filters or {})
    scoped_filters["__ticket_scope"] = "archive"
    return fetch_tickets(scoped_filters, sort, load_app_config())


def fetch_deleted_tickets(filters: Optional[Dict[str, str]] = None, sort: str = "newest") -> List[Dict[str, object]]:
    scoped_filters = dict(filters or {})
    scoped_filters["__ticket_scope"] = "deleted"
    return fetch_tickets(scoped_filters, sort, load_app_config())


def fetch_deleted_embedded_pages() -> List[Dict[str, object]]:
    return fetch_embedded_pages(deleted_only=True)


def ticket_matches_test_keyword(ticket: Dict[str, object]) -> bool:
    searchable_values = [
        ticket.get("ticket_no"),
        ticket.get("store_name"),
        ticket.get("submitter"),
        ticket.get("request_type"),
        ticket.get("brand"),
        ticket.get("product_name"),
        ticket.get("sku_barcode"),
        ticket.get("description"),
        ticket.get("handler_note"),
    ]
    haystack = " ".join(str(value or "") for value in searchable_values).lower()
    return any(keyword in haystack for keyword in TEST_DATA_KEYWORDS)


def cleanup_filters_from_form(
    store_name: str = "",
    submitter: str = "",
    keyword: str = "",
    date_start: str = "",
    date_end: str = "",
    incomplete_only: str = "",
    only_test: str = "",
) -> Dict[str, str]:
    return {
        "store_name": store_name.strip(),
        "submitter": submitter.strip(),
        "keyword": keyword.strip(),
        "date_start": date_start.strip(),
        "date_end": date_end.strip(),
        "incomplete_only": "1" if incomplete_only in {"1", "true", "on", "yes"} else "",
        "only_test": "1" if only_test in {"1", "true", "on", "yes"} else "",
    }


def cleanup_ticket_filters(filters: Dict[str, str]) -> Dict[str, str]:
    return {
        "store_name": str(filters.get("store_name") or "").strip(),
        "submitter": str(filters.get("submitter") or "").strip(),
        "keyword": str(filters.get("keyword") or "").strip(),
        "date_start": str(filters.get("date_start") or "").strip(),
        "date_end": str(filters.get("date_end") or "").strip(),
    }


def fetch_cleanup_candidates(filters: Dict[str, str]) -> List[Dict[str, object]]:
    tickets = fetch_tickets(cleanup_ticket_filters(filters), "newest", load_app_config())
    if filters.get("incomplete_only"):
        tickets = [ticket for ticket in tickets if ticket.get("status") != COMPLETED_STATUS]
    if filters.get("only_test"):
        tickets = [ticket for ticket in tickets if ticket_matches_test_keyword(ticket)]
    return tickets


def percent_value(count: int, total: int) -> int:
    if total <= 0:
        return 0
    return round(count * 100 / total)


def counter_rows(counter: Counter[str], total: int, limit: Optional[int] = None) -> List[Dict[str, object]]:
    rows = [
        {"label": label or "未填写", "count": count, "percent": percent_value(count, total)}
        for label, count in counter.most_common(limit)
    ]
    return rows


def fetch_dashboard_stats(filters: Dict[str, str]) -> Dict[str, object]:
    config = load_app_config()
    tickets = fetch_tickets(filters, "newest", config)
    total = len(tickets)
    today_prefix = datetime.now().date().isoformat()
    status_counter: Counter[str] = Counter(str(ticket.get("status") or "未填写") for ticket in tickets)
    type_counter: Counter[str] = Counter(str(ticket.get("request_type") or "未填写") for ticket in tickets)
    store_counter: Counter[str] = Counter()
    for ticket in tickets:
        store_names = ticket.get("store_names") or split_multi_value_text(ticket.get("store_name"))
        for store_name in store_names or ["未填写"]:
            store_counter[str(store_name or "未填写")] += 1
    handler_counter: Counter[str] = Counter(str(ticket.get("assigned_to") or "未指定") or "未指定" for ticket in tickets)
    urgency_counter: Counter[str] = Counter(str(ticket.get("urgency") or "未填写") for ticket in tickets)
    today_new_count = sum(1 for ticket in tickets if str(ticket.get("created_at") or "").startswith(today_prefix))
    overdue_count = sum(1 for ticket in tickets if ticket.get("due_status") == "已超时")
    due_today_count = sum(1 for ticket in tickets if ticket.get("due_status") == "今日到期")
    image_ticket_count = sum(1 for ticket in tickets if int(ticket.get("image_count") or 0) > 0)
    file_ticket_count = sum(1 for ticket in tickets if int(ticket.get("file_count") or 0) > 0)
    attachment_ticket_count = sum(
        1
        for ticket in tickets
        if int(ticket.get("image_count") or 0) > 0 or int(ticket.get("file_count") or 0) > 0
    )
    no_attachment_count = sum(
        1
        for ticket in tickets
        if int(ticket.get("image_count") or 0) == 0 and int(ticket.get("file_count") or 0) == 0
    )
    supplement_count = 0
    if tickets:
        ticket_ids = [int(ticket["id"]) for ticket in tickets]
        placeholders = ",".join("?" for _ in ticket_ids)
        with get_connection() as connection:
            row = connection.execute(
                f"SELECT COUNT(*) AS total FROM ticket_supplements WHERE ticket_id IN ({placeholders})",
                ticket_ids,
            ).fetchone()
        supplement_count = int(row["total"] or 0)
    cards = [
        {"label": "总工单数", "count": total, "percent": 100 if total else 0},
        {"label": "待处理数", "count": status_counter.get("待处理", 0), "percent": percent_value(status_counter.get("待处理", 0), total)},
        {"label": "处理中数", "count": status_counter.get("处理中", 0), "percent": percent_value(status_counter.get("处理中", 0), total)},
        {"label": "待门店补充数", "count": status_counter.get("待门店补充", 0), "percent": percent_value(status_counter.get("待门店补充", 0), total)},
        {"label": "已完成数", "count": status_counter.get("已完成", 0), "percent": percent_value(status_counter.get("已完成", 0), total)},
        {"label": "已驳回数", "count": status_counter.get("已驳回", 0), "percent": percent_value(status_counter.get("已驳回", 0), total)},
        {"label": "超时工单数", "count": overdue_count, "percent": percent_value(overdue_count, total)},
        {"label": "今日到期数", "count": due_today_count, "percent": percent_value(due_today_count, total)},
    ]
    attachment_rows = [
        {"label": "有图片工单数", "count": image_ticket_count, "percent": percent_value(image_ticket_count, total)},
        {"label": "有文件工单数", "count": file_ticket_count, "percent": percent_value(file_ticket_count, total)},
        {"label": "无附件工单数", "count": no_attachment_count, "percent": percent_value(no_attachment_count, total)},
    ]
    type_structure = [
        {"label": label, "count": type_counter.get(label, 0), "percent": percent_value(type_counter.get(label, 0), total)}
        for label in config.request_types
    ]
    return {
        "total": total,
        "today_new_count": today_new_count,
        "pending_count": status_counter.get("待处理", 0),
        "processing_count": status_counter.get("处理中", 0),
        "need_supplement_count": status_counter.get("待门店补充", 0),
        "completed_count": status_counter.get("已完成", 0),
        "overdue_count": overdue_count,
        "today_urgent_count": urgency_counter.get("当天必须处理", 0),
        "attachment_ticket_count": attachment_ticket_count,
        "store_supplement_count": supplement_count,
        "recent_tickets": tickets[:5],
        "cards": cards,
        "by_request_type": type_structure,
        "by_store": counter_rows(store_counter, total, limit=10),
        "by_handler": counter_rows(handler_counter, total),
        "by_status": counter_rows(status_counter, total),
        "by_urgency": counter_rows(urgency_counter, total),
        "attachments": attachment_rows,
    }


def build_excel(tickets: List[Dict[str, object]]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "门店需求工单"

    sheet.append(EXCEL_HEADERS)
    header_fill = PatternFill(fill_type="solid", fgColor="E8EEF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    image_map: Dict[int, List[str]] = {}
    file_name_map: Dict[int, List[str]] = {}
    file_url_map: Dict[int, List[str]] = {}
    if tickets:
        ids = [int(ticket["id"]) for ticket in tickets]
        placeholders = ",".join("?" for _ in ids)
        with get_connection() as connection:
            rows = connection.execute(
                f"SELECT ticket_id, image_path FROM ticket_images WHERE ticket_id IN ({placeholders}) ORDER BY id",
                ids,
            ).fetchall()
            file_rows = connection.execute(
                f"""
                SELECT id, ticket_id, original_filename
                FROM ticket_files
                WHERE ticket_id IN ({placeholders})
                ORDER BY id
                """,
                ids,
            ).fetchall()
        for row in rows:
            image_map.setdefault(int(row["ticket_id"]), []).append(protected_upload_url(str(row["image_path"])))
        for row in file_rows:
            ticket_id = int(row["ticket_id"])
            file_name_map.setdefault(ticket_id, []).append(str(row["original_filename"]))
            file_url_map.setdefault(ticket_id, []).append(protected_file_url(row["id"]))

    for ticket in tickets:
        sheet.append(
            [
                ticket.get("ticket_no"),
                ticket.get("created_at"),
                ticket.get("store_name"),
                ticket.get("submitter"),
                ticket.get("request_type"),
                ticket.get("urgency"),
                ticket.get("brand") or "",
                ticket.get("product_name") or "",
                ticket.get("sku_barcode") or "",
                ticket.get("quantity") if ticket.get("quantity") is not None else "",
                ticket.get("description"),
                "; ".join(image_map.get(int(ticket["id"]), [])),
                "; ".join(file_name_map.get(int(ticket["id"]), [])),
                "; ".join(file_url_map.get(int(ticket["id"]), [])),
                ticket.get("expected_finish_date") or "",
                ticket.get("status"),
                ticket.get("assigned_to") or "",
                ticket.get("handler_note") or "",
                ticket.get("closed_at") or "",
                ticket.get("updated_at"),
                processing_hours(ticket),
                overdue_text(ticket),
                ticket.get("due_status") or due_status_label(ticket),
            ]
        )

    widths = [22, 20, 16, 14, 14, 14, 16, 20, 20, 10, 38, 42, 26, 34, 18, 14, 14, 32, 20, 20, 16, 12, 14]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def create_app() -> FastAPI:
    load_env_file()
    warn_if_insecure_production_session()
    ensure_directories()
    init_db()

    app = FastAPI(title=load_app_config().app_name)
    templates = Jinja2Templates(directory=str(TEMPLATES_DIR))
    templates.env.globals["embedded_nav_pages"] = fetch_enabled_embedded_pages
    templates.env.globals["asset_version"] = current_asset_version()
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

    @app.get("/__version")
    def version_info() -> Dict[str, object]:
        config = load_app_config()
        return {
            "app": "store-request-tool",
            "name": "止痒 ERP",
            "main_file": str(BASE_DIR / "main.py"),
            "route_count": len(app.routes),
            "git_commit": current_git_commit(),
            "port": config.port,
            "required_missing_routes": required_missing_routes(app),
            "started_at": APP_STARTED_AT,
            "urls": public_access_urls(config.port),
        }

    @app.get("/healthz")
    def healthz() -> Dict[str, object]:
        database_ok = False
        try:
            with get_connection() as connection:
                connection.execute("SELECT 1").fetchone()
            database_ok = True
        except sqlite3.Error:
            database_ok = False
        upload_dir_ok = get_upload_dir().is_dir()
        try:
            get_embedded_pages_dir().mkdir(parents=True, exist_ok=True)
            embedded_pages_dir_ok = get_embedded_pages_dir().is_dir()
        except OSError:
            embedded_pages_dir_ok = False
        return {
            "ok": bool(database_ok and upload_dir_ok and embedded_pages_dir_ok),
            "database": database_ok,
            "upload_dir": upload_dir_ok,
            "embedded_pages_dir": embedded_pages_dir_ok,
            "route_count": len(app.routes),
        }

    @app.get("/api/handlers")
    def handlers_api(_admin: str = Depends(require_admin)) -> Dict[str, List[str]]:
        return {"handlers": load_app_config().handlers}

    @app.exception_handler(404)
    def admin_not_found_handler(request: Request, exc: HTTPException):
        is_admin_path = request.url.path.startswith("/admin")
        admin = current_admin_username(request) if is_admin_path else ""
        if is_admin_path and not admin:
            return RedirectResponse(url=login_redirect_location(request), status_code=303)
        admin_routes = [
            {"method": method, "path": path}
            for method, path in registered_route_pairs(app)
            if path.startswith("/admin")
        ]
        return templates.TemplateResponse(
            request,
            "not_found.html",
            {
                "request": request,
                "admin_user": admin,
                "csrf_token": current_csrf_token(request) if is_admin_path else "",
                "request_path": request.url.path,
                "request_method": request.method,
                "error_detail": str(getattr(exc, "detail", "") or ""),
                "admin_routes": admin_routes,
                "is_admin_path": is_admin_path,
                "recommended_urls": recommended_access_url_entries(),
            },
            status_code=404,
        )

    def make_legacy_admin_redirect(target_path: str):
        def legacy_admin_redirect(_admin: str = Depends(require_admin)) -> RedirectResponse:
            return RedirectResponse(url=target_path, status_code=303)

        return legacy_admin_redirect

    for legacy_path, target_path in LEGACY_ADMIN_REDIRECTS.items():
        app.add_api_route(
            legacy_path,
            make_legacy_admin_redirect(target_path),
            methods=["GET"],
            name=f"legacy_admin_redirect_{legacy_path.strip('/').replace('/', '_').replace('-', '_')}",
            include_in_schema=False,
        )

    def make_legacy_public_redirect(target_path: str):
        def legacy_public_redirect() -> RedirectResponse:
            return RedirectResponse(url=target_path, status_code=303)

        return legacy_public_redirect

    for legacy_path, target_path in LEGACY_PUBLIC_REDIRECTS.items():
        app.add_api_route(
            legacy_path,
            make_legacy_public_redirect(target_path),
            methods=["GET"],
            name=f"legacy_public_redirect_{legacy_path.strip('/').replace('/', '_').replace('-', '_')}",
            include_in_schema=False,
        )

    def ticket_create_form_context(
        config: AppConfig,
        values: Optional[Dict[str, object]] = None,
        error: str = "",
    ) -> Dict[str, object]:
        return {
            "stores": config.stores,
            "request_types": config.request_types,
            "urgency_levels": config.urgency_levels,
            "brands": config.brands,
            "image_accept": ",".join(f".{extension}" for extension in config.allowed_image_extensions),
            "file_accept": ",".join(f".{extension}" for extension in config.allowed_file_extensions),
            "max_image_count": config.max_image_count,
            "max_total_upload_mb": config.max_total_upload_mb,
            "max_image_mb": config.max_image_mb,
            "allowed_file_extensions": config.allowed_file_extensions,
            "max_file_count": config.max_file_count,
            "max_file_mb": config.max_file_mb,
            "max_total_file_upload_mb": config.max_total_file_upload_mb,
            "error": error,
            "values": values or {},
            "request_type_rules_json": json.dumps(load_request_type_rules(), ensure_ascii=False),
        }

    def render_submit_form(
        request: Request,
        status_code: int = 200,
        error: str = "",
        values: Optional[Dict[str, object]] = None,
    ) -> HTMLResponse:
        config = load_app_config()
        context = {"request": request}
        context.update(ticket_create_form_context(config, values=values, error=error))
        return templates.TemplateResponse(
            request,
            "submit.html",
            context,
            status_code=status_code,
        )

    def render_login_form(
        request: Request,
        status_code: int = 200,
        error: str = "",
        username: str = "",
        next_url: str = "/admin",
        logged_out: str = "",
    ) -> HTMLResponse:
        return templates.TemplateResponse(
            request,
            "login.html",
            {
                "request": request,
                "error": error,
                "username": username,
                "next_url": safe_admin_return_url(next_url),
                "logged_out": logged_out,
            },
            status_code=status_code,
        )

    def render_ticket_detail(
        request: Request,
        ticket_id: int,
        admin: str,
        saved: str = "",
        attachments_saved: str = "",
        upload_error: str = "",
        return_url: str = "",
        status_code: int = 200,
    ) -> HTMLResponse:
        ticket = fetch_ticket(ticket_id)
        if not ticket:
            raise HTTPException(status_code=404, detail="工单不存在")
        images = fetch_ticket_images(ticket_id)
        files = fetch_ticket_files(ticket_id)
        logs = fetch_ticket_logs(ticket_id)
        supplements = fetch_store_ticket_supplements(ticket_id)
        participants = fetch_ticket_participants(ticket_id)
        comments = fetch_ticket_comments(ticket_id)
        tasks = fetch_ticket_tasks(ticket_id)
        open_task_count = sum(1 for task in tasks if str(task.get("status") or "") != COMPLETED_STATUS)
        config = load_app_config()
        return templates.TemplateResponse(
            request,
            "ticket_detail.html",
            {
                "request": request,
                "ticket": ticket,
                "images": images,
                "files": files,
                "statuses": config.statuses,
                "handlers": detail_handler_options(config, ticket, admin),
                "logs": logs,
                "supplements": supplements,
                "participants": participants,
                "comments": comments,
                "tasks": tasks,
                "open_task_count": open_task_count,
                "close_prompt": should_prompt_close_ticket(ticket, tasks),
                "task_statuses": TASK_STATUSES,
                "saved": saved,
                "attachments_saved": attachments_saved,
                "upload_error": upload_error,
                "return_url": safe_admin_return_url(return_url),
                "admin_user": admin,
                "image_accept": ",".join(f".{extension}" for extension in config.allowed_image_extensions),
                "file_accept": ",".join(f".{extension}" for extension in config.allowed_file_extensions),
                "max_image_count": config.max_image_count,
                "max_image_mb": config.max_image_mb,
                "max_file_count": config.max_file_count,
                "max_file_mb": config.max_file_mb,
                "max_total_file_upload_mb": config.max_total_file_upload_mb,
                "csrf_token": current_csrf_token(request),
            },
            status_code=status_code,
        )

    def render_embedded_pages_admin(
        request: Request,
        admin: str,
        error: str = "",
        status_code: int = 200,
    ) -> HTMLResponse:
        config = load_app_config()
        return templates.TemplateResponse(
            request,
            "embedded_pages_admin.html",
            {
                "request": request,
                "pages": fetch_embedded_pages(),
                "error": error,
                "max_embedded_html_mb": config.max_embedded_html_mb,
                "max_embedded_zip_mb": config.max_embedded_zip_mb,
                "admin_user": admin,
                "csrf_token": current_csrf_token(request),
            },
            status_code=status_code,
        )

    def render_archive_page(
        request: Request,
        admin: str,
        filters: Optional[Dict[str, str]] = None,
        sort: str = "newest",
        page: int = 1,
        unarchived_count: int = 0,
        deleted_count: int = 0,
        error: str = "",
    ) -> HTMLResponse:
        config = load_app_config()
        active_filters = dict(filters or {})
        active_filters["__ticket_scope"] = "archive"
        tickets, pagination = fetch_ticket_page(active_filters, sort, config, page)
        return_url = safe_admin_return_url(request_path_with_query(request))
        for ticket in tickets:
            ticket["detail_url"] = build_ticket_detail_url(int(ticket["id"]), return_url)
        export_query = build_query_params(active_filters, sort)
        prev_query = build_query_params(active_filters, sort, pagination["prev_page"])
        next_query = build_query_params(active_filters, sort, pagination["next_page"])
        return templates.TemplateResponse(
            request,
            "archive.html",
            {
                "request": request,
                "tickets": tickets,
                "stores": config.stores,
                "request_types": config.request_types,
                "urgency_levels": config.urgency_levels,
                "statuses": config.statuses,
                "handlers": config.handlers,
                "due_status_options": DUE_STATUS_OPTIONS,
                "filters": active_filters,
                "sort": sort,
                "pagination": pagination,
                "export_url": "/admin/archive/export" + (f"?{export_query}" if export_query else ""),
                "prev_page_url": "/admin/archive" + (f"?{prev_query}" if prev_query else ""),
                "next_page_url": "/admin/archive" + (f"?{next_query}" if next_query else ""),
                "unarchived_count": unarchived_count,
                "deleted_count": deleted_count,
                "error": error,
                "admin_user": admin,
                "csrf_token": current_csrf_token(request),
            },
        )

    def render_trash_page(
        request: Request,
        admin: str,
        filters: Optional[Dict[str, str]] = None,
        sort: str = "newest",
        page: int = 1,
        restored_count: int = 0,
        hard_deleted_count: int = 0,
        error: str = "",
    ) -> HTMLResponse:
        config = load_app_config()
        active_filters = dict(filters or {})
        active_filters["__ticket_scope"] = "deleted"
        deleted_tickets, pagination = fetch_ticket_page(active_filters, sort, config, page)
        prev_query = build_query_params(active_filters, sort, pagination["prev_page"])
        next_query = build_query_params(active_filters, sort, pagination["next_page"])
        return templates.TemplateResponse(
            request,
            "trash.html",
            {
                "request": request,
                "deleted_tickets": deleted_tickets,
                "deleted_embedded_pages": fetch_deleted_embedded_pages(),
                "stores": config.stores,
                "request_types": config.request_types,
                "urgency_levels": config.urgency_levels,
                "statuses": config.statuses,
                "handlers": config.handlers,
                "due_status_options": DUE_STATUS_OPTIONS,
                "filters": active_filters,
                "sort": sort,
                "pagination": pagination,
                "prev_page_url": "/admin/trash" + (f"?{prev_query}" if prev_query else ""),
                "next_page_url": "/admin/trash" + (f"?{next_query}" if next_query else ""),
                "restored_count": restored_count,
                "hard_deleted_count": hard_deleted_count,
                "error": error,
                "admin_user": admin,
                "csrf_token": current_csrf_token(request),
            },
        )

    def render_employees_page(
        request: Request,
        admin: str,
        store_name: str = "",
        status: str = "",
        scope: str = "active",
        page: int = 1,
        error: str = "",
        success: str = "",
        status_code: int = 200,
    ) -> HTMLResponse:
        config = load_app_config()
        selected_scope = normalize_employee_record_scope(scope)
        all_employees = fetch_employees(store_name, status, selected_scope)
        grouped_all_employees = group_employees_by_role(all_employees)
        sorted_all_employees = [
            employee
            for group in grouped_all_employees
            for employee in list(group.get("employees") or [])
        ]
        page_size = 24
        total_count = len(sorted_all_employees)
        total_pages = max(1, (total_count + page_size - 1) // page_size)
        current_page = min(max(page, 1), total_pages)
        start_index = (current_page - 1) * page_size
        employees = sorted_all_employees[start_index : start_index + page_size]
        employee_role_groups = group_employees_by_role(employees)
        filters = {"store_name": store_name.strip(), "status": status.strip(), "scope": selected_scope}
        base_params = {key: value for key, value in filters.items() if value}
        return templates.TemplateResponse(
            request,
            "employees.html",
            {
                "request": request,
                "employees": employees,
                "employee_role_groups": employee_role_groups,
                "stores": config.stores,
                "statuses": EMPLOYEE_STATUSES,
                "filters": filters,
                "pagination": {
                    "current_page": current_page,
                    "total_pages": total_pages,
                    "total_count": total_count,
                    "page_size": page_size,
                    "prev_url": "/admin/employees?"
                    + urlencode({**base_params, "page": max(current_page - 1, 1)}),
                    "next_url": "/admin/employees?"
                    + urlencode({**base_params, "page": min(current_page + 1, total_pages)}),
                },
                "error": error,
                "success": success,
                "admin_user": admin,
                "csrf_token": current_csrf_token(request),
            },
            status_code=status_code,
        )

    def render_shift_types_page(
        request: Request,
        admin: str,
        error: str = "",
        success: str = "",
        store_names: Optional[List[str]] = None,
        status_filter: str = "all",
        global_scope: str = "all",
        data_scope: str = "current",
        status_code: int = 200,
    ) -> HTMLResponse:
        config = load_app_config()
        selected_stores, _is_all_stores, invalid_store = normalize_schedule_store_filter(
            store_names,
            "",
            config,
            default_to_first=False,
        )
        normalized_status = status_filter if status_filter in {"all", "active", "inactive"} else "all"
        normalized_global_scope = global_scope if global_scope in {"all", "global", "store"} else "all"
        normalized_data_scope = normalize_shift_data_scope(data_scope)
        hours_map = fetch_store_business_hours()
        active_only = normalized_status == "active" or normalized_data_scope == "active"
        shift_types = fetch_shift_types(
            include_inactive=normalized_status != "active",
            active_only=active_only,
            store_names=selected_stores or None,
            global_scope=normalized_global_scope,
            data_scope=normalized_data_scope,
        )
        if normalized_status == "inactive" or normalized_data_scope == "inactive":
            shift_types = [shift for shift in shift_types if int(shift.get("is_active") or 0) == 0]
        form_query_items: List[Tuple[str, str]] = []
        form_query_items.extend(("store_names", store) for store in selected_stores)
        if normalized_status != "all":
            form_query_items.append(("status", normalized_status))
        if normalized_global_scope != "all":
            form_query_items.append(("global_scope", normalized_global_scope))
        if normalized_data_scope != "current":
            form_query_items.append(("data_scope", normalized_data_scope))
        combined_error = combine_error_messages(error, "门店筛选参数无效，已忽略。" if invalid_store else "")
        return templates.TemplateResponse(
            request,
            "shift_types.html",
            {
                "request": request,
                "shift_types": shift_types,
                "stores": config.stores,
                "selected_store_names": selected_stores,
                "selected_store_count": len(selected_stores) if selected_stores else len(config.stores),
                "status_filter": normalized_status,
                "global_scope": normalized_global_scope,
                "data_scope": normalized_data_scope,
                "shift_data_scopes": SHIFT_DATA_SCOPES,
                "business_hours": hours_map,
                "selected_business_store": selected_stores[0] if len(selected_stores) == 1 else "",
                "selected_business_hours": hours_map.get(selected_stores[0], {}) if len(selected_stores) == 1 else {},
                "shift_types_form_query": urlencode(form_query_items),
                "error": combined_error,
                "success": success,
                "admin_user": admin,
                "csrf_token": current_csrf_token(request),
            },
            status_code=status_code,
        )

    def render_schedules_page(
        request: Request,
        admin: str,
        store_name: str = "",
        month: str = "",
        employee_status: str = "",
        view: str = "calendar",
        show_cross_store: bool = False,
        scope: str = "store",
        employee_scope: str = "all",
        employee_ids: Optional[List[int]] = None,
        shift_type_id: int = 0,
        store_names: Optional[List[str]] = None,
        employee_statuses: Optional[List[str]] = None,
        shift_type_ids: Optional[List[int]] = None,
        include_custom_shift: bool = False,
        invalid_store_filter: bool = False,
        saved: int = 0,
        saved_count: int = 0,
        created_count: int = 0,
        updated_count: int = 0,
        skipped_count: int = 0,
        deleted: int = 0,
        error: str = "",
        status_code: int = 200,
    ) -> HTMLResponse:
        config = load_app_config()
        selected_view = normalize_schedule_view(view)
        default_to_first = not (scope == "all" or (selected_view == "store-summary" and not store_name.strip() and store_names is None))
        selected_store_names, is_all_stores, invalid_store = normalize_schedule_store_filter(
            store_names,
            store_name,
            config,
            default_to_first=default_to_first,
        )
        selected_scope = "all" if is_all_stores or scope == "all" or (selected_view == "store-summary" and not selected_store_names) else "store"
        selected_store = selected_store_names[0] if len(selected_store_names) == 1 else ""
        combined_error = combine_error_messages(
            error,
            "门店筛选参数无效，已忽略。" if (invalid_store_filter or invalid_store) else "",
        )
        context = fetch_schedule_context(
            selected_store,
            month,
            employee_status,
            config,
            selected_view,
            show_cross_store,
            selected_scope,
            employee_scope,
            employee_ids,
            shift_type_id,
            store_names=selected_store_names,
            employee_statuses=employee_statuses,
            shift_type_ids=shift_type_ids,
            include_custom_shift=include_custom_shift,
            is_all_stores=selected_scope == "all",
        )
        return templates.TemplateResponse(
            request,
            "schedules.html",
            {
                "request": request,
                "context": context,
                "employee_statuses": EMPLOYEE_STATUSES,
                "saved": max(saved, 0),
                "saved_count": max(saved_count, 0),
                "created_count": max(created_count, 0),
                "updated_count": max(updated_count, 0),
                "skipped_count": max(skipped_count, 0),
                "deleted": max(deleted, 0),
                "error": combined_error,
                "admin_user": admin,
                "csrf_token": current_csrf_token(request),
            },
            status_code=status_code,
        )

    def render_store_schedule_page(
        request: Request,
        store_name: str = "",
        month: str = "",
        error: str = "",
        status_code: int = 200,
    ) -> HTMLResponse:
        config = load_app_config()
        selected_month = normalize_month(month)
        clean_store = store_name.strip()
        rows = fetch_schedule_rows(clean_store, selected_month) if clean_store else []
        return templates.TemplateResponse(
            request,
            "store_schedule.html",
            {
                "request": request,
                "stores": config.stores,
                "store_name": clean_store,
                "month": selected_month,
                "rows": rows,
                "error": error,
            },
            status_code=status_code,
        )

    def render_cleanup_page(
        request: Request,
        admin: str,
        filters: Optional[Dict[str, str]] = None,
        preview_tickets: Optional[List[Dict[str, object]]] = None,
        previewed: bool = False,
        deleted_count: int = 0,
        error: str = "",
        status_code: int = 200,
    ) -> HTMLResponse:
        config = load_app_config()
        active_filters = filters or cleanup_filters_from_form()
        return templates.TemplateResponse(
            request,
            "cleanup.html",
            {
                "request": request,
                "stores": config.stores,
                "filters": active_filters,
                "preview_tickets": preview_tickets or [],
                "previewed": previewed,
                "deleted_count": deleted_count,
                "error": error,
                "admin_user": admin,
                "csrf_token": current_csrf_token(request),
            },
            status_code=status_code,
        )

    def render_query_page(
        request: Request,
        filters: Optional[Dict[str, str]] = None,
        tickets: Optional[List[Dict[str, object]]] = None,
        pagination: Optional[Dict[str, int]] = None,
        error: str = "",
        searched: bool = False,
        status_code: int = 200,
    ) -> HTMLResponse:
        config = load_app_config()
        active_filters = filters or {}
        page_state = pagination or {
            "current_page": 1,
            "total_pages": 1,
            "total_count": 0,
            "page_size": config.store_query_page_size,
            "has_prev": 0,
            "has_next": 0,
            "prev_page": 1,
            "next_page": 1,
        }
        prev_query = build_public_query_params(active_filters, page_state["prev_page"])
        next_query = build_public_query_params(active_filters, page_state["next_page"])
        return templates.TemplateResponse(
            request,
            "query.html",
            {
                "request": request,
                "stores": config.stores,
                "statuses": config.statuses,
                "filters": active_filters,
                "tickets": tickets or [],
                "pagination": page_state,
                "prev_page_url": "/query" + (f"?{prev_query}" if prev_query else ""),
                "next_page_url": "/query" + (f"?{next_query}" if next_query else ""),
                "error": error,
                "searched": searched,
            },
            status_code=status_code,
        )

    def render_supplement_page(
        request: Request,
        ticket: Dict[str, object],
        store_name: str,
        return_url: str = "",
        error: str = "",
        success: bool = False,
        status_code: int = 200,
    ) -> HTMLResponse:
        config = load_app_config()
        query_url = store_query_list_return_url(store_name, return_url)
        detail_url = build_store_ticket_detail_url(int(ticket["id"]), store_name, query_url)
        supplement_url = build_store_ticket_supplement_url(int(ticket["id"]), store_name, query_url)
        return templates.TemplateResponse(
            request,
            "supplement.html",
            {
                "request": request,
                "ticket": ticket,
                "store_name": store_name,
                "error": error,
                "success": success,
                "query_url": query_url,
                "detail_url": detail_url,
                "supplement_url": supplement_url,
                "return_url": query_url,
                "image_accept": ",".join(f".{extension}" for extension in config.allowed_image_extensions),
                "file_accept": ",".join(f".{extension}" for extension in config.allowed_file_extensions),
                "max_image_count": config.max_image_count,
                "max_image_mb": config.max_image_mb,
                "max_file_count": config.max_file_count,
                "max_file_mb": config.max_file_mb,
                "max_total_file_upload_mb": config.max_total_file_upload_mb,
            },
            status_code=status_code,
        )

    def render_store_ticket_detail(
        request: Request,
        ticket: Dict[str, object],
        store_name: str,
        return_url: str = "",
        error: str = "",
        status_code: int = 200,
    ) -> HTMLResponse:
        query_url = store_query_list_return_url(store_name, return_url)
        detail_url = build_store_ticket_detail_url(int(ticket["id"]), store_name, query_url)
        supplement_url = build_store_ticket_supplement_url(int(ticket["id"]), store_name, query_url)
        supplements = fetch_store_ticket_supplements(int(ticket["id"]))
        comments = fetch_ticket_comments(int(ticket["id"]), public_only=True)
        return templates.TemplateResponse(
            request,
            "query_detail.html",
            {
                "request": request,
                "ticket": ticket,
                "store_name": store_name,
                "query_url": query_url,
                "detail_url": detail_url,
                "supplement_url": supplement_url,
                "attachment_counts": fetch_ticket_attachment_counts(int(ticket["id"])),
                "supplements": supplements,
                "comments": comments,
                "logs": fetch_store_visible_logs(int(ticket["id"])),
                "needs_store_supplement": str(ticket.get("status") or "") == "待门店补充",
                "handler_note": str(ticket.get("handler_note") or "").strip(),
                "error": error,
            },
            status_code=status_code,
        )

    @app.get("/admin/login", response_class=HTMLResponse)
    def admin_login_page(request: Request, next: str = Query("/admin"), logged_out: str = Query("")) -> HTMLResponse:
        next_url = safe_admin_return_url(next)
        if current_admin_username(request):
            return RedirectResponse(url=next_url, status_code=303)
        return render_login_form(request, next_url=next_url, logged_out=logged_out)

    @app.post("/admin/login", response_class=HTMLResponse)
    def admin_login(
        request: Request,
        username: str = Form(""),
        password: str = Form(""),
        next: str = Form("/admin"),
    ) -> HTMLResponse:
        next_url = safe_admin_return_url(next)
        authenticated_username = authenticate_admin(username, password)
        if not authenticated_username:
            return render_login_form(
                request,
                status_code=400,
                error="用户名或密码不正确。",
                username=username.strip(),
                next_url=next_url,
            )
        response = RedirectResponse(url=next_url, status_code=303)
        max_age = get_session_max_age_seconds()
        response.set_cookie(
            SESSION_COOKIE_NAME,
            create_admin_session(authenticated_username, max_age_seconds=max_age),
            httponly=True,
            samesite="lax",
            max_age=max_age,
            secure=session_cookie_secure(),
        )
        return response

    @app.post("/admin/logout")
    def admin_logout(request: Request, csrf_token: str = Form("")) -> RedirectResponse:
        if current_admin_username(request):
            require_admin_csrf(request, csrf_token)
        response = RedirectResponse(url="/admin/login?logged_out=1", status_code=303)
        response.delete_cookie(SESSION_COOKIE_NAME)
        return response

    @app.get("/admin/logout")
    def admin_logout_get() -> RedirectResponse:
        response = RedirectResponse(url="/admin/login?logged_out=1", status_code=303)
        response.delete_cookie(SESSION_COOKIE_NAME)
        return response

    @app.get("/", include_in_schema=False)
    def root() -> RedirectResponse:
        return RedirectResponse(url="/submit", status_code=303)

    @app.get("/submit", response_class=HTMLResponse)
    def submit_page(request: Request) -> HTMLResponse:
        return render_submit_form(request)

    @app.post("/submit", response_class=HTMLResponse)
    async def submit_ticket(
        request: Request,
        store_names: Optional[List[str]] = Form(None),
        brands: Optional[List[str]] = Form(None),
        brand_extra: str = Form(""),
        store_name: str = Form(""),
        submitter: str = Form(""),
        request_type: str = Form(""),
        urgency: str = Form(""),
        brand: str = Form(""),
        product_name: str = Form(""),
        sku_barcode: str = Form(""),
        quantity: str = Form(""),
        description: str = Form(""),
        expected_finish_date: str = Form(""),
        images: Optional[List[UploadFile]] = File(None),
        files: Optional[List[UploadFile]] = File(None),
    ) -> HTMLResponse:
        config = load_app_config()
        normalized_stores = normalize_store_names(store_names, store_name)
        normalized_brands = normalize_brand_names(brands, brand_extra, brand)
        store_display = join_display_values(normalized_stores)
        brand_display = join_display_values(normalized_brands)
        form_values = {
            "store_name": store_name,
            "store_names": normalized_stores,
            "submitter": submitter,
            "request_type": request_type,
            "urgency": urgency,
            "brand": brand,
            "brands": normalized_brands,
            "brand_extra": brand_extra,
            "product_name": product_name,
            "sku_barcode": sku_barcode,
            "quantity": quantity,
            "description": description,
            "expected_finish_date": expected_finish_date,
        }
        error = validate_submission(
            normalized_stores,
            submitter,
            request_type,
            urgency,
            quantity,
            description,
            config.stores,
            config.request_types,
            config.urgency_levels,
        )
        if error:
            return render_submit_form(request, status_code=400, error=error, values=form_values)

        try:
            prepared_images = await prepare_images(images, config)
            prepared_files = await prepare_files(files, config)
        except ValueError as exc:
            return render_submit_form(request, status_code=400, error=str(exc), values=form_values)

        rule_values = dict(form_values)
        rule_values["brand"] = brand_display
        rule_error = validate_request_type_rule(request_type, rule_values, prepared_images, prepared_files)
        if rule_error:
            return render_submit_form(request, status_code=400, error=rule_error, values=form_values)

        timestamp = now_text()
        quantity_value = int(quantity.strip()) if quantity.strip() else None
        try:
            ticket_id, ticket_no = create_ticket_with_images(
                {
                    "created_at": timestamp,
                    "store_name": store_display,
                    "store_names": normalized_stores,
                    "submitter": submitter.strip(),
                    "request_type": request_type,
                    "urgency": urgency,
                    "brand": brand_display,
                    "brands": normalized_brands,
                    "product_name": product_name.strip(),
                    "sku_barcode": sku_barcode.strip(),
                    "quantity": quantity_value,
                    "description": description.strip(),
                    "expected_finish_date": expected_finish_date.strip(),
                },
                prepared_images,
                config,
                prepared_files,
            )
            try:
                created_ticket = fetch_ticket(ticket_id)
                if created_ticket:
                    create_new_ticket_notification(created_ticket)
            except Exception:
                pass
        except RuntimeError as exc:
            return render_submit_form(request, status_code=500, error=str(exc), values=form_values)
        except OSError:
            return render_submit_form(request, status_code=500, error="附件保存失败，请稍后重试。", values=form_values)

        return templates.TemplateResponse(
            request,
            "submit_success.html",
            {"request": request, "ticket_no": ticket_no},
        )

    @app.post("/admin/tickets/create")
    async def admin_create_ticket(
        request: Request,
        admin: str = Depends(require_admin),
        store_names: Optional[List[str]] = Form(None),
        brands: Optional[List[str]] = Form(None),
        brand_extra: str = Form(""),
        store_name: str = Form(""),
        submitter: str = Form(""),
        request_type: str = Form(""),
        urgency: str = Form(""),
        brand: str = Form(""),
        product_name: str = Form(""),
        sku_barcode: str = Form(""),
        quantity: str = Form(""),
        description: str = Form(""),
        expected_finish_date: str = Form(""),
        images: Optional[List[UploadFile]] = File(None),
        files: Optional[List[UploadFile]] = File(None),
        csrf_token: str = Form(""),
    ) -> JSONResponse:
        require_admin_csrf(request, csrf_token)
        config = load_app_config()
        result = await create_ticket_from_submission(
            store_names=store_names,
            brands=brands,
            brand_extra=brand_extra,
            store_name=store_name,
            submitter=submitter,
            request_type=request_type,
            urgency=urgency,
            brand=brand,
            product_name=product_name,
            sku_barcode=sku_barcode,
            quantity=quantity,
            description=description,
            expected_finish_date=expected_finish_date,
            images=images,
            files=files,
            config=config,
        )
        if not result["ok"]:
            return JSONResponse(
                {
                    "ok": False,
                    "error": str(result.get("error") or ""),
                    "values": result.get("values") or {},
                },
                status_code=int(result.get("status_code") or 400),
            )
        return JSONResponse(
            {
                "ok": True,
                "ticket_id": result["ticket_id"],
                "ticket_no": result["ticket_no"],
                "message": f"已创建工单 {result['ticket_no']}。",
            }
        )

    @app.get("/query", response_class=HTMLResponse)
    def query_page(
        request: Request,
        store_name: str = Query(""),
        ticket_no: str = Query(""),
        submitter: str = Query(""),
        keyword: str = Query(""),
        status: str = Query(""),
        date_start: str = Query(""),
        date_end: str = Query(""),
        page: int = Query(1),
    ) -> HTMLResponse:
        config = load_app_config()
        raw_filters = {
            "store_name": store_name,
            "ticket_no": ticket_no,
            "submitter": submitter,
            "keyword": keyword,
            "status": status,
            "date_start": date_start,
            "date_end": date_end,
        }
        submitted = any(str(value or "").strip() for value in raw_filters.values())
        filters = normalize_store_query_filters(raw_filters, config)
        if not submitted:
            return render_query_page(request, filters=filters)
        if not filters.get("store_name") or filters.get("store_name") not in config.stores:
            return render_query_page(
                request,
                filters=filters,
                error="请选择门店后再查询。",
                searched=True,
                status_code=400,
            )
        if filters.get("status") and filters.get("status") not in config.statuses:
            return render_query_page(
                request,
                filters=filters,
                error="请选择有效状态。",
                searched=True,
                status_code=400,
            )
        tickets, pagination = fetch_store_query_page(filters, config, page)
        return_url = safe_query_return_url(request_path_with_query(request))
        for ticket in tickets:
            ticket["detail_url"] = build_store_ticket_detail_url(int(ticket["id"]), filters["store_name"], return_url)
            ticket["supplement_url"] = build_store_ticket_supplement_url(int(ticket["id"]), filters["store_name"], return_url)
        return render_query_page(request, filters=filters, tickets=tickets, pagination=pagination, searched=True)

    @app.post("/query")
    def query_submit(
        store_name: str = Form(""),
        ticket_no: str = Form(""),
        submitter: str = Form(""),
        keyword: str = Form(""),
        status: str = Form(""),
        date_start: str = Form(""),
        date_end: str = Form(""),
    ) -> RedirectResponse:
        params = {
            "store_name": store_name.strip(),
            "ticket_no": ticket_no.strip(),
            "submitter": submitter.strip(),
            "keyword": keyword.strip(),
            "status": status.strip(),
            "date_start": date_start.strip(),
            "date_end": date_end.strip(),
        }
        query = build_public_query_params(params)
        return RedirectResponse(url="/query" + (f"?{query}" if query else ""), status_code=303)

    @app.get("/schedule", response_class=HTMLResponse)
    def store_schedule_page(
        request: Request,
        store_name: str = Query(""),
        month: str = Query(""),
    ) -> HTMLResponse:
        try:
            return render_store_schedule_page(request, store_name, month)
        except ValueError as exc:
            return render_store_schedule_page(request, store_name, datetime.now().strftime("%Y-%m"), error=str(exc), status_code=400)

    @app.get("/query/ticket/{ticket_id}", response_class=HTMLResponse)
    def store_ticket_detail(
        request: Request,
        ticket_id: int,
        store_name: str = Query(""),
        return_url: str = Query(""),
    ) -> HTMLResponse:
        clean_store_name = store_name.strip()
        if not clean_store_name:
            return HTMLResponse("请选择门店后查看工单详情。", status_code=400)
        ticket = fetch_store_ticket(ticket_id, clean_store_name)
        if not ticket:
            raise HTTPException(status_code=404, detail="未找到该门店对应工单")
        return render_store_ticket_detail(request, ticket, clean_store_name, return_url=return_url)

    @app.post("/query/ticket/{ticket_id}/comments", response_class=HTMLResponse)
    def add_store_ticket_comment(
        request: Request,
        ticket_id: int,
        store_name: str = Form(""),
        author_name: str = Form(""),
        content: str = Form(""),
        return_url: str = Form(""),
    ) -> HTMLResponse:
        clean_store_name = store_name.strip()
        if not clean_store_name:
            return HTMLResponse("请选择门店后提交沟通内容。", status_code=400)
        ticket = fetch_store_ticket(ticket_id, clean_store_name)
        if not ticket:
            raise HTTPException(status_code=404, detail="未找到该门店对应工单")
        clean_author_name = author_name.strip() or "门店"
        if not content.strip():
            return render_store_ticket_detail(
                request,
                ticket,
                clean_store_name,
                return_url=return_url,
                error="请填写沟通内容。",
                status_code=400,
            )
        create_ticket_comment(
            ticket,
            "store",
            clean_author_name,
            content,
            "public",
            f"门店:{clean_author_name}",
        )
        detail_url = build_store_ticket_detail_url(ticket_id, clean_store_name, return_url)
        return RedirectResponse(url=detail_url, status_code=303)

    @app.get("/query/ticket/{ticket_id}/supplement", response_class=HTMLResponse)
    def supplement_page(
        request: Request,
        ticket_id: int,
        store_name: str = Query(""),
        return_url: str = Query(""),
    ) -> HTMLResponse:
        clean_store_name = store_name.strip()
        ticket = fetch_store_ticket(ticket_id, clean_store_name)
        if not ticket:
            raise HTTPException(status_code=403, detail="门店不匹配")
        return render_supplement_page(request, ticket, clean_store_name, return_url=return_url)

    @app.post("/query/ticket/{ticket_id}/supplement", response_class=HTMLResponse)
    async def submit_supplement(
        request: Request,
        ticket_id: int,
        store_name: str = Form(""),
        submitter: str = Form(""),
        note: str = Form(""),
        return_url: str = Form(""),
        images: Optional[List[UploadFile]] = File(None),
        files: Optional[List[UploadFile]] = File(None),
    ) -> HTMLResponse:
        clean_store_name = store_name.strip()
        ticket = fetch_store_ticket(ticket_id, clean_store_name)
        if not ticket:
            raise HTTPException(status_code=403, detail="门店不匹配")
        clean_submitter = submitter.strip()
        if not clean_submitter:
            return render_supplement_page(
                request,
                ticket,
                clean_store_name,
                return_url=return_url,
                error="请填写补充人。",
                status_code=400,
            )
        config = load_app_config()
        try:
            prepared_images = await prepare_images(images, config)
            prepared_files = await prepare_files(files, config)
        except ValueError as exc:
            return render_supplement_page(
                request,
                ticket,
                clean_store_name,
                return_url=return_url,
                error=str(exc),
                status_code=400,
            )
        if not note.strip() and not prepared_images and not prepared_files:
            return render_supplement_page(
                request,
                ticket,
                clean_store_name,
                return_url=return_url,
                error="请填写补充说明或上传附件。",
                status_code=400,
            )
        try:
            create_store_supplement(ticket, clean_submitter, note, prepared_images, prepared_files, config)
            try:
                updated_for_notification = fetch_ticket(ticket_id) or ticket
                create_store_supplement_notification(updated_for_notification, clean_submitter)
            except Exception:
                pass
        except OSError:
            return render_supplement_page(
                request,
                ticket,
                clean_store_name,
                return_url=return_url,
                error="附件保存失败，请稍后重试。",
                status_code=500,
            )
        updated_ticket = fetch_store_ticket(ticket_id, clean_store_name) or ticket
        return render_supplement_page(request, updated_ticket, clean_store_name, return_url=return_url, success=True)

    @app.get("/admin", response_class=HTMLResponse)
    def admin_page(
        request: Request,
        admin: str = Depends(require_admin),
        store_name: str = Query(""),
        request_type: str = Query(""),
        urgency: str = Query(""),
        status: str = Query(""),
        assigned_to: str = Query(""),
        date_start: str = Query(""),
        date_end: str = Query(""),
        keyword: str = Query(""),
        due_status: str = Query(""),
        sort: str = Query("newest"),
        page: int = Query(1),
        archived_count: int = Query(0),
        deleted_count: int = Query(0),
        error: str = Query(""),
    ) -> HTMLResponse:
        filters = {
            "store_name": store_name,
            "request_type": request_type,
            "urgency": urgency,
            "status": status,
            "assigned_to": assigned_to,
            "date_start": date_start,
            "date_end": date_end,
            "keyword": keyword,
            "due_status": due_status,
        }
        config = load_app_config()
        tickets, pagination = fetch_ticket_page(filters, sort, config, page)
        summary = fetch_ticket_summary(filters)
        return_url = safe_admin_return_url(request_path_with_query(request))
        for ticket in tickets:
            ticket["detail_url"] = f"/admin/ticket/{ticket['id']}?{urlencode({'return_url': return_url})}"
        export_query = build_query_params(filters, sort)
        export_url = "/admin/export" + (f"?{export_query}" if export_query else "")
        prev_query = build_query_params(filters, sort, pagination["prev_page"])
        next_query = build_query_params(filters, sort, pagination["next_page"])
        return templates.TemplateResponse(
            request,
            "admin.html",
            {
                "request": request,
                "tickets": tickets,
                "stores": config.stores,
                "request_types": config.request_types,
                "urgency_levels": config.urgency_levels,
                "statuses": config.statuses,
                "handlers": config.handlers,
                "due_status_options": DUE_STATUS_OPTIONS,
                "filters": filters,
                "sort": sort,
                "summary": summary,
                "export_url": export_url,
                "pagination": pagination,
                "prev_page_url": "/admin" + (f"?{prev_query}" if prev_query else ""),
                "next_page_url": "/admin" + (f"?{next_query}" if next_query else ""),
                "archived_count": max(archived_count, 0),
                "deleted_count": max(deleted_count, 0),
                "error": error,
                "admin_user": admin,
                "csrf_token": current_csrf_token(request),
                "brands": config.brands,
                "image_accept": ",".join(f".{extension}" for extension in config.allowed_image_extensions),
                "file_accept": ",".join(f".{extension}" for extension in config.allowed_file_extensions),
                "max_image_count": config.max_image_count,
                "max_total_upload_mb": config.max_total_upload_mb,
                "max_image_mb": config.max_image_mb,
                "allowed_file_extensions": config.allowed_file_extensions,
                "max_file_count": config.max_file_count,
                "max_file_mb": config.max_file_mb,
                "max_total_file_upload_mb": config.max_total_file_upload_mb,
                "values": {},
                "request_type_rules_json": json.dumps(load_request_type_rules(), ensure_ascii=False),
            },
        )

    @app.get("/admin/my-work", response_class=HTMLResponse)
    def admin_my_work(request: Request, admin: str = Depends(require_admin)) -> HTMLResponse:
        config = load_app_config()
        work = fetch_my_work(admin, config)
        return templates.TemplateResponse(
            request,
            "my_work.html",
            {
                "request": request,
                "work": work,
                "admin_user": admin,
                "csrf_token": current_csrf_token(request),
            },
        )

    @app.get("/admin/employees", response_class=HTMLResponse)
    def admin_employees(
        request: Request,
        admin: str = Depends(require_admin),
        store_name: str = Query(""),
        status: str = Query(""),
        scope: str = Query("active"),
        page: int = Query(1),
        error: str = Query(""),
        success: str = Query(""),
    ) -> HTMLResponse:
        return render_employees_page(request, admin, store_name, status, scope=scope, page=page, error=error, success=success)

    @app.post("/admin/employees")
    def create_employee_route(
        request: Request,
        admin: str = Depends(require_admin),
        employee_name: str = Form(""),
        store_name: str = Form(""),
        primary_store_name: str = Form(""),
        store_names: Optional[List[str]] = Form(None),
        role: str = Form(""),
        phone: str = Form(""),
        status: str = Form("在职"),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        try:
            create_employee(
                employee_name,
                store_name,
                role,
                phone,
                status,
                load_app_config(),
                store_names=store_names,
                primary_store_name=primary_store_name,
            )
        except (ValueError, HTTPException) as exc:
            return RedirectResponse(url="/admin/employees?" + urlencode({"error": form_error_message(exc)}), status_code=303)
        return RedirectResponse(url="/admin/employees?" + urlencode({"success": "已新增员工。"}), status_code=303)

    @app.post("/admin/employees/{employee_id}/update")
    def update_employee_route(
        request: Request,
        employee_id: int,
        admin: str = Depends(require_admin),
        employee_name: str = Form(""),
        store_name: str = Form(""),
        primary_store_name: str = Form(""),
        store_names: Optional[List[str]] = Form(None),
        role: str = Form(""),
        phone: str = Form(""),
        status: str = Form("在职"),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        try:
            update_employee(
                employee_id,
                employee_name,
                store_name,
                role,
                phone,
                status,
                load_app_config(),
                store_names=store_names,
                primary_store_name=primary_store_name,
            )
        except (ValueError, HTTPException) as exc:
            return RedirectResponse(url="/admin/employees?" + urlencode({"error": form_error_message(exc)}), status_code=303)
        return RedirectResponse(url="/admin/employees?" + urlencode({"success": "员工信息已保存。"}), status_code=303)

    @app.post("/admin/employees/{employee_id}/disable")
    def disable_employee_route(
        request: Request,
        employee_id: int,
        _admin: str = Depends(require_admin),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        try:
            disable_employee(employee_id)
        except HTTPException as exc:
            return RedirectResponse(url="/admin/employees?" + urlencode({"error": form_error_message(exc)}), status_code=303)
        return RedirectResponse(url="/admin/employees?" + urlencode({"success": "员工已标记为离职。"}), status_code=303)

    @app.post("/admin/employees/{employee_id}/archive")
    def archive_employee_route(
        request: Request,
        employee_id: int,
        admin: str = Depends(require_admin),
        archive_reason: str = Form(""),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        try:
            archive_employee(employee_id, admin, archive_reason)
        except (ValueError, HTTPException) as exc:
            return RedirectResponse(url="/admin/employees?" + urlencode({"error": form_error_message(exc)}), status_code=303)
        return RedirectResponse(url="/admin/employees?" + urlencode({"success": "员工已归档。"}), status_code=303)

    @app.post("/admin/employees/{employee_id}/unarchive")
    def unarchive_employee_route(
        request: Request,
        employee_id: int,
        _admin: str = Depends(require_admin),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        try:
            unarchive_employee(employee_id)
        except (ValueError, HTTPException) as exc:
            return RedirectResponse(url="/admin/employees?scope=archive&" + urlencode({"error": form_error_message(exc)}), status_code=303)
        return RedirectResponse(url="/admin/employees?" + urlencode({"success": "员工已取消归档。"}), status_code=303)

    @app.post("/admin/employees/{employee_id}/delete")
    def delete_employee_route(
        request: Request,
        employee_id: int,
        admin: str = Depends(require_admin),
        delete_reason: str = Form(""),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        try:
            soft_delete_employee(employee_id, admin, delete_reason)
        except (ValueError, HTTPException) as exc:
            return RedirectResponse(url="/admin/employees?" + urlencode({"error": form_error_message(exc)}), status_code=303)
        return RedirectResponse(url="/admin/employees?" + urlencode({"success": "员工已移入回收站。"}), status_code=303)

    @app.post("/admin/employees/{employee_id}/restore")
    def restore_employee_route(
        request: Request,
        employee_id: int,
        _admin: str = Depends(require_admin),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        try:
            restore_employee(employee_id)
        except (ValueError, HTTPException) as exc:
            return RedirectResponse(url="/admin/employees?scope=trash&" + urlencode({"error": form_error_message(exc)}), status_code=303)
        return RedirectResponse(url="/admin/employees?" + urlencode({"success": "员工已恢复。"}), status_code=303)

    @app.post("/admin/employees/{employee_id}/hard-delete")
    def hard_delete_employee_route(
        request: Request,
        employee_id: int,
        _admin: str = Depends(require_admin),
        confirm_delete: str = Form(""),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        if confirm_delete != "1":
            return RedirectResponse(url="/admin/employees?scope=trash&" + urlencode({"error": "请二次确认后再永久删除员工。"}), status_code=303)
        try:
            hard_delete_employee(employee_id)
        except (ValueError, HTTPException) as exc:
            return RedirectResponse(url="/admin/employees?scope=trash&" + urlencode({"error": form_error_message(exc)}), status_code=303)
        return RedirectResponse(url="/admin/employees?scope=trash&" + urlencode({"success": "员工已永久删除。"}), status_code=303)

    @app.get("/admin/shift-types", response_class=HTMLResponse)
    def admin_shift_types(
        request: Request,
        admin: str = Depends(require_admin),
        store_names: Optional[List[str]] = Query(None),
        status_filter: str = Query("all", alias="status"),
        global_scope: str = Query("all"),
        data_scope: str = Query("current"),
        error: str = Query(""),
        success: str = Query(""),
    ) -> HTMLResponse:
        return render_shift_types_page(
            request,
            admin,
            error=error,
            success=success,
            store_names=store_names,
            status_filter=status_filter,
            global_scope=global_scope,
            data_scope=data_scope,
        )

    @app.post("/admin/shift-types")
    def create_shift_type_route(
        request: Request,
        admin: str = Depends(require_admin),
        shift_scope: str = Form(""),
        store_name: str = Form(""),
        is_global: str = Form(""),
        shift_name: str = Form(""),
        start_time: str = Form(""),
        end_time: str = Form(""),
        duration_hours: str = Form("0"),
        color: str = Form(""),
        csrf_token: str = Form(""),
        return_store_names: Optional[List[str]] = Form(None),
        return_status: str = Form(""),
        return_global_scope: str = Form(""),
        return_data_scope: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        try:
            create_shift_type(
                shift_name,
                start_time,
                end_time,
                duration_hours,
                color,
                store_name=store_name,
                is_global=is_global_from_shift_scope(shift_scope, is_global),
                config=load_app_config(),
            )
        except (ValueError, HTTPException) as exc:
            return RedirectResponse(
                url=shift_types_redirect_url(request, return_store_names, return_status, return_global_scope, return_data_scope, error=form_error_message(exc)),
                status_code=303,
            )
        return RedirectResponse(
            url=shift_types_redirect_url(request, return_store_names, return_status, return_global_scope, return_data_scope, success="已新增班次。"),
            status_code=303,
        )

    @app.post("/admin/shift-types/{shift_type_id}/update")
    def update_shift_type_route(
        request: Request,
        shift_type_id: int,
        admin: str = Depends(require_admin),
        shift_scope: str = Form(""),
        store_name: str = Form(""),
        is_global: str = Form(""),
        shift_name: str = Form(""),
        start_time: str = Form(""),
        end_time: str = Form(""),
        duration_hours: str = Form("0"),
        color: str = Form(""),
        is_active: str = Form("0"),
        csrf_token: str = Form(""),
        return_store_names: Optional[List[str]] = Form(None),
        return_status: str = Form(""),
        return_global_scope: str = Form(""),
        return_data_scope: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        try:
            update_shift_type(
                shift_type_id,
                shift_name,
                start_time,
                end_time,
                duration_hours,
                color,
                is_active in {"1", "true", "on", "yes"},
                store_name=store_name,
                is_global=is_global_from_shift_scope(shift_scope, is_global),
                config=load_app_config(),
            )
        except (ValueError, HTTPException) as exc:
            return RedirectResponse(
                url=shift_types_redirect_url(request, return_store_names, return_status, return_global_scope, return_data_scope, error=form_error_message(exc)),
                status_code=303,
            )
        return RedirectResponse(
            url=shift_types_redirect_url(request, return_store_names, return_status, return_global_scope, return_data_scope, success="班次已保存。"),
            status_code=303,
        )

    @app.post("/admin/shift-types/business-hours")
    def update_store_business_hours_route(
        request: Request,
        admin: str = Depends(require_admin),
        store_name: str = Form(""),
        business_start_time: str = Form(""),
        business_end_time: str = Form(""),
        csrf_token: str = Form(""),
        return_store_names: Optional[List[str]] = Form(None),
        return_status: str = Form(""),
        return_global_scope: str = Form(""),
        return_data_scope: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        try:
            upsert_store_business_hours(store_name, business_start_time, business_end_time, load_app_config())
        except (ValueError, HTTPException) as exc:
            return RedirectResponse(
                url=shift_types_redirect_url(request, return_store_names, return_status, return_global_scope, return_data_scope, error=form_error_message(exc)),
                status_code=303,
            )
        return RedirectResponse(
            url=shift_types_redirect_url(
                request,
                return_store_names or [store_name],
                return_status,
                return_global_scope,
                return_data_scope,
                success="门店营业时间已保存。",
            ),
            status_code=303,
        )

    @app.post("/admin/shift-types/{shift_type_id}/disable")
    def disable_shift_type_route(
        request: Request,
        shift_type_id: int,
        _admin: str = Depends(require_admin),
        csrf_token: str = Form(""),
        return_store_names: Optional[List[str]] = Form(None),
        return_status: str = Form(""),
        return_global_scope: str = Form(""),
        return_data_scope: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        try:
            disable_shift_type(shift_type_id)
        except HTTPException as exc:
            return RedirectResponse(
                url=shift_types_redirect_url(request, return_store_names, return_status, return_global_scope, return_data_scope, error=form_error_message(exc)),
                status_code=303,
            )
        return RedirectResponse(
            url=shift_types_redirect_url(request, return_store_names, return_status, return_global_scope, return_data_scope, success="班次已停用。"),
            status_code=303,
        )

    @app.post("/admin/shift-types/{shift_type_id}/archive")
    def archive_shift_type_route(
        request: Request,
        shift_type_id: int,
        admin: str = Depends(require_admin),
        archive_reason: str = Form(""),
        csrf_token: str = Form(""),
        return_store_names: Optional[List[str]] = Form(None),
        return_status: str = Form(""),
        return_global_scope: str = Form(""),
        return_data_scope: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        try:
            archive_shift_type(shift_type_id, admin, archive_reason)
        except (ValueError, HTTPException) as exc:
            return RedirectResponse(
                url=shift_types_redirect_url(request, return_store_names, return_status, return_global_scope, return_data_scope, error=form_error_message(exc)),
                status_code=303,
            )
        return RedirectResponse(
            url=shift_types_redirect_url(request, return_store_names, return_status, return_global_scope, return_data_scope, success="班次已归档。"),
            status_code=303,
        )

    @app.post("/admin/shift-types/{shift_type_id}/delete")
    def delete_shift_type_route(
        request: Request,
        shift_type_id: int,
        admin: str = Depends(require_admin),
        delete_reason: str = Form(""),
        csrf_token: str = Form(""),
        return_store_names: Optional[List[str]] = Form(None),
        return_status: str = Form(""),
        return_global_scope: str = Form(""),
        return_data_scope: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        try:
            soft_delete_shift_type(shift_type_id, admin, delete_reason)
        except (ValueError, HTTPException) as exc:
            return RedirectResponse(
                url=shift_types_redirect_url(request, return_store_names, return_status, return_global_scope, return_data_scope, error=form_error_message(exc)),
                status_code=303,
            )
        return RedirectResponse(
            url=shift_types_redirect_url(request, return_store_names, return_status, return_global_scope, return_data_scope, success="班次已移入回收站。"),
            status_code=303,
        )

    @app.post("/admin/shift-types/{shift_type_id}/restore")
    def restore_shift_type_route(
        request: Request,
        shift_type_id: int,
        _admin: str = Depends(require_admin),
        csrf_token: str = Form(""),
        return_store_names: Optional[List[str]] = Form(None),
        return_status: str = Form(""),
        return_global_scope: str = Form(""),
        return_data_scope: str = Form("current"),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        try:
            restore_shift_type(shift_type_id)
        except (ValueError, HTTPException) as exc:
            return RedirectResponse(
                url=shift_types_redirect_url(request, return_store_names, return_status, return_global_scope, return_data_scope, error=form_error_message(exc)),
                status_code=303,
            )
        return RedirectResponse(
            url=shift_types_redirect_url(request, return_store_names, return_status, return_global_scope, "current", success="班次已恢复。"),
            status_code=303,
        )

    @app.post("/admin/shift-types/{shift_type_id}/hard-delete")
    def hard_delete_shift_type_route(
        request: Request,
        shift_type_id: int,
        _admin: str = Depends(require_admin),
        confirm_delete: str = Form(""),
        csrf_token: str = Form(""),
        return_store_names: Optional[List[str]] = Form(None),
        return_status: str = Form(""),
        return_global_scope: str = Form(""),
        return_data_scope: str = Form("trash"),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        if confirm_delete != "1":
            return RedirectResponse(
                url=shift_types_redirect_url(request, return_store_names, return_status, return_global_scope, return_data_scope, error="请二次确认后再永久删除班次。"),
                status_code=303,
            )
        try:
            hard_delete_shift_type(shift_type_id)
        except (ValueError, HTTPException) as exc:
            return RedirectResponse(
                url=shift_types_redirect_url(request, return_store_names, return_status, return_global_scope, return_data_scope, error=form_error_message(exc)),
                status_code=303,
            )
        return RedirectResponse(
            url=shift_types_redirect_url(request, return_store_names, return_status, return_global_scope, return_data_scope, success="班次已永久删除。"),
            status_code=303,
        )

    @app.get("/admin/schedules", response_class=HTMLResponse)
    def admin_schedules(
        request: Request,
        admin: str = Depends(require_admin),
        store_name: str = Query(""),
        store_names: Optional[List[str]] = Query(None),
        month: str = Query(""),
        employee_status: str = Query(""),
        employee_statuses: Optional[List[str]] = Query(None),
        view: str = Query(""),
        view_mode: str = Query(""),
        employee_scope: str = Query("all"),
        employee_scopes: Optional[List[str]] = Query(None),
        employee_ids: Optional[List[str]] = Query(None),
        shift_type_id: str = Query(""),
        shift_type_ids: Optional[List[str]] = Query(None),
        show_cross_store: str = Query(""),
        scope: str = Query("store"),
        saved: int = Query(0),
        saved_count: int = Query(0),
        created_count: int = Query(0),
        updated_count: int = Query(0),
        skipped_count: int = Query(0),
        deleted: int = Query(0),
        error: str = Query(""),
    ) -> HTMLResponse:
        parsed_employee_ids, has_invalid_employee_ids = parse_optional_int_list(employee_ids)
        parsed_shift_type_ids, include_custom_shift, has_invalid_shift_type_ids = parse_shift_filter_values(shift_type_ids, shift_type_id)
        parsed_shift_type_id = parsed_shift_type_ids[0] if len(parsed_shift_type_ids) == 1 and not include_custom_shift else 0
        parsed_employee_statuses, has_invalid_employee_statuses = normalize_employee_status_filters(employee_statuses, employee_status)
        normalized_employee_scope = normalize_employee_scope_filter(employee_scopes, employee_scope)
        config = load_app_config()
        _selected_store_names, _is_all_stores, has_invalid_store_names = normalize_schedule_store_filter(
            store_names,
            store_name,
            config,
            default_to_first=False,
        )
        parameter_error = ""
        if has_invalid_shift_type_ids:
            parameter_error = combine_error_messages(parameter_error, "班次筛选参数无效，已忽略。")
        if has_invalid_employee_ids:
            parameter_error = combine_error_messages(parameter_error, "员工筛选参数无效，已忽略。")
        if has_invalid_employee_statuses:
            parameter_error = combine_error_messages(parameter_error, "员工状态筛选参数无效，已忽略。")
        if has_invalid_store_names:
            parameter_error = combine_error_messages(parameter_error, "门店筛选参数无效，已忽略。")
        return render_schedules_page(
            request,
            admin,
            store_name,
            month,
            employee_status,
            view_mode or view or "calendar",
            show_cross_store in {"1", "true", "on", "yes"},
            scope,
            normalized_employee_scope,
            parsed_employee_ids,
            parsed_shift_type_id or 0,
            store_names=store_names,
            employee_statuses=parsed_employee_statuses,
            shift_type_ids=parsed_shift_type_ids,
            include_custom_shift=include_custom_shift,
            invalid_store_filter=has_invalid_store_names,
            saved=saved,
            saved_count=saved_count,
            created_count=created_count,
            updated_count=updated_count,
            skipped_count=skipped_count,
            deleted=deleted,
            error=combine_error_messages(error, parameter_error),
        )

    @app.post("/admin/schedules")
    def create_schedule_route(
        request: Request,
        admin: str = Depends(require_admin),
        store_name: str = Form(""),
        store_names: Optional[List[str]] = Form(None),
        employee_ids: Optional[List[str]] = Form(None),
        schedule_dates: Optional[List[str]] = Form(None),
        employee_id: str = Form(""),
        schedule_date: str = Form(""),
        shift_type_id: str = Form(""),
        schedule_mode: str = Form("shift"),
        custom_label: str = Form(""),
        custom_start_time: str = Form(""),
        custom_end_time: str = Form(""),
        custom_duration_hours: str = Form(""),
        note: str = Form(""),
        overwrite_existing: str = Form(""),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        month_for_return = datetime.now().strftime("%Y-%m")
        return_store = store_name.strip()
        try:
            config = load_app_config()
            selected_stores, is_all_stores, invalid_store = normalize_schedule_store_filter(
                store_names,
                store_name,
                config,
                default_to_first=False,
            )
            if invalid_store:
                raise ValueError("请选择有效门店。")
            if is_all_stores or len(selected_stores) != 1:
                raise ValueError("请选择具体门店后进行排班操作。请选择单个具体门店后进行排班操作。")
            return_store = selected_stores[0]
            clean_employee_ids = normalize_schedule_employee_ids(employee_ids, employee_id)
            clean_dates = normalize_schedule_dates(schedule_dates, schedule_date)
            clean_shift_type_id = parse_optional_positive_int(shift_type_id)
            clean_shift_type_id, custom_schedule = normalize_custom_schedule_payload(
                "custom" if shift_type_id.strip() == "custom" else schedule_mode,
                clean_shift_type_id or 0,
                custom_label,
                custom_start_time,
                custom_end_time,
                custom_duration_hours,
            )
            month_for_return = clean_dates[0][:7]
            is_legacy_single_submission = not employee_ids and not schedule_dates and bool(employee_id and schedule_date.strip())
            result = bulk_upsert_schedules(
                return_store,
                clean_employee_ids,
                clean_dates,
                clean_shift_type_id,
                note,
                is_legacy_single_submission or overwrite_existing in {"1", "true", "on", "yes"},
                admin,
                config.max_bulk_schedule_count,
                custom_schedule=custom_schedule,
            )
        except (ValueError, HTTPException) as exc:
            return RedirectResponse(url=schedule_redirect_url(return_store, month_for_return, error=form_error_message(exc)), status_code=303)
        return RedirectResponse(
            url=schedule_redirect_url(
                return_store,
                month_for_return,
                saved_count=result["saved_count"],
                created_count=result["created_count"],
                updated_count=result["updated_count"],
                skipped_count=result["skipped_count"],
            ),
            status_code=303,
        )

    @app.post("/admin/schedules/copy-day")
    def copy_schedule_day_route(
        request: Request,
        admin: str = Depends(require_admin),
        store_name: str = Form(""),
        target_date: str = Form(""),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        month_for_return = datetime.now().strftime("%Y-%m")
        try:
            result = copy_previous_day_schedules(store_name, target_date, admin)
            month_for_return = str(result["target_date"])[:7]
        except (ValueError, HTTPException) as exc:
            return RedirectResponse(url=schedule_redirect_url(store_name, month_for_return, error=form_error_message(exc)), status_code=303)
        return RedirectResponse(
            url=schedule_redirect_url(store_name, month_for_return, saved_count=int(result["copied_count"] or 0)),
            status_code=303,
        )

    @app.post("/admin/schedules/clear-employee")
    def clear_employee_schedule_route(
        request: Request,
        admin: str = Depends(require_admin),
        store_name: str = Form(""),
        employee_id: str = Form(""),
        month: str = Form(""),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        month_for_return = month.strip() or datetime.now().strftime("%Y-%m")
        try:
            clean_employee_id = parse_optional_positive_int(employee_id)
            if clean_employee_id is None:
                raise ValueError("员工不存在。")
            deleted_count = clear_employee_month_schedules(store_name, clean_employee_id, month_for_return, admin)
        except (ValueError, HTTPException) as exc:
            return RedirectResponse(url=schedule_redirect_url(store_name, month_for_return, error=form_error_message(exc)), status_code=303)
        return RedirectResponse(
            url=schedule_redirect_url(store_name, month_for_return, deleted=deleted_count),
            status_code=303,
        )

    @app.post("/admin/schedules/{schedule_id}/delete")
    def delete_schedule_route(
        request: Request,
        schedule_id: int,
        admin: str = Depends(require_admin),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        try:
            old_schedule = delete_schedule(schedule_id, admin)
        except HTTPException as exc:
            return RedirectResponse(url=schedule_redirect_url("", datetime.now().strftime("%Y-%m"), error=form_error_message(exc)), status_code=303)
        return RedirectResponse(
            url=schedule_redirect_url(str(old_schedule.get("store_name") or ""), str(old_schedule.get("schedule_date") or "")[:7], deleted=1),
            status_code=303,
        )

    @app.get("/admin/dashboard", response_class=HTMLResponse)
    def admin_dashboard(
        request: Request,
        admin: str = Depends(require_admin),
        store_name: str = Query(""),
        request_type: str = Query(""),
        status: str = Query(""),
        assigned_to: str = Query(""),
        date_start: str = Query(""),
        date_end: str = Query(""),
    ) -> HTMLResponse:
        config = load_app_config()
        filters = {
            "store_name": store_name,
            "request_type": request_type,
            "status": status,
            "assigned_to": assigned_to,
            "date_start": date_start,
            "date_end": date_end,
        }
        stats = fetch_dashboard_stats(filters)
        for ticket in stats["recent_tickets"]:
            ticket["detail_url"] = build_ticket_detail_url(int(ticket["id"]), "/admin/dashboard")
        return templates.TemplateResponse(
            request,
            "dashboard.html",
            {
                "request": request,
                "stores": config.stores,
                "request_types": config.request_types,
                "statuses": config.statuses,
                "handlers": config.handlers,
                "filters": filters,
                "stats": stats,
                "today_label": datetime.now().strftime("%Y-%m-%d"),
                "admin_user": admin,
                "csrf_token": current_csrf_token(request),
            },
        )

    @app.get("/admin/settings", response_class=HTMLResponse)
    def admin_settings(request: Request, admin: str = Depends(require_admin)) -> HTMLResponse:
        return templates.TemplateResponse(
            request,
            "settings.html",
            {
                "request": request,
                "admin_user": admin,
                "csrf_token": current_csrf_token(request),
                "config_files": [
                    "stores.json",
                    "request_types.json",
                    "urgency_levels.json",
                    "statuses.json",
                    "brands.json",
                    "handlers.json",
                    "system.json",
                    "request_type_rules.json",
                ],
            },
        )

    @app.get("/admin/account", response_class=HTMLResponse)
    def admin_account(request: Request, admin: str = Depends(require_admin)) -> HTMLResponse:
        return templates.TemplateResponse(
            request,
            "account.html",
            {
                "request": request,
                "admin_user": admin,
                "csrf_token": current_csrf_token(request),
            },
        )

    @app.get("/admin/system", response_class=HTMLResponse)
    def admin_system(request: Request, admin: str = Depends(require_admin)) -> HTMLResponse:
        return templates.TemplateResponse(
            request,
            "system.html",
            {
                "request": request,
                "admin_user": admin,
                "csrf_token": current_csrf_token(request),
                "database_path": str(get_db_path()),
                "upload_dir": str(get_upload_dir()),
                "config_dir": str(get_config_dir()),
            },
        )

    @app.get("/admin/route-health", response_class=HTMLResponse)
    def admin_route_health(request: Request, admin: str = Depends(require_admin)) -> HTMLResponse:
        route_pairs = registered_route_pairs(app)
        admin_get_routes = [path for method, path in route_pairs if method == "GET" and path.startswith("/admin")]
        admin_post_routes = [path for method, path in route_pairs if method == "POST" and path.startswith("/admin")]
        template_references = scan_template_admin_routes(app)
        missing_references = [item for item in template_references if not item["exists"]]
        navigation_references = [item for item in template_references if item["attr"] == "href"]
        form_references = [item for item in template_references if item["attr"] in {"action", "formaction"}]
        missing_navigation_references = [item for item in navigation_references if not item["exists"]]
        missing_form_references = [item for item in form_references if not item["exists"]]
        required_routes = required_route_items(app)
        missing_required_routes = [item for item in required_routes if not item["exists"]]
        return templates.TemplateResponse(
            request,
            "route_health.html",
            {
                "request": request,
                "admin_user": admin,
                "csrf_token": current_csrf_token(request),
                "main_file": str(BASE_DIR / "main.py"),
                "route_count": len(route_pairs),
                "admin_get_routes": admin_get_routes,
                "admin_post_routes": admin_post_routes,
                "required_routes": required_routes,
                "missing_required_routes": missing_required_routes,
                "template_references": template_references,
                "missing_references": missing_references,
                "navigation_references": navigation_references,
                "form_references": form_references,
                "missing_navigation_references": missing_navigation_references,
                "missing_form_references": missing_form_references,
                "legacy_redirects": LEGACY_ADMIN_REDIRECTS,
                "recommended_urls": recommended_access_url_entries(),
            },
        )

    @app.get("/admin/archive", response_class=HTMLResponse)
    def admin_archive(
        request: Request,
        admin: str = Depends(require_admin),
        store_name: str = Query(""),
        request_type: str = Query(""),
        urgency: str = Query(""),
        status: str = Query(""),
        assigned_to: str = Query(""),
        date_start: str = Query(""),
        date_end: str = Query(""),
        keyword: str = Query(""),
        due_status: str = Query(""),
        sort: str = Query("newest"),
        page: int = Query(1),
        unarchived_count: int = Query(0),
        deleted_count: int = Query(0),
        error: str = Query(""),
    ) -> HTMLResponse:
        filters = ticket_filters_from_params(
            store_name,
            request_type,
            urgency,
            status,
            assigned_to,
            date_start,
            date_end,
            keyword,
            due_status,
            scope="archive",
        )
        return render_archive_page(
            request,
            admin,
            filters=filters,
            sort=sort,
            page=page,
            unarchived_count=max(unarchived_count, 0),
            deleted_count=max(deleted_count, 0),
            error=error,
        )

    @app.get("/admin/trash", response_class=HTMLResponse)
    def admin_trash(
        request: Request,
        admin: str = Depends(require_admin),
        store_name: str = Query(""),
        request_type: str = Query(""),
        urgency: str = Query(""),
        status: str = Query(""),
        assigned_to: str = Query(""),
        date_start: str = Query(""),
        date_end: str = Query(""),
        keyword: str = Query(""),
        due_status: str = Query(""),
        sort: str = Query("newest"),
        page: int = Query(1),
        restored_count: int = Query(0),
        hard_deleted_count: int = Query(0),
        error: str = Query(""),
    ) -> HTMLResponse:
        filters = ticket_filters_from_params(
            store_name,
            request_type,
            urgency,
            status,
            assigned_to,
            date_start,
            date_end,
            keyword,
            due_status,
            scope="deleted",
        )
        return render_trash_page(
            request,
            admin,
            filters=filters,
            sort=sort,
            page=page,
            restored_count=max(restored_count, 0),
            hard_deleted_count=max(hard_deleted_count, 0),
            error=error,
        )

    @app.get("/admin/cleanup", response_class=HTMLResponse)
    def admin_cleanup(
        request: Request,
        admin: str = Depends(require_admin),
        deleted_count: int = Query(0),
    ) -> HTMLResponse:
        return render_cleanup_page(request, admin, deleted_count=max(deleted_count, 0))

    @app.post("/admin/cleanup/preview", response_class=HTMLResponse)
    def admin_cleanup_preview(
        request: Request,
        admin: str = Depends(require_admin),
        store_name: str = Form(""),
        submitter: str = Form(""),
        keyword: str = Form(""),
        date_start: str = Form(""),
        date_end: str = Form(""),
        incomplete_only: str = Form(""),
        only_test: str = Form(""),
        csrf_token: str = Form(""),
    ) -> HTMLResponse:
        require_admin_csrf(request, csrf_token)
        filters = cleanup_filters_from_form(
            store_name,
            submitter,
            keyword,
            date_start,
            date_end,
            incomplete_only,
            only_test,
        )
        tickets = fetch_cleanup_candidates(filters)
        return render_cleanup_page(request, admin, filters=filters, preview_tickets=tickets, previewed=True)

    @app.post("/admin/cleanup/delete")
    def admin_cleanup_delete(
        request: Request,
        admin: str = Depends(require_admin),
        store_name: str = Form(""),
        submitter: str = Form(""),
        keyword: str = Form(""),
        date_start: str = Form(""),
        date_end: str = Form(""),
        incomplete_only: str = Form(""),
        only_test: str = Form(""),
        confirm_cleanup: str = Form(""),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        if confirm_cleanup not in {"1", "true", "on", "yes"}:
            raise HTTPException(status_code=400, detail="请先确认批量清理。")
        filters = cleanup_filters_from_form(
            store_name,
            submitter,
            keyword,
            date_start,
            date_end,
            incomplete_only,
            only_test,
        )
        candidates = fetch_cleanup_candidates(filters)
        for ticket in candidates:
            soft_delete_ticket(int(ticket["id"]), admin, "批量清理测试数据")
        return RedirectResponse(url=f"/admin/cleanup?deleted_count={len(candidates)}", status_code=303)

    @app.get("/admin/embedded-pages", response_class=HTMLResponse)
    def embedded_pages_admin(request: Request, admin: str = Depends(require_admin)) -> HTMLResponse:
        return render_embedded_pages_admin(request, admin)

    @app.post("/admin/embedded-pages", response_class=HTMLResponse)
    async def create_embedded_page_route(
        request: Request,
        admin: str = Depends(require_admin),
        page_key: str = Form(""),
        title: str = Form(""),
        nav_label: str = Form(""),
        enabled: str = Form("0"),
        html_file: Optional[UploadFile] = File(None),
        csrf_token: str = Form(""),
    ) -> HTMLResponse:
        require_admin_csrf(request, csrf_token)
        config = load_app_config()
        try:
            upload = await prepare_embedded_upload_file(html_file, config)
            create_embedded_page(
                page_key,
                title,
                nav_label,
                enabled in {"1", "true", "on", "yes"},
                upload,
                admin,
            )
        except ValueError as exc:
            return render_embedded_pages_admin(request, admin, error=str(exc), status_code=400)
        return RedirectResponse(url="/admin/embedded-pages", status_code=303)

    @app.post("/admin/embedded-pages/{page_key}/replace", response_class=HTMLResponse)
    async def replace_embedded_page_route(
        request: Request,
        page_key: str,
        admin: str = Depends(require_admin),
        html_file: Optional[UploadFile] = File(None),
        csrf_token: str = Form(""),
    ) -> HTMLResponse:
        require_admin_csrf(request, csrf_token)
        config = load_app_config()
        try:
            upload = await prepare_embedded_upload_file(html_file, config)
            replace_embedded_page_file(page_key, upload, admin)
        except ValueError as exc:
            return render_embedded_pages_admin(request, admin, error=str(exc), status_code=400)
        return RedirectResponse(url="/admin/embedded-pages", status_code=303)

    @app.post("/admin/embedded-pages/{page_key}/toggle")
    def toggle_embedded_page_route(
        request: Request,
        page_key: str,
        admin: str = Depends(require_admin),
        enabled: str = Form("0"),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        set_embedded_page_enabled(page_key, enabled in {"1", "true", "on", "yes"}, admin)
        return RedirectResponse(url="/admin/embedded-pages", status_code=303)

    @app.post("/admin/embedded-pages/{page_key}/delete")
    def delete_embedded_page_route(
        request: Request,
        page_key: str,
        admin: str = Depends(require_admin),
        delete_reason: str = Form(""),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        soft_delete_embedded_page(page_key, admin, delete_reason)
        return RedirectResponse(url="/admin/embedded-pages", status_code=303)

    @app.post("/admin/embedded-pages/{page_key}/restore")
    def restore_embedded_page_route(
        request: Request,
        page_key: str,
        admin: str = Depends(require_admin),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        restore_embedded_page(page_key, admin)
        return RedirectResponse(url="/admin/trash", status_code=303)

    @app.post("/admin/embedded-pages/{page_key}/hard-delete")
    def hard_delete_embedded_page_route(
        request: Request,
        page_key: str,
        _admin: str = Depends(require_admin),
        confirm_delete: str = Form(""),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        if confirm_delete not in {"1", "true", "on", "yes"}:
            raise HTTPException(status_code=400, detail="请确认永久删除。")
        hard_delete_embedded_page(page_key)
        return RedirectResponse(url="/admin/trash", status_code=303)

    def embedded_page_content_response(page_key: str, resource_path: str, _admin: str) -> FileResponse:
        page = fetch_embedded_page(page_key, enabled_only=True)
        if not page:
            raise HTTPException(status_code=404, detail="嵌入页面不存在或未启用")
        target = resolve_embedded_resource(page, resource_path)
        if not target:
            raise HTTPException(status_code=404, detail="嵌入页面文件不存在")
        return FileResponse(
            target,
            media_type=embedded_media_type(target),
            headers={
                "X-Frame-Options": "SAMEORIGIN",
                "X-Content-Type-Options": "nosniff",
            },
        )

    @app.get("/admin/embed-content/{page_key}", response_class=FileResponse)
    def embedded_page_content_index(page_key: str, admin: str = Depends(require_admin)) -> FileResponse:
        return embedded_page_content_response(page_key, "index.html", admin)

    @app.get("/admin/embed-content/{page_key}/{resource_path:path}", response_class=FileResponse)
    def embedded_page_content(
        page_key: str,
        resource_path: str,
        admin: str = Depends(require_admin),
    ) -> FileResponse:
        return embedded_page_content_response(page_key, resource_path, admin)

    @app.get("/admin/embed/{page_key}", response_class=HTMLResponse)
    def embedded_page_view(
        request: Request,
        page_key: str,
        admin: str = Depends(require_admin),
    ) -> HTMLResponse:
        page = fetch_embedded_page(page_key, enabled_only=True)
        if not page:
            raise HTTPException(status_code=404, detail="嵌入页面不存在或未启用")
        entry_file = safe_embedded_resource_path(page.get("entry_file") or "index.html") or "index.html"
        content_url = f"/admin/embed-content/{page['page_key']}/{entry_file}?{urlencode({'v': page['version']})}"
        return templates.TemplateResponse(
            request,
            "embedded_page.html",
            {
                "request": request,
                "page": page,
                "content_url": content_url,
                "admin_user": admin,
                "csrf_token": current_csrf_token(request),
                "current_menu": f"embedded:{page['page_key']}",
            },
        )

    @app.get("/admin/api/notifications")
    def notifications_api(
        admin: str = Depends(require_admin),
        after_id: int = Query(0),
        limit: int = Query(20),
        unread_only: bool = Query(False),
    ) -> Dict[str, object]:
        notifications = fetch_notifications(
            admin,
            after_id=max(after_id, 0),
            limit=limit,
            unread_only=unread_only,
        )
        return {
            "unread_count": count_unread_notifications(admin),
            "latest_id": latest_notification_id(),
            "notifications": notifications,
        }

    @app.post("/admin/api/notifications/{event_id}/read")
    def mark_notification_read_api(
        request: Request,
        event_id: int,
        admin: str = Depends(require_admin),
        csrf_token: str = Form(""),
    ) -> Dict[str, object]:
        require_admin_csrf(request, csrf_token)
        if not mark_notification_read(admin, event_id):
            raise HTTPException(status_code=404, detail="消息不存在")
        return {
            "ok": True,
            "unread_count": count_unread_notifications(admin),
        }

    @app.post("/admin/api/notifications/read-all")
    def mark_all_notifications_read_api(
        request: Request,
        admin: str = Depends(require_admin),
        csrf_token: str = Form(""),
    ) -> Dict[str, object]:
        require_admin_csrf(request, csrf_token)
        mark_all_notifications_read(admin)
        return {
            "ok": True,
            "unread_count": count_unread_notifications(admin),
        }

    @app.get("/admin/ticket/{ticket_id}", response_class=HTMLResponse)
    def ticket_detail(
        request: Request,
        ticket_id: int,
        saved: str = Query(""),
        attachments_saved: str = Query(""),
        upload_error: str = Query(""),
        return_url: str = Query(""),
        admin: str = Depends(require_admin),
    ) -> HTMLResponse:
        return render_ticket_detail(
            request,
            ticket_id,
            admin,
            saved=saved,
            attachments_saved=attachments_saved,
            upload_error=upload_error,
            return_url=return_url,
        )

    @app.post("/admin/ticket/{ticket_id}/accept")
    def accept_ticket_route(
        request: Request,
        ticket_id: int,
        admin: str = Depends(require_admin),
        return_url: str = Form(""),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        accept_ticket(ticket_id, admin)
        if return_url.strip():
            return RedirectResponse(url=safe_admin_return_url(return_url), status_code=303)
        return RedirectResponse(url=build_ticket_detail_url(ticket_id, "/admin/my-work", saved="1"), status_code=303)

    @app.post("/admin/tickets/bulk-archive")
    def bulk_archive_tickets_route(
        request: Request,
        admin: str = Depends(require_admin),
        filters: Dict[str, str] = Depends(admin_ticket_filters_from_form),
        ticket_ids: Optional[List[int]] = Form(None),
        select_scope: str = Form("selected"),
        source_view: str = Form("active"),
        sort: str = Form("newest"),
        archive_reason: str = Form("批量归档"),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        filters["__ticket_scope"] = "active"
        ids = bulk_ticket_ids_from_scope(ticket_ids or [], select_scope, filters, sort, load_app_config())
        if not ids:
            return RedirectResponse(url=admin_redirect_url(source_view, error=BULK_SELECTION_REQUIRED_MESSAGE), status_code=303)
        archived_count = bulk_archive_tickets(ids, admin, archive_reason)
        return RedirectResponse(url=admin_redirect_url(source_view, archived_count=archived_count), status_code=303)

    @app.post("/admin/tickets/bulk-delete")
    def bulk_delete_tickets_route(
        request: Request,
        admin: str = Depends(require_admin),
        filters: Dict[str, str] = Depends(admin_ticket_filters_from_form),
        ticket_ids: Optional[List[int]] = Form(None),
        select_scope: str = Form("selected"),
        source_view: str = Form("active"),
        sort: str = Form("newest"),
        delete_reason: str = Form("批量移入回收站"),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        ids = bulk_ticket_ids_from_scope(ticket_ids or [], select_scope, filters, sort, load_app_config())
        if not ids:
            return RedirectResponse(url=admin_redirect_url(source_view, error=BULK_SELECTION_REQUIRED_MESSAGE), status_code=303)
        deleted_count = bulk_soft_delete_tickets(ids, admin, delete_reason)
        return RedirectResponse(url=admin_redirect_url(source_view, deleted_count=deleted_count), status_code=303)

    @app.post("/admin/tickets/bulk-unarchive")
    def bulk_unarchive_tickets_route(
        request: Request,
        admin: str = Depends(require_admin),
        filters: Dict[str, str] = Depends(admin_ticket_filters_from_form),
        ticket_ids: Optional[List[int]] = Form(None),
        select_scope: str = Form("selected"),
        source_view: str = Form("archive"),
        sort: str = Form("newest"),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        filters["__ticket_scope"] = "archive"
        ids = bulk_ticket_ids_from_scope(ticket_ids or [], select_scope, filters, sort, load_app_config())
        if not ids:
            return RedirectResponse(url=admin_redirect_url(source_view, error=BULK_SELECTION_REQUIRED_MESSAGE), status_code=303)
        unarchived_count = bulk_unarchive_tickets(ids, admin)
        return RedirectResponse(url=admin_redirect_url(source_view, unarchived_count=unarchived_count), status_code=303)

    @app.post("/admin/tickets/bulk-restore")
    def bulk_restore_tickets_route(
        request: Request,
        admin: str = Depends(require_admin),
        filters: Dict[str, str] = Depends(admin_ticket_filters_from_form),
        ticket_ids: Optional[List[int]] = Form(None),
        select_scope: str = Form("selected"),
        source_view: str = Form("trash"),
        sort: str = Form("newest"),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        filters["__ticket_scope"] = "deleted"
        ids = bulk_ticket_ids_from_scope(ticket_ids or [], select_scope, filters, sort, load_app_config())
        if not ids:
            return RedirectResponse(url=admin_redirect_url(source_view, error=BULK_SELECTION_REQUIRED_MESSAGE), status_code=303)
        restored_count = bulk_restore_tickets(ids, admin)
        return RedirectResponse(url=admin_redirect_url(source_view, restored_count=restored_count), status_code=303)

    @app.post("/admin/tickets/bulk-hard-delete")
    def bulk_hard_delete_tickets_route(
        request: Request,
        admin: str = Depends(require_admin),
        filters: Dict[str, str] = Depends(admin_ticket_filters_from_form),
        ticket_ids: Optional[List[int]] = Form(None),
        select_scope: str = Form("selected"),
        source_view: str = Form("trash"),
        sort: str = Form("newest"),
        confirm_delete: str = Form(""),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        if confirm_delete not in {"1", "true", "on", "yes"}:
            return RedirectResponse(url=admin_redirect_url(source_view, error="请二次确认永久删除"), status_code=303)
        filters["__ticket_scope"] = "deleted"
        ids = bulk_ticket_ids_from_scope(ticket_ids or [], select_scope, filters, sort, load_app_config())
        if not ids:
            return RedirectResponse(url=admin_redirect_url(source_view, error=BULK_SELECTION_REQUIRED_MESSAGE), status_code=303)
        hard_deleted_count = bulk_hard_delete_tickets(ids, admin)
        return RedirectResponse(url=admin_redirect_url(source_view, hard_deleted_count=hard_deleted_count), status_code=303)

    @app.post("/admin/ticket/{ticket_id}/archive")
    def archive_ticket_route(
        request: Request,
        ticket_id: int,
        admin: str = Depends(require_admin),
        archive_reason: str = Form("后台手动归档"),
        return_url: str = Form(""),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        archive_ticket(ticket_id, admin, archive_reason)
        return RedirectResponse(url=safe_admin_return_url(return_url or "/admin"), status_code=303)

    @app.post("/admin/ticket/{ticket_id}/unarchive")
    def unarchive_ticket_route(
        request: Request,
        ticket_id: int,
        admin: str = Depends(require_admin),
        return_url: str = Form(""),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        unarchive_ticket(ticket_id, admin)
        return RedirectResponse(url=safe_admin_return_url(return_url or "/admin/archive"), status_code=303)

    @app.post("/admin/ticket/{ticket_id}/delete")
    def delete_ticket_route(
        request: Request,
        ticket_id: int,
        admin: str = Depends(require_admin),
        delete_reason: str = Form(""),
        return_url: str = Form(""),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        soft_delete_ticket(ticket_id, admin, delete_reason)
        return RedirectResponse(url=safe_admin_return_url(return_url or "/admin"), status_code=303)

    @app.post("/admin/ticket/{ticket_id}/restore")
    def restore_ticket_route(
        request: Request,
        ticket_id: int,
        admin: str = Depends(require_admin),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        restore_ticket(ticket_id, admin)
        return RedirectResponse(url="/admin/trash", status_code=303)

    @app.post("/admin/ticket/{ticket_id}/hard-delete")
    def hard_delete_ticket_route(
        request: Request,
        ticket_id: int,
        admin: str = Depends(require_admin),
        confirm_delete: str = Form(""),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        if confirm_delete not in {"1", "true", "on", "yes"}:
            raise HTTPException(status_code=400, detail="请确认永久删除。")
        hard_delete_ticket(ticket_id, admin)
        return RedirectResponse(url="/admin/trash", status_code=303)

    @app.post("/admin/ticket/{ticket_id}/participants")
    def add_ticket_participant_route(
        request: Request,
        ticket_id: int,
        admin: str = Depends(require_admin),
        participant_type: str = Form(""),
        participant_name: str = Form(""),
        role: str = Form(""),
        return_url: str = Form(""),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        ticket = fetch_ticket(ticket_id)
        if not ticket:
            raise HTTPException(status_code=404, detail="工单不存在")
        try:
            create_ticket_participant(ticket, participant_type, participant_name, role, admin)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        return RedirectResponse(url=build_ticket_detail_url(ticket_id, return_url, saved="1"), status_code=303)

    @app.post("/admin/ticket/{ticket_id}/participant/{participant_id}/delete")
    def delete_ticket_participant_route(
        request: Request,
        ticket_id: int,
        participant_id: int,
        admin: str = Depends(require_admin),
        return_url: str = Form(""),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        soft_delete_ticket_child(
            "ticket_participants",
            participant_id,
            ticket_id,
            admin,
            "移除协作人",
            "participant_name",
        )
        return RedirectResponse(url=build_ticket_detail_url(ticket_id, return_url, saved="1"), status_code=303)

    @app.post("/admin/ticket/{ticket_id}/comments")
    def add_admin_ticket_comment(
        request: Request,
        ticket_id: int,
        admin: str = Depends(require_admin),
        content: str = Form(""),
        visibility: str = Form("public"),
        return_url: str = Form(""),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        ticket = fetch_ticket(ticket_id)
        if not ticket:
            raise HTTPException(status_code=404, detail="工单不存在")
        try:
            create_ticket_comment(ticket, "admin", admin, content, visibility, admin)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        return RedirectResponse(url=build_ticket_detail_url(ticket_id, return_url, saved="1"), status_code=303)

    @app.post("/admin/ticket/{ticket_id}/comment/{comment_id}/delete")
    def delete_ticket_comment_route(
        request: Request,
        ticket_id: int,
        comment_id: int,
        admin: str = Depends(require_admin),
        return_url: str = Form(""),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        soft_delete_ticket_child("ticket_comments", comment_id, ticket_id, admin, "删除评论", "content")
        return RedirectResponse(url=build_ticket_detail_url(ticket_id, return_url, saved="1"), status_code=303)

    @app.post("/admin/ticket/{ticket_id}/tasks")
    def add_ticket_task_route(
        request: Request,
        ticket_id: int,
        admin: str = Depends(require_admin),
        title: str = Form(""),
        assignee: str = Form(""),
        status: str = Form("待处理"),
        due_date: str = Form(""),
        return_url: str = Form(""),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        ticket = fetch_ticket(ticket_id)
        if not ticket:
            raise HTTPException(status_code=404, detail="工单不存在")
        try:
            create_ticket_task(ticket, title, assignee, status, due_date, admin)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        return RedirectResponse(url=build_ticket_detail_url(ticket_id, return_url, saved="1"), status_code=303)

    @app.post("/admin/ticket/{ticket_id}/task/{task_id}/delete")
    def delete_ticket_task_route(
        request: Request,
        ticket_id: int,
        task_id: int,
        admin: str = Depends(require_admin),
        return_url: str = Form(""),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        soft_delete_ticket_child("ticket_tasks", task_id, ticket_id, admin, "删除子任务", "title")
        return RedirectResponse(url=build_ticket_detail_url(ticket_id, return_url, saved="1"), status_code=303)

    @app.post("/admin/ticket/{ticket_id}/tasks/{task_id}")
    def update_ticket_task_route(
        request: Request,
        ticket_id: int,
        task_id: int,
        admin: str = Depends(require_admin),
        title: str = Form(""),
        assignee: str = Form(""),
        status: str = Form("待处理"),
        due_date: str = Form(""),
        return_url: str = Form(""),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        ticket = fetch_ticket(ticket_id)
        if not ticket:
            raise HTTPException(status_code=404, detail="工单不存在")
        try:
            update_ticket_task(ticket, task_id, title, assignee, status, due_date, admin)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        return RedirectResponse(url=build_ticket_detail_url(ticket_id, return_url, saved="1"), status_code=303)

    @app.post("/admin/ticket/{ticket_id}/supplement/{supplement_id}/delete")
    def delete_ticket_supplement_route(
        request: Request,
        ticket_id: int,
        supplement_id: int,
        admin: str = Depends(require_admin),
        return_url: str = Form(""),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        soft_delete_ticket_child(
            "ticket_supplements",
            supplement_id,
            ticket_id,
            admin,
            "隐藏门店补充记录",
            "note",
            "门店补充资料",
        )
        return RedirectResponse(url=build_ticket_detail_url(ticket_id, return_url, saved="1"), status_code=303)

    @app.post("/admin/ticket/{ticket_id}")
    def update_ticket(
        request: Request,
        ticket_id: int,
        admin: str = Depends(require_admin),
        status: str = Form(""),
        assigned_to: str = Form(""),
        handler_note: str = Form(""),
        return_url: str = Form(""),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        config = load_app_config()
        if status not in config.statuses:
            raise HTTPException(status_code=400, detail="状态不正确")
        assigned_to = assigned_to.strip()
        if assigned_to and config.handlers and assigned_to not in config.handlers:
            raise HTTPException(status_code=400, detail="处理人不正确")
        handler_note = handler_note.strip()
        timestamp = now_text()
        should_notify_need_supplement = False
        notification_ticket: Optional[Dict[str, object]] = None
        with get_connection() as connection:
            old_row = connection.execute(
                "SELECT * FROM tickets WHERE id = ? AND deleted_at IS NULL",
                (ticket_id,),
            ).fetchone()
            if not old_row:
                raise HTTPException(status_code=404, detail="工单不存在")
            old_ticket = dict(old_row)
            old_status = str(old_ticket.get("status") or "")
            old_assigned_to = str(old_ticket.get("assigned_to") or "")
            old_note = str(old_ticket.get("handler_note") or "")
            closed_at = old_ticket.get("closed_at")
            if status == COMPLETED_STATUS and not closed_at:
                closed_at = timestamp
            elif old_status == COMPLETED_STATUS and status != COMPLETED_STATUS:
                closed_at = None

            changed = status != old_status or assigned_to != old_assigned_to or handler_note != old_note
            if changed:
                connection.execute(
                    """
                    UPDATE tickets
                    SET status = ?, assigned_to = ?, handler_note = ?, closed_at = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (status, assigned_to, handler_note, closed_at, timestamp, ticket_id),
                )
                connection.execute(
                    """
                    INSERT INTO ticket_logs (
                        ticket_id, action, old_status, new_status, old_assigned_to,
                        new_assigned_to, note, operator, created_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        ticket_id,
                        "update",
                        old_status,
                        status,
                        old_assigned_to,
                        assigned_to,
                        handler_note,
                        admin,
                        timestamp,
                    ),
                )
                should_notify_need_supplement = old_status != status and status == "待门店补充"
                if should_notify_need_supplement:
                    notification_ticket = dict(old_ticket)
                    notification_ticket["status"] = status
                    notification_ticket["updated_at"] = timestamp
                    notification_ticket["assigned_to"] = assigned_to
        if should_notify_need_supplement and notification_ticket:
            try:
                create_need_store_supplement_notification(notification_ticket, admin)
            except Exception:
                pass
        return RedirectResponse(url=build_ticket_detail_url(ticket_id, return_url, saved="1"), status_code=303)

    @app.post("/admin/ticket/{ticket_id}/attachments", response_class=HTMLResponse)
    async def add_attachments(
        request: Request,
        ticket_id: int,
        admin: str = Depends(require_admin),
        new_images: Optional[List[UploadFile]] = File(None),
        new_files: Optional[List[UploadFile]] = File(None),
        return_url: str = Form(""),
        csrf_token: str = Form(""),
    ) -> HTMLResponse:
        require_admin_csrf(request, csrf_token)
        ticket = fetch_ticket(ticket_id)
        if not ticket:
            raise HTTPException(status_code=404, detail="工单不存在")
        config = load_app_config()
        try:
            prepared_images = await prepare_images(new_images, config)
            prepared_files = await prepare_files(new_files, config)
        except ValueError as exc:
            return render_ticket_detail(
                request,
                ticket_id,
                admin,
                upload_error=str(exc),
                return_url=return_url,
                status_code=400,
            )
        if not prepared_images and not prepared_files:
            return render_ticket_detail(
                request,
                ticket_id,
                admin,
                upload_error="请选择要上传的附件。",
                return_url=return_url,
                status_code=400,
            )
        try:
            add_ticket_attachments(ticket_id, str(ticket["ticket_no"]), prepared_images, prepared_files, admin)
        except OSError:
            return render_ticket_detail(
                request,
                ticket_id,
                admin,
                upload_error="附件保存失败，请稍后重试。",
                return_url=return_url,
                status_code=500,
            )
        return RedirectResponse(url=build_ticket_detail_url(ticket_id, return_url, attachments_saved="1"), status_code=303)

    @app.post("/admin/ticket/{ticket_id}/image/{image_id}/delete")
    def delete_ticket_image(
        request: Request,
        ticket_id: int,
        image_id: int,
        admin: str = Depends(require_admin),
        return_url: str = Form(""),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        image = fetch_ticket_image(image_id, ticket_id)
        if not image:
            raise HTTPException(status_code=404, detail="图片不存在")
        image_path = str(image.get("image_path") or "")
        physical_path = resolve_upload_path(image_filename(image_path))
        timestamp = now_text()
        with get_connection() as connection:
            connection.execute("DELETE FROM ticket_images WHERE id = ? AND ticket_id = ?", (image_id, ticket_id))
            connection.execute("UPDATE tickets SET updated_at = ? WHERE id = ?", (timestamp, ticket_id))
            connection.execute(
                """
                INSERT INTO ticket_logs (
                    ticket_id, action, old_status, new_status, old_assigned_to,
                    new_assigned_to, note, operator, created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (ticket_id, "删除图片", None, None, None, None, image_path, admin, timestamp),
            )
        if physical_path:
            try:
                physical_path.unlink(missing_ok=True)
            except OSError:
                pass
        return RedirectResponse(url=build_ticket_detail_url(ticket_id, return_url, saved="1"), status_code=303)

    @app.post("/admin/ticket/{ticket_id}/file/{file_id}/delete")
    def delete_ticket_file(
        request: Request,
        ticket_id: int,
        file_id: int,
        admin: str = Depends(require_admin),
        return_url: str = Form(""),
        csrf_token: str = Form(""),
    ) -> RedirectResponse:
        require_admin_csrf(request, csrf_token)
        ticket_file = fetch_ticket_file(file_id, ticket_id)
        if not ticket_file:
            raise HTTPException(status_code=404, detail="文件不存在")
        stored_filename = safe_uploaded_name(str(ticket_file.get("stored_filename") or ""))
        physical_path = resolve_upload_path(stored_filename)
        original_filename = str(ticket_file.get("original_filename") or "")
        timestamp = now_text()
        with get_connection() as connection:
            connection.execute("DELETE FROM ticket_files WHERE id = ? AND ticket_id = ?", (file_id, ticket_id))
            connection.execute("UPDATE tickets SET updated_at = ? WHERE id = ?", (timestamp, ticket_id))
            connection.execute(
                """
                INSERT INTO ticket_logs (
                    ticket_id, action, old_status, new_status, old_assigned_to,
                    new_assigned_to, note, operator, created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (ticket_id, "删除文件", None, None, None, None, original_filename, admin, timestamp),
            )
        if physical_path:
            try:
                physical_path.unlink(missing_ok=True)
            except OSError:
                pass
        return RedirectResponse(url=build_ticket_detail_url(ticket_id, return_url, saved="1"), status_code=303)

    @app.get("/admin/uploads/{filename:path}")
    def protected_upload(filename: str, _admin: str = Depends(require_admin)) -> FileResponse:
        target = resolve_upload_file(filename)
        if not target or not active_ticket_image_exists(filename):
            raise HTTPException(status_code=404, detail="图片不存在")
        return FileResponse(target)

    @app.get("/admin/files/{file_id}")
    def protected_file(file_id: int, _admin: str = Depends(require_admin)) -> FileResponse:
        ticket_file = fetch_ticket_file(file_id)
        if not ticket_file:
            raise HTTPException(status_code=404, detail="文件不存在")
        target = resolve_upload_file(safe_uploaded_name(str(ticket_file.get("stored_filename") or "")))
        if not target:
            raise HTTPException(status_code=404, detail="文件不存在")
        return FileResponse(
            target,
            media_type="application/octet-stream",
            filename=str(ticket_file.get("original_filename") or target.name),
            content_disposition_type="attachment",
        )

    @app.get("/admin/schedules/export")
    def export_schedules(
        _admin: str = Depends(require_admin),
        store_name: str = Query(""),
        store_names: Optional[List[str]] = Query(None),
        month: str = Query(""),
        shift_type_id: str = Query(""),
        shift_type_ids: Optional[List[str]] = Query(None),
        employee_statuses: Optional[List[str]] = Query(None),
    ) -> StreamingResponse:
        config = load_app_config()
        selected_month = normalize_month(month)
        selected_store_names, is_all_stores, _invalid_store = normalize_schedule_store_filter(
            store_names,
            store_name,
            config,
            default_to_first=False,
        )
        selected_shift_ids, include_custom_shift, _invalid_shift = parse_shift_filter_values(shift_type_ids, shift_type_id)
        selected_employee_statuses, _invalid_status = normalize_employee_status_filters(employee_statuses, "")
        rows = fetch_schedule_rows(
            selected_store_names[0] if len(selected_store_names) == 1 else "",
            selected_month,
            store_names=selected_store_names,
            employee_statuses=selected_employee_statuses,
            shift_type_ids=selected_shift_ids,
            include_custom_shift=include_custom_shift,
            is_all_stores=is_all_stores,
        )
        output = build_schedule_excel(rows)
        if not selected_store_names:
            filename_store = "全部门店"
        elif len(selected_store_names) == 1:
            filename_store = selected_store_names[0]
        else:
            filename_store = "多门店"
        filename = f"门店排班_{filename_store}_{selected_month}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
        )

    @app.get("/admin/archive/export")
    def export_archived_tickets(
        _admin: str = Depends(require_admin),
        store_name: str = Query(""),
        request_type: str = Query(""),
        urgency: str = Query(""),
        status: str = Query(""),
        assigned_to: str = Query(""),
        date_start: str = Query(""),
        date_end: str = Query(""),
        keyword: str = Query(""),
        due_status: str = Query(""),
        sort: str = Query("newest"),
    ) -> StreamingResponse:
        filters = ticket_filters_from_params(
            store_name,
            request_type,
            urgency,
            status,
            assigned_to,
            date_start,
            date_end,
            keyword,
            due_status,
            scope="archive",
        )
        config = load_app_config()
        output = build_excel(fetch_tickets(filters, sort, config))
        filename = f"{config.excel_filename_prefix}_归档_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
        )

    @app.get("/admin/export")
    def export_tickets(
        _admin: str = Depends(require_admin),
        store_name: str = Query(""),
        request_type: str = Query(""),
        urgency: str = Query(""),
        status: str = Query(""),
        assigned_to: str = Query(""),
        date_start: str = Query(""),
        date_end: str = Query(""),
        keyword: str = Query(""),
        due_status: str = Query(""),
        sort: str = Query("newest"),
    ) -> StreamingResponse:
        filters = {
            "store_name": store_name,
            "request_type": request_type,
            "urgency": urgency,
            "status": status,
            "assigned_to": assigned_to,
            "date_start": date_start,
            "date_end": date_end,
            "keyword": keyword,
            "due_status": due_status,
        }
        config = load_app_config()
        output = build_excel(fetch_tickets(filters, sort, config))
        filename = f"{config.excel_filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
        )

    return app


app = create_app()
