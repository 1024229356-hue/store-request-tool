import base64
import html
import importlib
import json
import os
import re
import sqlite3
import subprocess
import sys
import zipfile
from datetime import date, timedelta
from io import BytesIO
from pathlib import Path
from urllib.parse import quote, unquote, urlsplit

from fastapi.testclient import TestClient
from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parents[1]
ADMIN_AUTH = ("regional-admin", "very-secret-value")

VALID_PNG = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAIAAACQd1PeAAAADElEQVR4nGP4//8/AAX+Av4N70a4AAAAAElFTkSuQmCC"
)

DEFAULT_TEST_CONFIG = {
    "stores.json": ["南京门东店", "南昌万寿宫店", "山城巷店"],
    "request_types.json": ["建单需求", "审单需求", "商品异常", "缺货需求", "新品需求", "系统问题", "其他"],
    "urgency_levels.json": ["普通", "加急", "当天必须处理"],
    "statuses.json": ["待处理", "处理中", "待门店补充", "已完成", "已驳回"],
    "brands.json": [],
    "handlers.json": ["总部商品", "总部运营", "采购", "财务"],
    "system.json": {
        "app_name": "门店需求工单系统",
        "port": 8701,
        "max_image_mb": 10,
        "allowed_image_extensions": ["jpg", "jpeg", "png", "webp"],
        "default_status": "待处理",
        "excel_filename_prefix": "门店需求工单",
        "page_size": 50,
        "max_image_count": 5,
        "max_total_upload_mb": 30,
        "allowed_file_extensions": ["pdf", "doc", "docx", "xls", "xlsx", "csv", "txt", "zip", "rar"],
        "max_file_mb": 20,
        "max_file_count": 5,
        "max_total_file_upload_mb": 50,
        "max_embedded_html_mb": 20,
        "max_embedded_zip_mb": 100,
        "max_bulk_schedule_count": 200,
    },
}

EXPECTED_EXPORT_HEADERS = [
    "工单号",
    "提交时间",
    "门店",
    "提报人",
    "需求类型",
    "紧急程度",
    "品牌",
    "商品名称",
    "规格条码",
    "数量",
    "问题说明",
    "图片路径",
    "文件名称",
    "文件路径",
    "期望完成时间",
    "当前状态",
    "处理人",
    "处理备注",
    "完成时间",
    "最后更新时间",
    "处理时长小时",
    "是否超时",
    "时效状态",
]

ROUTE_ATTRIBUTE_RE = re.compile(r'\b(href|action|formaction)\s*=\s*(["\'])(.*?)\2', re.IGNORECASE | re.DOTALL)
FORM_METHOD_RE = re.compile(r'\bmethod\s*=\s*(["\']?)(get|post)\1', re.IGNORECASE)
JINJA_EXPRESSION_RE = re.compile(r"{{.*?}}", re.DOTALL)


def route_pattern(path):
    escaped = re.escape(path)
    escaped = re.sub(r"\\\{[^{}:]+:path\\\}", ".+", escaped)
    escaped = re.sub(r"\\\{[^{}]+\\\}", "[^/]+", escaped)
    return re.compile(f"^{escaped}$")


def route_exists(routes, method, path):
    for route_method, route_path in routes:
        if route_method == method and route_pattern(route_path).match(path):
            return True
    return False


def normalized_admin_path(raw_value):
    value = html.unescape(raw_value.strip())
    if not value.startswith("/admin"):
        return None
    value = JINJA_EXPRESSION_RE.sub("1", value)
    return urlsplit(value).path


def template_admin_route_references():
    references = []
    for template_path in sorted((PROJECT_DIR / "templates").glob("*.html")):
        text = template_path.read_text(encoding="utf-8")
        for match in ROUTE_ATTRIBUTE_RE.finditer(text):
            attr = match.group(1).lower()
            path = normalized_admin_path(match.group(3))
            if not path:
                continue
            method = "GET"
            if attr == "formaction":
                method = "POST"
            elif attr == "action":
                form_start = text.rfind("<form", 0, match.start())
                form_tag_end = text.find(">", form_start)
                form_tag = text[form_start:form_tag_end] if form_start >= 0 and form_tag_end >= form_start else ""
                method_match = FORM_METHOD_RE.search(form_tag)
                method = method_match.group(2).upper() if method_match else "GET"
            references.append((template_path.name, attr, path, method))
    return references


def write_config(config_dir, overrides=None):
    config_dir.mkdir(parents=True, exist_ok=True)
    config = dict(DEFAULT_TEST_CONFIG)
    config.update(overrides or {})
    for filename, value in config.items():
        (config_dir / filename).write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")


def build_client(
    tmp_path,
    monkeypatch,
    config_overrides=None,
    write_configs=True,
    admin_auth=ADMIN_AUTH,
    admin_users=None,
):
    config_dir = tmp_path / "config"
    if write_configs:
        write_config(config_dir, config_overrides)
    monkeypatch.setenv("STORE_REQUEST_DB_PATH", str(tmp_path / "tickets.db"))
    monkeypatch.setenv("STORE_REQUEST_UPLOAD_DIR", str(tmp_path / "uploads"))
    monkeypatch.setenv("STORE_REQUEST_CONFIG_DIR", str(config_dir))
    block_dotenv_admin_users = admin_users is None
    if admin_users is None:
        monkeypatch.setenv("ADMIN_USERS", "")
    else:
        monkeypatch.setenv("ADMIN_USERS", admin_users)
    monkeypatch.setenv("ADMIN_USERNAME", admin_auth[0])
    monkeypatch.setenv("ADMIN_PASSWORD", admin_auth[1])
    monkeypatch.setenv("SESSION_SECRET", "test-session-secret")
    monkeypatch.syspath_prepend(str(PROJECT_DIR))
    sys.modules.pop("main", None)
    main = importlib.import_module("main")
    if block_dotenv_admin_users:
        monkeypatch.delenv("ADMIN_USERS", raising=False)
    return TestClient(main.app), main


def submit_ticket(client, **overrides):
    data = {
        "store_name": "南京门东店",
        "submitter": "测试提报人",
        "request_type": "建单需求",
        "urgency": "加急",
        "brand": "测试品牌",
        "product_name": "测试商品",
        "sku_barcode": "690000000001",
        "quantity": "12",
        "description": "这是一条自动化测试工单",
        "expected_finish_date": "2026-07-04",
    }
    data.update(overrides)
    return client.post("/submit", data=data)


def submit_ticket_with_image(client, filename="issue.png", content=VALID_PNG, **overrides):
    data = {
        "store_name": "南京门东店",
        "submitter": "测试提报人",
        "request_type": "建单需求",
        "urgency": "加急",
        "brand": "测试品牌",
        "product_name": "测试商品",
        "sku_barcode": "690000000001",
        "quantity": "12",
        "description": "这是一条自动化测试工单",
        "expected_finish_date": "2026-07-04",
    }
    data.update(overrides)
    media_type = "image/png" if filename.endswith(".png") else "image/jpeg"
    return client.post("/submit", data=data, files=[("images", (filename, content, media_type))])


def submit_ticket_with_file(client, filename="need-list.xlsx", content=b"sku,qty\nA001,1\n", **overrides):
    data = {
        "store_name": "南京门东店",
        "submitter": "测试提报人",
        "request_type": "建单需求",
        "urgency": "加急",
        "brand": "测试品牌",
        "product_name": "测试商品",
        "sku_barcode": "690000000001",
        "quantity": "12",
        "description": "这是一条带普通附件的自动化测试工单",
        "expected_finish_date": "2026-07-04",
    }
    data.update(overrides)
    return client.post(
        "/submit",
        data=data,
        files=[("files", (filename, content, "application/octet-stream"))],
    )


def submit_ticket_with_image_and_file(
    client,
    image_filename="issue.png",
    file_filename="need-list.xlsx",
    file_content=b"sku,qty\nA001,1\n",
    **overrides,
):
    data = {
        "store_name": "南京门东店",
        "submitter": "测试提报人",
        "request_type": "建单需求",
        "urgency": "加急",
        "brand": "测试品牌",
        "product_name": "测试商品",
        "sku_barcode": "690000000001",
        "quantity": "12",
        "description": "这是一条同时带图片和普通附件的自动化测试工单",
        "expected_finish_date": "2026-07-04",
    }
    data.update(overrides)
    return client.post(
        "/submit",
        data=data,
        files=[
            ("images", (image_filename, VALID_PNG, "image/png")),
            ("files", (file_filename, file_content, "application/octet-stream")),
        ],
    )


def rows_for(tmp_path, table_name):
    with sqlite3.connect(tmp_path / "tickets.db") as connection:
        connection.row_factory = sqlite3.Row
        return [dict(row) for row in connection.execute(f"SELECT * FROM {table_name} ORDER BY id")]


def login_admin(client, username=ADMIN_AUTH[0], password=ADMIN_AUTH[1], follow_redirects=False):
    return client.post(
        "/admin/login",
        data={"username": username, "password": password},
        follow_redirects=follow_redirects,
    )


def assert_login_success(response):
    assert response.status_code == 303
    assert response.headers["location"].startswith("/admin")
    set_cookie = response.headers.get("set-cookie", "")
    assert "admin_session=" in set_cookie
    assert "HttpOnly" in set_cookie
    assert "samesite=lax" in set_cookie.lower()


def logged_in_client(client, username=ADMIN_AUTH[0], password=ADMIN_AUTH[1]):
    assert_login_success(login_admin(client, username, password))
    return client


def csrf_token_for(client, path="/admin"):
    response = client.get(path)
    assert response.status_code == 200
    match = re.search(r'name="csrf_token" value="([^"]+)"', response.text)
    assert match, response.text[:500]
    return match.group(1)


def admin_post(client, url, data=None, **kwargs):
    payload = dict(data or {})
    if "csrf_token" not in payload:
        payload["csrf_token"] = csrf_token_for(client)
    return client.post(url, data=payload, **kwargs)


def assert_html_forbidden(response):
    assert response.status_code == 403
    assert "text/html" in response.headers.get("content-type", "")
    assert "application/json" not in response.headers.get("content-type", "")
    assert "权限不足" in response.text


def ticket_action_logs(tmp_path, ticket_id):
    return [
        row
        for row in rows_for(tmp_path, "ticket_logs")
        if int(row["ticket_id"]) == int(ticket_id)
    ]


def basic_auth_headers(username=ADMIN_AUTH[0], password=ADMIN_AUTH[1]):
    credentials = base64.b64encode(f"{username}:{password}".encode("utf-8")).decode("ascii")
    return {"Authorization": f"Basic {credentials}"}


def make_zip_bytes(entries):
    output = BytesIO()
    with zipfile.ZipFile(output, "w") as archive:
        for name, content in entries.items():
            archive.writestr(name, content)
    return output.getvalue()


def test_submit_page_exposes_image_and_file_upload_inputs(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)

    submit_page = client.get("/submit")

    assert submit_page.status_code == 200
    assert "止痒门店需求提报" in submit_page.text
    assert "请尽量一次性补充完整信息" in submit_page.text
    assert "基础信息" in submit_page.text
    assert "商品信息" in submit_page.text
    assert "问题说明" in submit_page.text
    assert "附件上传" in submit_page.text
    assert 'name="images"' in submit_page.text
    assert "data-image-input" in submit_page.text
    assert "data-image-preview" in submit_page.text
    assert 'type="button"' in submit_page.text
    assert "data-clear-images" in submit_page.text
    assert 'name="files"' in submit_page.text
    assert "data-file-input" in submit_page.text
    assert "data-file-preview" in submit_page.text
    assert "data-clear-files" in submit_page.text
    assert "文件上传" in submit_page.text
    assert re.search(r'/static/style\.css\?v=[A-Za-z0-9._-]+', submit_page.text)
    assert re.search(r'/static/app\.js\?v=[A-Za-z0-9._-]+', submit_page.text)


def test_submit_success_page_highlights_ticket_number_and_copy_action(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)

    response = submit_ticket(client)

    assert response.status_code == 200
    assert "提交成功" in response.text
    assert "请截图保存工单号" in response.text
    assert "success-ticket-no" in response.text
    assert "复制工单号" in response.text
    assert "data-copy-ticket" in response.text
    assert "已提交成功，请截图保存工单号。后续可通过门店查询查看处理进度。" in response.text
    assert "查询工单" in response.text
    assert "继续提交" in response.text
    assert "后台入口" in response.text
    assert re.search(r'/static/style\.css\?v=[A-Za-z0-9._-]+', response.text)
    assert re.search(r'/static/app\.js\?v=[A-Za-z0-9._-]+', response.text)


def test_upload_script_uses_independent_state_and_clear_selectors():
    script = (PROJECT_DIR / "static" / "app.js").read_text(encoding="utf-8")

    assert "DOMContentLoaded" in script
    assert "selectedImages" in script
    assert "selectedFiles" in script
    assert "[data-clear-images]" in script
    assert "[data-clear-files]" in script
    assert "event.preventDefault()" in script
    assert "event.stopPropagation()" in script
    assert "new DataTransfer()" in script
    assert "[data-copy-ticket]" in script
    assert "navigator.clipboard" in script
    assert "[data-extra-image-input]" in script
    assert "[data-extra-file-input]" in script


def test_submit_page_exposes_multi_store_and_multi_brand_choices(tmp_path, monkeypatch):
    client, _ = build_client(
        tmp_path,
        monkeypatch,
        config_overrides={"brands.json": ["止痒", "雨过山", "logaloga"]},
    )

    submit_page = client.get("/submit")

    assert submit_page.status_code == 200
    assert 'name="store_names"' in submit_page.text
    assert 'type="checkbox"' in submit_page.text
    assert "可选择一个或多个门店。" in submit_page.text
    assert 'name="brands"' in submit_page.text
    assert 'name="brand_extra"' in submit_page.text
    assert "未在列表中的品牌，可手动输入" in submit_page.text
    assert "choice-grid" in submit_page.text
    assert "choice-chip" in submit_page.text


def test_submit_multi_store_multi_brand_persists_relations_queries_and_export(tmp_path, monkeypatch):
    client, _ = build_client(
        tmp_path,
        monkeypatch,
        config_overrides={"brands.json": ["止痒", "雨过山", "logaloga"]},
    )
    response = client.post(
        "/submit",
        data={
            "store_names": ["南京门东店", "南昌万寿宫店"],
            "submitter": "多选测试",
            "request_type": "缺货需求",
            "urgency": "加急",
            "brands": ["止痒", "雨过山"],
            "brand_extra": "logaloga，止痒；手动品牌",
            "product_name": "多门店商品",
            "sku_barcode": "SKU-MULTI",
            "quantity": "8",
            "description": "多门店多品牌自动化测试工单",
            "expected_finish_date": "2026-07-04",
        },
    )
    assert response.status_code == 200

    tickets = rows_for(tmp_path, "tickets")
    assert len(tickets) == 1
    assert tickets[0]["store_name"] == "南京门东店、南昌万寿宫店"
    assert tickets[0]["brand"] == "止痒、雨过山、logaloga、手动品牌"
    ticket_stores = rows_for(tmp_path, "ticket_stores")
    ticket_brands = rows_for(tmp_path, "ticket_brands")
    assert [row["store_name"] for row in ticket_stores] == ["南京门东店", "南昌万寿宫店"]
    assert [row["brand"] for row in ticket_brands] == ["止痒", "雨过山", "logaloga", "手动品牌"]

    logged_in_client(client)
    admin_filter = client.get("/admin?store_name=南昌万寿宫店")
    assert admin_filter.status_code == 200
    assert "多门店多品牌自动化测试工单" in admin_filter.text
    assert "南京门东店、南昌万寿宫店" in admin_filter.text
    assert "止痒、雨过山、logaloga、手动品牌" in admin_filter.text

    dashboard = client.get("/admin/dashboard")
    assert dashboard.status_code == 200
    assert "南京门东店" in dashboard.text
    assert "南昌万寿宫店" in dashboard.text

    query_nanjing = client.get("/query?store_name=南京门东店")
    assert query_nanjing.status_code == 200
    assert "多门店多品牌自动化测试工单" in query_nanjing.text
    query_nanchang = client.get("/query?store_name=南昌万寿宫店")
    assert query_nanchang.status_code == 200
    assert "多门店多品牌自动化测试工单" in query_nanchang.text

    detail_allowed = client.get("/query/ticket/1?store_name=南昌万寿宫店")
    assert detail_allowed.status_code == 200
    assert "tag-chip" in detail_allowed.text
    assert "南京门东店" in detail_allowed.text
    assert "南昌万寿宫店" in detail_allowed.text
    assert "止痒" in detail_allowed.text
    assert "雨过山" in detail_allowed.text
    assert "logaloga" in detail_allowed.text
    assert "手动品牌" in detail_allowed.text
    detail_forbidden = client.get("/query/ticket/1?store_name=山城巷店")
    assert detail_forbidden.status_code == 404

    export_response = client.get("/admin/export?store_name=南昌万寿宫店")
    assert export_response.status_code == 200
    workbook = load_workbook(BytesIO(export_response.content))
    sheet = workbook.active
    assert sheet["C2"].value == "南京门东店、南昌万寿宫店"
    assert sheet["G2"].value == "止痒、雨过山、logaloga、手动品牌"


def test_multi_store_and_brand_validation_errors(tmp_path, monkeypatch):
    rules = {
        "建单需求": {
            "required_fields": ["brand", "product_name", "quantity"],
            "require_any_attachment": False,
        }
    }
    client, _ = build_client(tmp_path, monkeypatch, config_overrides={"request_type_rules.json": rules})

    missing_store = client.post(
        "/submit",
        data={
            "submitter": "测试",
            "request_type": "系统问题",
            "urgency": "普通",
            "description": "没有选择门店",
        },
    )
    assert missing_store.status_code == 400
    assert "请至少选择一个门店。" in missing_store.text

    invalid_store = client.post(
        "/submit",
        data={
            "store_names": ["不存在门店"],
            "submitter": "测试",
            "request_type": "系统问题",
            "urgency": "普通",
            "description": "无效门店",
        },
    )
    assert invalid_store.status_code == 400
    assert "请选择有效门店。" in invalid_store.text

    missing_brand_is_allowed = client.post(
        "/submit",
        data={
            "store_names": ["南京门东店"],
            "submitter": "测试",
            "request_type": "建单需求",
            "urgency": "普通",
            "product_name": "规则商品",
            "quantity": "1",
            "description": "品牌必填规则",
        },
    )
    assert missing_brand_is_allowed.status_code == 200
    assert rows_for(tmp_path, "tickets")[-1]["brand"] == ""


def test_admin_comments_visibility_logs_and_notifications(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    submit_ticket(client, store_name="南京门东店", description="需要协作沟通")

    unauthenticated = client.post(
        "/admin/ticket/1/comments",
        data={"content": "未登录不允许评论", "visibility": "public"},
        follow_redirects=False,
    )
    assert unauthenticated.status_code == 303
    assert unauthenticated.headers["location"].startswith("/admin/login")

    logged_in_client(client)
    public_comment = admin_post(
        client,
        "/admin/ticket/1/comments",
        data={"content": "门店可见回复：总部已经联系供应商。", "visibility": "public"},
        follow_redirects=False,
    )
    assert public_comment.status_code == 303
    internal_comment = admin_post(
        client,
        "/admin/ticket/1/comments",
        data={"content": "内部备注：供应商报价暂不外发。", "visibility": "internal"},
        follow_redirects=False,
    )
    assert internal_comment.status_code == 303

    comments = rows_for(tmp_path, "ticket_comments")
    assert [comment["visibility"] for comment in comments] == ["public", "internal"]
    assert comments[0]["author_type"] == "admin"
    assert comments[0]["author_name"] == ADMIN_AUTH[0]
    assert comments[1]["content"] == "内部备注：供应商报价暂不外发。"

    logs = rows_for(tmp_path, "ticket_logs")
    assert [log["action"] for log in logs[-2:]] == ["新增评论", "新增评论"]
    assert "门店可见" in logs[-2]["note"]
    assert "内部备注" in logs[-1]["note"]

    events = rows_for(tmp_path, "notification_events")
    assert [event["event_type"] for event in events] == ["new_ticket", "ticket_comment", "ticket_comment"]
    assert events[-2]["title"] == "新增门店可见回复"
    assert events[-1]["title"] == "新增内部备注"

    admin_detail = client.get("/admin/ticket/1")
    assert admin_detail.status_code == 200
    assert "沟通记录" in admin_detail.text
    assert "门店可见回复：总部已经联系供应商。" in admin_detail.text
    assert "内部备注：供应商报价暂不外发。" in admin_detail.text
    assert "内部备注" in admin_detail.text

    store_detail = client.get("/query/ticket/1?store_name=南京门东店")
    assert store_detail.status_code == 200
    assert "可见沟通记录" in store_detail.text
    assert "总部回复" in store_detail.text
    assert "门店可见回复：总部已经联系供应商。" in store_detail.text
    assert "内部备注：供应商报价暂不外发。" not in store_detail.text


def test_store_public_comment_validates_store_and_notifies_admin(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    submit_ticket(client, store_name="南京门东店", description="门店需要补充沟通")

    mismatch = client.post(
        "/query/ticket/1/comments",
        data={"store_name": "南昌万寿宫店", "author_name": "小李", "content": "其他门店不能评论"},
    )
    assert mismatch.status_code == 404

    empty = client.post(
        "/query/ticket/1/comments",
        data={"store_name": "南京门东店", "author_name": "小李", "content": "  "},
    )
    assert empty.status_code == 400
    assert "请填写沟通内容。" in empty.text

    response = client.post(
        "/query/ticket/1/comments",
        data={"store_name": "南京门东店", "author_name": "小李", "content": "门店回复：已经补充到货照片。"},
        follow_redirects=False,
    )
    assert response.status_code == 303
    assert response.headers["location"].startswith("/query/ticket/1")

    comments = rows_for(tmp_path, "ticket_comments")
    assert len(comments) == 1
    assert comments[0]["author_type"] == "store"
    assert comments[0]["author_name"] == "小李"
    assert comments[0]["visibility"] == "public"
    assert comments[0]["content"] == "门店回复：已经补充到货照片。"

    logs = rows_for(tmp_path, "ticket_logs")
    assert logs[-1]["action"] == "门店评论"
    assert logs[-1]["operator"] == "门店:小李"

    events = rows_for(tmp_path, "notification_events")
    assert [event["event_type"] for event in events] == ["new_ticket", "ticket_comment"]
    assert events[-1]["created_by"] == "门店:小李"
    assert events[-1]["title"] == "门店新增沟通"

    detail = client.get("/query/ticket/1?store_name=南京门东店")
    assert detail.status_code == 200
    assert "门店回复：已经补充到货照片。" in detail.text
    assert "小李" in detail.text


def test_admin_tasks_lifecycle_logs_notifications_and_participants(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    submit_ticket(client, store_name="南京门东店", description="需要拆子任务")

    unauthenticated = client.post(
        "/admin/ticket/1/tasks",
        data={"title": "联系供应商确认库存", "assignee": "总部商品", "status": "待处理"},
        follow_redirects=False,
    )
    assert unauthenticated.status_code == 303
    assert unauthenticated.headers["location"].startswith("/admin/login")

    logged_in_client(client)
    participant = admin_post(
        client,
        "/admin/ticket/1/participants",
        data={"participant_type": "team", "participant_name": "总部商品", "role": "协作处理"},
        follow_redirects=False,
    )
    assert participant.status_code == 303

    create_task = admin_post(
        client,
        "/admin/ticket/1/tasks",
        data={
            "title": "联系供应商确认库存",
            "assignee": "总部商品",
            "status": "待处理",
            "due_date": "2026-07-06",
        },
        follow_redirects=False,
    )
    assert create_task.status_code == 303

    update_task = admin_post(
        client,
        "/admin/ticket/1/tasks/1",
        data={
            "title": "联系供应商确认库存",
            "assignee": "采购",
            "status": "已完成",
            "due_date": "2026-07-06",
        },
        follow_redirects=False,
    )
    assert update_task.status_code == 303

    participants = rows_for(tmp_path, "ticket_participants")
    assert participants[0]["participant_type"] == "team"
    assert participants[0]["participant_name"] == "总部商品"
    assert participants[0]["role"] == "协作处理"

    tasks = rows_for(tmp_path, "ticket_tasks")
    assert len(tasks) == 1
    assert tasks[0]["title"] == "联系供应商确认库存"
    assert tasks[0]["assignee"] == "采购"
    assert tasks[0]["status"] == "已完成"
    assert tasks[0]["due_date"] == "2026-07-06"
    assert tasks[0]["completed_at"]

    logs = rows_for(tmp_path, "ticket_logs")
    assert [log["action"] for log in logs[-3:]] == ["新增协作人", "新增子任务", "更新子任务"]
    assert "总部商品" in logs[-3]["note"]
    assert "联系供应商确认库存" in logs[-2]["note"]
    assert "已完成" in logs[-1]["note"]

    events = rows_for(tmp_path, "notification_events")
    assert [event["event_type"] for event in events] == [
        "new_ticket",
        "ticket_participant",
        "ticket_task",
        "ticket_task",
    ]
    assert events[-2]["title"] == "新增子任务"
    assert events[-1]["title"] == "子任务已更新"

    admin_detail = client.get("/admin/ticket/1")
    assert admin_detail.status_code == 200
    assert "协作人区域" in admin_detail.text
    assert "子任务区域" in admin_detail.text
    assert "总部商品" in admin_detail.text
    assert "联系供应商确认库存" in admin_detail.text
    assert "已完成" in admin_detail.text


def test_admin_my_work_shows_assigned_tickets_and_user_tasks(tmp_path, monkeypatch):
    client, main = build_client(
        tmp_path,
        monkeypatch,
        admin_users="admin:123456,caigou:123456",
        config_overrides={"handlers.json": ["admin", "caigou", "总部商品"]},
    )
    submit_ticket(client, description="admin 负责的协作工单")
    submit_ticket(client, description="采购子任务工单")

    assert_login_success(login_admin(client, "admin", "123456"))
    admin_post(
        client,
        "/admin/ticket/1",
        data={"status": "处理中", "assigned_to": "admin", "handler_note": "admin 接手"},
        follow_redirects=False,
    )
    admin_post(
        client,
        "/admin/ticket/1/tasks",
        data={"title": "跟进供应商报价", "assignee": "admin", "status": "待处理"},
        follow_redirects=False,
    )
    admin_post(
        client,
        "/admin/ticket/2/tasks",
        data={"title": "采购确认库存", "assignee": "caigou", "status": "待处理"},
        follow_redirects=False,
    )

    admin_work = client.get("/admin/my-work")
    assert admin_work.status_code == 200
    assert "我的待办" in admin_work.text
    assert "负责工单" in admin_work.text
    assert "子任务" in admin_work.text
    assert "admin 负责的协作工单" in admin_work.text
    assert "跟进供应商报价" in admin_work.text
    assert "采购确认库存" not in admin_work.text

    caigou_client = TestClient(main.app)
    assert_login_success(login_admin(caigou_client, "caigou", "123456"))
    caigou_work = caigou_client.get("/admin/my-work")
    assert caigou_work.status_code == 200
    assert "采购确认库存" in caigou_work.text
    assert "跟进供应商报价" not in caigou_work.text


def test_admin_users_are_valid_handler_options_for_detail_update(tmp_path, monkeypatch):
    client, _ = build_client(
        tmp_path,
        monkeypatch,
        config_overrides={"handlers.json": ["总部运营"]},
        admin_users=f"{ADMIN_AUTH[0]}:{ADMIN_AUTH[1]},liuhao:pw,newaccount:pw",
    )
    logged_in_client(client)
    submit_ticket(client, description="登录账号也能作为处理人的工单")

    detail_page = client.get("/admin/ticket/1")
    api_response = client.get("/api/handlers")

    assert detail_page.status_code == 200
    assert api_response.status_code == 200
    for handler in ("总部运营", ADMIN_AUTH[0], "liuhao", "newaccount"):
        assert f'value="{handler}"' in detail_page.text
        assert handler in api_response.json()["handlers"]

    update = admin_post(
        client,
        "/admin/ticket/1",
        data={"status": "处理中", "assigned_to": "liuhao", "handler_note": "liuhao 接手处理"},
        follow_redirects=False,
    )

    assert update.status_code == 303
    assert rows_for(tmp_path, "tickets")[0]["assigned_to"] == "liuhao"
    assert ticket_action_logs(tmp_path, 1)[-1]["new_assigned_to"] == "liuhao"


def test_handlers_fallback_to_admin_users_when_config_file_missing(tmp_path, monkeypatch):
    client, main = build_client(
        tmp_path,
        monkeypatch,
        admin_users=f"{ADMIN_AUTH[0]}:{ADMIN_AUTH[1]},liuhao:pw,newaccount:pw",
    )
    (tmp_path / "config" / "handlers.json").unlink()
    logged_in_client(client)

    handlers = main.load_app_config().handlers
    response = client.get("/api/handlers")

    assert handlers == [ADMIN_AUTH[0], "liuhao", "newaccount"]
    assert response.status_code == 200
    assert response.json()["handlers"] == handlers
    assert "总部运营" not in handlers


def role_id_by_name(tmp_path, role_name):
    with sqlite3.connect(tmp_path / "tickets.db") as connection:
        row = connection.execute(
            "SELECT id FROM admin_roles WHERE role_name = ?",
            (role_name,),
        ).fetchone()
    assert row is not None
    return int(row[0])


def user_id_by_username(tmp_path, username):
    with sqlite3.connect(tmp_path / "tickets.db") as connection:
        row = connection.execute(
            "SELECT id FROM admin_users WHERE username = ?",
            (username,),
        ).fetchone()
    assert row is not None
    return int(row[0])


def create_test_admin_user(
    tmp_path,
    username,
    display_name,
    password="123456",
    role_name="运营管理",
    allow_login=1,
    is_active=1,
    is_assignable=0,
):
    import main

    role_id = role_id_by_name(tmp_path, role_name)
    timestamp = "2026-07-06 10:00:00"
    with sqlite3.connect(tmp_path / "tickets.db") as connection:
        cursor = connection.execute(
            """
            INSERT INTO admin_users (
                username, display_name, password_hash, role_id, employee_id,
                allow_login, participate_schedule, is_active, is_assignable,
                data_scope, store_names, created_at, updated_at, created_by, updated_by
            )
            VALUES (?, ?, ?, ?, NULL, ?, 0, ?, ?, 'all', '', ?, ?, 'pytest', 'pytest')
            """,
            (
                username,
                display_name,
                main.hash_password(password),
                role_id,
                int(allow_login),
                int(is_active),
                int(is_assignable),
                timestamp,
                timestamp,
            ),
        )
        return int(cursor.lastrowid)


def create_role_with_permissions(tmp_path, role_name, permissions):
    with sqlite3.connect(tmp_path / "tickets.db") as connection:
        timestamp = "2026-07-06 10:00:00"
        cursor = connection.execute(
            """
            INSERT INTO admin_roles (role_name, description, is_system, created_at, updated_at)
            VALUES (?, ?, 0, ?, ?)
            """,
            (role_name, "test role", timestamp, timestamp),
        )
        role_id = int(cursor.lastrowid)
        for permission in permissions:
            connection.execute(
                """
                INSERT INTO admin_role_permissions (role_id, permission_key, created_at)
                VALUES (?, ?, ?)
                """,
                (role_id, permission, timestamp),
            )
    return role_id


def update_admin_user_access(tmp_path, username, role_id, data_scope="all", store_names=""):
    with sqlite3.connect(tmp_path / "tickets.db") as connection:
        connection.execute(
            """
            UPDATE admin_users
            SET role_id = ?, data_scope = ?, store_names = ?, is_active = 1
            WHERE username = ?
            """,
            (int(role_id), data_scope, store_names, username),
        )


def set_ticket_assignment(tmp_path, ticket_id, assigned_to):
    with sqlite3.connect(tmp_path / "tickets.db") as connection:
        connection.execute(
            "UPDATE tickets SET assigned_to = ?, updated_at = ? WHERE id = ?",
            (assigned_to, "2026-07-06 10:00:00", int(ticket_id)),
        )


def insert_assignment_rule(
    tmp_path,
    request_type="",
    brand="",
    store_name="",
    default_handler="",
    default_role_id=None,
    priority=100,
    enabled=1,
    note="pytest",
):
    timestamp = "2026-07-06 10:00:00"
    with sqlite3.connect(tmp_path / "tickets.db") as connection:
        cursor = connection.execute(
            """
            INSERT INTO ticket_assignment_rules (
                request_type, brand, store_name, default_handler, default_role_id,
                priority, enabled, note, created_at, updated_at, created_by, updated_by
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'pytest', 'pytest')
            """,
            (
                request_type,
                brand,
                store_name,
                default_handler,
                default_role_id,
                int(priority),
                int(enabled),
                note,
                timestamp,
                timestamp,
            ),
        )
    return int(cursor.lastrowid)


def insert_sla_rule(
    tmp_path,
    request_type="",
    urgency_level="",
    due_hours=24,
    enabled=1,
    note="pytest",
):
    timestamp = "2026-07-06 10:00:00"
    with sqlite3.connect(tmp_path / "tickets.db") as connection:
        cursor = connection.execute(
            """
            INSERT INTO ticket_sla_rules (
                request_type, urgency_level, due_hours, enabled, note,
                created_at, updated_at, created_by, updated_by
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, 'pytest', 'pytest')
            """,
            (
                request_type,
                urgency_level,
                int(due_hours),
                int(enabled),
                note,
                timestamp,
                timestamp,
            ),
        )
    return int(cursor.lastrowid)


def insert_request_type_template(
    tmp_path,
    request_type,
    template_name="",
    description="",
    recommended_fields="",
    required_description=1,
    image_required=0,
    file_required=0,
    expected_finish_required=0,
    field_help_text="",
    enabled=1,
    sort_order=100,
):
    timestamp = "2026-07-06 10:00:00"
    with sqlite3.connect(tmp_path / "tickets.db") as connection:
        cursor = connection.execute(
            """
            INSERT INTO request_type_templates (
                request_type, template_name, description, recommended_fields,
                required_description, image_required, file_required, expected_finish_required,
                field_help_text, enabled, sort_order, created_at, updated_at, created_by, updated_by
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'pytest', 'pytest')
            """,
            (
                request_type,
                template_name or request_type,
                description,
                recommended_fields,
                int(required_description),
                int(image_required),
                int(file_required),
                int(expected_finish_required),
                field_help_text,
                int(enabled),
                int(sort_order),
                timestamp,
                timestamp,
            ),
        )
    return int(cursor.lastrowid)


def test_rbac_tables_env_migration_password_hashes_and_login_logs(tmp_path, monkeypatch):
    client, _ = build_client(
        tmp_path,
        monkeypatch,
        admin_users="admin:123456,caigou:123456",
    )

    with sqlite3.connect(tmp_path / "tickets.db") as connection:
        tables = {
            row[0]
            for row in connection.execute(
                "SELECT name FROM sqlite_master WHERE type = 'table'"
            ).fetchall()
        }
    assert {
        "admin_users",
        "admin_roles",
        "admin_role_permissions",
        "admin_login_logs",
        "admin_operation_logs",
    }.issubset(tables)

    roles = {row["id"]: row["role_name"] for row in rows_for(tmp_path, "admin_roles")}
    users = {row["username"]: row for row in rows_for(tmp_path, "admin_users")}
    assert roles[users["admin"]["role_id"]] == "系统管理员"
    assert roles[users["caigou"]["role_id"]] == "运营管理"
    assert users["admin"]["password_hash"].startswith("pbkdf2_sha256$")
    assert users["admin"]["password_hash"] != "123456"
    assert users["admin"]["is_active"] == 1
    assert users["admin"]["is_assignable"] == 1
    assert users["admin"]["data_scope"] == "all"

    bad_login = login_admin(client, "admin", "wrong-password")
    assert bad_login.status_code == 400
    assert_login_success(login_admin(client, "admin", "123456"))

    with sqlite3.connect(tmp_path / "tickets.db") as connection:
        connection.execute("UPDATE admin_users SET is_active = 0 WHERE username = 'caigou'")
    disabled = login_admin(client, "caigou", "123456")
    assert disabled.status_code == 400

    login_logs = rows_for(tmp_path, "admin_login_logs")
    assert [row["username"] for row in login_logs] == ["admin", "admin", "caigou"]
    assert [row["success"] for row in login_logs] == [0, 1, 0]
    assert "wrong-password" not in json.dumps(login_logs, ensure_ascii=False)
    assert rows_for(tmp_path, "admin_users")[0]["last_login_at"]


def test_admin_access_safeguard_restores_env_first_admin_and_permissions(tmp_path, monkeypatch):
    client, main = build_client(
        tmp_path,
        monkeypatch,
        admin_users="admin:123456,ops:123456",
    )
    system_role_id = role_id_by_name(tmp_path, "系统管理员")
    ops_role_id = role_id_by_name(tmp_path, "运营管理")
    with sqlite3.connect(tmp_path / "tickets.db") as connection:
        connection.execute("DELETE FROM admin_role_permissions WHERE role_id = ?", (system_role_id,))
        connection.execute(
            """
            UPDATE admin_users
            SET role_id = ?, is_active = 0, is_assignable = 0, data_scope = 'stores', store_names = '南京门东店'
            WHERE username = 'admin'
            """,
            (ops_role_id,),
        )

    main.ensure_admin_access_safeguard()

    users = {row["username"]: row for row in rows_for(tmp_path, "admin_users")}
    roles = {row["id"]: row["role_name"] for row in rows_for(tmp_path, "admin_roles")}
    admin = users["admin"]
    assert admin["is_active"] == 1
    assert admin["is_assignable"] == 1
    assert admin["data_scope"] == "all"
    assert admin["store_names"] == ""
    assert roles[admin["role_id"]] == "系统管理员"
    with sqlite3.connect(tmp_path / "tickets.db") as connection:
        permissions = {
            row[0]
            for row in connection.execute(
                "SELECT permission_key FROM admin_role_permissions WHERE role_id = ?",
                (system_role_id,),
            ).fetchall()
        }
    assert permissions == set(main.ADMIN_PERMISSION_KEYS)
    assert_login_success(login_admin(client, "admin", "123456"))
    assert client.get("/admin/dashboard").status_code == 200
    assert client.get("/admin/account").status_code == 200
    assert client.get("/admin/permission-overview").status_code == 200


def test_repair_admin_access_script_is_present_and_secret_safe():
    script_path = PROJECT_DIR / "repair_admin_access.py"
    assert script_path.exists()
    script = script_path.read_text(encoding="utf-8")
    assert "ensure_admin_access_safeguard" in script
    assert "reconfigure" in script
    assert "print(password" not in script
    assert "DELETE FROM" not in script.upper()


def test_account_management_crud_safety_and_password_reset(tmp_path, monkeypatch):
    client, _ = build_client(
        tmp_path,
        monkeypatch,
        admin_users="admin:123456",
    )
    assert_login_success(login_admin(client, "admin", "123456"))
    csrf_token = csrf_token_for(client, "/admin/account")
    ops_role_id = role_id_by_name(tmp_path, "运营管理")

    account_page = client.get("/admin/account")
    assert account_page.status_code == 200
    assert "账号列表" in account_page.text
    assert "新增账号" in account_page.text
    assert "password_hash" not in account_page.text

    create = client.post(
        "/admin/account/users",
        data={
            "csrf_token": csrf_token,
            "username": "buyer01",
            "display_name": "采购01",
            "password": "newpass1",
            "password_confirm": "newpass1",
            "role_id": str(ops_role_id),
            "data_scope": "stores",
            "store_names": "南京门东店",
            "is_assignable": "1",
            "is_active": "1",
        },
        follow_redirects=False,
    )
    assert create.status_code == 303
    buyer = next(row for row in rows_for(tmp_path, "admin_users") if row["username"] == "buyer01")
    assert buyer["display_name"] == "采购01"
    assert buyer["role_id"] == ops_role_id
    assert buyer["data_scope"] == "stores"
    assert buyer["store_names"] == "南京门东店"
    assert buyer["is_assignable"] == 1
    assert_login_success(login_admin(client, "buyer01", "newpass1"))

    assert_login_success(login_admin(client, "admin", "123456"))
    buyer_id = user_id_by_username(tmp_path, "buyer01")
    csrf_token = csrf_token_for(client, "/admin/account")
    update = client.post(
        f"/admin/account/users/{buyer_id}/update",
        data={
            "csrf_token": csrf_token,
            "display_name": "采购专员01",
            "role_id": str(ops_role_id),
            "data_scope": "assigned",
            "store_names": "",
            "is_assignable": "",
            "is_active": "1",
        },
        follow_redirects=False,
    )
    assert update.status_code == 303
    buyer = next(row for row in rows_for(tmp_path, "admin_users") if row["username"] == "buyer01")
    assert buyer["display_name"] == "采购专员01"
    assert buyer["data_scope"] == "assigned"
    assert buyer["is_assignable"] == 0

    disable = client.post(
        f"/admin/account/users/{buyer_id}/disable",
        data={"csrf_token": csrf_token_for(client, "/admin/account")},
        follow_redirects=False,
    )
    assert disable.status_code == 303
    assert next(row for row in rows_for(tmp_path, "admin_users") if row["username"] == "buyer01")["is_active"] == 0
    assert login_admin(client, "buyer01", "newpass1").status_code == 400

    enable = client.post(
        f"/admin/account/users/{buyer_id}/enable",
        data={"csrf_token": csrf_token_for(client, "/admin/account")},
        follow_redirects=False,
    )
    assert enable.status_code == 303

    reset = client.post(
        f"/admin/account/users/{buyer_id}/reset-password",
        data={
            "csrf_token": csrf_token_for(client, "/admin/account"),
            "password": "reset123",
            "password_confirm": "reset123",
        },
        follow_redirects=False,
    )
    assert reset.status_code == 303
    assert login_admin(client, "buyer01", "newpass1").status_code == 400
    assert_login_success(login_admin(client, "buyer01", "reset123"))

    assert_login_success(login_admin(client, "admin", "123456"))
    admin_id = user_id_by_username(tmp_path, "admin")
    self_disable = client.post(
        f"/admin/account/users/{admin_id}/disable",
        data={"csrf_token": csrf_token_for(client, "/admin/account")},
    )
    assert self_disable.status_code == 400
    assert "不能停用当前登录账号" in self_disable.text
    assert "最后一个系统管理员" in self_disable.text

    operation_logs = rows_for(tmp_path, "admin_operation_logs")
    assert [row["action"] for row in operation_logs] == [
        "account.create",
        "account.update",
        "account.disable",
        "account.enable",
        "account.reset_password",
    ]
    assert "reset123" not in json.dumps(operation_logs, ensure_ascii=False)


def test_person_account_schema_and_legacy_records_remain_usable(tmp_path, monkeypatch):
    client, main = build_client(
        tmp_path,
        monkeypatch,
        admin_users="admin:123456,legacyops:123456",
    )
    logged_in_client(client, "admin", "123456")
    legacy_employee_id = main.create_employee("旧排班员工", "南京门东店", "店员", "13800000001", "在职", main.load_app_config())

    with sqlite3.connect(tmp_path / "tickets.db") as connection:
        admin_columns = {row[1] for row in connection.execute("PRAGMA table_info(admin_users)").fetchall()}
        employee_columns = {row[1] for row in connection.execute("PRAGMA table_info(employees)").fetchall()}

    assert {"employee_id", "allow_login", "participate_schedule"}.issubset(admin_columns)
    assert {"user_id", "allow_login", "participate_schedule"}.issubset(employee_columns)

    users = {row["username"]: row for row in rows_for(tmp_path, "admin_users")}
    assert users["admin"]["allow_login"] == 1
    assert users["legacyops"]["allow_login"] == 1
    assert users["legacyops"]["participate_schedule"] == 0
    assert_login_success(login_admin(client, "legacyops", "123456"))
    assert_login_success(login_admin(client, "admin", "123456"))

    employees = {row["employee_name"]: row for row in rows_for(tmp_path, "employees")}
    assert employees["旧排班员工"]["allow_login"] == 0
    assert employees["旧排班员工"]["participate_schedule"] == 1
    schedule_page = client.get("/admin/schedules?store_name=南京门东店&month=2026-07")
    assert schedule_page.status_code == 200
    assert "旧排班员工" in schedule_page.text
    assert rows_for(tmp_path, "employees")[0]["id"] == legacy_employee_id


def test_person_account_management_creates_login_schedule_and_combined_people(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch, admin_users="admin:123456")
    assert_login_success(login_admin(client, "admin", "123456"))
    ops_role_id = role_id_by_name(tmp_path, "运营管理")

    page = client.get("/admin/account")
    assert page.status_code == 200
    assert "人员与账号管理" in page.text
    assert "总人员数" in page.text
    assert "允许登录后台" in page.text
    assert "参与排班" in page.text

    login_only = client.post(
        "/admin/account/users",
        data={
            "csrf_token": csrf_token_for(client, "/admin/account"),
            "employee_name": "总部采购",
            "phone": "13800001001",
            "employee_role": "采购专员",
            "allow_login": "1",
            "username": "hqbuyer",
            "password": "buyer123",
            "password_confirm": "buyer123",
            "role_id": str(ops_role_id),
            "data_scope": "all",
            "is_assignable": "1",
            "is_active": "1",
            "participate_schedule": "",
        },
        follow_redirects=False,
    )
    assert login_only.status_code == 303
    buyer_user = next(row for row in rows_for(tmp_path, "admin_users") if row["username"] == "hqbuyer")
    assert buyer_user["allow_login"] == 1
    assert buyer_user["participate_schedule"] == 0
    assert buyer_user["employee_id"] is None
    assert_login_success(login_admin(client, "hqbuyer", "buyer123"))
    assert "hqbuyer" in client.get("/api/handlers").json()["handlers"]
    assert "总部采购" not in client.get("/admin/schedules?store_name=南京门东店&month=2026-07").text

    assert_login_success(login_admin(client, "admin", "123456"))
    schedule_only = client.post(
        "/admin/account/users",
        data={
            "csrf_token": csrf_token_for(client, "/admin/account"),
            "employee_name": "普通店员",
            "phone": "13800001002",
            "employee_role": "店员",
            "primary_store_name": "南京门东店",
            "schedule_store_names": ["南京门东店", "南昌万寿宫店"],
            "allow_login": "0",
            "participate_schedule": "1",
            "employee_status": "在职",
            "is_assignable": "1",
        },
        follow_redirects=False,
    )
    assert schedule_only.status_code == 303
    staff_employee = next(row for row in rows_for(tmp_path, "employees") if row["employee_name"] == "普通店员")
    assert staff_employee["allow_login"] == 0
    assert staff_employee["participate_schedule"] == 1
    assert staff_employee["user_id"] is None
    assert not any(row["username"] == "普通店员" for row in rows_for(tmp_path, "admin_users"))
    assert "普通店员" in client.get("/admin/schedules?store_name=南京门东店&month=2026-07").text
    assert "普通店员" not in client.get("/api/handlers").json()["handlers"]

    combined = client.post(
        "/admin/account/users",
        data={
            "csrf_token": csrf_token_for(client, "/admin/account"),
            "employee_name": "门店店长",
            "phone": "13800001003",
            "employee_role": "店长",
            "primary_store_name": "南京门东店",
            "schedule_store_names": ["南京门东店"],
            "allow_login": "1",
            "username": "storemanager",
            "password": "manager123",
            "password_confirm": "manager123",
            "role_id": str(ops_role_id),
            "data_scope": "stores",
            "store_names": "南京门东店",
            "is_assignable": "1",
            "is_active": "1",
            "participate_schedule": "1",
            "employee_status": "在职",
        },
        follow_redirects=False,
    )
    assert combined.status_code == 303
    manager_user = next(row for row in rows_for(tmp_path, "admin_users") if row["username"] == "storemanager")
    manager_employee = next(row for row in rows_for(tmp_path, "employees") if row["employee_name"] == "门店店长")
    assert manager_user["allow_login"] == 1
    assert manager_user["participate_schedule"] == 1
    assert manager_user["employee_id"] == manager_employee["id"]
    assert manager_employee["user_id"] == manager_user["id"]
    assert manager_employee["allow_login"] == 1
    assert manager_employee["participate_schedule"] == 1
    assert_login_success(login_admin(client, "storemanager", "manager123"))
    assert_login_success(login_admin(client, "admin", "123456"))
    assert "门店店长" in client.get("/admin/schedules?store_name=南京门东店&month=2026-07").text
    assert "storemanager" in client.get("/api/handlers").json()["handlers"]


def test_person_status_flags_control_login_schedule_handlers_and_password_reset(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch, admin_users="admin:123456")
    assert_login_success(login_admin(client, "admin", "123456"))
    ops_role_id = role_id_by_name(tmp_path, "运营管理")
    create = client.post(
        "/admin/account/users",
        data={
            "csrf_token": csrf_token_for(client, "/admin/account"),
            "employee_name": "可变店长",
            "employee_role": "店长",
            "phone": "13800002001",
            "primary_store_name": "南京门东店",
            "schedule_store_names": ["南京门东店"],
            "allow_login": "1",
            "username": "mutablemanager",
            "password": "mutable123",
            "password_confirm": "mutable123",
            "role_id": str(ops_role_id),
            "data_scope": "all",
            "is_assignable": "1",
            "is_active": "1",
            "participate_schedule": "1",
            "employee_status": "在职",
        },
        follow_redirects=False,
    )
    assert create.status_code == 303
    user_id = user_id_by_username(tmp_path, "mutablemanager")
    employee_id = next(row["id"] for row in rows_for(tmp_path, "employees") if row["employee_name"] == "可变店长")

    assert_login_success(login_admin(client, "mutablemanager", "mutable123"))
    assert_login_success(login_admin(client, "admin", "123456"))
    update = client.post(
        f"/admin/account/users/{user_id}/update",
        data={
            "csrf_token": csrf_token_for(client, "/admin/account"),
            "employee_name": "可变店长",
            "employee_role": "店长",
            "phone": "13800002001",
            "primary_store_name": "南京门东店",
            "schedule_store_names": ["南京门东店"],
            "allow_login": "0",
            "username": "mutablemanager",
            "role_id": str(ops_role_id),
            "data_scope": "all",
            "store_names": "",
            "is_assignable": "1",
            "is_active": "1",
            "participate_schedule": "1",
            "employee_status": "在职",
        },
        follow_redirects=False,
    )
    assert update.status_code == 303
    user = next(row for row in rows_for(tmp_path, "admin_users") if row["id"] == user_id)
    assert user["allow_login"] == 0
    assert user["is_assignable"] == 0
    assert login_admin(client, "mutablemanager", "mutable123").status_code == 400
    assert "可变店长" in client.get("/admin/schedules?store_name=南京门东店&month=2026-07").text
    assert "mutablemanager" not in client.get("/api/handlers").json()["handlers"]

    reset_blocked = client.post(
        f"/admin/account/users/{user_id}/reset-password",
        data={
            "csrf_token": csrf_token_for(client, "/admin/account"),
            "password": "newpass1",
            "password_confirm": "newpass1",
        },
    )
    assert reset_blocked.status_code == 400
    assert "未允许登录" in reset_blocked.text

    leave = client.post(
        f"/admin/account/users/{user_id}/update",
        data={
            "csrf_token": csrf_token_for(client, "/admin/account"),
            "employee_name": "可变店长",
            "employee_role": "店长",
            "phone": "13800002001",
            "primary_store_name": "南京门东店",
            "schedule_store_names": ["南京门东店"],
            "allow_login": "1",
            "username": "mutablemanager",
            "role_id": str(ops_role_id),
            "data_scope": "all",
            "store_names": "",
            "is_assignable": "1",
            "is_active": "1",
            "participate_schedule": "1",
            "employee_status": "离职",
        },
        follow_redirects=False,
    )
    assert leave.status_code == 303
    user = next(row for row in rows_for(tmp_path, "admin_users") if row["id"] == user_id)
    employee = next(row for row in rows_for(tmp_path, "employees") if row["id"] == employee_id)
    assert user["allow_login"] == 0
    assert user["is_active"] == 0
    assert user["is_assignable"] == 0
    assert user["participate_schedule"] == 0
    assert employee["allow_login"] == 0
    assert employee["participate_schedule"] == 0
    assert employee["status"] == "离职"
    assert login_admin(client, "mutablemanager", "mutable123").status_code == 400
    assert "可变店长" not in client.get("/admin/schedules?store_name=南京门东店&month=2026-07").text
    assert "mutablemanager" not in client.get("/api/handlers").json()["handlers"]


def test_employees_page_is_schedule_people_view_with_unified_management_hint(tmp_path, monkeypatch):
    client, main = build_client(tmp_path, monkeypatch)
    logged_in_client(client)
    main.create_employee("排班视图员工", "南京门东店", "店员", "", "在职", main.load_app_config())
    with sqlite3.connect(tmp_path / "tickets.db") as connection:
        connection.execute(
            """
            UPDATE employees
            SET participate_schedule = 0
            WHERE employee_name = '排班视图员工'
            """
        )

    page = client.get("/admin/employees")

    assert page.status_code == 200
    assert "排班人员视图" in page.text
    assert "人员资料请统一到“人员与账号管理”维护" in page.text
    assert 'href="/admin/account?filter=participate_schedule"' in page.text
    assert "排班视图员工" not in page.text
    assert 'data-drawer-open="employee-create-drawer"' not in page.text


def test_personnel_governance_detects_candidates_and_ignore_roundtrip(tmp_path, monkeypatch):
    client, main = build_client(tmp_path, monkeypatch, admin_users="admin:123456")
    logged_in_client(client, "admin", "123456")
    user_id = create_test_admin_user(tmp_path, "13800003001", "治理店长", is_assignable=1)
    employee_id = main.create_employee(
        "治理店长",
        "南京门东店",
        "店长",
        "13800003001",
        "在职",
        main.load_app_config(),
    )
    create_test_admin_user(tmp_path, "lisi_ops", "李四")
    main.create_employee("李四", "南昌万寿宫店", "店员", "", "在职", main.load_app_config())
    create_test_admin_user(tmp_path, "wangwu_3003", "王五")
    main.create_employee("小王五", "山城巷店", "店员", "13900003003", "在职", main.load_app_config())

    with sqlite3.connect(tmp_path / "tickets.db") as connection:
        tables = {row[0] for row in connection.execute("SELECT name FROM sqlite_master WHERE type = 'table'")}
    assert "personnel_match_ignores" in tables

    page = client.get("/admin/personnel-governance")
    assert page.status_code == 200
    assert "人员数据治理" in page.text
    assert "高置信" in page.text
    assert "中置信" in page.text
    assert "低置信" in page.text
    assert "治理店长" in page.text
    assert "13800003001" in page.text

    ignored = client.post(
        "/admin/personnel-governance/ignore",
        data={
            "csrf_token": csrf_token_for(client, "/admin/personnel-governance"),
            "user_id": str(user_id),
            "employee_id": str(employee_id),
            "reason": "测试暂不合并",
        },
        follow_redirects=False,
    )
    assert ignored.status_code == 303
    pending_page = client.get("/admin/personnel-governance?filter=unprocessed")
    assert "测试暂不合并" not in pending_page.text
    ignored_page = client.get("/admin/personnel-governance?filter=ignored")
    assert ignored_page.status_code == 200
    assert "测试暂不合并" in ignored_page.text

    unignored = client.post(
        "/admin/personnel-governance/unignore",
        data={
            "csrf_token": csrf_token_for(client, "/admin/personnel-governance?filter=ignored"),
            "user_id": str(user_id),
            "employee_id": str(employee_id),
        },
        follow_redirects=False,
    )
    assert unignored.status_code == 303
    assert "治理店长" in client.get("/admin/personnel-governance?filter=unprocessed").text
    operation_actions = [row["action"] for row in rows_for(tmp_path, "admin_operation_logs")]
    assert "personnel.match.ignore" in operation_actions
    assert "personnel.match.unignore" in operation_actions


def test_personnel_governance_link_unlink_preserves_data_and_prevents_one_to_many(tmp_path, monkeypatch):
    client, main = build_client(
        tmp_path,
        monkeypatch,
        admin_users="admin:123456",
        config_overrides={"handlers.json": ["传统处理人"]},
    )
    logged_in_client(client, "admin", "123456")
    user_id = create_test_admin_user(tmp_path, "governance_manager", "治理店长", password="link123", is_assignable=1)
    second_user_id = create_test_admin_user(tmp_path, "other_manager", "其他店长", password="other123", is_assignable=1)
    employee_id = main.create_employee(
        "治理店长",
        "南京门东店",
        "店长",
        "13800003011",
        "在职",
        main.load_app_config(),
    )
    second_employee_id = main.create_employee(
        "其他店长",
        "南昌万寿宫店",
        "店长",
        "13800003012",
        "在职",
        main.load_app_config(),
    )
    with sqlite3.connect(tmp_path / "tickets.db") as connection:
        shift_id = connection.execute("SELECT id FROM shift_types ORDER BY id LIMIT 1").fetchone()[0]
        connection.execute(
            """
            INSERT INTO store_schedules (store_name, employee_id, schedule_date, shift_type_id, created_by, created_at, updated_at)
            VALUES ('南京门东店', ?, '2026-07-06', ?, 'pytest', '2026-07-06 10:00:00', '2026-07-06 10:00:00')
            """,
            (employee_id, shift_id),
        )

    linked = client.post(
        "/admin/personnel-governance/link",
        data={
            "csrf_token": csrf_token_for(client, "/admin/personnel-governance"),
            "user_id": str(user_id),
            "employee_id": str(employee_id),
        },
        follow_redirects=False,
    )
    assert linked.status_code == 303
    users = {row["username"]: row for row in rows_for(tmp_path, "admin_users")}
    employees = {row["id"]: row for row in rows_for(tmp_path, "employees")}
    assert users["governance_manager"]["employee_id"] == employee_id
    assert employees[employee_id]["user_id"] == user_id
    people = [
        person
        for person in main.fetch_admin_users_for_account_page()
        if person.get("username") == "governance_manager" or person.get("employee_id") == employee_id
    ]
    assert len(people) == 1
    handlers = client.get("/api/handlers").json()["handlers"]
    assert handlers.count("governance_manager") == 1
    schedule_page = client.get("/admin/schedules?store_name=南京门东店&month=2026-07")
    assert schedule_page.status_code == 200
    assert schedule_page.text.count("治理店长") >= 1

    one_to_many = client.post(
        "/admin/personnel-governance/link",
        data={
            "csrf_token": csrf_token_for(client, "/admin/personnel-governance"),
            "user_id": str(second_user_id),
            "employee_id": str(employee_id),
        },
    )
    assert one_to_many.status_code == 400
    assert "已关联" in one_to_many.text
    many_to_one = client.post(
        "/admin/personnel-governance/link",
        data={
            "csrf_token": csrf_token_for(client, "/admin/personnel-governance"),
            "user_id": str(user_id),
            "employee_id": str(second_employee_id),
        },
    )
    assert many_to_one.status_code == 400
    assert "已关联" in many_to_one.text

    unlinked = client.post(
        "/admin/personnel-governance/unlink",
        data={
            "csrf_token": csrf_token_for(client, "/admin/personnel-governance"),
            "user_id": str(user_id),
            "employee_id": str(employee_id),
        },
        follow_redirects=False,
    )
    assert unlinked.status_code == 303
    users = {row["username"]: row for row in rows_for(tmp_path, "admin_users")}
    employees = {row["id"]: row for row in rows_for(tmp_path, "employees")}
    assert users["governance_manager"]["employee_id"] is None
    assert employees[employee_id]["user_id"] is None
    assert rows_for(tmp_path, "store_schedules")[0]["employee_id"] == employee_id
    assert_login_success(login_admin(client, "governance_manager", "link123"))
    assert_login_success(login_admin(client, "admin", "123456"))
    assert "治理店长" in client.get("/admin/schedules?store_name=南京门东店&month=2026-07").text
    operation_actions = [row["action"] for row in rows_for(tmp_path, "admin_operation_logs")]
    assert "personnel.link" in operation_actions
    assert "personnel.unlink" in operation_actions


def test_personnel_governance_permissions_detail_and_permission_overview(tmp_path, monkeypatch):
    client, main = build_client(tmp_path, monkeypatch, admin_users="admin:123456,viewer:123456")
    logged_in_client(client, "admin", "123456")
    user_id = create_test_admin_user(tmp_path, "detail_user", "详情人员", is_assignable=1)
    employee_id = main.create_employee(
        "详情人员",
        "南京门东店",
        "店长",
        "13800003021",
        "在职",
        main.load_app_config(),
    )
    role_id = create_role_with_permissions(tmp_path, "governance-viewer", ["account.view"])
    update_admin_user_access(tmp_path, "viewer", role_id)
    admin_post(client, "/admin/logout", follow_redirects=False)
    assert_login_success(login_admin(client, "viewer", "123456"))

    page = client.get("/admin/personnel-governance")
    assert page.status_code == 200
    assert "人员数据治理" in page.text
    assert_html_forbidden(
        client.post(
            "/admin/personnel-governance/link",
            data={
                "csrf_token": csrf_token_for(client, "/admin/personnel-governance"),
                "user_id": str(user_id),
                "employee_id": str(employee_id),
            },
        )
    )
    assert_html_forbidden(
        client.post(
            "/admin/personnel-governance/ignore",
            data={
                "csrf_token": csrf_token_for(client, "/admin/personnel-governance"),
                "user_id": str(user_id),
                "employee_id": str(employee_id),
                "reason": "无权限测试",
            },
        )
    )
    assert rows_for(tmp_path, "personnel_match_ignores") == []

    detail = client.get(f"/admin/account/detail?user_id={user_id}&employee_id={employee_id}")
    assert detail.status_code == 200
    assert "人员详情" in detail.text
    assert "登录账号信息" in detail.text
    assert "业务身份" in detail.text
    assert "治理记录" in detail.text

    context = main.permission_overview_context(main.app)
    assert context["high_risk_uncontrolled_count"] == 0
    route_paths = {item["path"] for item in context["route_items"]}
    assert "/admin/personnel-governance/link" in route_paths
    assert "/admin/personnel-governance/ignore" in route_paths


def test_account_management_requires_account_permissions(tmp_path, monkeypatch):
    client, _ = build_client(
        tmp_path,
        monkeypatch,
        admin_users="admin:123456,caigou:123456",
    )

    assert_login_success(login_admin(client, "caigou", "123456"))
    denied_page = client.get("/admin/account")
    assert denied_page.status_code == 403
    assert "权限不足" in denied_page.text
    assert "application/json" not in denied_page.headers.get("content-type", "")

    denied_create = client.post(
        "/admin/account/users",
        data={
            "csrf_token": csrf_token_for(client, "/admin"),
            "username": "buyer02",
            "display_name": "采购02",
            "password": "newpass1",
            "password_confirm": "newpass1",
            "role_id": str(role_id_by_name(tmp_path, "运营管理")),
            "data_scope": "all",
        },
    )
    assert denied_create.status_code == 403
    assert "权限不足" in denied_create.text
    assert not any(row["username"] == "buyer02" for row in rows_for(tmp_path, "admin_users"))


def test_database_assignable_admin_users_extend_handler_candidates(tmp_path, monkeypatch):
    client, _ = build_client(
        tmp_path,
        monkeypatch,
        config_overrides={"handlers.json": ["总部运营"]},
        admin_users="admin:123456,liuhao:123456,caigou:123456,yunying:123456",
    )
    with sqlite3.connect(tmp_path / "tickets.db") as connection:
        connection.execute("UPDATE admin_users SET is_assignable = 0 WHERE username = 'caigou'")
        connection.execute("UPDATE admin_users SET is_active = 0 WHERE username = 'yunying'")
    assert_login_success(login_admin(client, "admin", "123456"))
    submit_ticket(client, description="数据库账号处理人测试工单")

    detail_page = client.get("/admin/ticket/1")
    handlers = client.get("/api/handlers").json()["handlers"]

    assert detail_page.status_code == 200
    assert "总部运营" in handlers
    assert "admin" in handlers
    assert "liuhao" in handlers
    assert "caigou" not in handlers
    assert "yunying" not in handlers
    assert 'value="liuhao"' in detail_page.text
    assert 'value="caigou"' not in detail_page.text
    assert 'value="yunying"' not in detail_page.text

    update = admin_post(
        client,
        "/admin/ticket/1",
        data={"status": "处理中", "assigned_to": "liuhao", "handler_note": "liuhao 接手处理"},
        follow_redirects=False,
    )

    assert update.status_code == 303
    assert rows_for(tmp_path, "tickets")[0]["assigned_to"] == "liuhao"
    assert ticket_action_logs(tmp_path, 1)[-1]["new_assigned_to"] == "liuhao"


def test_phase2_ticket_permissions_and_data_scope_are_enforced(tmp_path, monkeypatch):
    client, _ = build_client(
        tmp_path,
        monkeypatch,
        admin_users="admin:123456,blocked:123456,storeuser:123456,assigneduser:123456",
    )
    no_ticket_role = create_role_with_permissions(tmp_path, "no-ticket-access", [])
    store_role = create_role_with_permissions(tmp_path, "store-ticket-exporter", ["ticket.view", "ticket.export"])
    assigned_role = create_role_with_permissions(tmp_path, "assigned-ticket-exporter", ["ticket.view", "ticket.export"])
    update_admin_user_access(tmp_path, "blocked", no_ticket_role)
    update_admin_user_access(tmp_path, "storeuser", store_role, data_scope="stores", store_names="南京门东店")
    update_admin_user_access(tmp_path, "assigneduser", assigned_role, data_scope="assigned")

    submit_ticket(client, store_name="南京门东店", description="phase2 visible nanjing ticket")
    submit_ticket(client, store_name="南昌万寿宫店", description="phase2 hidden nanchang ticket")
    set_ticket_assignment(tmp_path, 1, "assigneduser")
    set_ticket_assignment(tmp_path, 2, "someone-else")

    assert_login_success(login_admin(client, "blocked", "123456"))
    denied_list = client.get("/admin")
    assert denied_list.status_code == 403
    assert "权限不足" in denied_list.text
    assert "application/json" not in denied_list.headers.get("content-type", "")
    denied_export = client.get("/admin/export")
    assert denied_export.status_code == 403

    assert_login_success(login_admin(client, "storeuser", "123456"))
    list_page = client.get("/admin")
    assert list_page.status_code == 200
    assert "phase2 visible nanjing ticket" in list_page.text
    assert "phase2 hidden nanchang ticket" not in list_page.text
    bypass_page = client.get("/admin?store_name=南昌万寿宫店")
    assert bypass_page.status_code == 200
    assert "phase2 hidden nanchang ticket" not in bypass_page.text

    store_export = client.get("/admin/export")
    assert store_export.status_code == 200
    workbook = load_workbook(BytesIO(store_export.content))
    descriptions = {row[10].value for row in workbook.active.iter_rows(min_row=2)}
    assert "phase2 visible nanjing ticket" in descriptions
    assert "phase2 hidden nanchang ticket" not in descriptions

    assert_login_success(login_admin(client, "assigneduser", "123456"))
    assigned_page = client.get("/admin")
    assert assigned_page.status_code == 200
    assert "phase2 visible nanjing ticket" in assigned_page.text
    assert "phase2 hidden nanchang ticket" not in assigned_page.text
    assigned_export = client.get("/admin/export")
    assert assigned_export.status_code == 200
    assigned_workbook = load_workbook(BytesIO(assigned_export.content))
    assigned_descriptions = {row[10].value for row in assigned_workbook.active.iter_rows(min_row=2)}
    assert assigned_descriptions == {"phase2 visible nanjing ticket"}

    operation_logs = rows_for(tmp_path, "admin_operation_logs")
    assert any(row["action"] == "ticket.export" and row["username"] == "storeuser" for row in operation_logs)


def test_phase2_schedule_employee_and_roles_permissions_are_enforced(tmp_path, monkeypatch):
    client, main = build_client(
        tmp_path,
        monkeypatch,
        admin_users="admin:123456,limited:123456,ops:123456",
    )
    limited_role = create_role_with_permissions(tmp_path, "phase2-limited", [])
    update_admin_user_access(tmp_path, "limited", limited_role)

    no_login_client = TestClient(main.app)
    no_login_response = no_login_client.get("/admin/roles", follow_redirects=False)
    assert no_login_response.status_code == 303
    assert no_login_response.headers["location"].startswith("/admin/login")

    assert_login_success(login_admin(client, "limited", "123456"))
    assert client.get("/admin/schedules").status_code == 403
    assert client.get("/admin/employees").status_code == 403
    assert client.get("/admin/schedules/export").status_code == 403

    assert_login_success(login_admin(client, "admin", "123456"))
    roles_page = client.get("/admin/roles")
    assert roles_page.status_code == 200
    assert "ticket.view" in roles_page.text
    ops_role_id = role_id_by_name(tmp_path, "运营管理")
    grant = client.post(
        f"/admin/roles/{ops_role_id}/permissions",
        data={
            "csrf_token": csrf_token_for(client, "/admin/roles"),
            "permissions": ["ticket.view", "schedule.view", "employee.view"],
        },
        follow_redirects=False,
    )
    assert grant.status_code == 303

    assert_login_success(login_admin(client, "ops", "123456"))
    assert client.get("/admin/schedules").status_code == 200
    assert client.get("/admin/employees").status_code == 200
    assert client.get("/admin/export").status_code == 403

    assert_login_success(login_admin(client, "admin", "123456"))
    revoke = client.post(
        f"/admin/roles/{ops_role_id}/permissions",
        data={"csrf_token": csrf_token_for(client, "/admin/roles"), "permissions": ["ticket.view"]},
        follow_redirects=False,
    )
    assert revoke.status_code == 303

    assert_login_success(login_admin(client, "ops", "123456"))
    assert client.get("/admin/schedules").status_code == 403
    assert client.get("/admin/employees").status_code == 403

    operation_logs = rows_for(tmp_path, "admin_operation_logs")
    assert any(row["action"] == "role.permissions.update" for row in operation_logs)


def test_phase3_permission_service_scope_audit_and_overview(tmp_path, monkeypatch):
    client, main = build_client(
        tmp_path,
        monkeypatch,
        admin_users="admin:123456,storeuser:123456",
    )
    store_role = create_role_with_permissions(
        tmp_path,
        "phase3-store-governance",
        ["ticket.view", "ticket.export", "schedule.view", "schedule.export", "role.view"],
    )
    update_admin_user_access(tmp_path, "storeuser", store_role, data_scope="stores", store_names="南京门东店")

    submit_ticket(client, store_name="南京门东店", description="phase3 visible ticket")
    submit_ticket(client, store_name="南昌万寿宫店", description="phase3 hidden ticket")
    visible_employee_id = main.create_employee("phase3 visible employee", "南京门东店", "店员", "", "在职", main.load_app_config())
    hidden_employee_id = main.create_employee("phase3 hidden employee", "南昌万寿宫店", "店员", "", "在职", main.load_app_config())
    visible_shift_id = main.create_shift_type("phase3 visible shift", "09:00", "18:00", "8", "#2563eb", store_name="南京门东店", config=main.load_app_config())
    hidden_shift_id = main.create_shift_type("phase3 hidden shift", "09:00", "18:00", "8", "#dc2626", store_name="南昌万寿宫店", config=main.load_app_config())
    main.bulk_upsert_schedules("南京门东店", [visible_employee_id], ["2026-07-07"], visible_shift_id, "", False, "admin", 20)
    main.bulk_upsert_schedules("南昌万寿宫店", [hidden_employee_id], ["2026-07-07"], hidden_shift_id, "", False, "admin", 20)

    user = main.fetch_admin_user_by_username("storeuser")
    assert user is not None
    assert isinstance(main.permission_service, main.PermissionService)
    assert main.permission_service.has_permission(user, "ticket.view")
    assert "ticket.export" in main.permission_service.get_user_permissions(user)
    assert "ticket.view" in main.permission_service.get_role_permissions(store_role)

    scoped_filters = main.apply_data_scope({}, user)
    scoped_tickets = main.fetch_tickets(scoped_filters, "newest", main.load_app_config())
    assert {ticket["description"] for ticket in scoped_tickets} == {"phase3 visible ticket"}
    assert main.permission_service.filter_query_by_scope({}, user) == scoped_filters

    scoped_schedule_query = main.apply_data_scope({"__scope_kind": "schedule", "store_names": ["南京门东店", "南昌万寿宫店"], "is_all_stores": False}, user)
    assert scoped_schedule_query["store_names"] == ["南京门东店"]

    main.audit_service.record_operation("storeuser", "phase3.audit", "unit", "1", {"password": "secret", "safe": True}, None)
    audit_rows = rows_for(tmp_path, "admin_operation_logs")
    assert any(row["action"] == "phase3.audit" and '"safe": true' in row["detail"] for row in audit_rows)
    assert "secret" not in json.dumps(audit_rows, ensure_ascii=False)

    assert_login_success(login_admin(client, "storeuser", "123456"))
    list_page = client.get("/admin")
    assert list_page.status_code == 200
    assert "phase3 visible ticket" in list_page.text
    assert "phase3 hidden ticket" not in list_page.text

    dashboard = client.get("/admin/dashboard")
    assert dashboard.status_code == 200
    assert "phase3 visible ticket" in dashboard.text
    assert "phase3 hidden ticket" not in dashboard.text

    ticket_export = client.get("/admin/export")
    assert ticket_export.status_code == 200
    ticket_workbook = load_workbook(BytesIO(ticket_export.content))
    exported_ticket_descriptions = {row[10].value for row in ticket_workbook.active.iter_rows(min_row=2)}
    assert exported_ticket_descriptions == {"phase3 visible ticket"}

    schedule_page = client.get("/admin/schedules?store_scope=all&month=2026-07")
    assert schedule_page.status_code == 200
    assert "phase3 visible employee" in schedule_page.text
    assert "phase3 hidden employee" not in schedule_page.text

    schedule_export = client.get("/admin/schedules/export?store_names=南京门东店&store_names=南昌万寿宫店&month=2026-07")
    assert schedule_export.status_code == 200
    schedule_workbook = load_workbook(BytesIO(schedule_export.content))
    exported_schedule_stores = {row[0].value for row in schedule_workbook.active.iter_rows(min_row=2)}
    exported_schedule_employees = {row[1].value for row in schedule_workbook.active.iter_rows(min_row=2)}
    assert exported_schedule_stores == {"南京门东店"}
    assert exported_schedule_employees == {"phase3 visible employee"}

    assert_login_success(login_admin(client, "admin", "123456"))
    overview = client.get("/admin/permission-overview")
    assert overview.status_code == 200
    assert "PermissionService" in overview.text
    assert "data_scope" in overview.text
    assert "ticket.view" in overview.text
    assert "未接入权限系统" in overview.text


def test_phase3_p0_routes_use_html_permission_denials(tmp_path, monkeypatch):
    client, main = build_client(
        tmp_path,
        monkeypatch,
        admin_users="admin:123456,limited:123456,governed:123456",
    )
    limited_role = create_role_with_permissions(tmp_path, "phase3-no-access", [])
    governed_role = create_role_with_permissions(
        tmp_path,
        "phase3-p0-access",
        ["embedded.view", "shift.view", "config.view", "system.view", "system.route_health", "role.view"],
    )
    update_admin_user_access(tmp_path, "limited", limited_role)
    update_admin_user_access(tmp_path, "governed", governed_role)

    unauthenticated = TestClient(main.app).get("/admin/permission-overview", follow_redirects=False)
    assert unauthenticated.status_code == 303
    assert unauthenticated.headers["location"].startswith("/admin/login")

    assert_login_success(login_admin(client, "limited", "123456"))
    for path in (
        "/admin/embedded-pages",
        "/admin/shift-types",
        "/admin/settings",
        "/admin/system",
        "/admin/route-health",
        "/admin/permission-overview",
    ):
        denied = client.get(path)
        assert denied.status_code == 403, path
        assert "text/html" in denied.headers.get("content-type", ""), path
        assert "application/json" not in denied.headers.get("content-type", ""), path
        assert "权限不足" in denied.text

    assert_login_success(login_admin(client, "governed", "123456"))
    for path in (
        "/admin/embedded-pages",
        "/admin/shift-types",
        "/admin/settings",
        "/admin/system",
        "/admin/route-health",
        "/admin/permission-overview",
    ):
        allowed = client.get(path)
        assert allowed.status_code == 200, path


def test_phase3_permission_and_audit_legacy_paths_are_consolidated():
    source = (PROJECT_DIR / "main.py").read_text(encoding="utf-8")
    assert "class PermissionService" in source
    assert "class AuditService" in source
    assert "def apply_data_scope(" in source
    assert "if not has_permission(" not in source
    assert source.count("INSERT INTO admin_operation_logs") == 1


def test_phase4_write_routes_require_permissions_and_hide_buttons(tmp_path, monkeypatch):
    client, main = build_client(
        tmp_path,
        monkeypatch,
        admin_users="admin:123456,employee_viewer:123456,schedule_viewer:123456,ticket_viewer:123456,trash_viewer:123456,cleanup_viewer:123456",
    )
    employee_role = create_role_with_permissions(tmp_path, "phase4-employee-view", ["employee.view"])
    schedule_role = create_role_with_permissions(tmp_path, "phase4-schedule-view", ["schedule.view"])
    ticket_role = create_role_with_permissions(tmp_path, "phase4-ticket-view", ["ticket.view"])
    trash_role = create_role_with_permissions(tmp_path, "phase4-trash-view", ["ticket.view_trash"])
    cleanup_role = create_role_with_permissions(tmp_path, "phase4-cleanup-view", ["ticket.view"])
    update_admin_user_access(tmp_path, "employee_viewer", employee_role)
    update_admin_user_access(tmp_path, "schedule_viewer", schedule_role)
    update_admin_user_access(tmp_path, "ticket_viewer", ticket_role)
    update_admin_user_access(tmp_path, "trash_viewer", trash_role)
    update_admin_user_access(tmp_path, "cleanup_viewer", cleanup_role)

    employee_id = main.create_employee("phase4 existing employee", "南京门东店", "店员", "", "在职", main.load_app_config())
    shift_id = main.create_shift_type("phase4 schedule shift", "09:00", "18:00", "8", "#2563eb", store_name="南京门东店", config=main.load_app_config())
    main.bulk_upsert_schedules("南京门东店", [employee_id], ["2026-07-08"], shift_id, "", False, "admin", 20)
    submit_ticket(client, description="phase4 collaboration ticket")
    submit_ticket(client, description="phase4 trash ticket")
    main.soft_delete_ticket(2, "admin", "phase4 trash setup")

    assert_login_success(login_admin(client, "employee_viewer", "123456"))
    employees_page = client.get("/admin/employees")
    assert employees_page.status_code == 200
    assert 'data-drawer-open="employee-create-drawer"' not in employees_page.text
    assert f'action="/admin/employees/{employee_id}/update"' not in employees_page.text
    assert f'action="/admin/employees/{employee_id}/delete"' not in employees_page.text
    employee_csrf = csrf_token_for(client, "/admin/employees")
    assert_html_forbidden(
        client.post(
            "/admin/employees",
            data={
                "csrf_token": employee_csrf,
                "employee_name": "phase4 blocked employee",
                "store_name": "南京门东店",
                "primary_store_name": "南京门东店",
                "store_names": ["南京门东店"],
                "role": "店员",
                "status": "在职",
            },
        )
    )
    assert_html_forbidden(
        client.post(
            f"/admin/employees/{employee_id}/update",
            data={
                "csrf_token": employee_csrf,
                "employee_name": "phase4 blocked update",
                "store_name": "南京门东店",
                "primary_store_name": "南京门东店",
                "store_names": ["南京门东店"],
                "role": "店员",
                "status": "在职",
            },
        )
    )
    assert_html_forbidden(client.post(f"/admin/employees/{employee_id}/delete", data={"csrf_token": employee_csrf}))
    assert not any(row["employee_name"] == "phase4 blocked employee" for row in rows_for(tmp_path, "employees"))

    assert_login_success(login_admin(client, "schedule_viewer", "123456"))
    schedules_page = client.get("/admin/schedules?store_names=南京门东店&month=2026-07")
    assert schedules_page.status_code == 200
    assert 'method="post" action="/admin/schedules"' not in schedules_page.text
    assert 'action="/admin/schedules/1/delete"' not in schedules_page.text
    schedule_csrf = csrf_token_for(client, "/admin/schedules?store_names=南京门东店&month=2026-07")
    assert_html_forbidden(
        client.post(
            "/admin/schedules",
            data={
                "csrf_token": schedule_csrf,
                "store_name": "南京门东店",
                "store_names": ["南京门东店"],
                "employee_ids": [str(employee_id)],
                "schedule_dates": ["2026-07-09"],
                "shift_type_id": str(shift_id),
            },
        )
    )
    assert_html_forbidden(client.post("/admin/schedules/1/delete", data={"csrf_token": schedule_csrf}))

    assert_login_success(login_admin(client, "ticket_viewer", "123456"))
    detail_page = client.get("/admin/ticket/1")
    assert detail_page.status_code == 200
    assert 'action="/admin/ticket/1/comments"' not in detail_page.text
    assert 'action="/admin/ticket/1/participants"' not in detail_page.text
    ticket_csrf = csrf_token_for(client, "/admin/ticket/1")
    assert_html_forbidden(client.post("/admin/ticket/1/comments", data={"csrf_token": ticket_csrf, "content": "blocked internal comment"}))
    assert_html_forbidden(
        client.post(
            "/admin/ticket/1/participants",
            data={"csrf_token": ticket_csrf, "participant_type": "team", "participant_name": "blocked partner", "role": "协作处理"},
        )
    )

    assert_login_success(login_admin(client, "trash_viewer", "123456"))
    trash_page = client.get("/admin/trash")
    assert trash_page.status_code == 200
    assert 'formaction="/admin/tickets/bulk-hard-delete"' not in trash_page.text
    assert 'form="hard-delete-ticket-2"' not in trash_page.text
    trash_csrf = csrf_token_for(client, "/admin/trash")
    assert_html_forbidden(client.post("/admin/ticket/2/hard-delete", data={"csrf_token": trash_csrf, "confirm_delete": "1"}))

    assert_login_success(login_admin(client, "cleanup_viewer", "123456"))
    assert_html_forbidden(client.post("/admin/cleanup/preview", data={"csrf_token": csrf_token_for(client, "/admin")}))


def test_phase4_audit_service_records_key_write_operations(tmp_path, monkeypatch):
    client, main = build_client(tmp_path, monkeypatch, admin_users="admin:123456")
    assert_login_success(login_admin(client, "admin", "123456"))

    employee_csrf = csrf_token_for(client, "/admin/employees")
    create_employee_response = client.post(
        "/admin/employees",
        data={
            "csrf_token": employee_csrf,
            "employee_name": "phase4 audited employee",
            "store_name": "南京门东店",
            "primary_store_name": "南京门东店",
            "store_names": ["南京门东店"],
            "role": "店员",
            "status": "在职",
        },
        follow_redirects=False,
    )
    assert create_employee_response.status_code == 303
    employee_id = rows_for(tmp_path, "employees")[0]["id"]
    update_employee_response = client.post(
        f"/admin/employees/{employee_id}/update",
        data={
            "csrf_token": employee_csrf,
            "employee_name": "phase4 audited employee updated",
            "store_name": "南京门东店",
            "primary_store_name": "南京门东店",
            "store_names": ["南京门东店"],
            "role": "店员",
            "status": "在职",
        },
        follow_redirects=False,
    )
    assert update_employee_response.status_code == 303

    shift_id = main.create_shift_type("phase4 audited shift", "09:00", "18:00", "8", "#2563eb", store_name="南京门东店", config=main.load_app_config())
    schedule_csrf = csrf_token_for(client, "/admin/schedules?store_names=南京门东店&month=2026-07")
    create_schedule_response = client.post(
        "/admin/schedules",
        data={
            "csrf_token": schedule_csrf,
            "store_name": "南京门东店",
            "store_names": ["南京门东店"],
            "employee_ids": [str(employee_id)],
            "schedule_dates": ["2026-07-10"],
            "shift_type_id": str(shift_id),
        },
        follow_redirects=False,
    )
    assert create_schedule_response.status_code == 303
    delete_schedule_response = client.post("/admin/schedules/1/delete", data={"csrf_token": schedule_csrf}, follow_redirects=False)
    assert delete_schedule_response.status_code == 303

    submit_ticket(client, description="phase4 audited collaboration")
    ticket_csrf = csrf_token_for(client, "/admin/ticket/1")
    comment_response = client.post(
        "/admin/ticket/1/comments",
        data={"csrf_token": ticket_csrf, "content": "phase4 audited comment", "visibility": "internal"},
        follow_redirects=False,
    )
    participant_response = client.post(
        "/admin/ticket/1/participants",
        data={"csrf_token": ticket_csrf, "participant_type": "team", "participant_name": "phase4 audited partner", "role": "协作处理"},
        follow_redirects=False,
    )
    assert comment_response.status_code == 303
    assert participant_response.status_code == 303

    actions = [row["action"] for row in rows_for(tmp_path, "admin_operation_logs")]
    assert "employee.create" in actions
    assert "employee.update" in actions
    assert "schedule.create" in actions
    assert "schedule.delete" in actions
    assert "ticket.comment.create" in actions
    assert "ticket.participant.create" in actions


def test_phase4_permission_overview_has_no_high_risk_uncontrolled_posts(tmp_path, monkeypatch):
    client, main = build_client(tmp_path, monkeypatch, admin_users="admin:123456")
    context = main.permission_overview_context(main.app)
    assert context["high_risk_uncontrolled_count"] == 0
    assert all(
        route["coverage_status"] != "uncontrolled"
        for route in context["admin_route_items"]
        if route["method"] == "POST"
    )

    assert_login_success(login_admin(client, "admin", "123456"))
    overview = client.get("/admin/permission-overview")
    assert overview.status_code == 200
    assert "剩余未接入路由数量" in overview.text
    assert "高风险 POST 未接入：0" in overview.text
    assert "登录保护 API" in overview.text


def test_p1a_ticket_rule_tables_auto_assignment_priority_and_audit(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch, admin_users="admin:123456")
    active_user_id = create_test_admin_user(tmp_path, "buyer_active", "采购可指派", is_assignable=1)
    create_test_admin_user(tmp_path, "buyer_disabled", "停用采购", is_active=0, is_assignable=1)
    create_test_admin_user(tmp_path, "buyer_no_login", "无登录采购", allow_login=0, is_assignable=1)
    create_test_admin_user(tmp_path, "buyer_not_assignable", "不可指派采购", is_assignable=0)
    role_id = role_id_by_name(tmp_path, "运营管理")

    with sqlite3.connect(tmp_path / "tickets.db") as connection:
        tables = {row[0] for row in connection.execute("SELECT name FROM sqlite_master WHERE type = 'table'")}
        assignment_columns = {row[1] for row in connection.execute("PRAGMA table_info(ticket_assignment_rules)")}
    assert "ticket_assignment_rules" in tables
    assert {
        "request_type",
        "brand",
        "store_name",
        "default_handler",
        "default_role_id",
        "priority",
        "enabled",
        "note",
    }.issubset(assignment_columns)

    insert_assignment_rule(tmp_path, request_type="建单需求", default_handler="buyer_disabled", priority=1)
    insert_assignment_rule(tmp_path, request_type="建单需求", default_handler="buyer_no_login", priority=2)
    insert_assignment_rule(tmp_path, request_type="建单需求", default_handler="buyer_not_assignable", priority=3)
    insert_assignment_rule(tmp_path, request_type="建单需求", default_role_id=role_id, priority=4)
    insert_assignment_rule(
        tmp_path,
        request_type="建单需求",
        brand="测试品牌",
        store_name="南京门东店",
        default_handler="buyer_active",
        priority=100,
    )
    insert_assignment_rule(tmp_path, request_type="系统问题", default_handler="buyer_active", enabled=0)

    response = submit_ticket(client, request_type="建单需求", brand="测试品牌", description="门店类型品牌最高优先级")
    assert response.status_code == 200
    first_ticket = rows_for(tmp_path, "tickets")[-1]
    assert first_ticket["assigned_to"] == "buyer_active"
    logs = ticket_action_logs(tmp_path, first_ticket["id"])
    assert any(row["action"] == "自动分派" and row["new_assigned_to"] == "buyer_active" for row in logs)
    operations = rows_for(tmp_path, "admin_operation_logs")
    assert any(row["action"] == "ticket.auto_assign" and str(first_ticket["id"]) == str(row["target_id"]) for row in operations)

    response = submit_ticket(client, request_type="建单需求", brand="其他品牌", description="类型规则角色兜底")
    assert response.status_code == 200
    second_ticket = rows_for(tmp_path, "tickets")[-1]
    assert second_ticket["assigned_to"] == "buyer_active"

    response = submit_ticket(client, request_type="系统问题", description="停用规则不生效")
    assert response.status_code == 200
    disabled_rule_ticket = rows_for(tmp_path, "tickets")[-1]
    assert disabled_rule_ticket["assigned_to"] in ("", None)
    assert active_user_id > 0


def test_p1a_sla_rules_deadlines_statuses_export_and_dashboard_scope(tmp_path, monkeypatch):
    client, main = build_client(tmp_path, monkeypatch, admin_users="admin:123456")
    fixed_now = "2026-07-06 10:00:00"
    monkeypatch.setattr(main, "now_text", lambda: fixed_now)
    insert_sla_rule(tmp_path, urgency_level="加急", due_hours=24)
    insert_sla_rule(tmp_path, request_type="系统问题", due_hours=72)

    assert "ticket_sla_rules" in {
        row[0]
        for row in sqlite3.connect(tmp_path / "tickets.db")
        .execute("SELECT name FROM sqlite_master WHERE type = 'table'")
        .fetchall()
    }

    response = submit_ticket(client, expected_finish_date="", urgency="加急", description="SLA 自动截止")
    assert response.status_code == 200
    auto_sla_ticket = rows_for(tmp_path, "tickets")[-1]
    assert auto_sla_ticket["expected_finish_date"] == "2026-07-07 10:00:00"

    response = submit_ticket(
        client,
        expected_finish_date="2026-07-10",
        urgency="加急",
        description="手动期望时间优先",
    )
    assert response.status_code == 200
    manual_ticket = rows_for(tmp_path, "tickets")[-1]
    assert manual_ticket["expected_finish_date"] == "2026-07-10"

    operations = rows_for(tmp_path, "admin_operation_logs")
    assert any(row["action"] == "ticket.sla_deadline.auto" for row in operations)
    assert main.due_status_label({"expected_finish_date": "", "status": "待处理"}) == "未设置"
    assert main.due_status_label({"expected_finish_date": date.today().isoformat(), "status": "待处理"}) == "今日到期"
    assert main.due_status_label({"expected_finish_date": (date.today() - timedelta(days=1)).isoformat(), "status": "待处理"}) == "已超时"
    assert (
        main.due_status_label(
            {"expected_finish_date": "2026-07-07", "status": "已完成", "closed_at": "2026-07-07 09:00:00"}
        )
        == "按时完成"
    )
    assert (
        main.due_status_label(
            {"expected_finish_date": "2026-07-07", "status": "已完成", "closed_at": "2026-07-08 09:00:00"}
        )
        == "超时完成"
    )

    with sqlite3.connect(tmp_path / "tickets.db") as connection:
        connection.execute(
            "UPDATE tickets SET store_name = ?, expected_finish_date = ?, status = ?, closed_at = ? WHERE id = ?",
            ("南京门东店", date.today().isoformat(), "已完成", date.today().isoformat(), int(auto_sla_ticket["id"])),
        )
        connection.execute(
            "UPDATE ticket_stores SET store_name = ? WHERE ticket_id = ?",
            ("南京门东店", int(auto_sla_ticket["id"])),
        )
        connection.execute(
            "UPDATE tickets SET store_name = ?, expected_finish_date = ?, status = ? WHERE id = ?",
            ("南昌万寿宫店", (date.today() - timedelta(days=1)).isoformat(), "待处理", int(manual_ticket["id"])),
        )
        connection.execute(
            "UPDATE ticket_stores SET store_name = ? WHERE ticket_id = ?",
            ("南昌万寿宫店", int(manual_ticket["id"])),
        )

    store_role = create_role_with_permissions(tmp_path, "p1a-store-viewer", ["ticket.view", "ticket.export"])
    create_test_admin_user(tmp_path, "storeviewer", "门店范围账号", password="scope123")
    update_admin_user_access(tmp_path, "storeviewer", store_role, data_scope="stores", store_names="南京门东店")
    assert_login_success(login_admin(client, "storeviewer", "scope123"))

    dashboard = client.get("/admin/dashboard")
    assert dashboard.status_code == 200
    assert "SLA 完成率" in dashboard.text
    assert "平均处理时长" in dashboard.text
    assert "SLA 自动截止" in dashboard.text
    assert "手动期望时间优先" not in dashboard.text

    workbook = load_workbook(BytesIO(client.get("/admin/export").content))
    headers = [cell.value for cell in workbook.active[1]]
    assert "时效状态" in headers
    exported_descriptions = {row[10].value for row in workbook.active.iter_rows(min_row=2)}
    assert exported_descriptions == {"SLA 自动截止"}


def test_p1a_ticket_rules_page_permissions_and_overview(tmp_path, monkeypatch):
    client, main = build_client(tmp_path, monkeypatch, admin_users="admin:123456")
    logged_in_client(client, "admin", "123456")
    page = client.get("/admin/ticket-rules")
    assert page.status_code == 200
    assert "自动分派规则" in page.text
    assert "SLA 规则" in page.text
    assert "测试匹配" in page.text

    response = admin_post(
        client,
        "/admin/ticket-rules/assignment",
        data={
            "request_type": "缺货需求",
            "default_handler": "admin",
            "priority": "20",
            "enabled": "1",
        },
        follow_redirects=False,
    )
    assert response.status_code == 303
    assert rows_for(tmp_path, "ticket_assignment_rules")[-1]["request_type"] == "缺货需求"

    response = admin_post(
        client,
        "/admin/ticket-rules/sla",
        data={"urgency_level": "加急", "due_hours": "24", "enabled": "1"},
        follow_redirects=False,
    )
    assert response.status_code == 303
    assert rows_for(tmp_path, "ticket_sla_rules")[-1]["due_hours"] == 24
    assert {"ticket.assignment_rule.update", "ticket.sla_rule.update"}.issubset(
        {row["action"].replace(".create", ".update") for row in rows_for(tmp_path, "admin_operation_logs")}
    )

    readonly_role = create_role_with_permissions(
        tmp_path,
        "p1a-rule-readonly",
        ["ticket.assignment_rule.view", "ticket.sla_rule.view"],
    )
    create_test_admin_user(tmp_path, "ruleviewer", "规则只读", password="readonly123")
    update_admin_user_access(tmp_path, "ruleviewer", readonly_role)
    assert_login_success(login_admin(client, "ruleviewer", "readonly123"))
    readonly_page = client.get("/admin/ticket-rules")
    assert readonly_page.status_code == 200
    token = csrf_token_for(client, "/admin/ticket-rules")
    denied_assignment = client.post(
        "/admin/ticket-rules/assignment",
        data={"csrf_token": token, "request_type": "缺货需求", "priority": "1"},
    )
    assert_html_forbidden(denied_assignment)
    denied_sla = client.post(
        "/admin/ticket-rules/sla",
        data={"csrf_token": token, "urgency_level": "加急", "due_hours": "12"},
    )
    assert_html_forbidden(denied_sla)

    context = main.permission_overview_context(main.app)
    assert context["high_risk_uncontrolled_count"] == 0
    route_paths = {(row["method"], row["path"]) for row in context["route_items"]}
    assert ("GET", "/admin/ticket-rules") in route_paths
    assert ("POST", "/admin/ticket-rules/assignment") in route_paths
    assert ("POST", "/admin/ticket-rules/sla") in route_paths


def test_p1b_request_type_templates_schema_defaults_and_submit_ui(tmp_path, monkeypatch):
    config = {
        "request_types.json": ["建采购单", "缺货补货", "瑕疵/破损/漏发", "审盘点单", "排班问题", "其他", "无模板类型"],
    }
    client, _ = build_client(tmp_path, monkeypatch, admin_users="admin:123456", config_overrides=config)
    with sqlite3.connect(tmp_path / "tickets.db") as connection:
        tables = {row[0] for row in connection.execute("SELECT name FROM sqlite_master WHERE type = 'table'")}
        columns = {row[1] for row in connection.execute("PRAGMA table_info(request_type_templates)")}
    assert "request_type_templates" in tables
    assert {
        "request_type",
        "recommended_fields",
        "required_description",
        "image_required",
        "file_required",
        "expected_finish_required",
        "field_help_text",
        "enabled",
        "sort_order",
    }.issubset(columns)
    assert not {"brand_required", "product_name_required", "barcode_required", "quantity_required"}.intersection(columns)

    templates = {row["request_type"]: row for row in rows_for(tmp_path, "request_type_templates")}
    assert {"建采购单", "缺货补货", "瑕疵/破损/漏发", "审盘点单", "排班问题", "其他"}.issubset(templates)
    assert templates["瑕疵/破损/漏发"]["image_required"] == 1
    assert templates["审盘点单"]["file_required"] == 1
    assert "商品信息不完整也可以提交" in templates["缺货补货"]["field_help_text"]

    page = client.get("/submit")
    assert page.status_code == 200
    assert "data-request-type-template-card" in page.text
    assert "建议填写信息" in page.text
    assert "必填信息" in page.text
    assert "建议填写" in page.text
    assert "商品名称 <strong>*</strong>" not in page.text
    assert "规格条码 <strong>*</strong>" not in page.text
    assert "数量 <strong>*</strong>" not in page.text

    logged_in_client(client, "admin", "123456")
    rules_page = client.get("/admin/ticket-rules")
    assert rules_page.status_code == 200
    assert "工单类型模板" in rules_page.text


def test_p1b_product_fields_are_recommended_and_template_validation_controls_non_product_fields(tmp_path, monkeypatch):
    config = {
        "request_types.json": ["建采购单", "缺货补货", "瑕疵/破损/漏发", "审盘点单", "排班问题", "其他", "无模板类型"],
    }
    client, _ = build_client(tmp_path, monkeypatch, admin_users="admin:123456", config_overrides=config)

    product_blank = submit_ticket(
        client,
        request_type="缺货补货",
        brand="",
        brand_extra="",
        product_name="",
        sku_barcode="",
        quantity="",
        description="商品信息都为空但说明完整",
    )
    assert product_blank.status_code == 200
    ticket = rows_for(tmp_path, "tickets")[-1]
    assert ticket["request_type"] == "缺货补货"
    assert ticket["brand"] == ""
    assert ticket["product_name"] == ""
    assert ticket["sku_barcode"] == ""
    assert ticket["quantity"] is None

    image_missing = submit_ticket(
        client,
        request_type="瑕疵/破损/漏发",
        description="图片必填但没有上传",
    )
    assert image_missing.status_code == 400
    assert "瑕疵/破损/漏发必须上传至少一张图片。" in image_missing.text

    image_ok = submit_ticket_with_image(
        client,
        request_type="瑕疵/破损/漏发",
        description="已上传现场图片",
    )
    assert image_ok.status_code == 200

    file_missing = submit_ticket(client, request_type="审盘点单", description="附件必填但没有上传")
    assert file_missing.status_code == 400
    assert "审盘点单必须上传至少一个文件附件。" in file_missing.text

    file_ok = submit_ticket_with_file(client, request_type="审盘点单", description="已上传盘点附件")
    assert file_ok.status_code == 200

    description_missing = submit_ticket(client, request_type="排班问题", description="")
    assert description_missing.status_code == 400
    assert "请填写问题说明。" in description_missing.text

    insert_request_type_template(
        tmp_path,
        "无模板类型",
        required_description=0,
        expected_finish_required=1,
        field_help_text="这个类型需要期望完成时间。",
    )
    expected_missing = submit_ticket(client, request_type="无模板类型", description="有说明但缺期望完成时间", expected_finish_date="")
    assert expected_missing.status_code == 400
    assert "无模板类型必须填写期望完成时间。" in expected_missing.text

    with sqlite3.connect(tmp_path / "tickets.db") as connection:
        connection.execute("UPDATE request_type_templates SET enabled = 0 WHERE request_type = ?", ("无模板类型",))
    disabled_template = submit_ticket(client, request_type="无模板类型", description="停用模板不再拦截", expected_finish_date="")
    assert disabled_template.status_code == 200


def test_p1b_template_permissions_crud_audit_and_permission_overview(tmp_path, monkeypatch):
    client, main = build_client(tmp_path, monkeypatch, admin_users="admin:123456")
    logged_in_client(client, "admin", "123456")
    page = client.get("/admin/ticket-rules")
    assert page.status_code == 200
    assert "工单类型模板" in page.text
    assert "新增模板" in page.text

    response = admin_post(
        client,
        "/admin/ticket-rules/templates",
        data={
            "request_type": "系统问题",
            "template_name": "系统问题模板",
            "recommended_fields": "品牌,商品名称",
            "required_description": "1",
            "image_required": "0",
            "file_required": "0",
            "expected_finish_required": "1",
            "field_help_text": "请说明系统页面和异常截图。",
            "enabled": "1",
            "sort_order": "12",
        },
        follow_redirects=False,
    )
    assert response.status_code == 303
    template = rows_for(tmp_path, "request_type_templates")[-1]
    assert template["template_name"] == "系统问题模板"
    assert template["expected_finish_required"] == 1

    toggle = admin_post(
        client,
        f"/admin/ticket-rules/templates/{template['id']}/toggle",
        data={"enabled": "0"},
        follow_redirects=False,
    )
    assert toggle.status_code == 303
    assert rows_for(tmp_path, "request_type_templates")[-1]["enabled"] == 0
    assert any(row["action"] in {"ticket.template.create", "ticket.template.toggle"} for row in rows_for(tmp_path, "admin_operation_logs"))

    readonly_role = create_role_with_permissions(tmp_path, "p1b-template-viewer", ["ticket.template.view"])
    create_test_admin_user(tmp_path, "templateviewer", "模板只读", password="viewer123")
    update_admin_user_access(tmp_path, "templateviewer", readonly_role)
    assert_login_success(login_admin(client, "templateviewer", "viewer123"))
    readonly_page = client.get("/admin/ticket-rules")
    assert readonly_page.status_code == 200
    assert "工单类型模板" in readonly_page.text
    token = csrf_token_for(client, "/admin/ticket-rules")
    denied = client.post(
        "/admin/ticket-rules/templates",
        data={"csrf_token": token, "request_type": "其他", "template_name": "只读不能保存"},
    )
    assert_html_forbidden(denied)

    no_view_role = create_role_with_permissions(tmp_path, "p1b-no-template-view", ["ticket.assignment_rule.view"])
    create_test_admin_user(tmp_path, "notemplate", "无模板权限", password="noview123")
    update_admin_user_access(tmp_path, "notemplate", no_view_role)
    assert_login_success(login_admin(client, "notemplate", "noview123"))
    no_view_page = client.get("/admin/ticket-rules")
    assert no_view_page.status_code == 200
    assert "工单类型模板" not in no_view_page.text

    context = main.permission_overview_context(main.app)
    assert context["high_risk_uncontrolled_count"] == 0
    route_paths = {(row["method"], row["path"]) for row in context["route_items"]}
    assert ("POST", "/admin/ticket-rules/templates") in route_paths
    assert ("POST", "/admin/ticket-rules/templates/{template_id}/toggle") in route_paths


def test_p1b_template_validation_keeps_auto_assignment_and_sla(tmp_path, monkeypatch):
    config = {"request_types.json": ["缺货补货"]}
    client, main = build_client(tmp_path, monkeypatch, admin_users="admin:123456", config_overrides=config)
    monkeypatch.setattr(main, "now_text", lambda: "2026-07-06 10:00:00")
    create_test_admin_user(tmp_path, "buyer_active", "采购可指派", is_assignable=1)
    insert_assignment_rule(tmp_path, request_type="缺货补货", default_handler="buyer_active")
    insert_sla_rule(tmp_path, request_type="缺货补货", due_hours=24)

    response = submit_ticket(
        client,
        request_type="缺货补货",
        brand="",
        product_name="",
        sku_barcode="",
        quantity="",
        expected_finish_date="",
        description="模板校验通过后继续执行 P1A",
    )
    assert response.status_code == 200
    ticket = rows_for(tmp_path, "tickets")[-1]
    assert ticket["assigned_to"] == "buyer_active"
    assert ticket["expected_finish_date"] == "2026-07-07 10:00:00"
    actions = [row["action"] for row in rows_for(tmp_path, "admin_operation_logs")]
    assert "ticket.auto_assign" in actions
    assert "ticket.sla_deadline.auto" in actions


def test_phase4_templates_use_has_perm_for_privileged_controls():
    expectations = {
        "templates/base_admin.html": ["account.view", "role.view", "system.route_health", "ticket.view_trash", "ticket.delete"],
        "templates/admin.html": ["ticket.export", "ticket.archive", "ticket.delete"],
        "templates/ticket_detail.html": ["ticket.comment", "ticket.assign", "ticket.update", "ticket.delete", "ticket.archive", "ticket.hard_delete"],
        "templates/employees.html": ["employee.create", "employee.update", "employee.archive", "employee.delete", "employee.hard_delete"],
        "templates/schedules.html": ["schedule.create", "schedule.delete", "schedule.export"],
        "templates/trash.html": ["ticket.restore", "ticket.hard_delete", "ticket.delete"],
        "templates/cleanup.html": ["ticket.delete"],
    }
    for relative_path, permissions in expectations.items():
        text = (PROJECT_DIR / relative_path).read_text(encoding="utf-8")
        for permission in permissions:
            assert f"has_perm('{permission}')" in text, f"{relative_path} missing {permission}"


def test_ticket_detail_workbench_comment_modes_and_close_prompt(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    submit_ticket(client, description="协作工作台提示工单")
    logged_in_client(client)
    admin_post(
        client,
        "/admin/ticket/1/tasks",
        data={"title": "完成收尾确认", "assignee": ADMIN_AUTH[0], "status": "已完成"},
        follow_redirects=False,
    )

    detail = client.get("/admin/ticket/1")
    assert detail.status_code == 200
    assert "协作工作台" in detail.text
    assert "谁负责" in detail.text
    assert "谁协作" in detail.text
    assert "下一步" in detail.text
    assert "门店可见回复" in detail.text
    assert "内部备注" in detail.text
    assert 'option value="public"' in detail.text
    assert 'option value="internal"' in detail.text
    assert "所有子任务已完成" in detail.text
    assert "是否关闭工单" in detail.text


def test_notification_workflow_actions_and_accept_ticket(tmp_path, monkeypatch):
    client, _ = build_client(
        tmp_path,
        monkeypatch,
        admin_users="admin:123456,caigou:123456",
        config_overrides={"handlers.json": ["admin", "caigou", "总部商品"]},
    )
    submit_ticket(client, description="待接单提醒工单")

    assert_login_success(login_admin(client, "admin", "123456"))
    payload = client.get("/admin/api/notifications").json()
    notification = payload["notifications"][0]
    labels = [action["label"] for action in notification["actions"]]
    assert labels == ["查看工单", "接单", "回复", "标记已读"]
    accept_action = next(action for action in notification["actions"] if action["label"] == "接单")
    assert accept_action["method"] == "post"
    assert accept_action["url"] == "/admin/ticket/1/accept"

    accept = client.post(
        "/admin/ticket/1/accept",
        data={"csrf_token": csrf_token_for(client), "return_url": "/admin/my-work"},
        follow_redirects=False,
    )
    assert accept.status_code == 303
    assert accept.headers["location"].startswith("/admin/my-work")
    assert rows_for(tmp_path, "tickets")[0]["assigned_to"] == "admin"
    assert ticket_action_logs(tmp_path, 1)[-1]["action"] == "接单"

    work = client.get("/admin/my-work")
    assert "待接单提醒工单" in work.text


def test_store_need_supplement_highlight_and_supplement_reminds_handler(tmp_path, monkeypatch):
    client, _ = build_client(
        tmp_path,
        monkeypatch,
        admin_users="admin:123456",
        config_overrides={"handlers.json": ["admin", "总部商品"]},
    )
    submit_ticket(client, store_name="南京门东店", description="需要门店补充的工单")
    assert_login_success(login_admin(client, "admin", "123456"))
    admin_post(
        client,
        "/admin/ticket/1",
        data={"status": "待门店补充", "assigned_to": "admin", "handler_note": "请补充现场照片和到货说明"},
        follow_redirects=False,
    )

    store_detail = client.get("/query/ticket/1?store_name=南京门东店")
    assert store_detail.status_code == 200
    assert "需要补充资料" in store_detail.text
    assert "请补充现场照片和到货说明" in store_detail.text

    supplement = client.post(
        "/query/ticket/1/supplement",
        data={"store_name": "南京门东店", "submitter": "小李", "note": "已补充现场照片"},
    )
    assert supplement.status_code == 200
    events = rows_for(tmp_path, "notification_events")
    assert events[-1]["event_type"] == "store_supplement"
    assert events[-1]["title"] == "门店补充资料"
    assert "admin" in events[-1]["content"]


def test_cookie_login_logout_switch_account_and_operator_log(tmp_path, monkeypatch):
    client, _ = build_client(
        tmp_path,
        monkeypatch,
        admin_auth=("legacy-admin", "legacy-secret"),
        admin_users="admin:123456,caigou:123456",
    )

    unauthenticated = client.get("/admin", follow_redirects=False)
    assert unauthenticated.status_code == 303
    assert unauthenticated.headers["location"].startswith("/admin/login")

    login_page = client.get("/admin/login")
    assert login_page.status_code == 200
    assert "止痒工单后台登录" in login_page.text
    assert "用户名" in login_page.text
    assert "密码" in login_page.text

    bad_login = client.post("/admin/login", data={"username": "admin", "password": "wrong"})
    assert bad_login.status_code == 400
    assert "用户名或密码不正确" in bad_login.text

    assert_login_success(login_admin(client, "admin", "123456"))
    admin_page = client.get("/admin")
    assert admin_page.status_code == 200
    assert "当前账号：admin" in admin_page.text
    assert "退出登录" in admin_page.text
    assert "切换账号" in admin_page.text

    logout = admin_post(client, "/admin/logout", follow_redirects=False)
    assert logout.status_code == 303
    assert logout.headers["location"].startswith("/admin/login")
    assert "logged_out=1" in logout.headers["location"]
    assert "admin_session=" in logout.headers.get("set-cookie", "")
    logout_page = client.get(logout.headers["location"])
    assert logout_page.status_code == 200
    assert "\u5df2\u9000\u51fa\u767b\u5f55\uff0c\u8bf7\u91cd\u65b0\u767b\u5f55\u3002" in logout_page.text
    after_logout = client.get("/admin", follow_redirects=False)
    assert after_logout.status_code == 303
    assert after_logout.headers["location"].startswith("/admin/login")

    basic_after_logout = client.get(
        "/admin",
        headers=basic_auth_headers("admin", "123456"),
        follow_redirects=False,
    )
    assert basic_after_logout.status_code == 303
    assert basic_after_logout.headers["location"].startswith("/admin/login")

    assert_login_success(login_admin(client, "caigou", "123456"))
    caigou_page = client.get("/admin")
    assert caigou_page.status_code == 200
    assert "\u5f53\u524d\u8d26\u53f7\uff1acaigou" in caigou_page.text
    response = submit_ticket(client)
    assert response.status_code == 200
    update_response = admin_post(
        client,
        "/admin/ticket/1",
        data={"status": "处理中", "assigned_to": "采购", "handler_note": "采购账号处理"},
        follow_redirects=False,
    )
    assert update_response.status_code == 303

    logs = rows_for(tmp_path, "ticket_logs")
    assert len(logs) == 1
    assert logs[0]["operator"] == "caigou"


def test_legacy_admin_credentials_can_login_with_cookie_when_admin_users_missing(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch, admin_auth=("legacy-admin", "legacy-secret"))

    assert_login_success(login_admin(client, "legacy-admin", "legacy-secret"))
    assert client.get("/admin").status_code == 200
    admin_post(client, "/admin/logout", follow_redirects=False)
    wrong_password = login_admin(client, "legacy-admin", "wrong-password")
    assert wrong_password.status_code == 400
    assert "用户名或密码不正确" in wrong_password.text


def test_basic_auth_header_does_not_authenticate_admin_or_protected_files(tmp_path, monkeypatch):
    client, main = build_client(
        tmp_path,
        monkeypatch,
        admin_auth=("legacy-admin", "legacy-secret"),
        admin_users="admin:123456,caigou:123456",
    )
    submit_ticket_with_image_and_file(client)
    uploaded_image = next((tmp_path / "uploads").glob("*.png"))
    basic_headers = basic_auth_headers("admin", "123456")
    fresh_client = TestClient(main.app)

    admin_page = fresh_client.get("/admin", headers=basic_headers, follow_redirects=False)
    assert admin_page.status_code == 303
    assert admin_page.headers["location"].startswith("/admin/login")

    login_page = fresh_client.get("/admin/login", headers=basic_headers, follow_redirects=False)
    assert login_page.status_code == 200
    assert "login-form" in login_page.text

    assert fresh_client.get("/admin/export", headers=basic_headers).status_code == 401
    assert fresh_client.get(f"/admin/uploads/{uploaded_image.name}", headers=basic_headers).status_code == 401
    assert fresh_client.get("/admin/files/1", headers=basic_headers).status_code == 401


def test_admin_detail_can_add_images_and_files_as_supplemental_attachments(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)
    response = submit_ticket(client)
    assert response.status_code == 200

    upload_response = client.post(
        "/admin/ticket/1/attachments",
        data={"csrf_token": csrf_token_for(client)},
        files=[
            ("new_images", ("proof.png", VALID_PNG, "image/png")),
            ("new_files", ("reply.xlsx", b"sku,qty\nA001,1\n", "application/octet-stream")),
        ],
        follow_redirects=False,
    )
    assert upload_response.status_code == 303
    assert "attachments_saved=1" in upload_response.headers["location"]

    images = rows_for(tmp_path, "ticket_images")
    files = rows_for(tmp_path, "ticket_files")
    logs = rows_for(tmp_path, "ticket_logs")
    assert len(images) == 1
    assert len(files) == 1
    assert files[0]["original_filename"] == "reply.xlsx"
    assert logs[-1]["action"] == "补充附件"
    assert logs[-1]["operator"] == ADMIN_AUTH[0]
    assert "新增图片 1 张，文件 1 个" in logs[-1]["note"]

    detail_page = client.get("/admin/ticket/1")
    assert detail_page.status_code == 200
    assert "proof" in detail_page.text or "查看图片" in detail_page.text
    assert "reply.xlsx" in detail_page.text
    assert "补充附件" in detail_page.text


def test_admin_detail_rejects_empty_supplemental_attachment_upload(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)
    response = submit_ticket(client)
    assert response.status_code == 200

    upload_response = admin_post(client, "/admin/ticket/1/attachments")

    assert upload_response.status_code == 400
    assert "请选择要上传的附件" in upload_response.text
    assert rows_for(tmp_path, "ticket_images") == []
    assert rows_for(tmp_path, "ticket_files") == []


def test_admin_summary_counts_all_filtered_results_not_current_page(tmp_path, monkeypatch):
    client, _ = build_client(
        tmp_path,
        monkeypatch,
        {
            "system.json": {
                "app_name": "门店需求工单系统",
                "port": 8701,
                "max_image_mb": 10,
                "allowed_image_extensions": ["jpg", "jpeg", "png", "webp"],
                "default_status": "待处理",
                "excel_filename_prefix": "门店需求工单",
                "page_size": 2,
                "max_image_count": 5,
                "max_total_upload_mb": 30,
                "allowed_file_extensions": ["pdf", "doc", "docx", "xls", "xlsx", "csv", "txt", "zip", "rar"],
                "max_file_mb": 20,
                "max_file_count": 5,
                "max_total_file_upload_mb": 50,
            }
        },
    )
    logged_in_client(client)
    submit_ticket(client, urgency="当天必须处理", description="第一个")
    submit_ticket(client, urgency="普通", description="第二个")
    submit_ticket(client, urgency="普通", description="第三个")
    admin_post(client, "/admin/ticket/2", data={"status": "处理中", "assigned_to": "采购", "handler_note": "处理中"})
    admin_post(client, "/admin/ticket/3", data={"status": "已完成", "assigned_to": "采购", "handler_note": "已完成"})

    admin_page = client.get("/admin?page=1")

    assert admin_page.status_code == 200
    assert "统计基于当前筛选条件，不受分页影响。" in admin_page.text
    assert 'data-summary-total="3"' in admin_page.text
    assert 'data-summary-pending="1"' in admin_page.text
    assert 'data-summary-processing="1"' in admin_page.text
    assert 'data-summary-today-urgent="1"' in admin_page.text
    assert 'data-summary-completed="1"' in admin_page.text


def test_admin_table_summary_uses_inner_clamped_div_instead_of_table_cell_display(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)
    submit_ticket(client, description="第一行\n第二行\n第三行，应该被两行省略")

    admin_page = client.get("/admin")

    assert admin_page.status_code == 200
    assert '<td class="summary-cell"' in admin_page.text
    assert '<div class="summary-text">' in admin_page.text
    assert "table-summary" not in admin_page.text


def test_detail_return_url_is_preserved_and_sanitized(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)
    submit_ticket(client)

    admin_page = client.get(f"/admin?status={quote('待处理')}&keyword={quote('自动化')}")
    assert admin_page.status_code == 200
    assert "return_url=" in admin_page.text

    safe_detail = client.get(f"/admin/ticket/1?return_url={quote('/admin?status=待处理&keyword=自动化', safe='')}")
    assert safe_detail.status_code == 200
    assert 'href="/admin?status=待处理&amp;keyword=自动化"' in safe_detail.text

    unsafe_detail = client.get("/admin/ticket/1?return_url=https%3A%2F%2Fevil.example%2Fadmin")
    assert unsafe_detail.status_code == 200
    assert 'href="/admin"' in unsafe_detail.text
    assert "evil.example" not in unsafe_detail.text


def test_admin_page_renders_polished_header_stats_and_table_badges(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    submit_ticket(client, urgency="当天必须处理", description="今天必须处理的陈列问题")
    submit_ticket(client, urgency="普通", description="普通补货需求")

    logged_in_client(client)
    admin_page = client.get("/admin")

    assert admin_page.status_code == 200
    assert "止痒工单后台" in admin_page.text
    assert "统一查看、筛选、处理门店需求" in admin_page.text
    assert "统计概览" in admin_page.text
    assert "当前筛选结果" in admin_page.text
    assert "待处理" in admin_page.text
    assert "当天必须处理" in admin_page.text
    assert "筛选条件" in admin_page.text
    assert "attachment-badge" in admin_page.text
    assert "summary-text" in admin_page.text


def test_detail_page_renders_grouped_layout_and_handler_hint(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    response = submit_ticket_with_image_and_file(client)
    assert response.status_code == 200

    logged_in_client(client)
    detail_page = client.get("/admin/ticket/1")

    assert detail_page.status_code == 200
    assert "基础信息" in detail_page.text
    assert "商品信息" in detail_page.text
    assert "时间信息" in detail_page.text
    assert "总部处理面板" in detail_page.text
    assert "状态改为已完成会记录完成时间" in detail_page.text
    assert "查看图片" in detail_page.text
    assert "下载文件" in detail_page.text
    assert "timeline-log" in detail_page.text
    assert re.search(r'/static/style\.css\?v=[A-Za-z0-9._-]+', detail_page.text)


def test_admin_users_allows_multiple_accounts_and_logs_actual_operator(tmp_path, monkeypatch):
    client, _ = build_client(
        tmp_path,
        monkeypatch,
        admin_auth=("legacy-admin", "legacy-secret"),
        admin_users=" admin : 123456 , caigou : 123456 , baduser: , :badpass ",
    )

    assert_login_success(login_admin(client, "admin", "123456"))
    admin_page = client.get("/admin")
    assert admin_page.status_code == 200
    assert "\u5f53\u524d\u8d26\u53f7\uff1aadmin" in admin_page.text
    admin_post(client, "/admin/logout", follow_redirects=False)
    assert login_admin(client, "caigou", "wrong-password").status_code == 400
    assert login_admin(client, "legacy-admin", "legacy-secret").status_code == 400
    assert_login_success(login_admin(client, "caigou", "123456"))
    caigou_page = client.get("/admin")
    assert caigou_page.status_code == 200
    assert "\u5f53\u524d\u8d26\u53f7\uff1acaigou" in caigou_page.text

    response = submit_ticket(client)
    assert response.status_code == 200

    update_response = admin_post(
        client,
        "/admin/ticket/1",
        data={"status": "处理中", "assigned_to": "采购", "handler_note": "采购账号处理"},
        follow_redirects=False,
    )
    assert update_response.status_code == 303

    logs = rows_for(tmp_path, "ticket_logs")
    assert len(logs) == 1
    assert logs[0]["operator"] == "caigou"


def test_legacy_admin_credentials_still_work_when_admin_users_missing(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch, admin_auth=("legacy-admin", "legacy-secret"))

    assert_login_success(login_admin(client, "legacy-admin", "legacy-secret"))
    assert client.get("/admin").status_code == 200
    admin_post(client, "/admin/logout", follow_redirects=False)
    assert login_admin(client, "legacy-admin", "wrong-password").status_code == 400


def test_file_upload_download_detail_admin_and_export(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    file_content = b"sku,qty\nA001,1\n"

    response = submit_ticket_with_file(client, filename="采购清单.xlsx", content=file_content)
    assert response.status_code == 200

    ticket_files = rows_for(tmp_path, "ticket_files")
    assert len(ticket_files) == 1
    ticket_file = ticket_files[0]
    assert ticket_file["original_filename"] == "采购清单.xlsx"
    assert ticket_file["file_ext"] == "xlsx"
    assert ticket_file["file_size"] == len(file_content)
    assert ticket_file["stored_filename"].startswith("FILE-REQ-")
    assert ticket_file["stored_filename"].endswith(".xlsx")
    assert ticket_file["stored_filename"] != "采购清单.xlsx"
    assert ticket_file["file_path"].startswith("uploads/")
    assert (tmp_path / ticket_file["file_path"]).read_bytes() == file_content

    assert client.get("/admin/files/1").status_code == 401
    logged_in_client(client)
    download_response = client.get("/admin/files/1")
    assert download_response.status_code == 200
    assert download_response.content == file_content
    assert download_response.headers["content-disposition"].lower().startswith("attachment")
    assert "%E9%87%87%E8%B4%AD%E6%B8%85%E5%8D%95.xlsx" in download_response.headers["content-disposition"]
    assert client.get("/files/1").status_code == 404

    detail_page = client.get("/admin/ticket/1")
    assert detail_page.status_code == 200
    assert "文件附件" in detail_page.text
    assert "采购清单.xlsx" in detail_page.text
    assert "/admin/files/1" in detail_page.text

    admin_page = client.get("/admin")
    assert admin_page.status_code == 200
    assert "图片 0 / 文件 1" in admin_page.text

    export_response = client.get("/admin/export")
    workbook = load_workbook(BytesIO(export_response.content))
    sheet = workbook.active
    assert [cell.value for cell in sheet[1]] == EXPECTED_EXPORT_HEADERS
    assert sheet["M2"].value == "采购清单.xlsx"
    assert sheet["N2"].value == "/admin/files/1"


def test_submit_admin_security_lifecycle_export_and_persistence(tmp_path, monkeypatch):
    client, main = build_client(tmp_path, monkeypatch)

    submit_page = client.get("/submit")
    assert submit_page.status_code == 200
    assert "南京门东店" in submit_page.text

    response = submit_ticket_with_image(client)
    assert response.status_code == 200
    assert "已提交" in response.text
    assert "请截图保存工单号" in response.text
    ticket_no = re.search(r"REQ-\d{8}-0001", response.text).group(0)

    tickets = rows_for(tmp_path, "tickets")
    assert tickets[0]["assigned_to"] in ("", None)
    assert tickets[0]["closed_at"] in ("", None)

    uploaded_files = list((tmp_path / "uploads").glob("*.png"))
    assert len(uploaded_files) == 1
    protected_path = f"/admin/uploads/{uploaded_files[0].name}"

    assert client.get(f"/uploads/{uploaded_files[0].name}").status_code != 200
    assert client.get(protected_path).status_code == 401
    logged_in_client(client)
    protected_response = client.get(protected_path)
    assert protected_response.status_code == 200
    assert protected_response.content.startswith(b"\x89PNG")
    assert client.get("/admin/uploads/../tickets.db").status_code == 404

    unauthenticated_client = TestClient(main.app)
    unauthenticated_admin = unauthenticated_client.get("/admin", follow_redirects=False)
    assert unauthenticated_admin.status_code == 303
    assert unauthenticated_admin.headers["location"].startswith("/admin/login")
    admin_page = client.get("/admin")
    assert admin_page.status_code == 200
    assert ticket_no in admin_page.text
    assert "处理人" in admin_page.text
    assert "图片数" in admin_page.text

    assert unauthenticated_client.get("/admin/ticket/1", follow_redirects=False).status_code == 303
    assert unauthenticated_client.post("/admin/ticket/1", data={"status": "处理中"}, follow_redirects=False).status_code == 303
    update_response = admin_post(
        client,
        "/admin/ticket/1",
        data={"status": "已完成", "assigned_to": "采购", "handler_note": "已安排总部同事处理"},
        follow_redirects=False,
    )
    assert update_response.status_code == 303

    tickets = rows_for(tmp_path, "tickets")
    assert tickets[0]["status"] == "已完成"
    assert tickets[0]["assigned_to"] == "采购"
    assert tickets[0]["closed_at"]
    logs = rows_for(tmp_path, "ticket_logs")
    assert len(logs) == 1
    assert logs[0]["old_status"] == "待处理"
    assert logs[0]["new_status"] == "已完成"
    assert logs[0]["new_assigned_to"] == "采购"
    assert logs[0]["operator"] == ADMIN_AUTH[0]

    detail_page = client.get("/admin/ticket/1")
    assert detail_page.status_code == 200
    assert "已安排总部同事处理" in detail_page.text
    assert "完成时间" in detail_page.text
    assert "处理日志" in detail_page.text
    assert protected_path in detail_page.text

    reopen_response = admin_post(
        client,
        "/admin/ticket/1",
        data={"status": "处理中", "assigned_to": "采购", "handler_note": "重新打开"},
        follow_redirects=False,
    )
    assert reopen_response.status_code == 303
    tickets = rows_for(tmp_path, "tickets")
    assert tickets[0]["closed_at"] in ("", None)
    assert len(rows_for(tmp_path, "ticket_logs")) == 2

    export_response = client.get("/admin/export")
    assert export_response.status_code == 200
    assert export_response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(export_response.content))
    sheet = workbook.active
    assert [cell.value for cell in sheet[1]] == EXPECTED_EXPORT_HEADERS
    assert sheet["A2"].value == ticket_no
    assert sheet["L2"].value == protected_path
    assert sheet["M2"].value in ("", None)
    assert sheet["N2"].value in ("", None)
    assert sheet["O2"].value == "2026-07-04"
    assert sheet["P2"].value == "处理中"
    assert sheet["Q2"].value == "采购"
    assert sheet["R2"].value == "重新打开"
    assert isinstance(sheet["U2"].value, (int, float))
    assert sheet["V2"].value in ("是", "否")

    monkeypatch.setenv("ADMIN_USERS", "")
    restarted_client = TestClient(main.create_app())
    monkeypatch.delenv("ADMIN_USERS", raising=False)
    assert_login_success(login_admin(restarted_client))
    restarted_admin_page = restarted_client.get("/admin")
    assert restarted_admin_page.status_code == 200
    assert ticket_no in restarted_admin_page.text


def test_upload_validation_rejects_fake_images_count_and_total_size(tmp_path, monkeypatch):
    client, _ = build_client(
        tmp_path,
        monkeypatch,
        {
            "system.json": {
                "app_name": "门店需求工单系统",
                "port": 8701,
                "max_image_mb": 10,
                "allowed_image_extensions": ["jpg", "jpeg", "png", "webp"],
                "default_status": "待处理",
                "excel_filename_prefix": "门店需求工单",
                "page_size": 50,
                "max_image_count": 1,
                "max_total_upload_mb": 1,
            }
        },
    )

    fake_jpg = submit_ticket_with_image(client, filename="fake.jpg", content=b"not an image")
    assert fake_jpg.status_code == 400
    assert "图片文件无法识别" in fake_jpg.text

    too_many = client.post(
        "/submit",
        data={
            "store_name": "南京门东店",
            "submitter": "测试提报人",
            "request_type": "建单需求",
            "urgency": "普通",
            "description": "超过图片数量限制",
        },
        files=[
            ("images", ("one.png", VALID_PNG, "image/png")),
            ("images", ("two.png", VALID_PNG, "image/png")),
        ],
    )
    assert too_many.status_code == 400
    assert "最多上传 1 张图片" in too_many.text

    too_large_total = submit_ticket_with_image(client, filename="large.png", content=VALID_PNG + b"x" * (1024 * 1024 + 1))
    assert too_large_total.status_code == 400
    assert "图片总大小不能超过 1MB" in too_large_total.text
    assert "REQ-" not in client.get("/admin").text


def test_file_upload_validation_rejects_extension_count_single_and_total_size(tmp_path, monkeypatch):
    limited_file_config = {
        "system.json": {
            "app_name": "门店需求工单系统",
            "port": 8701,
            "max_image_mb": 10,
            "allowed_image_extensions": ["jpg", "jpeg", "png", "webp"],
            "default_status": "待处理",
            "excel_filename_prefix": "门店需求工单",
            "page_size": 50,
            "max_image_count": 5,
            "max_total_upload_mb": 30,
            "allowed_file_extensions": ["txt"],
            "max_file_mb": 1,
            "max_file_count": 1,
            "max_total_file_upload_mb": 1,
        }
    }
    client, _ = build_client(tmp_path / "limited", monkeypatch, limited_file_config)

    executable = submit_ticket_with_file(client, filename="danger.exe", content=b"MZ")
    assert executable.status_code == 400
    assert "文件仅支持 txt 格式" in executable.text

    too_many = client.post(
        "/submit",
        data={
            "store_name": "南京门东店",
            "submitter": "测试提报人",
            "request_type": "建单需求",
            "urgency": "普通",
            "description": "超过文件数量限制",
        },
        files=[
            ("files", ("one.txt", b"one", "text/plain")),
            ("files", ("two.txt", b"two", "text/plain")),
        ],
    )
    assert too_many.status_code == 400
    assert "最多上传 1 个文件" in too_many.text

    single_size_config = dict(limited_file_config)
    single_size_config["system.json"] = dict(limited_file_config["system.json"], max_file_count=5, max_total_file_upload_mb=5)
    single_client, _ = build_client(tmp_path / "single", monkeypatch, single_size_config)
    too_large = submit_ticket_with_file(single_client, filename="large.txt", content=b"x" * (1024 * 1024 + 1))
    assert too_large.status_code == 400
    assert "单个文件不能超过 1MB" in too_large.text

    total_size_config = dict(limited_file_config)
    total_size_config["system.json"] = dict(limited_file_config["system.json"], max_file_mb=2, max_file_count=5)
    total_client, _ = build_client(tmp_path / "total", monkeypatch, total_size_config)
    too_large_total = total_client.post(
        "/submit",
        data={
            "store_name": "南京门东店",
            "submitter": "测试提报人",
            "request_type": "建单需求",
            "urgency": "普通",
            "description": "超过文件总大小限制",
        },
        files=[
            ("files", ("one.txt", b"x" * 600_000, "text/plain")),
            ("files", ("two.txt", b"x" * 600_000, "text/plain")),
        ],
    )
    assert too_large_total.status_code == 400
    assert "文件总大小不能超过 1MB" in too_large_total.text


def test_admin_can_delete_file_and_image_attachments_and_logs_actions(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    response = submit_ticket_with_image_and_file(client)
    assert response.status_code == 200

    image = rows_for(tmp_path, "ticket_images")[0]
    ticket_file = rows_for(tmp_path, "ticket_files")[0]
    image_path = tmp_path / image["image_path"]
    file_path = tmp_path / ticket_file["file_path"]
    assert image_path.exists()
    assert file_path.exists()
    logged_in_client(client)

    file_path.unlink()
    file_delete = admin_post(
        client,
        f"/admin/ticket/1/file/{ticket_file['id']}/delete",
        follow_redirects=False,
    )
    assert file_delete.status_code == 303
    assert rows_for(tmp_path, "ticket_files") == []

    image_path.unlink()
    image_delete = admin_post(
        client,
        f"/admin/ticket/1/image/{image['id']}/delete",
        follow_redirects=False,
    )
    assert image_delete.status_code == 303
    assert rows_for(tmp_path, "ticket_images") == []

    logs = rows_for(tmp_path, "ticket_logs")
    assert [log["action"] for log in logs] == ["删除文件", "删除图片"]
    assert all(log["operator"] == ADMIN_AUTH[0] for log in logs)

    detail_page = client.get("/admin/ticket/1")
    assert detail_page.status_code == 200
    assert "未上传图片" in detail_page.text
    assert "未上传文件附件" in detail_page.text
    assert "删除文件" in detail_page.text
    assert "删除图片" in detail_page.text


def test_ticket_soft_delete_restore_hard_delete_and_filters(tmp_path, monkeypatch):
    client, main = build_client(tmp_path, monkeypatch)
    submit_ticket_with_image_and_file(
        client,
        image_filename="delete-me.png",
        file_filename="delete-me.xlsx",
        description="Codex 待删除测试工单",
    )
    ticket_no = rows_for(tmp_path, "tickets")[0]["ticket_no"]
    image_path = tmp_path / rows_for(tmp_path, "ticket_images")[0]["image_path"]
    file_path = tmp_path / rows_for(tmp_path, "ticket_files")[0]["file_path"]
    assert image_path.exists()
    assert file_path.exists()

    logged_in_client(client)
    assert ticket_no in client.get("/admin").text
    assert ticket_no in client.get("/query?store_name=南京门东店").text
    workbook = load_workbook(BytesIO(client.get("/admin/export").content))
    assert ticket_no in {cell.value for cell in workbook.active["A"]}

    no_login = TestClient(main.app)
    unauthenticated_delete = no_login.post("/admin/ticket/1/delete", data={"delete_reason": "no login"}, follow_redirects=False)
    assert unauthenticated_delete.status_code == 303
    assert unauthenticated_delete.headers["location"].startswith("/admin/login")

    missing_csrf = client.post("/admin/ticket/1/delete", data={"delete_reason": "missing csrf"}, follow_redirects=False)
    assert missing_csrf.status_code == 403

    delete_response = admin_post(
        client,
        "/admin/ticket/1/delete",
        data={"delete_reason": "测试数据清理"},
        follow_redirects=False,
    )
    assert delete_response.status_code == 303
    assert delete_response.headers["location"] == "/admin"
    deleted_ticket = rows_for(tmp_path, "tickets")[0]
    assert deleted_ticket["deleted_at"]
    assert deleted_ticket["deleted_by"] == ADMIN_AUTH[0]
    assert deleted_ticket["delete_reason"] == "测试数据清理"

    assert ticket_no not in client.get("/admin").text
    assert ticket_no not in client.get("/query?store_name=南京门东店").text
    deleted_workbook = load_workbook(BytesIO(client.get("/admin/export").content))
    assert ticket_no not in {cell.value for cell in deleted_workbook.active["A"]}
    trash_page = client.get("/admin/trash")
    assert trash_page.status_code == 200
    assert ticket_no in trash_page.text
    assert "测试数据清理" in trash_page.text

    restore_response = admin_post(client, "/admin/ticket/1/restore", follow_redirects=False)
    assert restore_response.status_code == 303
    assert rows_for(tmp_path, "tickets")[0]["deleted_at"] is None
    assert ticket_no in client.get("/admin").text

    admin_post(client, "/admin/ticket/1/delete", data={"delete_reason": "永久删除前置"}, follow_redirects=False)
    hard_delete = admin_post(
        client,
        "/admin/ticket/1/hard-delete",
        data={"confirm_delete": "1"},
        follow_redirects=False,
    )
    assert hard_delete.status_code == 303
    assert rows_for(tmp_path, "tickets") == []
    assert rows_for(tmp_path, "ticket_images") == []
    assert rows_for(tmp_path, "ticket_files") == []
    assert not image_path.exists()
    assert not file_path.exists()


def test_admin_list_exposes_bulk_selection_controls(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    submit_ticket(client, description="bulk selection ticket A")
    submit_ticket(client, description="bulk selection ticket B")
    logged_in_client(client)

    page = client.get("/admin")

    assert page.status_code == 200
    assert 'class="bulk-action-bar' in page.text
    assert 'data-bulk-toolbar' in page.text
    assert 'data-select-current-page' in page.text
    assert 'data-clear-selection' in page.text
    assert 'data-select-filtered' in page.text
    assert 'data-filtered-count="2"' in page.text
    assert 'class="ticket-select-all"' in page.text
    assert 'class="ticket-select"' in page.text
    assert 'class="ticket-select-cell"' in page.text
    assert 'name="ticket_ids"' in page.text
    assert 'value="1"' in page.text
    assert 'value="2"' in page.text
    assert 'action="/admin/tickets/bulk-archive"' in page.text
    assert 'action="/admin/tickets/bulk-delete"' in page.text
    assert 'name="select_scope" value="selected"' in page.text
    assert 'name="csrf_token"' in page.text


def test_archive_trash_routes_and_sidebar_navigation(tmp_path, monkeypatch):
    client, main = build_client(tmp_path, monkeypatch)

    unauth_archive = TestClient(main.app).get("/admin/archive", follow_redirects=False)
    assert unauth_archive.status_code == 303
    assert unauth_archive.headers["location"].startswith("/admin/login")
    unauth_trash = TestClient(main.app).get("/admin/trash", follow_redirects=False)
    assert unauth_trash.status_code == 303
    assert unauth_trash.headers["location"].startswith("/admin/login")

    logged_in_client(client)
    admin_page = client.get("/admin")
    assert admin_page.status_code == 200
    assert 'href="/admin/archive"' in admin_page.text
    assert 'href="/admin/trash"' in admin_page.text

    archive_page = client.get("/admin/archive")
    assert archive_page.status_code == 200
    assert "归档工单" in archive_page.text
    assert 'href="/admin/archive"' in archive_page.text
    trash_page = client.get("/admin/trash")
    assert trash_page.status_code == 200
    assert "回收站" in trash_page.text
    assert 'href="/admin/trash"' in trash_page.text


def test_bulk_selection_css_keeps_checkboxes_and_toolbar_compact():
    style = (PROJECT_DIR / "static" / "style.css").read_text(encoding="utf-8")

    assert ".ticket-select-cell" in style
    assert ".ticket-select,\n.ticket-select-all" in style
    assert "width: 18px;" in style
    assert "height: 18px;" in style
    assert "min-height: 0;" in style
    assert ".bulk-action-bar" in style
    assert "min-height: 34px;" in style


def test_soft_neumorphic_design_tokens_and_component_contract():
    style = (PROJECT_DIR / "static" / "style.css").read_text(encoding="utf-8")

    for token in (
        "--app-bg: #eef3f9;",
        "--app-bg-2: #f6f9fd;",
        "--panel: #f8fbff;",
        "--panel-glass: rgba(255, 255, 255, 0.72);",
        "--primary-gradient: linear-gradient(135deg, #5b6cff 0%, #22d3ee 100%);",
        "--shadow-neu:",
        "--shadow-inset:",
        "--radius-xl: 28px;",
        "--control-h: 42px;",
    ):
        assert token in style

    for selector in (
        ".neu-card",
        ".primary-button",
        ".ghost-button",
        ".secondary-button",
        ".danger-button",
        ".modal-backdrop",
        ".modal-card",
        ".drawer-overlay",
        ".drawer-panel",
        ".toast-container",
        ".toast-card",
    ):
        assert selector in style

    assert re.search(r'input\[type="checkbox"\][^{]*{[^}]*width:\s*18px;[^}]*height:\s*18px;', style, re.DOTALL)
    assert ".primary-button:hover" in style
    assert ":focus-visible" in style


def test_soft_neumorphic_templates_expose_workbench_sections(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    submit_ticket(client, description="soft neumorphic ui contract")
    logged_in_client(client)

    dashboard = client.get("/admin/dashboard")
    assert dashboard.status_code == 200
    for fragment in (
        "dashboard-workbench-hero",
        "dashboard-metric-grid",
        "metric-icon",
        "dashboard-activity-grid",
    ):
        assert fragment in dashboard.text

    admin_page = client.get("/admin")
    assert admin_page.status_code == 200
    assert "neumorphic-stat-card" in admin_page.text
    assert "bulk-action-bar" in admin_page.text
    assert "ticket-no-link" in admin_page.text

    schedule_page = client.get("/admin/schedules?store_name=å—äº¬é—¨ä¸œåº—&month=2026-07")
    assert schedule_page.status_code == 200
    for fragment in (
        "schedule-hero-card",
        "schedule-filter-card",
        "schedule-bulk-card",
        "schedule-view-card",
        "schedule-view-tabs",
    ):
        assert fragment in schedule_page.text

    create_schedule_employee(client, "NeoCardEmp", role="Crew")
    employees_page = client.get("/admin/employees")
    assert employees_page.status_code == 200
    assert "employee-role-group" in employees_page.text
    assert "employee-card" in employees_page.text
    assert "drawer-panel" in employees_page.text

    shifts_page = client.get("/admin/shift-types")
    assert shifts_page.status_code == 200
    assert "shift-type-card" in shifts_page.text
    assert "shift-time-axis" in shifts_page.text


def test_static_assets_are_versioned_and_core_pages_do_not_404(tmp_path, monkeypatch):
    client, main = build_client(tmp_path, monkeypatch)
    submit_ticket(client, description="core page smoke")

    for public_path in ("/submit", "/query", "/schedule", "/__version", "/healthz"):
        response = client.get(public_path, follow_redirects=False)
        assert response.status_code == 200

    unauthenticated = TestClient(main.app).get("/admin/dashboard", follow_redirects=False)
    assert unauthenticated.status_code == 303
    assert unauthenticated.headers["location"].startswith("/admin/login")

    logged_in_client(client)
    for admin_path in (
        "/admin/dashboard",
        "/admin",
        "/admin/ticket/1",
        "/admin/archive",
        "/admin/trash",
        "/admin/employees",
        "/admin/shift-types",
        "/admin/schedules",
        "/admin/embedded-pages",
    ):
        response = client.get(admin_path, follow_redirects=False)
        assert response.status_code == 200
        if response.headers.get("content-type", "").startswith("text/html"):
            assert re.search(r'/static/style\.css\?v=[A-Za-z0-9._-]+', response.text)
            assert re.search(r'/static/app\.js\?v=[A-Za-z0-9._-]+', response.text)


def assert_native_bulk_form_wraps_checkbox(page_text, expected_actions):
    form_start = page_text.index('<form id="bulk-ticket-form"')
    form_tag_end = page_text.index(">", form_start)
    form_end = page_text.index("</form>", form_start)
    checkbox_index = page_text.index('name="ticket_ids"', form_start)
    assert checkbox_index < form_end
    form_tag = page_text[form_start:form_tag_end]
    fragment = page_text[form_start:form_end]
    assert 'name="csrf_token"' in fragment
    assert 'name="select_scope" value="selected"' in fragment
    assert 'name="source_view"' in fragment
    assert 'data-bulk-form' in fragment
    assert any(f'action="{action}"' in form_tag for action in expected_actions)
    for action in expected_actions:
        assert f'formaction="{action}"' in fragment or f'action="{action}"' in fragment
    assert fragment.count("data-bulk-submit") == len(expected_actions)


def test_admin_javascript_initialization_is_page_isolated():
    app_js = (PROJECT_DIR / "static" / "app.js").read_text(encoding="utf-8")

    assert "initializeStoreRequestApp" in app_js
    assert "try {" in app_js
    assert "catch (error)" in app_js
    assert "console.error" in app_js
    assert "safeLocalStorageGet" in app_js
    assert 'window.localStorage.getItem("storeRequestDesktopNotifications")' not in app_js
    assert "if (!notificationRoot)" in app_js
    assert "if (!scheduleBulkForm" in app_js


def test_admin_template_links_and_form_actions_have_registered_routes(tmp_path, monkeypatch):
    _, main = build_client(tmp_path, monkeypatch)
    routes = {
        (method, route.path)
        for route in main.app.routes
        for method in (getattr(route, "methods", None) or [])
        if hasattr(route, "path")
    }

    missing = [
        f"{template_name} {attr} {method} {path}"
        for template_name, attr, path, method in template_admin_route_references()
        if not route_exists(routes, method, path)
    ]

    assert missing == []


def test_legacy_admin_paths_redirect_safely_after_login(tmp_path, monkeypatch):
    client, main = build_client(tmp_path, monkeypatch)
    legacy_paths = {
        "/admin/personnel": "/admin/employees",
        "/admin/staff": "/admin/employees",
        "/admin/employee": "/admin/employees",
        "/admin/schedule": "/admin/schedules",
        "/admin/store-schedule": "/admin/schedules",
        "/admin/shift-type": "/admin/shift-types",
        "/admin/archive-list": "/admin/archive",
        "/admin/recycle": "/admin/trash",
        "/admin/trashes": "/admin/trash",
    }

    unauthenticated = TestClient(main.app).get("/admin/personnel", follow_redirects=False)
    assert unauthenticated.status_code == 303
    assert unauthenticated.headers["location"].startswith("/admin/login")

    logged_in_client(client)
    for old_path, new_path in legacy_paths.items():
        response = client.get(old_path, follow_redirects=False)
        assert response.status_code == 303
        assert response.headers["location"] == new_path


def test_admin_route_health_page_reports_registered_routes_and_template_scan(tmp_path, monkeypatch):
    client, main = build_client(tmp_path, monkeypatch)
    unauthenticated = TestClient(main.app).get("/admin/route-health", follow_redirects=False)
    assert unauthenticated.status_code == 303
    assert unauthenticated.headers["location"].startswith("/admin/login")

    logged_in_client(client)
    page = client.get("/admin/route-health")
    assert page.status_code == 200
    assert "data-route-health" in page.text
    assert "data-route-health-navigation" in page.text
    assert "data-route-health-forms" in page.text
    assert "未匹配导航项" in page.text
    assert "未匹配 form action" in page.text
    assert "/admin/my-work" in page.text
    assert "/admin/tickets/bulk-archive" in page.text
    assert "route-health-missing" not in page.text


def test_base_admin_navigation_uses_only_canonical_admin_paths():
    text = (PROJECT_DIR / "templates" / "base_admin.html").read_text(encoding="utf-8")
    allowed_paths = {
        "/admin/dashboard",
        "/admin/my-work",
        "/admin",
        "/admin/archive",
        "/admin/trash",
        "/admin/cleanup",
        "/admin/employees",
        "/admin/shift-types",
            "/admin/schedules",
            "/admin/settings",
            "/admin/ticket-rules",
            "/admin/account",
        "/admin/personnel-governance",
        "/admin/roles",
        "/admin/route-health",
        "/admin/system",
        "/admin/embedded-pages",
    }
    forbidden_paths = {
        "/admin/personnel",
        "/admin/staff",
        "/admin/schedule",
        "/admin/archive-list",
        "/admin/recycle",
    }
    hrefs = set(re.findall(r'href="(/admin[^"]*)"', text))

    assert not forbidden_paths.intersection(hrefs)
    for href in hrefs:
        if "{{" in href:
            continue
        assert href in allowed_paths


def test_admin_core_pages_return_200_after_login(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)

    for path in (
        "/admin/dashboard",
        "/admin/my-work",
        "/admin",
        "/admin/archive",
        "/admin/trash",
        "/admin/cleanup",
        "/admin/employees",
        "/admin/shift-types",
        "/admin/schedules",
        "/admin/settings",
        "/admin/account",
        "/admin/personnel-governance",
        "/admin/system",
        "/admin/embedded-pages",
        "/admin/route-health",
    ):
        response = client.get(path)
        assert response.status_code == 200, path
        assert '{"detail":"Not Found"}' not in response.text


def test_version_endpoint_reports_runtime_route_health(tmp_path, monkeypatch):
    client, main = build_client(tmp_path, monkeypatch)
    required_paths = [
        "/admin/my-work",
        "/admin/archive",
        "/admin/trash",
        "/admin/employees",
        "/admin/shift-types",
        "/admin/schedules",
        "/admin/tickets/bulk-archive",
        "/admin/tickets/bulk-delete",
    ]

    response = client.get("/__version")

    assert response.status_code == 200
    payload = response.json()
    assert payload["app"] == "store-request-tool"
    assert Path(payload["main_file"]).resolve() == (PROJECT_DIR / "main.py").resolve()
    assert payload["route_count"] == len(main.app.routes)
    assert payload["required_missing_routes"] == []
    assert set(required_paths).issubset({route.path for route in main.app.routes if hasattr(route, "path")})
    assert payload["git_commit"] == main.APP_GIT_COMMIT
    monkeypatch.setattr(main, "current_git_commit", lambda: "newer-disk-commit")
    assert client.get("/__version").json()["git_commit"] == main.APP_GIT_COMMIT
    assert "started_at" in payload
    assert ".env" not in json.dumps(payload, ensure_ascii=False)


def test_runtime_health_endpoints_report_fixed_access_contract(tmp_path, monkeypatch):
    client, main = build_client(tmp_path, monkeypatch)

    version = client.get("/__version")
    health = client.get("/healthz")

    assert version.status_code == 200
    payload = version.json()
    assert payload["app"] == "store-request-tool"
    assert payload["name"] == "止痒 ERP"
    assert Path(payload["main_file"]).resolve() == (PROJECT_DIR / "main.py").resolve()
    assert payload["git_commit"]
    assert payload["route_count"] == len(main.app.routes)
    assert payload["port"] == 8701
    assert payload["started_at"]
    assert payload["required_missing_routes"] == []
    assert payload["urls"]["submit"] == "http://127.0.0.1:8701/submit"
    assert payload["urls"]["query"] == "http://127.0.0.1:8701/query"
    assert payload["urls"]["admin_login"] == "http://127.0.0.1:8701/admin/login"
    assert payload["urls"]["dashboard"] == "http://127.0.0.1:8701/admin/dashboard"
    assert payload["urls"]["tickets"] == "http://127.0.0.1:8701/admin"
    assert payload["urls"]["schedules"] == "http://127.0.0.1:8701/admin/schedules"
    assert payload["urls"]["employees"] == "http://127.0.0.1:8701/admin/employees"
    assert payload["urls"]["shift_types"] == "http://127.0.0.1:8701/admin/shift-types"
    assert payload["urls"]["embedded_pages"] == "http://127.0.0.1:8701/admin/embedded-pages"
    assert payload["urls"]["route_health"] == "http://127.0.0.1:8701/admin/route-health"
    assert ".env" not in json.dumps(payload, ensure_ascii=False)
    assert ADMIN_AUTH[1] not in json.dumps(payload, ensure_ascii=False)

    assert health.status_code == 200
    health_payload = health.json()
    assert health_payload["ok"] is True
    assert health_payload["database"] is True
    assert health_payload["upload_dir"] is True
    assert health_payload["embedded_pages_dir"] is True
    assert health_payload["route_count"] == len(main.app.routes)


def test_fixed_access_urls_never_return_404_when_unauthenticated(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    fixed_paths = [
        "/",
        "/submit",
        "/query",
        "/admin/login",
        "/admin/dashboard",
        "/admin",
        "/admin/my-work",
        "/admin/archive",
        "/admin/trash",
        "/admin/cleanup",
        "/admin/employees",
        "/admin/shift-types",
        "/admin/schedules",
        "/admin/settings",
        "/admin/account",
        "/admin/system",
        "/admin/embedded-pages",
        "/__version",
        "/healthz",
    ]

    for path in fixed_paths:
        response = client.get(path, follow_redirects=False)
        assert response.status_code in {200, 303}, path
        assert response.status_code != 404, path


def test_legacy_public_and_admin_urls_redirect_with_303(tmp_path, monkeypatch):
    client, main = build_client(tmp_path, monkeypatch)
    admin_legacy_paths = {
        "/admin/personnel": "/admin/employees",
        "/admin/staff": "/admin/employees",
        "/admin/employee": "/admin/employees",
        "/admin/schedule": "/admin/schedules",
        "/admin/store-schedule": "/admin/schedules",
        "/admin/shift-type": "/admin/shift-types",
        "/admin/archive-list": "/admin/archive",
        "/admin/recycle": "/admin/trash",
        "/admin/trashes": "/admin/trash",
        "/admin/home": "/admin/dashboard",
        "/admin/index": "/admin/dashboard",
        "/admin/tickets": "/admin",
        "/admin/orders": "/admin",
    }
    public_legacy_paths = {
        "/ticket": "/query",
        "/tickets": "/query",
        "/new": "/submit",
        "/create": "/submit",
        "/form": "/submit",
    }

    for old_path in admin_legacy_paths:
        unauthenticated = TestClient(main.app).get(old_path, follow_redirects=False)
        assert unauthenticated.status_code == 303
        assert unauthenticated.headers["location"].startswith("/admin/login")

    logged_in_client(client)
    for old_path, new_path in admin_legacy_paths.items():
        response = client.get(old_path, follow_redirects=False)
        assert response.status_code == 303, old_path
        assert response.headers["location"] == new_path

    public_client = TestClient(main.app)
    for old_path, new_path in public_legacy_paths.items():
        response = public_client.get(old_path, follow_redirects=False)
        assert response.status_code == 303, old_path
        assert response.headers["location"] == new_path


def test_not_found_pages_are_html_diagnostics_for_admin_and_public(tmp_path, monkeypatch):
    client, main = build_client(tmp_path, monkeypatch)
    unauthenticated = TestClient(main.app).get("/admin/old-missing-link", follow_redirects=False)
    assert unauthenticated.status_code == 303
    assert unauthenticated.headers["location"].startswith("/admin/login")

    logged_in_client(client)
    admin_missing = client.get("/admin/old-missing-link")
    public_missing = client.get("/old-public-link")

    assert admin_missing.status_code == 404
    assert "text/html" in admin_missing.headers["content-type"]
    assert "data-not-found-page" in admin_missing.text
    assert "data-admin-not-found" in admin_missing.text
    assert "/admin/old-missing-link" in admin_missing.text
    assert "页面地址已变更" in admin_missing.text
    assert "/admin/dashboard" in admin_missing.text
    assert "/admin" in admin_missing.text
    assert "/admin/schedules" in admin_missing.text
    assert "/admin/route-health" in admin_missing.text
    assert "/__version" in admin_missing.text
    assert '{"detail":"Not Found"}' not in admin_missing.text

    assert public_missing.status_code == 404
    assert "text/html" in public_missing.headers["content-type"]
    assert "data-not-found-page" in public_missing.text
    assert "返回提交工单" in public_missing.text
    assert "返回查询工单" in public_missing.text
    assert '{"detail":"Not Found"}' not in public_missing.text


def test_access_url_docs_run_bat_and_asset_versions_are_stable():
    access_urls = PROJECT_DIR / "ACCESS_URLS.md"
    assert access_urls.exists()
    access_content = access_urls.read_text(encoding="utf-8")
    for text in (
        "http://127.0.0.1:8701/submit",
        "http://127.0.0.1:8701/query",
        "http://127.0.0.1:8701/admin/login",
        "http://127.0.0.1:8701/admin/dashboard",
        "http://127.0.0.1:8701/admin",
        "http://127.0.0.1:8701/admin/schedules",
        "http://127.0.0.1:8701/admin/route-health",
        "http://127.0.0.1:8701/__version",
        "http://127.0.0.1:8701/healthz",
    ):
        assert text in access_content

    readme = (PROJECT_DIR / "README.md").read_text(encoding="utf-8")
    assert "本地启动与更新" in readme
    assert "restart.bat" in readme
    assert "check_runtime.bat" in readme
    assert "访问地址不变原则" in readme
    assert "/__version" in readme
    assert "Ctrl+F5" in readme
    assert "兼容地址" in readme

    run_bat = (PROJECT_DIR / "run.bat").read_text(encoding="utf-8")
    assert "Store Request Tool startup" in run_bat
    assert "PORT=8701" in run_bat
    assert "PID" in run_bat
    assert "WARNING" in run_bat
    assert "当前 8701 已被 PID" in run_bat
    assert "这可能是旧服务未关闭" in run_bat
    assert "请执行 restart.bat" in run_bat
    assert "http://127.0.0.1:8701/submit" in run_bat
    assert "http://127.0.0.1:8701/admin/route-health" in run_bat
    assert "python -m uvicorn main:app --host 127.0.0.1 --port 8701" in run_bat

    for template_name in ("base_admin.html", "base_public.html", "login.html"):
        template = (PROJECT_DIR / "templates" / template_name).read_text(encoding="utf-8")
        assert "/static/style.css?v={{ asset_version }}" in template
        assert "ui20260704" not in template
    assert "/static/app.js?v={{ asset_version }}" in (PROJECT_DIR / "templates" / "base_admin.html").read_text(encoding="utf-8")
    assert "/static/app.js?v={{ asset_version }}" in (PROJECT_DIR / "templates" / "base_public.html").read_text(encoding="utf-8")


def test_route_health_reports_required_routes_and_zero_missing(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)

    page = client.get("/admin/route-health")

    assert page.status_code == 200
    assert 'data-missing-required-count="0"' in page.text
    assert "后台路由正常" in page.text
    assert "必须存在的关键路由" in page.text
    assert "/admin/dashboard" in page.text
    assert "/admin/embedded-pages" in page.text
    assert "当前推荐访问地址" in page.text
    assert "http://127.0.0.1:8701/admin/route-health" in page.text


def test_admin_not_found_uses_diagnostic_html(tmp_path, monkeypatch):
    client, main = build_client(tmp_path, monkeypatch)
    unauthenticated = TestClient(main.app).get("/admin/old-missing-link", follow_redirects=False)
    assert unauthenticated.status_code == 303
    assert unauthenticated.headers["location"].startswith("/admin/login")

    logged_in_client(client)
    response = client.get("/admin/old-missing-link")

    assert response.status_code == 404
    assert "text/html" in response.headers["content-type"]
    assert "data-admin-error" in response.text
    assert "data-admin-not-found" in response.text
    assert "/admin/old-missing-link" in response.text
    assert "GET" in response.text
    assert "你可能访问了旧链接" in response.text
    assert "/admin/dashboard" in response.text
    assert "/admin/route-health" in response.text
    assert '{"detail":"Not Found"}' not in response.text


def test_run_bat_prints_startup_route_diagnostics():
    content = (PROJECT_DIR / "run.bat").read_text(encoding="utf-8")
    raw_content = (PROJECT_DIR / "run.bat").read_bytes()

    assert "PYTHONUTF8=1" in content
    assert "PYTHONIOENCODING=utf-8" in content
    assert "CURRENT_DIR=" in content
    assert "GIT_COMMIT=" in content
    assert "PYTHON_EXE=" in content
    assert "MAIN_FILE=" in content
    assert "ROUTE_COUNT=" in content
    assert "MISSING_ROUTES=" in content
    assert "Critical routes missing. Do not continue startup." in content
    assert "python -m uvicorn main:app --host 127.0.0.1 --port 8701" in content
    assert "startup_check.py" in content
    assert "请执行 restart.bat" in content
    assert "pause" in content.lower()
    assert raw_content.count(b"\r\n") > 20
    assert raw_content.count(b"\n") == raw_content.count(b"\r\n")


def test_runtime_restart_scripts_are_project_scoped_and_version_aware():
    stop_bat = PROJECT_DIR / "stop.bat"
    restart_bat = PROJECT_DIR / "restart.bat"
    check_runtime_bat = PROJECT_DIR / "check_runtime.bat"

    for script in (stop_bat, restart_bat, check_runtime_bat):
        assert script.exists(), script.name
        text = script.read_text(encoding="utf-8")
        assert "taskkill /IM python.exe" not in text
        assert "taskkill /IM python" not in text

    stop_text = stop_bat.read_text(encoding="utf-8")
    assert "Get-NetTCPConnection" in stop_text
    assert "Get-CimInstance Win32_Process" in stop_text
    assert "store_request_tool" in stop_text
    assert "uvicorn main:app" in stop_text
    assert r"D:\需求小程序\store_request_tool" in stop_text
    assert "taskkill /PID" in stop_text
    assert "--no-pause" in stop_text
    assert "无法确认这是本项目进程" in stop_text
    assert "8701 未运行" in stop_text
    assert "pause" in stop_text.lower()

    restart_text = restart_bat.read_text(encoding="utf-8")
    assert 'call "%~dp0stop.bat" --no-pause' in restart_text
    assert "timeout /t 1" in restart_text
    assert 'call "%~dp0run.bat"' in restart_text
    assert "git rev-parse --short HEAD" in restart_text
    assert "C:\\Program Files\\Git\\cmd\\git.exe" in restart_text
    for url in (
        "http://127.0.0.1:8701/healthz",
        "http://127.0.0.1:8701/__version",
        "http://127.0.0.1:8701/admin/login",
        "http://127.0.0.1:8701/admin/dashboard",
    ):
        assert url in restart_text

    check_text = check_runtime_bat.read_text(encoding="utf-8")
    assert "git rev-parse --short HEAD" in check_text
    assert "C:\\Program Files\\Git\\cmd\\git.exe" in check_text
    assert "http://127.0.0.1:8701/__version" in check_text
    assert "ConvertFrom-Json" in check_text
    assert "运行态一致" in check_text
    assert "运行态不一致，请执行 restart.bat" in check_text
    assert "服务未启动" in check_text
    assert "pause" in check_text.lower()


def test_startup_diagnostics_scripts_are_present_and_safe(tmp_path, monkeypatch):
    startup_check = PROJECT_DIR / "startup_check.py"
    run_debug = PROJECT_DIR / "run_debug.bat"
    gitignore = (PROJECT_DIR / ".gitignore").read_text(encoding="utf-8")

    assert startup_check.exists()
    startup_text = startup_check.read_text(encoding="utf-8")
    for package_name in ("fastapi", "uvicorn", "jinja2", "openpyxl", "PIL"):
        assert package_name in startup_text
    for route_path in (
        "/submit",
        "/query",
        "/admin/login",
        "/admin/dashboard",
        "/admin",
        "/admin/schedules",
        "/admin/employees",
        "/admin/shift-types",
        "/__version",
        "/healthz",
    ):
        assert route_path in startup_text

    debug_text = run_debug.read_text(encoding="utf-8")
    debug_bytes = run_debug.read_bytes()
    assert "logs\\startup.log" in debug_text
    assert "startup_check.py" in debug_text
    assert "uvicorn" in debug_text
    assert "pause" in debug_text.lower()
    assert "2>&1" in debug_text
    assert debug_bytes.count(b"\r\n") > 20
    assert debug_bytes.count(b"\n") == debug_bytes.count(b"\r\n")
    assert "logs/" in gitignore

    env = os.environ.copy()
    env["STORE_REQUEST_DB_PATH"] = str(tmp_path / "tickets.db")
    env["STORE_REQUEST_UPLOAD_DIR"] = str(tmp_path / "uploads")
    env["STORE_REQUEST_CONFIG_DIR"] = str(tmp_path / "config")
    env["ADMIN_USERNAME"] = ADMIN_AUTH[0]
    env["ADMIN_PASSWORD"] = ADMIN_AUTH[1]
    env["SESSION_SECRET"] = "test-session-secret"
    result = subprocess.run(
        [sys.executable, str(startup_check), "--format", "json"],
        cwd=PROJECT_DIR,
        env=env,
        text=True,
        encoding="utf-8",
        capture_output=True,
        timeout=30,
        check=False,
    )
    assert result.returncode == 0, result.stdout + result.stderr
    payload = json.loads(result.stdout)
    assert payload["ok"] is True
    assert payload["main_imported"] is True
    assert payload["app_exists"] is True
    assert payload["missing_routes"] == []
    assert payload["route_count"] > 0


def test_modal_open_and_close_controls_do_not_submit_or_navigate():
    for template_name in ("admin.html", "employees.html", "schedules.html", "shift_types.html"):
        text = (PROJECT_DIR / "templates" / template_name).read_text(encoding="utf-8")
        for match in re.finditer(r"<(?P<tag>\w+)\b[^>]*\bdata-modal-open\b[^>]*>", text, re.IGNORECASE):
            tag = match.group(0)
            assert match.group("tag").lower() == "button", f"{template_name}: {tag}"
            assert re.search(r'\btype\s*=\s*(["\'])button\1', tag, re.IGNORECASE), f"{template_name}: {tag}"
        for match in re.finditer(r"<(?P<tag>\w+)\b[^>]*\bdata-modal-close\b[^>]*>", text, re.IGNORECASE):
            tag = match.group(0)
            assert match.group("tag").lower() == "button", f"{template_name}: {tag}"
            assert re.search(r'\btype\s*=\s*(["\'])button\1', tag, re.IGNORECASE), f"{template_name}: {tag}"
        assert not re.search(r"<a\b[^>]*\bdata-modal-open\b", text, re.IGNORECASE), template_name


def test_admin_ticket_create_uses_modal_form_not_page_link(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch, config_overrides={"brands.json": ["ModalBrand"]})
    logged_in_client(client)

    page = client.get("/admin")

    assert page.status_code == 200
    assert 'data-modal="ticket-create-modal"' in page.text
    assert "data-modal-overlay" in page.text
    assert 'data-modal-open="ticket-create-modal"' in page.text
    assert 'data-admin-ticket-create-form' in page.text
    assert 'action="/admin/tickets/create"' in page.text
    assert 'method="post"' in page.text
    assert 'enctype="multipart/form-data"' in page.text
    assert 'name="csrf_token"' in page.text
    for field_name in (
        "store_names",
        "submitter",
        "request_type",
        "urgency",
        "brands",
        "brand_extra",
        "product_name",
        "sku_barcode",
        "quantity",
        "description",
        "expected_finish_date",
        "images",
        "files",
    ):
        assert f'name="{field_name}"' in page.text
    assert re.search(
        r"<button\b[^>]*\btype=(['\"])button\1[^>]*\bdata-modal-open=(['\"])ticket-create-modal\2",
        page.text,
        re.IGNORECASE,
    )
    assert not re.search(r"<a\b[^>]*\bdata-modal-open\b", page.text, re.IGNORECASE)


def test_admin_ticket_create_json_success_and_inline_error(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)
    store_name = DEFAULT_TEST_CONFIG["stores.json"][0]
    request_type = DEFAULT_TEST_CONFIG["request_types.json"][0]
    urgency = DEFAULT_TEST_CONFIG["urgency_levels.json"][0]

    response = admin_post(
        client,
        "/admin/tickets/create",
        data={
            "store_names": [store_name],
            "submitter": "Admin modal",
            "request_type": request_type,
            "urgency": urgency,
            "description": "created from admin modal",
        },
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/json")
    payload = response.json()
    assert payload["ok"] is True
    assert payload["ticket_id"] == 1
    assert payload["ticket_no"]
    assert "message" in payload
    tickets = rows_for(tmp_path, "tickets")
    assert len(tickets) == 1
    assert tickets[0]["submitter"] == "Admin modal"
    assert tickets[0]["description"] == "created from admin modal"

    error_response = admin_post(
        client,
        "/admin/tickets/create",
        data={
            "submitter": "",
            "request_type": request_type,
            "urgency": urgency,
            "description": "",
        },
    )

    assert error_response.status_code == 400
    assert error_response.headers["content-type"].startswith("application/json")
    error_payload = error_response.json()
    assert error_payload["ok"] is False
    assert error_payload["error"]
    assert "detail" not in error_payload


def test_submit_page_remains_compatible_after_ticket_create_modal(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)

    page = client.get("/submit")
    response = submit_ticket(client, description="standalone submit still works")

    assert page.status_code == 200
    assert 'action="/submit"' in page.text
    assert response.status_code == 200
    assert "submit_success" in response.text or "ticket_no" in response.text or rows_for(tmp_path, "tickets")
    assert rows_for(tmp_path, "tickets")[0]["description"] == "standalone submit still works"


def test_ticket_create_modal_js_and_css_contract():
    app_js = (PROJECT_DIR / "static" / "app.js").read_text(encoding="utf-8")
    style_css = (PROJECT_DIR / "static" / "style.css").read_text(encoding="utf-8")

    assert "window.openModal" in app_js
    assert "window.closeModal" in app_js
    assert "data-admin-ticket-create-form" in app_js
    assert "fetch(form.action" in app_js
    assert "data-modal-error" in app_js
    assert "window.location.reload" in app_js
    assert "Escape" in app_js
    assert "data-modal-overlay" in app_js
    for selector in (
        ".modal-overlay",
        ".modal-container",
        ".modal-header",
        ".modal-body",
        ".modal-footer",
    ):
        assert selector in style_css


def test_bulk_pages_use_native_form_submission_without_dynamic_ticket_inputs(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    submit_ticket(client, description="native active ticket")
    submit_ticket(client, description="native archived ticket")
    submit_ticket(client, description="native trash ticket")
    logged_in_client(client)

    admin_post(
        client,
        "/admin/tickets/bulk-archive",
        data={"ticket_ids": ["2"], "select_scope": "selected"},
        follow_redirects=False,
    )
    admin_post(
        client,
        "/admin/tickets/bulk-delete",
        data={"ticket_ids": ["3"], "select_scope": "selected"},
        follow_redirects=False,
    )

    assert_native_bulk_form_wraps_checkbox(
        client.get("/admin").text,
        ["/admin/tickets/bulk-archive", "/admin/tickets/bulk-delete"],
    )
    assert_native_bulk_form_wraps_checkbox(
        client.get("/admin/archive").text,
        ["/admin/tickets/bulk-unarchive", "/admin/tickets/bulk-delete"],
    )
    assert_native_bulk_form_wraps_checkbox(
        client.get("/admin/trash").text,
        ["/admin/tickets/bulk-restore", "/admin/tickets/bulk-hard-delete"],
    )
    assert 'formaction="/admin/tickets/bulk-restore" formnovalidate' in client.get("/admin/trash").text

    app_js = (PROJECT_DIR / "static" / "app.js").read_text(encoding="utf-8")
    assert "appendSelectedTicketInputs" not in app_js
    assert "data-generated-ticket-id" not in app_js


def test_bulk_empty_selection_redirects_back_with_friendly_error(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    submit_ticket(client, description="empty bulk active")
    submit_ticket(client, description="empty bulk archived")
    submit_ticket(client, description="empty bulk trash")
    logged_in_client(client)
    admin_post(client, "/admin/tickets/bulk-archive", data={"ticket_ids": ["2"]}, follow_redirects=False)
    admin_post(client, "/admin/tickets/bulk-delete", data={"ticket_ids": ["3"]}, follow_redirects=False)

    cases = [
        ("/admin/tickets/bulk-archive", {"source_view": "active"}, "/admin", "请选择要操作的工单"),
        ("/admin/tickets/bulk-delete", {"source_view": "active"}, "/admin", "请选择要操作的工单"),
        ("/admin/tickets/bulk-unarchive", {"source_view": "archive"}, "/admin/archive", "请选择要操作的工单"),
        ("/admin/tickets/bulk-restore", {"source_view": "trash"}, "/admin/trash", "请选择要操作的工单"),
        (
            "/admin/tickets/bulk-hard-delete",
            {"source_view": "trash", "confirm_delete": "1"},
            "/admin/trash",
            "请选择要操作的工单",
        ),
    ]

    for path, data, expected_prefix, message in cases:
        response = admin_post(client, path, data=data, follow_redirects=False)
        assert response.status_code == 303
        location = unquote(response.headers["location"])
        assert location.startswith(f"{expected_prefix}?")
        assert f"error={message}" in location
        page = client.get(response.headers["location"])
        assert page.status_code == 200
        assert message in page.text


def test_bulk_archive_selected_ticket_moves_to_archive_and_can_unarchive(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    submit_ticket(client, description="archive selected ticket")
    submit_ticket(client, description="active neighbor ticket")
    ticket_no = rows_for(tmp_path, "tickets")[0]["ticket_no"]
    store_name = rows_for(tmp_path, "tickets")[0]["store_name"]
    logged_in_client(client)

    archive = admin_post(
        client,
        "/admin/tickets/bulk-archive",
        data={"ticket_ids": ["1"], "select_scope": "selected", "archive_reason": "batch archive reason"},
        follow_redirects=False,
    )

    assert archive.status_code == 303
    assert archive.headers["location"] == "/admin?archived_count=1"
    tickets = rows_for(tmp_path, "tickets")
    assert tickets[0]["archived_at"]
    assert tickets[0]["archived_by"] == ADMIN_AUTH[0]
    assert tickets[0]["archive_reason"] == "batch archive reason"
    assert tickets[1]["archived_at"] is None
    assert "归档工单" in [row["action"] for row in ticket_action_logs(tmp_path, 1)]

    admin_page = client.get("/admin")
    assert "archive selected ticket" not in admin_page.text
    assert "active neighbor ticket" in admin_page.text
    archive_page = client.get("/admin/archive")
    assert archive_page.status_code == 200
    assert ticket_no in archive_page.text
    assert "batch archive reason" in archive_page.text
    query_page = client.get("/query", params={"store_name": store_name, "keyword": "archive selected"})
    assert query_page.status_code == 200
    assert ticket_no in query_page.text
    assert "已归档" in query_page.text

    unarchive = admin_post(
        client,
        "/admin/tickets/bulk-unarchive",
        data={"ticket_ids": ["1"], "select_scope": "selected"},
        follow_redirects=False,
    )

    assert unarchive.status_code == 303
    assert unarchive.headers["location"] == "/admin/archive?unarchived_count=1"
    assert rows_for(tmp_path, "tickets")[0]["archived_at"] is None
    assert ticket_no in client.get("/admin").text
    assert ticket_no not in client.get("/admin/archive").text
    assert "取消归档" in [row["action"] for row in ticket_action_logs(tmp_path, 1)]


def test_bulk_ticket_operations_require_csrf(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    submit_ticket(client, description="bulk csrf ticket")
    logged_in_client(client)

    missing_csrf = client.post(
        "/admin/tickets/bulk-archive",
        data={"ticket_ids": ["1"], "select_scope": "selected"},
        follow_redirects=False,
    )

    assert missing_csrf.status_code == 403


def test_bulk_archive_filtered_scope_and_archive_export(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    submit_ticket(client, description="filtered archive one")
    submit_ticket(client, description="filtered archive two")
    submit_ticket(client, description="regular active ticket")
    logged_in_client(client)

    archive = admin_post(
        client,
        "/admin/tickets/bulk-archive",
        data={
            "select_scope": "filtered",
            "keyword": "filtered archive",
            "archive_reason": "filtered archive reason",
        },
        follow_redirects=False,
    )

    assert archive.status_code == 303
    assert archive.headers["location"] == "/admin?archived_count=2"
    tickets = rows_for(tmp_path, "tickets")
    assert tickets[0]["archived_at"]
    assert tickets[1]["archived_at"]
    assert tickets[2]["archived_at"] is None
    archived_ticket_nos = {tickets[0]["ticket_no"], tickets[1]["ticket_no"]}

    active_export = load_workbook(BytesIO(client.get("/admin/export?keyword=filtered+archive").content))
    assert active_export.active.max_row == 1
    archive_page = client.get("/admin/archive?keyword=filtered+archive")
    assert archived_ticket_nos.issubset(set(re.findall(r"REQ-\d{8}-\d{4}", archive_page.text)))
    archived_export = load_workbook(BytesIO(client.get("/admin/archive/export?keyword=filtered+archive").content))
    values = {cell.value for cell in archived_export.active["K"]}
    assert "filtered archive one" in values
    assert "filtered archive two" in values


def test_bulk_delete_restore_and_hard_delete_preserve_archive_state(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    submit_ticket_with_image_and_file(
        client,
        image_filename="bulk-delete.png",
        file_filename="bulk-delete.xlsx",
        description="bulk delete active ticket",
    )
    submit_ticket(client, description="bulk delete archived ticket")
    active_ticket_no = rows_for(tmp_path, "tickets")[0]["ticket_no"]
    archived_ticket_no = rows_for(tmp_path, "tickets")[1]["ticket_no"]
    image_path = tmp_path / rows_for(tmp_path, "ticket_images")[0]["image_path"]
    file_path = tmp_path / rows_for(tmp_path, "ticket_files")[0]["file_path"]
    logged_in_client(client)

    admin_post(
        client,
        "/admin/tickets/bulk-archive",
        data={"ticket_ids": ["2"], "select_scope": "selected", "archive_reason": "archive before trash"},
        follow_redirects=False,
    )
    delete = admin_post(
        client,
        "/admin/tickets/bulk-delete",
        data={"ticket_ids": ["1", "2"], "select_scope": "selected", "delete_reason": "bulk trash reason"},
        follow_redirects=False,
    )

    assert delete.status_code == 303
    assert delete.headers["location"] == "/admin?deleted_count=2"
    tickets = rows_for(tmp_path, "tickets")
    assert tickets[0]["deleted_at"]
    assert tickets[1]["deleted_at"]
    assert tickets[1]["archived_at"]
    assert active_ticket_no not in client.get("/admin").text
    assert archived_ticket_no not in client.get("/admin/archive").text
    trash = client.get("/admin/trash")
    assert trash.status_code == 200
    assert active_ticket_no in trash.text
    assert archived_ticket_no in trash.text
    assert 'name="ticket_ids"' in trash.text
    assert 'action="/admin/tickets/bulk-restore"' in trash.text
    assert 'formaction="/admin/tickets/bulk-hard-delete"' in trash.text
    assert "建议先执行 backup.sh" in trash.text
    assert "移入回收站" in [row["action"] for row in ticket_action_logs(tmp_path, 1)]

    restore = admin_post(
        client,
        "/admin/tickets/bulk-restore",
        data={"ticket_ids": ["1", "2"], "select_scope": "selected"},
        follow_redirects=False,
    )

    assert restore.status_code == 303
    restored_tickets = rows_for(tmp_path, "tickets")
    assert restored_tickets[0]["deleted_at"] is None
    assert restored_tickets[0]["archived_at"] is None
    assert restored_tickets[1]["deleted_at"] is None
    assert restored_tickets[1]["archived_at"]
    assert active_ticket_no in client.get("/admin").text
    assert archived_ticket_no in client.get("/admin/archive").text
    assert "恢复工单" in [row["action"] for row in ticket_action_logs(tmp_path, 1)]

    admin_post(
        client,
        "/admin/tickets/bulk-delete",
        data={"ticket_ids": ["1"], "select_scope": "selected", "delete_reason": "hard delete setup"},
        follow_redirects=False,
    )
    hard_delete = admin_post(
        client,
        "/admin/tickets/bulk-hard-delete",
        data={"ticket_ids": ["1"], "confirm_delete": "1"},
        follow_redirects=False,
    )

    assert hard_delete.status_code == 303
    assert [row["id"] for row in rows_for(tmp_path, "tickets")] == [2]
    assert rows_for(tmp_path, "ticket_images") == []
    assert rows_for(tmp_path, "ticket_files") == []
    assert not image_path.exists()
    assert not file_path.exists()


def test_schedule_admin_pages_auth_defaults_and_navigation(tmp_path, monkeypatch):
    client, main = build_client(tmp_path, monkeypatch)

    unauth_schedules = TestClient(main.app).get("/admin/schedules", follow_redirects=False)
    assert unauth_schedules.status_code == 303
    assert unauth_schedules.headers["location"].startswith("/admin/login")
    unauth_employees = TestClient(main.app).get("/admin/employees", follow_redirects=False)
    assert unauth_employees.status_code == 303
    assert unauth_employees.headers["location"].startswith("/admin/login")

    shift_types = rows_for(tmp_path, "shift_types")
    assert [row["shift_name"] for row in shift_types] == ["早班", "晚班", "全天", "休息"]
    assert [row["duration_hours"] for row in shift_types] == [8.0, 8.0, 12.5, 0.0]

    logged_in_client(client)
    schedules = client.get("/admin/schedules")
    assert schedules.status_code == 200
    assert "门店排班" in schedules.text
    assert 'href="/admin/schedules"' in schedules.text
    assert 'href="/admin/employees"' in schedules.text
    assert 'href="/admin/shift-types"' in schedules.text
    assert client.get("/admin/employees").status_code == 200
    assert client.get("/admin/shift-types").status_code == 200

    public_query = client.get("/query")
    assert public_query.status_code == 200
    assert 'href="/schedule"' in public_query.text


def create_schedule_employee(client, name, store_name="南京门东店", role="店员", status="在职", store_names=None):
    selected_stores = store_names or [store_name]
    return admin_post(
        client,
        "/admin/employees",
        data={
            "employee_name": name,
            "store_name": store_name,
            "primary_store_name": store_name,
            "store_names": selected_stores,
            "role": role,
            "phone": "",
            "status": status,
        },
        follow_redirects=False,
    )


def test_employee_store_map_filters_schedule_employees_and_allows_cross_store_staff(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)
    create_schedule_employee(client, "跨店员工", store_names=["南京门东店", "南昌万寿宫店"])
    create_schedule_employee(client, "山城员工", store_name="山城巷店")

    employee_store_map = rows_for(tmp_path, "employee_store_map")
    assert [(row["employee_id"], row["store_name"]) for row in employee_store_map] == [
        (1, "南京门东店"),
        (1, "南昌万寿宫店"),
        (2, "山城巷店"),
    ]

    nanjing_page = client.get("/admin/schedules?store_name=南京门东店&month=2026-07")
    assert nanjing_page.status_code == 200
    assert "跨店员工" in nanjing_page.text
    assert "山城员工" not in nanjing_page.text

    nanchang_page = client.get("/admin/schedules?store_name=南昌万寿宫店&month=2026-07")
    assert nanchang_page.status_code == 200
    assert "跨店员工" in nanchang_page.text
    assert "山城员工" not in nanchang_page.text

    shancheng_page = client.get("/admin/schedules?store_name=山城巷店&month=2026-07")
    assert shancheng_page.status_code == 200
    assert "跨店员工" not in shancheng_page.text
    assert "山城员工" in shancheng_page.text

    cross_store_schedule = admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南昌万寿宫店",
            "employee_ids": ["1"],
            "schedule_dates": ["2026-07-06"],
            "shift_type_id": "1",
            "overwrite_existing": "1",
        },
        follow_redirects=False,
    )
    assert cross_store_schedule.status_code == 303
    schedules = rows_for(tmp_path, "store_schedules")
    assert [(row["store_name"], row["employee_id"]) for row in schedules] == [("南昌万寿宫店", 1)]

    unrelated_store_schedule = admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南京门东店",
            "employee_ids": ["2"],
            "schedule_dates": ["2026-07-07"],
            "shift_type_id": "1",
            "overwrite_existing": "1",
        },
        follow_redirects=False,
    )
    assert unrelated_store_schedule.status_code == 303
    assert "员工必须绑定该门店" in client.get(unrelated_store_schedule.headers["location"]).text
    assert len(rows_for(tmp_path, "store_schedules")) == 1


def test_employee_primary_store_bindings_and_archive_trash_workflow(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)

    employee_page = client.get("/admin/employees")
    assert employee_page.status_code == 200
    assert "排班人员视图" in employee_page.text
    assert "人员资料请统一到“人员与账号管理”维护" in employee_page.text
    assert 'href="/admin/account?filter=participate_schedule"' in employee_page.text
    assert 'data-drawer-open="employee-create-drawer"' not in employee_page.text

    create_response = admin_post(
        client,
        "/admin/employees",
        data={
            "employee_name": "主店南京支援南昌",
            "primary_store_name": "南京门东店",
            "store_names": ["南昌万寿宫店"],
            "role": "店员",
            "phone": "13800000000",
            "status": "在职",
        },
        follow_redirects=False,
    )
    assert create_response.status_code == 303
    employee = rows_for(tmp_path, "employees")[0]
    assert employee["store_name"] == "南京门东店"
    assert employee["primary_store_name"] == "南京门东店"
    assert [(row["employee_id"], row["store_name"]) for row in rows_for(tmp_path, "employee_store_map")] == [
        (1, "南京门东店"),
        (1, "南昌万寿宫店"),
    ]

    archive_response = admin_post(client, "/admin/employees/1/archive", data={"archive_reason": "历史员工"}, follow_redirects=False)
    assert archive_response.status_code == 303
    archived_employee = rows_for(tmp_path, "employees")[0]
    assert archived_employee["archived_at"]
    assert archived_employee["archived_by"] == ADMIN_AUTH[0]
    assert archived_employee["archive_reason"] == "历史员工"
    assert "主店南京支援南昌" not in client.get("/admin/employees").text
    assert "主店南京支援南昌" not in client.get("/admin/schedules?store_name=南京门东店&month=2026-07").text
    assert "主店南京支援南昌" in client.get("/admin/employees?scope=archive").text

    unarchive_response = admin_post(client, "/admin/employees/1/unarchive", follow_redirects=False)
    assert unarchive_response.status_code == 303
    assert rows_for(tmp_path, "employees")[0]["archived_at"] is None
    assert "主店南京支援南昌" in client.get("/admin/employees").text

    delete_response = admin_post(client, "/admin/employees/1/delete", data={"delete_reason": "测试清理"}, follow_redirects=False)
    assert delete_response.status_code == 303
    deleted_employee = rows_for(tmp_path, "employees")[0]
    assert deleted_employee["deleted_at"]
    assert deleted_employee["deleted_by"] == ADMIN_AUTH[0]
    assert "主店南京支援南昌" not in client.get("/admin/employees").text
    assert "主店南京支援南昌" in client.get("/admin/employees?scope=trash").text

    restore_response = admin_post(client, "/admin/employees/1/restore", follow_redirects=False)
    assert restore_response.status_code == 303
    assert rows_for(tmp_path, "employees")[0]["deleted_at"] is None

    admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南京门东店",
            "employee_ids": ["1"],
            "schedule_dates": ["2026-07-06"],
            "shift_type_id": "1",
            "overwrite_existing": "1",
        },
        follow_redirects=False,
    )
    admin_post(client, "/admin/employees/1/delete", follow_redirects=False)
    hard_delete_blocked = admin_post(
        client,
        "/admin/employees/1/hard-delete",
        data={"confirm_delete": "1"},
        follow_redirects=False,
    )
    assert hard_delete_blocked.status_code == 303
    assert "该员工已有排班记录，建议归档而不是永久删除" in client.get(hard_delete_blocked.headers["location"]).text
    assert rows_for(tmp_path, "employees")


def test_schedule_page_exposes_bulk_employee_and_date_controls(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)
    create_schedule_employee(client, "王早班", role="店员")
    create_schedule_employee(client, "张晚班", role="值班员")

    page = client.get("/admin/schedules?store_name=南京门东店&month=2026-07")

    assert page.status_code == 200
    assert "schedule-bulk-form" in page.text
    assert 'name="employee_ids"' in page.text
    assert 'name="schedule_dates"' in page.text
    assert 'value="2026-07-01"' in page.text
    assert 'value="2026-07-31"' in page.text
    assert "data-select-all-employees" in page.text
    assert "data-clear-employees" in page.text
    assert "data-select-all-schedule-dates" in page.text
    assert "data-select-weekdays" in page.text
    assert "data-select-weekends" in page.text
    assert "data-clear-schedule-dates" in page.text
    assert "schedule-bulk-summary" in page.text
    assert "王早班" in page.text
    assert "值班员" in page.text


def test_schedule_page_has_saas_layers_and_three_views(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)
    store_name = DEFAULT_TEST_CONFIG["stores.json"][0]
    create_schedule_employee(client, "EmpMorning", store_name=store_name, role="Crew")
    create_schedule_employee(client, "EmpEvening", store_name=store_name, role="Lead")
    admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": store_name,
            "employee_ids": ["1", "2"],
            "schedule_dates": ["2026-07-06", "2026-07-07"],
            "shift_type_id": "1",
            "note": "ops-view",
            "overwrite_existing": "1",
        },
        follow_redirects=False,
    )

    calendar_page = client.get(
        "/admin/schedules",
        params={"store_name": store_name, "month": "2026-07", "view": "calendar"},
    )

    assert calendar_page.status_code == 200
    assert "schedule-workbench" in calendar_page.text
    assert "schedule-action-layer" in calendar_page.text
    assert "schedule-filter-layer" in calendar_page.text
    assert "schedule-data-layer" in calendar_page.text
    assert "schedule-view-switch" in calendar_page.text
    assert 'data-schedule-view="calendar"' in calendar_page.text
    assert "view_mode=table" in calendar_page.text
    assert "view_mode=calendar" in calendar_page.text
    assert "view_mode=employee" in calendar_page.text
    assert "view_mode=store-summary" in calendar_page.text
    assert "schedule-calendar-grid" in calendar_page.text
    assert "schedule-calendar-day" in calendar_page.text
    assert "schedule-calendar-shift" in calendar_page.text
    assert "data-schedule-edit" in calendar_page.text
    assert re.search(r"<button\b[^>]*\btype=(['\"])button\1[^>]*\bdata-schedule-edit\b", calendar_page.text)

    employee_page = client.get(
        "/admin/schedules",
        params={"store_name": store_name, "month": "2026-07", "view": "employee"},
    )

    assert employee_page.status_code == 200
    assert 'data-schedule-view="employee"' in employee_page.text
    assert "schedule-employee-view" in employee_page.text
    assert "schedule-employee-card" in employee_page.text
    assert "employee-timeline" in employee_page.text
    assert "metric-hours" in employee_page.text
    assert "metric-days" in employee_page.text


def test_schedule_cross_store_toggle_tags_external_staff(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)
    current_store = DEFAULT_TEST_CONFIG["stores.json"][0]
    other_store = DEFAULT_TEST_CONFIG["stores.json"][2]
    create_schedule_employee(client, "LocalEmp", store_name=current_store)
    create_schedule_employee(client, "OtherStoreEmp", store_name=other_store)
    create_schedule_employee(client, "SharedEmp", store_name=DEFAULT_TEST_CONFIG["stores.json"][1], store_names=[current_store, DEFAULT_TEST_CONFIG["stores.json"][1]])

    default_page = client.get("/admin/schedules", params={"store_name": current_store, "month": "2026-07"})
    primary_page = client.get("/admin/schedules", params={"store_name": current_store, "month": "2026-07", "employee_scope": "primary"})
    support_page = client.get("/admin/schedules", params={"store_name": current_store, "month": "2026-07", "employee_scope": "support"})

    assert default_page.status_code == 200
    assert "LocalEmp" in default_page.text
    assert "SharedEmp" in default_page.text
    assert "OtherStoreEmp" not in default_page.text
    assert primary_page.status_code == 200
    assert "LocalEmp" in primary_page.text
    assert "SharedEmp" not in primary_page.text
    assert support_page.status_code == 200
    assert "SharedEmp" in support_page.text
    assert "LocalEmp" not in support_page.text
    assert "OtherStoreEmp" not in support_page.text
    assert "store-tag-primary" in support_page.text
    assert "store-tag-available" in support_page.text
    assert "cross-store-readonly" not in support_page.text


def test_schedule_workbench_supports_view_mode_employee_filter_and_store_summary(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)
    create_schedule_employee(client, "南京主店", store_name="南京门东店")
    create_schedule_employee(client, "南昌支援南京", store_name="南昌万寿宫店", store_names=["南昌万寿宫店", "南京门东店"])
    create_schedule_employee(client, "山城员工", store_name="山城巷店")
    admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南京门东店",
            "employee_ids": ["1", "2"],
            "schedule_dates": ["2026-07-06", "2026-07-07"],
            "shift_type_id": "1",
            "overwrite_existing": "1",
        },
        follow_redirects=False,
    )
    admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "山城巷店",
            "employee_ids": ["3"],
            "schedule_dates": ["2026-07-06"],
            "shift_type_id": "2",
            "overwrite_existing": "1",
        },
        follow_redirects=False,
    )

    calendar_page = client.get("/admin/schedules?store_name=南京门东店&month=2026-07&view_mode=calendar")
    assert calendar_page.status_code == 200
    assert 'data-schedule-view="calendar"' in calendar_page.text
    assert "schedule-calendar-summary" in calendar_page.text
    assert "早班 2 人" in calendar_page.text
    assert "南京主店" in calendar_page.text
    assert "山城员工" not in calendar_page.text

    employee_page = client.get("/admin/schedules?store_name=南京门东店&month=2026-07&view_mode=employee&employee_ids=2")
    assert employee_page.status_code == 200
    assert 'data-schedule-view="employee"' in employee_page.text
    assert 'name="employee_ids"' in employee_page.text
    assert "南昌支援南京" in employee_page.text
    assert employee_page.text.count('<article class="schedule-employee-card') == 1
    assert "休息天数" in employee_page.text

    store_summary_page = client.get("/admin/schedules?store_name=&month=2026-07&view_mode=store-summary")
    assert store_summary_page.status_code == 200
    assert 'data-schedule-view="store-summary"' in store_summary_page.text
    assert "store-summary-calendar" in store_summary_page.text
    assert "请选择具体门店后进行排班操作" in store_summary_page.text
    assert "南京门东店 2 人" in store_summary_page.text
    assert "山城巷店 1 人" in store_summary_page.text

    all_store_post = admin_post(
        client,
        "/admin/schedules",
        data={"store_name": "", "employee_ids": ["1"], "schedule_dates": ["2026-07-08"], "shift_type_id": "1"},
        follow_redirects=False,
    )
    assert all_store_post.status_code == 303
    assert "请选择具体门店后进行排班操作" in client.get(all_store_post.headers["location"]).text

    all_store_export = client.get("/admin/schedules/export?store_name=&month=2026-07")
    assert all_store_export.status_code == 200
    assert "%E5%85%A8%E9%83%A8%E9%97%A8%E5%BA%97" in all_store_export.headers["content-disposition"]
    workbook = load_workbook(BytesIO(all_store_export.content))
    rows = list(workbook.active.iter_rows(values_only=True))
    row_prefixes = {row[:8] for row in rows}
    assert ("南京门东店", "南京主店", "2026-07-06", "周一", "早班", "09:30-17:30", "8 小时", "否") in row_prefixes
    assert ("山城巷店", "山城员工", "2026-07-06", "周一", "晚班", "14:00-22:00", "8 小时", "否") in row_prefixes


def test_schedule_filters_ignore_empty_numeric_query_values_without_json(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)
    create_schedule_employee(client, "FilterSafeEmp")

    empty_shift = client.get(
        "/admin/schedules?store_name=南京门东店&month=2026-07&view_mode=calendar&shift_type_id="
    )
    assert empty_shift.status_code == 200
    assert "text/html" in empty_shift.headers["content-type"]
    assert '{"detail":' not in empty_shift.text
    assert 'data-schedule-view="calendar"' in empty_shift.text

    invalid_shift = client.get(
        "/admin/schedules?store_name=南京门东店&month=2026-07&view_mode=calendar&shift_type_id=abc"
    )
    assert invalid_shift.status_code == 200
    assert "text/html" in invalid_shift.headers["content-type"]
    assert "班次筛选参数无效" in invalid_shift.text
    assert "alert-error" in invalid_shift.text
    assert '{"detail":' not in invalid_shift.text

    invalid_employee = client.get(
        "/admin/schedules?store_name=南京门东店&month=2026-07&view_mode=employee&employee_ids=abc"
    )
    assert invalid_employee.status_code == 200
    assert "员工筛选参数无效" in invalid_employee.text
    assert '{"detail":' not in invalid_employee.text

    empty_shift_post = admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南京门东店",
            "employee_ids": ["1"],
            "schedule_dates": ["2026-07-06"],
            "shift_type_id": "",
        },
        follow_redirects=False,
    )
    assert empty_shift_post.status_code == 303
    empty_shift_post_page = client.get(empty_shift_post.headers["location"])
    assert "请选择班次" in empty_shift_post_page.text
    assert '{"detail":' not in empty_shift_post_page.text


def test_schedule_layout_hash_and_preserve_scroll_contract(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)
    create_schedule_employee(client, "LayoutEmp")

    page = client.get("/admin/schedules?store_name=南京门东店&month=2026-07&view_mode=calendar")

    assert page.status_code == 200
    for expected in (
        "schedule-overview-layer",
        "schedule-filter-layer",
        "schedule-operation-layer",
        "schedule-view-layer",
        "schedule-hero-card",
        "schedule-filter-card",
        "schedule-bulk-card",
        "schedule-view-card",
        'id="schedule-filters"',
        'id="bulk-schedule"',
        'id="schedule-view"',
        'id="schedule-calendar-view"',
        "schedule-filter-grid",
        "schedule-filter-actions",
        "data-preserve-scroll",
    ):
        assert expected in page.text
    assert "运营排班工作台" in page.text
    assert "筛选排班" in page.text
    assert "批量排班" in page.text
    assert "排班查看" in page.text
    assert "已选 1 个门店" in page.text
    assert "已选 0 名员工" in page.text
    assert "已选 0 个班次" in page.text

    filter_form_start = page.text.index('<form class="schedule-filter-grid')
    filter_form_end = page.text.index("</form>", filter_form_start)
    filter_form = page.text[filter_form_start:filter_form_end]
    assert 'method="get"' in filter_form
    assert 'type="submit">查看排班' in filter_form
    assert 'href="/admin/schedules#schedule-view"' in filter_form

    assert "view_mode=calendar#schedule-calendar-view" in page.text
    assert "view_mode=employee#employee-schedule-view" in page.text
    assert "view_mode=table#schedule-table-view" in page.text
    assert "view_mode=store-summary#store-summary-view" in page.text

    table_page = client.get("/admin/schedules?store_name=南京门东店&month=2026-07&view_mode=table")
    assert table_page.status_code == 200
    assert "精细查看可横向滚动，推荐日常使用日历视图或员工视图。" in table_page.text
    assert "schedule-table-scroll" in table_page.text

    employee_page = client.get("/admin/schedules?store_name=南京门东店&month=2026-07&view_mode=employee")
    assert employee_page.status_code == 200
    assert "员工筛选" in employee_page.text
    assert 'name="employee_ids"' in employee_page.text
    assert 'id="employee-schedule-view"' in employee_page.text

    multi_store_page = client.get(
        "/admin/schedules",
        params=[("store_names", "南京门东店"), ("store_names", "南昌万寿宫店"), ("month", "2026-07")],
    )
    assert multi_store_page.status_code == 200
    assert "当前为多门店查看模式，请选择单个门店后进行排班操作。" in multi_store_page.text
    assert "schedule-bulk-form" not in multi_store_page.text

    app_js = (PROJECT_DIR / "static" / "app.js").read_text(encoding="utf-8")
    assert "data-preserve-scroll" in app_js
    assert "sessionStorage" in app_js
    assert "window.location.hash" in app_js


def test_bulk_schedule_creates_updates_and_skips_existing_rows(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)
    create_schedule_employee(client, "王早班")
    create_schedule_employee(client, "张晚班")

    create_bulk = admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南京门东店",
            "employee_ids": ["1", "2"],
            "schedule_dates": ["2026-07-06", "2026-07-07", "2026-07-08"],
            "shift_type_id": "1",
            "note": "批量早班",
            "overwrite_existing": "1",
        },
        follow_redirects=False,
    )

    assert create_bulk.status_code == 303
    assert "store_name=%E5%8D%97%E4%BA%AC%E9%97%A8%E4%B8%9C%E5%BA%97" in create_bulk.headers["location"]
    assert "month=2026-07" in create_bulk.headers["location"]
    assert "saved_count=6" in create_bulk.headers["location"]
    schedules = sorted(rows_for(tmp_path, "store_schedules"), key=lambda row: (row["employee_id"], row["schedule_date"]))
    assert len(schedules) == 6
    assert {row["shift_type_id"] for row in schedules} == {1}
    assert {row["note"] for row in schedules} == {"批量早班"}
    success_page = client.get(create_bulk.headers["location"])
    assert "已保存 6 条排班" in success_page.text
    assert "新增 6 条" in success_page.text

    update_bulk = admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南京门东店",
            "employee_ids": ["1", "2"],
            "schedule_dates": ["2026-07-06", "2026-07-07", "2026-07-08"],
            "shift_type_id": "2",
            "note": "批量晚班",
            "overwrite_existing": "1",
        },
        follow_redirects=False,
    )

    assert update_bulk.status_code == 303
    assert "updated_count=6" in update_bulk.headers["location"]
    updated = rows_for(tmp_path, "store_schedules")
    assert len(updated) == 6
    assert {row["shift_type_id"] for row in updated} == {2}
    assert {row["note"] for row in updated} == {"批量晚班"}

    skip_bulk = admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南京门东店",
            "employee_ids": ["1", "2"],
            "schedule_dates": ["2026-07-06", "2026-07-07", "2026-07-08"],
            "shift_type_id": "3",
            "note": "不覆盖",
        },
        follow_redirects=False,
    )

    assert skip_bulk.status_code == 303
    assert "skipped_count=6" in skip_bulk.headers["location"]
    skipped = rows_for(tmp_path, "store_schedules")
    assert len(skipped) == 6
    assert {row["shift_type_id"] for row in skipped} == {2}
    assert "跳过 6 条" in client.get(skip_bulk.headers["location"]).text


def test_schedule_quick_copy_previous_day_and_clear_employee_month(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)
    store_name = DEFAULT_TEST_CONFIG["stores.json"][0]
    create_schedule_employee(client, "QuickA", store_name=store_name)
    create_schedule_employee(client, "QuickB", store_name=store_name)
    admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": store_name,
            "employee_ids": ["1", "2"],
            "schedule_dates": ["2026-07-06"],
            "shift_type_id": "1",
            "note": "copy-source",
            "overwrite_existing": "1",
        },
        follow_redirects=False,
    )

    copy_response = admin_post(
        client,
        "/admin/schedules/copy-day",
        data={"store_name": store_name, "target_date": "2026-07-07"},
        follow_redirects=False,
    )

    assert copy_response.status_code == 303
    assert "month=2026-07" in copy_response.headers["location"]
    copied_rows = sorted(rows_for(tmp_path, "store_schedules"), key=lambda row: (row["employee_id"], row["schedule_date"]))
    assert [(row["employee_id"], row["schedule_date"]) for row in copied_rows] == [
        (1, "2026-07-06"),
        (1, "2026-07-07"),
        (2, "2026-07-06"),
        (2, "2026-07-07"),
    ]
    assert all(row["note"] == "copy-source" for row in copied_rows)

    clear_response = admin_post(
        client,
        "/admin/schedules/clear-employee",
        data={"store_name": store_name, "employee_id": "1", "month": "2026-07"},
        follow_redirects=False,
    )

    assert clear_response.status_code == 303
    remaining_rows = rows_for(tmp_path, "store_schedules")
    assert {(row["employee_id"], row["schedule_date"]) for row in remaining_rows} == {(2, "2026-07-06"), (2, "2026-07-07")}
    assert any(row["action"] == "删除排班" for row in rows_for(tmp_path, "schedule_logs"))


def test_bulk_schedule_validation_errors_return_schedule_page(tmp_path, monkeypatch):
    client, _ = build_client(
        tmp_path,
        monkeypatch,
        config_overrides={"system.json": dict(DEFAULT_TEST_CONFIG["system.json"], max_bulk_schedule_count=2)},
    )
    logged_in_client(client)
    create_schedule_employee(client, "王早班")
    create_schedule_employee(client, "张晚班")
    create_schedule_employee(client, "跨店员工", store_name="南昌万寿宫店")

    missing_employees = admin_post(
        client,
        "/admin/schedules",
        data={"store_name": "南京门东店", "schedule_dates": ["2026-07-06"], "shift_type_id": "1"},
        follow_redirects=False,
    )
    assert missing_employees.status_code == 303
    missing_employees_page = client.get(missing_employees.headers["location"])
    assert "请选择至少 1 名员工" in missing_employees_page.text
    assert "alert-error" in missing_employees_page.text
    assert '{"detail":' not in missing_employees_page.text

    missing_dates = admin_post(
        client,
        "/admin/schedules",
        data={"store_name": "南京门东店", "employee_ids": ["1"], "shift_type_id": "1"},
        follow_redirects=False,
    )
    assert missing_dates.status_code == 303
    assert "请选择至少 1 个日期" in client.get(missing_dates.headers["location"]).text

    too_many = admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南京门东店",
            "employee_ids": ["1", "2"],
            "schedule_dates": ["2026-07-06", "2026-07-07"],
            "shift_type_id": "1",
            "overwrite_existing": "1",
        },
        follow_redirects=False,
    )
    assert too_many.status_code == 303
    assert "一次最多批量生成 2 条排班，请减少员工或日期数量。" in client.get(too_many.headers["location"]).text

    wrong_store = admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南京门东店",
            "employee_ids": ["3"],
            "schedule_dates": ["2026-07-06"],
            "shift_type_id": "1",
        },
        follow_redirects=False,
    )
    assert wrong_store.status_code == 303
    assert "员工必须绑定该门店" in client.get(wrong_store.headers["location"]).text


def test_bulk_schedule_security_keeps_public_schedule_read_only(tmp_path, monkeypatch):
    client, main = build_client(tmp_path, monkeypatch)
    logged_in_client(client)
    create_schedule_employee(client, "王早班")

    missing_csrf = client.post(
        "/admin/schedules",
        data={
            "store_name": "南京门东店",
            "employee_ids": ["1"],
            "schedule_dates": ["2026-07-06"],
            "shift_type_id": "1",
        },
        follow_redirects=False,
    )
    assert missing_csrf.status_code == 403

    unauthenticated = TestClient(main.app).post(
        "/admin/schedules",
        data={
            "store_name": "南京门东店",
            "employee_ids": ["1"],
            "schedule_dates": ["2026-07-06"],
            "shift_type_id": "1",
        },
        follow_redirects=False,
    )
    assert unauthenticated.status_code == 303
    assert unauthenticated.headers["location"].startswith("/admin/login")

    public_post = client.post(
        "/schedule",
        data={
            "store_name": "南京门东店",
            "employee_ids": ["1"],
            "schedule_dates": ["2026-07-06"],
            "shift_type_id": "1",
        },
        follow_redirects=False,
    )
    assert public_post.status_code == 405
    assert rows_for(tmp_path, "store_schedules") == []


def test_employee_and_shift_type_management_lifecycle(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)

    create_employee = admin_post(
        client,
        "/admin/employees",
        data={
            "employee_name": "张小排",
            "store_name": "南京门东店",
            "role": "店员",
            "phone": "13800000000",
            "status": "在职",
        },
        follow_redirects=False,
    )
    assert create_employee.status_code == 303
    employees = rows_for(tmp_path, "employees")
    assert len(employees) == 1
    assert employees[0]["employee_name"] == "张小排"
    assert employees[0]["store_name"] == "南京门东店"
    assert employees[0]["status"] == "在职"
    assert [(row["employee_id"], row["store_name"]) for row in rows_for(tmp_path, "employee_store_map")] == [(1, "南京门东店")]

    employee_page = client.get("/admin/employees")
    assert employee_page.status_code == 200
    assert 'name="store_names"' in employee_page.text
    assert "绑定门店" in employee_page.text
    assert "主门店：南京门东店" in employee_page.text
    assert "暂无跨店绑定" in employee_page.text

    update_employee = admin_post(
        client,
        "/admin/employees/1/update",
        data={
            "employee_name": "张小排",
            "primary_store_name": "南京门东店",
            "store_names": ["南京门东店", "南昌万寿宫店"],
            "role": "值班员",
            "phone": "13900000000",
            "status": "在职",
        },
        follow_redirects=False,
    )
    assert update_employee.status_code == 303
    assert rows_for(tmp_path, "employees")[0]["role"] == "值班员"
    assert [(row["employee_id"], row["store_name"]) for row in rows_for(tmp_path, "employee_store_map")] == [
        (1, "南京门东店"),
        (1, "南昌万寿宫店"),
    ]
    updated_employee_page = client.get("/admin/employees")
    assert updated_employee_page.text.count("主门店：南京门东店") == 1
    assert updated_employee_page.text.count("store-tag-available\">南京门东店") == 0
    assert updated_employee_page.text.count("store-tag-available\">南昌万寿宫店") == 1

    switch_primary = admin_post(
        client,
        "/admin/employees/1/update",
        data={
            "employee_name": "张小排",
            "primary_store_name": "山城巷店",
            "store_names": ["南昌万寿宫店"],
            "role": "值班员",
            "phone": "13900000000",
            "status": "在职",
        },
        follow_redirects=False,
    )
    assert switch_primary.status_code == 303
    assert {
        (row["employee_id"], row["store_name"]) for row in rows_for(tmp_path, "employee_store_map")
    } == {
        (1, "南昌万寿宫店"),
        (1, "山城巷店"),
    }
    switched_employee_page = client.get("/admin/employees")
    assert "主门店：山城巷店" in switched_employee_page.text
    assert "store-tag-available\">山城巷店" not in switched_employee_page.text
    assert "store-tag-available\">南昌万寿宫店" in switched_employee_page.text

    disable_employee = admin_post(client, "/admin/employees/1/disable", follow_redirects=False)
    assert disable_employee.status_code == 303
    assert rows_for(tmp_path, "employees")[0]["status"] == "离职"

    create_shift = admin_post(
        client,
        "/admin/shift-types",
        data={
            "shift_scope": "global",
            "shift_name": "中班",
            "start_time": "12:00",
            "end_time": "20:00",
            "duration_hours": "8",
            "color": "#7c3aed",
            "is_global": "1",
        },
        follow_redirects=False,
    )
    assert create_shift.status_code == 303
    assert any(row["shift_name"] == "中班" for row in rows_for(tmp_path, "shift_types"))

    update_shift = admin_post(
        client,
        "/admin/shift-types/5/update",
        data={
            "shift_name": "中班",
            "start_time": "12:30",
            "end_time": "20:30",
            "duration_hours": "8",
            "color": "#2563eb",
            "is_active": "1",
            "is_global": "1",
        },
        follow_redirects=False,
    )
    assert update_shift.status_code == 303
    shift = rows_for(tmp_path, "shift_types")[4]
    assert shift["start_time"] == "12:30"
    assert shift["is_active"] == 1

    disable_shift = admin_post(client, "/admin/shift-types/5/disable", follow_redirects=False)
    assert disable_shift.status_code == 303
    assert rows_for(tmp_path, "shift_types")[4]["is_active"] == 0


def test_employee_management_uses_cards_drawer_and_store_tags(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)
    create_schedule_employee(
        client,
        "CardEmp",
        role="Lead",
        store_names=[DEFAULT_TEST_CONFIG["stores.json"][0], DEFAULT_TEST_CONFIG["stores.json"][1]],
    )

    page = client.get("/admin/employees")

    assert page.status_code == 200
    assert "employee-card-grid" in page.text
    assert "employee-card" in page.text
    assert "employee-store-tags" in page.text
    assert "employee-status-pill" in page.text
    assert "employee-drawer" in page.text
    assert "drawer-panel" in page.text
    assert "data-drawer-open" in page.text
    assert "data-drawer-close" in page.text
    assert "data-feedback-form" in page.text
    assert re.search(r"<button\b[^>]*\btype=(['\"])button\1[^>]*\bdata-drawer-open\b", page.text)
    assert "CardEmp" in page.text
    assert "Lead" in page.text
    assert "data-employee-store-form" in page.text
    assert "data-primary-store-select" in page.text
    assert "data-primary-store-note" in page.text


def test_employee_cards_show_saved_secondary_stores_role_groups_and_schedule_links(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)

    create_schedule_employee(
        client,
        "南京店长",
        role="店长",
        store_name="南京门东店",
        store_names=["南京门东店", "南昌万寿宫店"],
    )
    create_schedule_employee(client, "山城店员", role="店员", store_name="山城巷店")
    create_schedule_employee(client, "周末兼职", role="兼职", store_name="南昌万寿宫店")
    create_schedule_employee(client, "区域经理A", role="区域经理", store_name="南京门东店")
    create_schedule_employee(client, "未设角色", role="", store_name="南京门东店")
    create_schedule_employee(client, "训练师", role="训练师", store_name="山城巷店")

    page = client.get("/admin/employees")

    assert page.status_code == 200
    assert "employee-role-group" in page.text
    assert "店长 / 经理" in page.text
    assert "店员" in page.text
    assert "兼职" in page.text
    assert "区域经理" in page.text
    assert "未设置角色" in page.text
    assert "其他角色" in page.text
    assert page.text.index("南京店长") < page.text.index("山城店员")
    assert "可排门店：南昌万寿宫店" in page.text
    assert "store-tag-available\">南京门东店" not in page.text
    assert 'class="ghost-button compact" href="/admin/schedules?view_mode=employee&amp;employee_ids=1&amp;store_scope=employee-all-stores#employee-schedule-view"' in page.text
    assert ">查看排班</a>" in page.text

    filtered = client.get("/admin/employees?store_name=南昌万寿宫店")
    assert filtered.status_code == 200
    assert "南京店长" in filtered.text
    assert "可排门店：南昌万寿宫店" in filtered.text
    assert "employee-role-group" in filtered.text


def test_shift_type_page_uses_card_timeline_structure(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)

    page = client.get("/admin/shift-types")

    assert page.status_code == 200
    assert "shift-type-grid" in page.text
    assert "shift-type-card" in page.text
    assert "shift-time-axis" in page.text
    assert "shift-color-dot" in page.text
    assert "data-draggable-shift" in page.text
    assert "data-feedback-form" in page.text
    assert "shift-status-pill" in page.text
    assert 'name="shift_scope" value="store"' in page.text
    assert 'name="shift_scope" value="global"' in page.text
    assert "门店班次" in page.text
    assert "通用班次" in page.text


def test_shift_type_archive_delete_restore_and_hard_delete_rules(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)

    create_response = admin_post(
        client,
        "/admin/shift-types",
        data={
            "shift_scope": "store",
            "store_name": "南京门东店",
            "shift_name": "测试可删除班次",
            "start_time": "11:00",
            "end_time": "15:00",
            "duration_hours": "4",
            "color": "#2563eb",
        },
        follow_redirects=False,
    )
    assert create_response.status_code == 303
    shift_id = rows_for(tmp_path, "shift_types")[-1]["id"]

    page = client.get("/admin/shift-types")
    assert page.status_code == 200
    for action in ("archive", "delete"):
        assert f'action="/admin/shift-types/{shift_id}/{action}"' in page.text
    assert "数据范围" in page.text
    assert "归档班次" in page.text
    assert "回收站班次" in page.text

    archive_response = admin_post(
        client,
        f"/admin/shift-types/{shift_id}/archive",
        data={"archive_reason": "测试归档"},
        follow_redirects=False,
    )
    assert archive_response.status_code == 303
    archived_shift = [row for row in rows_for(tmp_path, "shift_types") if row["id"] == shift_id][0]
    assert archived_shift["archived_at"]
    assert archived_shift["archived_by"] == ADMIN_AUTH[0]
    assert "测试可删除班次" not in client.get("/admin/shift-types").text
    assert "测试可删除班次" in client.get("/admin/shift-types?data_scope=archive").text

    restore_archived = admin_post(client, f"/admin/shift-types/{shift_id}/restore", follow_redirects=False)
    assert restore_archived.status_code == 303
    assert [row for row in rows_for(tmp_path, "shift_types") if row["id"] == shift_id][0]["archived_at"] is None

    delete_response = admin_post(
        client,
        f"/admin/shift-types/{shift_id}/delete",
        data={"delete_reason": "测试删除"},
        follow_redirects=False,
    )
    assert delete_response.status_code == 303
    deleted_shift = [row for row in rows_for(tmp_path, "shift_types") if row["id"] == shift_id][0]
    assert deleted_shift["deleted_at"]
    assert deleted_shift["deleted_by"] == ADMIN_AUTH[0]
    assert "测试可删除班次" not in client.get("/admin/shift-types").text
    trash_page = client.get("/admin/shift-types?data_scope=trash")
    assert "测试可删除班次" in trash_page.text
    assert f'action="/admin/shift-types/{shift_id}/restore"' in trash_page.text
    assert f'action="/admin/shift-types/{shift_id}/hard-delete"' in trash_page.text

    restore_deleted = admin_post(client, f"/admin/shift-types/{shift_id}/restore", follow_redirects=False)
    assert restore_deleted.status_code == 303
    assert [row for row in rows_for(tmp_path, "shift_types") if row["id"] == shift_id][0]["deleted_at"] is None

    create_schedule_employee(client, "历史排班员工", store_name="南京门东店")
    admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南京门东店",
            "employee_ids": ["1"],
            "schedule_dates": ["2026-07-06"],
            "shift_type_id": str(shift_id),
            "overwrite_existing": "1",
        },
        follow_redirects=False,
    )
    admin_post(client, f"/admin/shift-types/{shift_id}/delete", follow_redirects=False)
    hard_delete_blocked = admin_post(
        client,
        f"/admin/shift-types/{shift_id}/hard-delete",
        data={"confirm_delete": "1"},
        follow_redirects=False,
    )
    assert hard_delete_blocked.status_code == 303
    blocked_page = client.get(hard_delete_blocked.headers["location"])
    assert "该班次已有排班记录，建议停用或归档，不允许永久删除。" in blocked_page.text
    assert [row for row in rows_for(tmp_path, "shift_types") if row["id"] == shift_id]


def test_store_specific_shift_types_business_hours_and_schedule_visibility(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)

    default_shifts = rows_for(tmp_path, "shift_types")
    assert default_shifts
    assert all(row["is_global"] == 1 for row in default_shifts)
    assert all(row["store_name"] in ("", None) for row in default_shifts)

    save_hours = admin_post(
        client,
        "/admin/shift-types/business-hours",
        data={
            "store_name": "南京门东店",
            "business_start_time": "09:30",
            "business_end_time": "22:00",
        },
        follow_redirects=False,
    )
    assert save_hours.status_code == 303
    assert rows_for(tmp_path, "store_business_hours")[0]["business_start_time"] == "09:30"

    first_store_shift = admin_post(
        client,
        "/admin/shift-types",
        data={
            "shift_scope": "store",
            "store_name": "南京门东店",
            "shift_name": "早班",
            "start_time": "09:30",
            "end_time": "17:30",
            "duration_hours": "8",
            "color": "#2563eb",
        },
        follow_redirects=False,
    )
    second_store_shift = admin_post(
        client,
        "/admin/shift-types",
        data={
            "shift_scope": "store",
            "store_name": "南昌万寿宫店",
            "shift_name": "早班",
            "start_time": "10:00",
            "end_time": "18:00",
            "duration_hours": "8",
            "color": "#059669",
        },
        follow_redirects=False,
    )

    assert first_store_shift.status_code == 303
    assert second_store_shift.status_code == 303
    shift_rows = rows_for(tmp_path, "shift_types")
    store_named_morning = [row for row in shift_rows if row["shift_name"] == "早班" and row["is_global"] == 0]
    assert {row["store_name"] for row in store_named_morning} == {"南京门东店", "南昌万寿宫店"}

    shift_page = client.get("/admin/shift-types?store_names=南京门东店&global_scope=store&status=active")
    assert shift_page.status_code == 200
    assert 'name="store_names"' in shift_page.text
    assert "store-business-hours-card" in shift_page.text
    assert "09:30" in shift_page.text
    assert "当前门店未配置营业时间，请先在班次设置中维护。" not in shift_page.text
    assert "全部门店可用" not in shift_page.text

    store_shift_id = store_named_morning[0]["id"]
    update_store_shift = admin_post(
        client,
        f"/admin/shift-types/{store_shift_id}/update?store_names=南京门东店&status=active",
        data={
            "shift_scope": "store",
            "store_name": "南京门东店",
            "shift_name": "早班",
            "start_time": "09:45",
            "end_time": "17:45",
            "duration_hours": "8",
            "color": "#2563eb",
            "is_active": "1",
        },
        follow_redirects=False,
    )
    assert update_store_shift.status_code == 303
    assert unquote(update_store_shift.headers["location"]).startswith("/admin/shift-types?store_names=南京门东店&status=active")
    saved_store_shift = [row for row in rows_for(tmp_path, "shift_types") if row["id"] == store_shift_id][0]
    assert saved_store_shift["is_global"] == 0
    assert saved_store_shift["store_name"] == "南京门东店"
    assert saved_store_shift["start_time"] == "09:45"

    global_shift = admin_post(
        client,
        "/admin/shift-types?store_names=南京门东店",
        data={
            "shift_scope": "global",
            "store_name": "南京门东店",
            "shift_name": "门店共享午休",
            "start_time": "13:00",
            "end_time": "14:00",
            "duration_hours": "1",
            "color": "#64748b",
        },
        follow_redirects=False,
    )
    assert global_shift.status_code == 303
    saved_global_shift = rows_for(tmp_path, "shift_types")[-1]
    assert saved_global_shift["is_global"] == 1
    assert saved_global_shift["store_name"] in ("", None)

    missing_store = admin_post(
        client,
        "/admin/shift-types",
        data={
            "shift_scope": "store",
            "shift_name": "缺门店班次",
            "start_time": "09:00",
            "end_time": "10:00",
            "duration_hours": "1",
            "color": "#2563eb",
        },
        follow_redirects=False,
    )
    assert missing_store.status_code == 303
    missing_store_page = client.get(missing_store.headers["location"])
    assert "请选择班次所属门店" in missing_store_page.text
    assert "alert-error" in missing_store_page.text
    assert '{"detail":' not in missing_store_page.text

    create_schedule_employee(client, "南京排班员工", store_name="南京门东店")
    create_schedule_employee(client, "南昌排班员工", store_name="南昌万寿宫店")

    nanjing_page = client.get("/admin/schedules?store_names=南京门东店&month=2026-07")
    nanchang_page = client.get("/admin/schedules?store_names=南昌万寿宫店&month=2026-07")
    assert nanjing_page.status_code == 200
    assert nanchang_page.status_code == 200
    assert "09:45-17:45" in nanjing_page.text
    assert "10:00-18:00" not in nanjing_page.text
    assert "10:00-18:00" in nanchang_page.text
    assert "09:45-17:45" not in nanchang_page.text
    assert "09:30-17:30" in nanchang_page.text  # 通用旧班次仍兼容显示


def test_custom_schedule_time_duration_cross_midnight_and_bad_input_html(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)
    create_schedule_employee(client, "自定义排班员工")

    custom_response = admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南京门东店",
            "employee_ids": ["1"],
            "schedule_dates": ["2026-07-02"],
            "schedule_mode": "custom",
            "custom_label": "加班",
            "custom_start_time": "18:00",
            "custom_end_time": "21:00",
            "note": "盘点支援",
            "overwrite_existing": "1",
        },
        follow_redirects=False,
    )
    overnight_response = admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南京门东店",
            "employee_ids": ["1"],
            "schedule_dates": ["2026-07-03"],
            "schedule_mode": "custom",
            "custom_label": "夜间盘点",
            "custom_start_time": "22:00",
            "custom_end_time": "02:00",
            "overwrite_existing": "1",
        },
        follow_redirects=False,
    )

    assert custom_response.status_code == 303
    assert overnight_response.status_code == 303
    schedules = sorted(rows_for(tmp_path, "store_schedules"), key=lambda row: row["schedule_date"])
    assert [(row["is_custom_time"], row["custom_label"], row["custom_duration_hours"]) for row in schedules] == [
        (1, "加班", 3.0),
        (1, "夜间盘点", 4.0),
    ]

    table_page = client.get("/admin/schedules?store_names=南京门东店&month=2026-07&view_mode=table")
    calendar_page = client.get("/admin/schedules?store_names=南京门东店&month=2026-07&view_mode=calendar")
    assert "shift-badge custom-shift-badge" in table_page.text
    assert "加班" in table_page.text
    assert "custom-shift-badge" in calendar_page.text
    assert "夜间盘点" in calendar_page.text

    bad_custom = admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南京门东店",
            "employee_ids": ["1"],
            "schedule_dates": ["2026-07-04"],
            "schedule_mode": "custom",
            "custom_label": "缺时间",
            "custom_start_time": "",
            "custom_end_time": "",
        },
        follow_redirects=False,
    )
    assert bad_custom.status_code == 303
    bad_page = client.get(bad_custom.headers["location"])
    assert "自定义时间必须填写开始时间和结束时间" in bad_page.text
    assert "alert-error" in bad_page.text
    assert '{"detail":' not in bad_page.text


def test_schedule_calendar_uses_real_weekday_and_holiday_config(tmp_path, monkeypatch):
    client, main = build_client(
        tmp_path,
        monkeypatch,
        config_overrides={
            "holidays.json": {
                "2026-07-01": "香港特别行政区成立纪念日",
                "2026-07-04": "测试节假日",
            }
        },
    )
    logged_in_client(client)
    create_schedule_employee(client, "日历员工")

    calendar_days = main.build_month_calendar(2026, 7)
    july_first = next(day for day in calendar_days if day["date"] == "2026-07-01")
    assert july_first["weekday_index"] == 2
    assert july_first["weekday_label"] == "周三"
    assert july_first["weekday"] == "周三"
    assert july_first["is_holiday"] is True
    assert july_first["holiday_name"] == "香港特别行政区成立纪念日"

    page = client.get("/admin/schedules?store_names=南京门东店&month=2026-07&view_mode=calendar")
    assert page.status_code == 200
    assert 'id="schedule-calendar-view"' in page.text
    assert 'class="schedule-date-calendar"' in page.text
    assert 'data-date="2026-07-01"' in page.text
    assert 'data-weekday-label="周三"' in page.text
    assert "香港特别行政区成立纪念日" in page.text
    assert "schedule-date-card" in page.text
    assert "holiday" in page.text
    assert "data-select-holidays" in page.text


def test_custom_schedule_break_manual_duration_and_display_time(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)
    create_schedule_employee(client, "时间展示员工")

    with_break = admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南京门东店",
            "employee_ids": ["1"],
            "schedule_dates": ["2026-07-05"],
            "schedule_mode": "custom",
            "custom_label": "盘点",
            "custom_start_time": "09:00",
            "custom_end_time": "18:00",
            "custom_break_minutes": "60",
            "overwrite_existing": "1",
        },
        follow_redirects=False,
    )
    manual_duration = admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南京门东店",
            "employee_ids": ["1"],
            "schedule_dates": ["2026-07-06"],
            "schedule_mode": "custom",
            "custom_label": "夜间盘点",
            "custom_start_time": "22:00",
            "custom_end_time": "02:00",
            "custom_break_minutes": "60",
            "custom_duration_hours": "3.5",
            "overwrite_existing": "1",
        },
        follow_redirects=False,
    )

    assert with_break.status_code == 303
    assert manual_duration.status_code == 303
    schedules = sorted(rows_for(tmp_path, "store_schedules"), key=lambda row: row["schedule_date"])
    assert [(row["custom_label"], row["custom_duration_hours"]) for row in schedules] == [("盘点", 8.0), ("夜间盘点", 3.5)]

    table_page = client.get("/admin/schedules?store_names=南京门东店&month=2026-07&view_mode=table")
    calendar_page = client.get("/admin/schedules?store_names=南京门东店&month=2026-07&view_mode=calendar")
    employee_page = client.get("/admin/schedules?store_names=南京门东店&month=2026-07&view_mode=employee&employee_ids=1")
    assert "09:00-18:00" in table_page.text
    assert "22:00-02:00（跨天）" in table_page.text
    assert "3.5 小时" in table_page.text
    assert "schedule-time-line" in calendar_page.text
    assert "22:00-02:00（跨天）" in calendar_page.text
    assert "employee-schedule-store" in employee_page.text
    assert "南京门东店 · 夜间盘点 · 22:00-02:00（跨天） · 3.5 小时" in employee_page.text


def test_employee_schedule_link_uses_all_bound_stores_scope(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)
    create_schedule_employee(client, "跨店视图员工", store_names=["南京门东店", "南昌万寿宫店"])

    admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南京门东店",
            "employee_ids": ["1"],
            "schedule_dates": ["2026-07-07"],
            "shift_type_id": "1",
            "overwrite_existing": "1",
        },
        follow_redirects=False,
    )
    admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南昌万寿宫店",
            "employee_ids": ["1"],
            "schedule_dates": ["2026-07-07"],
            "schedule_mode": "custom",
            "custom_label": "支援",
            "custom_start_time": "18:00",
            "custom_end_time": "21:00",
            "overwrite_existing": "1",
        },
        follow_redirects=False,
    )

    employee_page = client.get("/admin/employees")
    assert "/admin/schedules?view_mode=employee&amp;employee_ids=1&amp;store_scope=employee-all-stores#employee-schedule-view" in employee_page.text

    schedule_page = client.get(
        "/admin/schedules?view_mode=employee&employee_ids=1&store_scope=employee-all-stores&month=2026-07"
    )
    assert schedule_page.status_code == 200
    assert "employee-all-store-summary" in schedule_page.text
    assert "南京门东店" in schedule_page.text
    assert "南昌万寿宫店" in schedule_page.text
    assert "早班" in schedule_page.text
    assert "支援" in schedule_page.text


def test_schedule_daily_overview_is_simplified_and_hour_based(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)
    create_schedule_employee(client, "每日概览员工")
    admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南京门东店",
            "employee_ids": ["1"],
            "schedule_dates": ["2026-07-01"],
            "shift_type_id": "1",
            "overwrite_existing": "1",
        },
        follow_redirects=False,
    )

    page = client.get("/admin/schedules?store_names=南京门东店&month=2026-07&view_mode=calendar")
    assert page.status_code == 200
    assert "每日排班概览" in page.text
    assert "每天每个门店排班人数" not in page.text
    assert "daily-overview-card" in page.text
    assert "1 人" in page.text
    assert "8 小时" in page.text


def test_schedule_multi_select_filters_dashboard_and_exports(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)
    month = date.today().strftime("%Y-%m")
    first_day = f"{month}-01"
    second_day = f"{month}-02"
    create_schedule_employee(client, "南京看板员工", store_name="南京门东店")
    create_schedule_employee(client, "南昌看板员工", store_name="南昌万寿宫店")
    create_schedule_employee(client, "离职看板员工", store_name="山城巷店", status="离职")

    admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南京门东店",
            "employee_ids": ["1"],
            "schedule_dates": [first_day],
            "shift_type_id": "1",
            "overwrite_existing": "1",
        },
        follow_redirects=False,
    )
    admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南昌万寿宫店",
            "employee_ids": ["2"],
            "schedule_dates": [second_day],
            "schedule_mode": "custom",
            "custom_label": "加班",
            "custom_start_time": "18:00",
            "custom_end_time": "21:00",
            "overwrite_existing": "1",
        },
        follow_redirects=False,
    )
    admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南京门东店",
            "employee_ids": ["1"],
            "schedule_dates": [f"{month}-03"],
            "shift_type_id": "4",
            "overwrite_existing": "1",
        },
        follow_redirects=False,
    )

    page = client.get(
        "/admin/schedules",
        params=[
            ("store_names", "南京门东店"),
            ("store_names", "南昌万寿宫店"),
            ("month", month),
            ("view_mode", "store-summary"),
            ("employee_statuses", "在职"),
            ("shift_type_ids", "1"),
            ("shift_type_ids", "custom"),
            ("shift_type_ids", ""),
            ("employee_ids", ""),
        ],
    )
    assert page.status_code == 200
    assert "text/html" in page.headers["content-type"]
    assert '{"detail":' not in page.text
    assert 'name="store_names"' in page.text
    assert 'name="employee_statuses"' in page.text
    assert 'name="shift_type_ids"' in page.text
    assert "已选 2 个门店" in page.text
    assert "当前为多门店查看模式，请选择单个门店后进行排班操作。" in page.text
    assert "排班看板" in page.text
    assert "当月已排工时" in page.text
    assert "截至当前日期已排工时" in page.text
    assert "当月排班人次" in page.text
    assert "当月排班人数" in page.text
    assert "自定义/加班班次数" in page.text
    assert "休息班次数" in page.text
    assert "南京门东店" in page.text
    assert "南昌万寿宫店" in page.text

    multi_store_post = admin_post(
        client,
        "/admin/schedules",
        data={
            "store_names": ["南京门东店", "南昌万寿宫店"],
            "employee_ids": ["1"],
            "schedule_dates": [first_day],
            "shift_type_id": "1",
        },
        follow_redirects=False,
    )
    assert multi_store_post.status_code == 303
    assert "请选择单个具体门店后进行排班操作" in client.get(multi_store_post.headers["location"]).text

    multi_export = client.get(
        "/admin/schedules/export",
        params=[("store_names", "南京门东店"), ("store_names", "南昌万寿宫店"), ("month", month)],
    )
    all_export = client.get(f"/admin/schedules/export?month={month}")
    single_export = client.get(f"/admin/schedules/export?store_names=南京门东店&month={month}")
    assert multi_export.status_code == 200
    assert all_export.status_code == 200
    assert single_export.status_code == 200
    assert quote(f"门店排班_多门店_{month}.xlsx") in multi_export.headers["content-disposition"]
    assert quote(f"门店排班_全部门店_{month}.xlsx") in all_export.headers["content-disposition"]
    assert quote(f"门店排班_南京门东店_{month}.xlsx") in single_export.headers["content-disposition"]

    workbook = load_workbook(BytesIO(multi_export.content))
    rows = list(workbook.active.iter_rows(values_only=True))
    assert rows[0][:9] == ("门店", "员工", "日期", "星期", "班次", "排班时间", "工时", "是否自定义", "备注")
    assert rows[0][-3:] == ("是否冲突", "工时异常", "异常说明")
    assert any(row[:8] == ("南京门东店", "南京看板员工", first_day, "周三", "早班", "09:30-17:30", "8 小时", "否") for row in rows)
    assert any(row[:8] == ("南昌万寿宫店", "南昌看板员工", second_day, "周四", "加班", "18:00-21:00", "3 小时", "是") for row in rows)


def test_store_specific_shift_cannot_be_used_for_other_store(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)
    create_schedule_employee(client, "南京员工", store_name="南京门东店")
    admin_post(
        client,
        "/admin/shift-types",
        data={
            "store_name": "南昌万寿宫店",
            "shift_name": "南昌专属晚班",
            "start_time": "15:00",
            "end_time": "23:00",
            "duration_hours": "8",
            "color": "#7c3aed",
        },
        follow_redirects=False,
    )
    nanchang_shift_id = rows_for(tmp_path, "shift_types")[-1]["id"]

    response = admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南京门东店",
            "employee_ids": ["1"],
            "schedule_dates": ["2026-07-08"],
            "shift_type_id": str(nanchang_shift_id),
        },
        follow_redirects=False,
    )

    assert response.status_code == 303
    page = client.get(response.headers["location"])
    assert "班次不属于当前门店" in page.text
    assert rows_for(tmp_path, "store_schedules") == []


def test_schedule_conflict_detection_hours_warnings_override_and_export(tmp_path, monkeypatch):
    client, main = build_client(tmp_path, monkeypatch)
    logged_in_client(client)
    create_schedule_employee(client, "冲突员工", store_names=["南京门东店", "南昌万寿宫店"])

    assert "schedule.override_conflict" in main.ADMIN_PERMISSION_KEYS
    assert "schedule.override_conflict" in main.permission_service.get_user_permissions(main.fetch_admin_user_by_username(ADMIN_AUTH[0]))

    created = admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南京门东店",
            "employee_id": "1",
            "schedule_date": "2026-07-08",
            "shift_type_id": "1",
            "note": "早班",
        },
        follow_redirects=False,
    )
    assert created.status_code == 303
    assert len(rows_for(tmp_path, "store_schedules")) == 1

    blocked = admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南昌万寿宫店",
            "employee_id": "1",
            "schedule_date": "2026-07-08",
            "shift_type_id": "1",
            "note": "跨店重叠",
        },
        follow_redirects=False,
    )
    assert blocked.status_code == 303
    blocked_page = client.get(blocked.headers["location"])
    assert "排班冲突" in blocked_page.text
    assert "冲突员工" in blocked_page.text
    assert "南昌万寿宫店" in blocked_page.text
    assert len(rows_for(tmp_path, "store_schedules")) == 1

    readonly_role = create_role_with_permissions(tmp_path, "schedule-no-override", ["schedule.view", "schedule.create", "schedule.update"])
    update_admin_user_access(tmp_path, ADMIN_AUTH[0], readonly_role)
    no_override_page = client.get("/admin/schedules?store_name=南京门东店&month=2026-07")
    assert "忽略冲突并保存" not in no_override_page.text

    still_blocked = admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南昌万寿宫店",
            "employee_id": "1",
            "schedule_date": "2026-07-08",
            "shift_type_id": "1",
            "override_conflict": "1",
            "csrf_token": csrf_token_for(client, "/admin/schedules?store_name=南京门东店&month=2026-07"),
        },
        follow_redirects=False,
    )
    assert still_blocked.status_code == 303
    assert "排班冲突" in client.get(still_blocked.headers["location"]).text
    assert len(rows_for(tmp_path, "store_schedules")) == 1

    system_role_id = next(row["id"] for row in rows_for(tmp_path, "admin_roles") if row["role_name"] == "系统管理员")
    update_admin_user_access(tmp_path, ADMIN_AUTH[0], system_role_id)
    override_page = client.get("/admin/schedules?store_name=南京门东店&month=2026-07")
    assert "忽略冲突并保存" in override_page.text
    overridden = admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南昌万寿宫店",
            "employee_id": "1",
            "schedule_date": "2026-07-08",
            "shift_type_id": "1",
            "override_conflict": "1",
            "note": "允许覆盖",
        },
        follow_redirects=False,
    )
    assert overridden.status_code == 303
    assert len(rows_for(tmp_path, "store_schedules")) == 2
    operations = rows_for(tmp_path, "admin_operation_logs")
    assert any(row["action"] == "schedule.override_conflict" and "override_conflict" in row["detail"] for row in operations)

    custom_overlap = admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南京门东店",
            "employee_id": "1",
            "schedule_date": "2026-07-08",
            "schedule_mode": "custom",
            "shift_type_id": "custom",
            "custom_label": "自定义重叠",
            "custom_start_time": "10:00",
            "custom_end_time": "12:00",
        },
        follow_redirects=False,
    )
    assert "排班冲突" in client.get(custom_overlap.headers["location"]).text
    assert len(rows_for(tmp_path, "store_schedules")) == 2

    admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南京门东店",
            "employee_id": "1",
            "schedule_date": "2026-07-09",
            "schedule_mode": "custom",
            "shift_type_id": "custom",
            "custom_label": "跨天",
            "custom_start_time": "23:00",
            "custom_end_time": "01:00",
        },
        follow_redirects=False,
    )
    cross_day_block = admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南昌万寿宫店",
            "employee_id": "1",
            "schedule_date": "2026-07-10",
            "schedule_mode": "custom",
            "shift_type_id": "custom",
            "custom_label": "次日凌晨",
            "custom_start_time": "00:30",
            "custom_end_time": "08:00",
        },
        follow_redirects=False,
    )
    assert "排班冲突" in client.get(cross_day_block.headers["location"]).text

    rest_shift = admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南昌万寿宫店",
            "employee_id": "1",
            "schedule_date": "2026-07-09",
            "shift_type_id": "4",
        },
        follow_redirects=False,
    )
    assert rest_shift.status_code == 303

    page = client.get("/admin/schedules?store_names=南京门东店&store_names=南昌万寿宫店&month=2026-07&view_mode=calendar")
    assert page.status_code == 200
    assert "冲突" in page.text
    assert "异常" in page.text
    assert "跨店排班" in page.text

    employee_page = client.get("/admin/schedules?store_names=南京门东店&store_names=南昌万寿宫店&month=2026-07&view_mode=employee")
    assert "本月工时" in employee_page.text
    assert "跨店排班" in employee_page.text

    export_response = client.get("/admin/schedules/export?store_names=南京门东店&store_names=南昌万寿宫店&month=2026-07")
    workbook = load_workbook(BytesIO(export_response.content))
    rows = list(workbook.active.iter_rows(values_only=True))
    assert rows[0][-3:] == ("是否冲突", "工时异常", "异常说明")
    assert any(row[-3] == "是" and "跨店排班" in (row[-1] or "") for row in rows[1:])

    overview = client.get("/admin/permission-overview")
    assert overview.status_code == 200
    assert "schedule.override_conflict" in overview.text
    assert "高风险 POST 未接入：0" in overview.text


def test_schedule_validation_service_overlap_rest_and_hour_rules(tmp_path, monkeypatch):
    client, main = build_client(tmp_path, monkeypatch)
    logged_in_client(client)
    create_schedule_employee(client, "规则员工")

    morning = {"employee_id": 1, "schedule_date": "2026-07-08", "start_time": "09:00", "end_time": "18:00", "shift_name": "早班"}
    evening = {"employee_id": 1, "schedule_date": "2026-07-08", "start_time": "18:00", "end_time": "22:00", "shift_name": "晚班"}
    overlap = {"employee_id": 1, "schedule_date": "2026-07-08", "start_time": "13:00", "end_time": "20:00", "shift_name": "中班"}
    rest = {"employee_id": 1, "schedule_date": "2026-07-08", "start_time": "", "end_time": "", "shift_name": "休息"}

    assert not main.schedule_validation_service.time_ranges_overlap(
        main.schedule_validation_service.normalize_schedule_time(morning),
        main.schedule_validation_service.normalize_schedule_time(evening),
    )
    assert main.schedule_validation_service.time_ranges_overlap(
        main.schedule_validation_service.normalize_schedule_time(morning),
        main.schedule_validation_service.normalize_schedule_time(overlap),
    )
    assert main.schedule_validation_service.normalize_schedule_time(rest)["is_rest"]

    yellow = main.schedule_validation_service.build_schedule_warnings(
        1,
        "2026-07",
        [
            {**morning, "duration_hours": 11, "store_name": "南京门东店", "employee_name": "规则员工"},
        ],
    )
    red = main.schedule_validation_service.build_schedule_warnings(
        1,
        "2026-07",
        [
            {**morning, "duration_hours": 13, "store_name": "南京门东店", "employee_name": "规则员工"},
        ],
    )
    assert any(item["level"] == "warning" and "单日工时偏高" in item["message"] for item in yellow)
    assert any(item["level"] == "danger" and "单日工时严重偏高" in item["message"] for item in red)

    continuous_rows = [
        {
            "employee_id": 1,
            "employee_name": "规则员工",
            "schedule_date": f"2026-07-{day:02d}",
            "shift_name": "早班",
            "store_name": "南京门东店",
            "duration_hours": 8,
            "start_time": "09:00",
            "end_time": "18:00",
        }
        for day in range(1, 8)
    ]
    continuous = main.schedule_validation_service.build_schedule_warnings(1, "2026-07", continuous_rows)
    assert any("连续排班超过 6 天" in item["message"] for item in continuous)
    interrupted = main.schedule_validation_service.build_schedule_warnings(
        1,
        "2026-07",
        [*continuous_rows[:3], {**continuous_rows[3], "shift_name": "休息", "duration_hours": 0}, *continuous_rows[4:]],
    )
    assert not any("连续排班超过 6 天" in item["message"] for item in interrupted)


def test_admin_schedule_interaction_js_feedback_contract():
    app_js = (PROJECT_DIR / "static" / "app.js").read_text(encoding="utf-8")
    style_css = (PROJECT_DIR / "static" / "style.css").read_text(encoding="utf-8")

    assert "data-feedback-form" in app_js
    assert "data-loading-label" in app_js
    assert "data-drawer-open" in app_js
    assert "data-drawer-close" in app_js
    assert "data-employee-store-form" in app_js
    assert "syncEmployeePrimaryStoreChoices" in app_js
    assert "data-shift-scope-form" in app_js
    assert "updateShiftScopeForm" in app_js
    assert "data-filter-toggle" in app_js
    assert "showAppToast" in app_js
    for selector in (
        ".ops-card",
        ".schedule-calendar-grid",
        ".schedule-employee-card",
        ".employee-card-grid",
        ".drawer-panel",
        ".shift-type-card",
    ):
        assert selector in style_css


def test_schedule_module_form_errors_redirect_with_page_alerts(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)

    employee_error = admin_post(client, "/admin/employees", data={"employee_name": ""}, follow_redirects=False)
    assert employee_error.status_code == 303
    assert unquote(employee_error.headers["location"]).startswith("/admin/employees?error=请填写员工姓名")
    employee_page = client.get(employee_error.headers["location"])
    assert employee_page.status_code == 200
    assert "请填写员工姓名" in employee_page.text
    assert "alert-error" in employee_page.text

    shift_error = admin_post(client, "/admin/shift-types", data={"shift_name": ""}, follow_redirects=False)
    assert shift_error.status_code == 303
    assert unquote(shift_error.headers["location"]).startswith("/admin/shift-types?error=请填写班次名称")
    shift_page = client.get(shift_error.headers["location"])
    assert shift_page.status_code == 200
    assert "请填写班次名称" in shift_page.text
    assert "alert-error" in shift_page.text

    schedule_error = admin_post(client, "/admin/schedules", data={}, follow_redirects=False)
    assert schedule_error.status_code == 303
    schedule_location = unquote(schedule_error.headers["location"])
    assert schedule_location.startswith("/admin/schedules?")
    assert "error=" in schedule_location
    schedule_page = client.get(schedule_error.headers["location"])
    assert schedule_page.status_code == 200
    assert "alert-error" in schedule_page.text

    update_missing_employee = admin_post(
        client,
        "/admin/employees/404/update",
        data={"employee_name": "ghost", "store_name": "南京门东店", "status": "在职"},
        follow_redirects=False,
    )
    assert update_missing_employee.status_code == 303
    assert "/admin/employees?error=" in update_missing_employee.headers["location"]

    update_missing_shift = admin_post(
        client,
        "/admin/shift-types/404/update",
        data={"shift_name": "ghost", "duration_hours": "8", "is_active": "1"},
        follow_redirects=False,
    )
    assert update_missing_shift.status_code == 303
    assert "/admin/shift-types?error=" in update_missing_shift.headers["location"]

    delete_missing_schedule = admin_post(client, "/admin/schedules/404/delete", follow_redirects=False)
    assert delete_missing_schedule.status_code == 303
    delete_location = unquote(delete_missing_schedule.headers["location"])
    assert delete_location.startswith("/admin/schedules?")
    assert "error=" in delete_location


def test_schedule_create_update_validation_public_view_export_and_logs(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)
    admin_post(
        client,
        "/admin/employees",
        data={"employee_name": "王早班", "store_name": "南京门东店", "role": "店员", "phone": "", "status": "在职"},
        follow_redirects=False,
    )
    admin_post(
        client,
        "/admin/employees",
        data={"employee_name": "李离职", "store_name": "南京门东店", "role": "店员", "phone": "", "status": "离职"},
        follow_redirects=False,
    )
    admin_post(client, "/admin/shift-types/4/disable", follow_redirects=False)

    create_schedule = admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南京门东店",
            "employee_id": "1",
            "schedule_date": "2026-07-06",
            "shift_type_id": "1",
            "note": "开店",
        },
        follow_redirects=False,
    )
    assert create_schedule.status_code == 303
    schedules = rows_for(tmp_path, "store_schedules")
    assert len(schedules) == 1
    assert schedules[0]["employee_id"] == 1
    assert schedules[0]["shift_type_id"] == 1
    assert schedules[0]["note"] == "开店"

    update_schedule = admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南京门东店",
            "employee_id": "1",
            "schedule_date": "2026-07-06",
            "shift_type_id": "2",
            "note": "改晚班",
        },
        follow_redirects=False,
    )
    assert update_schedule.status_code == 303
    schedules = rows_for(tmp_path, "store_schedules")
    assert len(schedules) == 1
    assert schedules[0]["shift_type_id"] == 2
    assert schedules[0]["note"] == "改晚班"

    logs = rows_for(tmp_path, "schedule_logs")
    assert [row["action"] for row in logs] == ["新增排班", "更新排班"]

    inactive_employee = admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南京门东店",
            "employee_id": "2",
            "schedule_date": "2026-07-07",
            "shift_type_id": "1",
            "note": "不允许",
        },
        follow_redirects=False,
    )
    assert inactive_employee.status_code == 303
    inactive_employee_page = client.get(inactive_employee.headers["location"])
    assert "员工必须是在职状态" in inactive_employee_page.text

    inactive_shift = admin_post(
        client,
        "/admin/schedules",
        data={
            "store_name": "南京门东店",
            "employee_id": "1",
            "schedule_date": "2026-07-07",
            "shift_type_id": "4",
            "note": "停用班次",
        },
        follow_redirects=False,
    )
    assert inactive_shift.status_code == 303
    inactive_shift_page = client.get(inactive_shift.headers["location"])
    assert "班次必须启用" in inactive_shift_page.text

    public_page = client.get("/schedule")
    assert public_page.status_code == 200
    assert "门店排班" in public_page.text
    public_store = client.get("/schedule?store_name=南京门东店&month=2026-07")
    assert public_store.status_code == 200
    assert "王早班" in public_store.text
    assert "晚班" in public_store.text
    assert "改晚班" in public_store.text

    admin_page = client.get("/admin/schedules?store_name=南京门东店&month=2026-07")
    assert admin_page.status_code == 200
    assert "当前月份总工时" in admin_page.text
    assert "每日排班概览" in admin_page.text
    assert "王早班" in admin_page.text

    export_response = client.get("/admin/schedules/export?store_name=南京门东店&month=2026-07")
    assert export_response.status_code == 200
    assert "%E9%97%A8%E5%BA%97%E6%8E%92%E7%8F%AD_%E5%8D%97%E4%BA%AC%E9%97%A8%E4%B8%9C%E5%BA%97_2026-07.xlsx" in export_response.headers["content-disposition"]
    workbook = load_workbook(BytesIO(export_response.content))
    values = list(workbook.active.iter_rows(values_only=True))
    assert values[0][:9] == ("门店", "员工", "日期", "星期", "班次", "排班时间", "工时", "是否自定义", "备注")
    assert values[0][-3:] == ("是否冲突", "工时异常", "异常说明")
    assert any(row[:9] == ("南京门东店", "王早班", "2026-07-06", "周一", "晚班", "14:00-22:00", "8 小时", "否", "改晚班") for row in values)

    delete_schedule = admin_post(client, "/admin/schedules/1/delete", follow_redirects=False)
    assert delete_schedule.status_code == 303
    assert rows_for(tmp_path, "store_schedules") == []
    assert rows_for(tmp_path, "schedule_logs")[-1]["action"] == "删除排班"


def test_admin_can_soft_delete_collaboration_items_and_supplements(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    submit_ticket(client)
    logged_in_client(client)

    admin_post(
        client,
        "/admin/ticket/1/comments",
        data={"content": "公开评论会被删除", "visibility": "public"},
        follow_redirects=False,
    )
    admin_post(
        client,
        "/admin/ticket/1/tasks",
        data={"title": "删除前子任务", "assignee": "采购", "status": "待处理", "due_date": "2026-07-06"},
        follow_redirects=False,
    )
    admin_post(
        client,
        "/admin/ticket/1/participants",
        data={"participant_type": "team", "participant_name": "临时协作人甲", "role": "协作处理"},
        follow_redirects=False,
    )
    client.post(
        "/query/ticket/1/supplement",
        data={"store_name": "南京门东店", "submitter": "小李", "note": "这条补充会被隐藏"},
    )
    assert "公开评论会被删除" in client.get("/admin/ticket/1").text
    assert "删除前子任务" in client.get("/admin/ticket/1").text
    assert "临时协作人甲" in client.get("/admin/ticket/1").text
    assert "这条补充会被隐藏" in client.get("/query/ticket/1?store_name=南京门东店").text

    comment_id = rows_for(tmp_path, "ticket_comments")[0]["id"]
    task_id = rows_for(tmp_path, "ticket_tasks")[0]["id"]
    participant_id = rows_for(tmp_path, "ticket_participants")[0]["id"]
    supplement_id = rows_for(tmp_path, "ticket_supplements")[0]["id"]

    assert admin_post(client, f"/admin/ticket/1/comment/{comment_id}/delete", follow_redirects=False).status_code == 303
    assert admin_post(client, f"/admin/ticket/1/task/{task_id}/delete", follow_redirects=False).status_code == 303
    assert admin_post(client, f"/admin/ticket/1/participant/{participant_id}/delete", follow_redirects=False).status_code == 303
    assert admin_post(client, f"/admin/ticket/1/supplement/{supplement_id}/delete", follow_redirects=False).status_code == 303

    admin_detail = client.get("/admin/ticket/1").text
    store_detail = client.get("/query/ticket/1?store_name=南京门东店").text
    assert "公开评论会被删除" not in admin_detail
    assert "删除前子任务" not in admin_detail
    assert "临时协作人甲" not in admin_detail
    assert "这条补充会被隐藏" not in store_detail
    assert rows_for(tmp_path, "ticket_comments")[0]["deleted_at"]
    assert rows_for(tmp_path, "ticket_tasks")[0]["deleted_at"]
    assert rows_for(tmp_path, "ticket_participants")[0]["deleted_at"]
    assert rows_for(tmp_path, "ticket_supplements")[0]["deleted_at"]
    actions = [row["action"] for row in rows_for(tmp_path, "ticket_logs")]
    assert "删除评论" in actions
    assert "删除子任务" in actions
    assert "移除协作人" in actions
    assert "隐藏门店补充记录" in actions


def test_embedded_page_soft_delete_restore_and_hard_delete(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)
    csrf_token = csrf_token_for(client, "/admin/embedded-pages")
    create = client.post(
        "/admin/embedded-pages",
        data={
            "csrf_token": csrf_token,
            "page_key": "trash-report",
            "title": "待删除扩展页",
            "nav_label": "待删除扩展",
            "enabled": "1",
        },
        files={"html_file": ("report.html", b"<h1>trash report</h1>", "text/html")},
        follow_redirects=False,
    )
    assert create.status_code == 303
    page_dir = tmp_path / "embedded_pages" / "trash-report"
    assert page_dir.is_dir()
    assert "待删除扩展" in client.get("/admin/dashboard").text
    assert client.get("/admin/embed/trash-report").status_code == 200

    delete = admin_post(
        client,
        "/admin/embedded-pages/trash-report/delete",
        data={"delete_reason": "测试扩展页清理"},
        follow_redirects=False,
    )
    assert delete.status_code == 303
    deleted_page = rows_for(tmp_path, "embedded_pages")[0]
    assert deleted_page["deleted_at"]
    assert deleted_page["enabled"] == 0
    assert "待删除扩展" not in client.get("/admin/dashboard").text
    assert client.get("/admin/embed/trash-report").status_code == 404
    assert "trash-report" in client.get("/admin/trash").text

    restore = admin_post(client, "/admin/embedded-pages/trash-report/restore", follow_redirects=False)
    assert restore.status_code == 303
    restored_page = rows_for(tmp_path, "embedded_pages")[0]
    assert restored_page["deleted_at"] is None
    assert restored_page["enabled"] == 1
    assert client.get("/admin/embed/trash-report").status_code == 200

    admin_post(client, "/admin/embedded-pages/trash-report/delete", follow_redirects=False)
    hard_delete = admin_post(
        client,
        "/admin/embedded-pages/trash-report/hard-delete",
        data={"confirm_delete": "1"},
        follow_redirects=False,
    )
    assert hard_delete.status_code == 303
    assert rows_for(tmp_path, "embedded_pages") == []
    assert not page_dir.exists()


def test_cleanup_preview_and_delete_soft_deletes_matching_test_tickets(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    submit_ticket(client, description="Codex smoke cleanup ticket")
    submit_ticket(client, description="正式运营工单")
    logged_in_client(client)

    cleanup_page = client.get("/admin/cleanup")
    assert cleanup_page.status_code == 200
    assert "预览将删除的工单" in cleanup_page.text

    preview = admin_post(
        client,
        "/admin/cleanup/preview",
        data={"keyword": "Codex", "only_test": "1"},
    )
    assert preview.status_code == 200
    assert "Codex smoke cleanup ticket" in preview.text
    assert "正式运营工单" not in preview.text

    delete = admin_post(
        client,
        "/admin/cleanup/delete",
        data={"keyword": "Codex", "only_test": "1", "confirm_cleanup": "1"},
        follow_redirects=False,
    )
    assert delete.status_code == 303
    tickets = rows_for(tmp_path, "tickets")
    assert tickets[0]["deleted_at"]
    assert tickets[0]["delete_reason"] == "批量清理测试数据"
    assert tickets[1]["deleted_at"] is None
    admin_page = client.get("/admin").text
    assert "Codex smoke cleanup ticket" not in admin_page
    assert "正式运营工单" in admin_page


def test_admin_pagination_filters_handlers_and_keyword_search(tmp_path, monkeypatch):
    client, _ = build_client(
        tmp_path,
        monkeypatch,
        {
            "system.json": {
                "app_name": "门店需求工单系统",
                "port": 8701,
                "max_image_mb": 10,
                "allowed_image_extensions": ["jpg", "jpeg", "png", "webp"],
                "default_status": "待处理",
                "excel_filename_prefix": "门店需求工单",
                "page_size": 2,
                "max_image_count": 5,
                "max_total_upload_mb": 30,
            },
            "handlers.json": ["张三", "李四"],
        },
    )

    for index in range(3):
        response = submit_ticket(client, description=f"分页测试工单 {index}", product_name=f"测试商品{index}")
        assert response.status_code == 200

    logged_in_client(client)
    admin_post(
        client,
        "/admin/ticket/1",
        data={"status": "处理中", "assigned_to": "张三", "handler_note": "分配给张三"},
    )

    page_1 = client.get("/admin?page=1&sort=newest")
    assert page_1.status_code == 200
    assert "当前第 1 页" in page_1.text
    assert "共 2 页" in page_1.text
    assert "共 3 条工单" in page_1.text
    assert "sort=newest" in page_1.text
    assert "page=2" in page_1.text

    page_2 = client.get("/admin?page=2&sort=newest")
    assert page_2.status_code == 200
    assert "当前第 2 页" in page_2.text

    assigned_filter = client.get(f"/admin?assigned_to={quote('张三')}&keyword={quote('张三')}")
    assert assigned_filter.status_code == 200
    assert "分配给张三" in assigned_filter.text or "张三" in assigned_filter.text
    assert "共 1 条工单" in assigned_filter.text

    export_response = client.get("/admin/export?page=2")
    workbook = load_workbook(BytesIO(export_response.content))
    assert workbook.active.max_row == 4


def test_export_without_data_still_contains_headers(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)

    assert client.get("/admin/export").status_code == 401

    logged_in_client(client)
    response = client.get("/admin/export")
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active

    assert [cell.value for cell in sheet[1]] == EXPECTED_EXPORT_HEADERS
    assert sheet.max_row == 1


def test_legacy_database_is_migrated_without_losing_existing_rows(tmp_path, monkeypatch):
    config_dir = tmp_path / "config"
    write_config(config_dir)
    db_path = tmp_path / "tickets.db"
    with sqlite3.connect(db_path) as connection:
        connection.execute(
            """
            CREATE TABLE tickets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ticket_no TEXT NOT NULL UNIQUE,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                store_name TEXT NOT NULL,
                submitter TEXT NOT NULL,
                request_type TEXT NOT NULL,
                urgency TEXT NOT NULL,
                brand TEXT,
                product_name TEXT,
                sku_barcode TEXT,
                quantity INTEGER,
                description TEXT NOT NULL,
                expected_finish_date TEXT,
                status TEXT NOT NULL,
                handler_note TEXT
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE ticket_images (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ticket_id INTEGER NOT NULL,
                image_path TEXT NOT NULL,
                uploaded_at TEXT NOT NULL
            )
            """
        )
        connection.execute(
            """
            INSERT INTO tickets (
                ticket_no, created_at, updated_at, store_name, submitter, request_type, urgency,
                brand, product_name, sku_barcode, quantity, description, expected_finish_date, status, handler_note
            )
            VALUES (
                'REQ-20260703-0001', '2026-07-03 10:00:00', '2026-07-03 10:00:00',
                '南京门东店', '旧数据', '建单需求', '普通', '', '旧商品', '', 1,
                '旧库里的工单', '2026-07-04', '待处理', ''
            )
            """
        )
        connection.execute(
            "INSERT INTO ticket_images (ticket_id, image_path, uploaded_at) VALUES (1, 'uploads/old.png', '2026-07-03 10:00:00')"
        )

    monkeypatch.setenv("STORE_REQUEST_DB_PATH", str(db_path))
    monkeypatch.setenv("STORE_REQUEST_UPLOAD_DIR", str(tmp_path / "uploads"))
    monkeypatch.setenv("STORE_REQUEST_CONFIG_DIR", str(config_dir))
    monkeypatch.setenv("ADMIN_USERS", "")
    monkeypatch.setenv("ADMIN_USERNAME", ADMIN_AUTH[0])
    monkeypatch.setenv("ADMIN_PASSWORD", ADMIN_AUTH[1])
    monkeypatch.syspath_prepend(str(PROJECT_DIR))
    sys.modules.pop("main", None)
    main = importlib.import_module("main")
    monkeypatch.delenv("ADMIN_USERS", raising=False)
    client = TestClient(main.app)

    with sqlite3.connect(db_path) as connection:
        soft_delete_tables = [
            "tickets",
            "ticket_supplements",
            "ticket_participants",
            "ticket_comments",
            "ticket_tasks",
            "embedded_pages",
            "shift_types",
        ]
        migrated_columns = {
            table: {row[1] for row in connection.execute(f"PRAGMA table_info({table})").fetchall()}
            for table in soft_delete_tables
        }
        ticket_columns = migrated_columns["tickets"]
        embedded_columns = migrated_columns["embedded_pages"]
        tables = {row[0] for row in connection.execute("SELECT name FROM sqlite_master WHERE type = 'table'").fetchall()}
        ticket_count = connection.execute("SELECT COUNT(*) FROM tickets").fetchone()[0]
        image_count = connection.execute("SELECT COUNT(*) FROM ticket_images").fetchone()[0]
        migrated_stores = connection.execute("SELECT ticket_id, store_name FROM ticket_stores ORDER BY id").fetchall()
        migrated_brands = connection.execute("SELECT ticket_id, brand FROM ticket_brands ORDER BY id").fetchall()

    assert "assigned_to" in ticket_columns
    assert "closed_at" in ticket_columns
    assert "ticket_logs" in tables
    assert "ticket_files" in tables
    assert "ticket_stores" in tables
    assert "ticket_brands" in tables
    assert "ticket_participants" in tables
    assert "ticket_comments" in tables
    assert "ticket_tasks" in tables
    assert "embedded_pages" in tables
    assert {"storage_type", "entry_file", "file_size"}.issubset(embedded_columns)
    assert {"archived_at", "archived_by", "archive_reason"}.issubset(ticket_columns)
    assert {"archived_at", "archived_by", "archive_reason"}.issubset(migrated_columns["shift_types"])
    for table in soft_delete_tables:
        assert {"deleted_at", "deleted_by", "delete_reason"}.issubset(migrated_columns[table])
    assert ticket_count == 1
    assert image_count == 1
    assert migrated_stores == [(1, "南京门东店")]
    assert migrated_brands == []

    assert_login_success(login_admin(client))
    admin_page = client.get("/admin")
    assert admin_page.status_code == 200
    assert "旧库里的工单" in admin_page.text


def test_legacy_employee_store_name_is_backfilled_to_employee_store_map(tmp_path, monkeypatch):
    config_dir = tmp_path / "config"
    write_config(config_dir)
    db_path = tmp_path / "tickets.db"
    with sqlite3.connect(db_path) as connection:
        connection.execute(
            """
            CREATE TABLE employees (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_name TEXT NOT NULL,
                store_name TEXT NOT NULL,
                role TEXT,
                phone TEXT,
                status TEXT NOT NULL DEFAULT '在职',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        connection.execute(
            """
            INSERT INTO employees (
                employee_name, store_name, role, phone, status, created_at, updated_at
            )
            VALUES ('旧南京员工', '南京门东店', '店员', '', '在职', '2026-07-03 10:00:00', '2026-07-03 10:00:00')
            """
        )
        connection.execute(
            """
            INSERT INTO employees (
                employee_name, store_name, role, phone, status, created_at, updated_at
            )
            VALUES ('旧南昌员工', '南昌万寿宫店', '店员', '', '在职', '2026-07-03 10:00:00', '2026-07-03 10:00:00')
            """
        )

    monkeypatch.setenv("STORE_REQUEST_DB_PATH", str(db_path))
    monkeypatch.setenv("STORE_REQUEST_UPLOAD_DIR", str(tmp_path / "uploads"))
    monkeypatch.setenv("STORE_REQUEST_CONFIG_DIR", str(config_dir))
    monkeypatch.setenv("ADMIN_USERS", "")
    monkeypatch.setenv("ADMIN_USERNAME", ADMIN_AUTH[0])
    monkeypatch.setenv("ADMIN_PASSWORD", ADMIN_AUTH[1])
    monkeypatch.setenv("SESSION_SECRET", "test-session-secret")
    monkeypatch.syspath_prepend(str(PROJECT_DIR))
    sys.modules.pop("main", None)
    main = importlib.import_module("main")
    monkeypatch.delenv("ADMIN_USERS", raising=False)

    with sqlite3.connect(db_path) as connection:
        pairs = connection.execute(
            "SELECT employee_id, store_name FROM employee_store_map ORDER BY employee_id, store_name"
        ).fetchall()

    assert pairs == [(1, "南京门东店"), (2, "南昌万寿宫店")]

    client = TestClient(main.app)
    assert_login_success(login_admin(client))
    nanjing_page = client.get("/admin/schedules?store_name=南京门东店&month=2026-07")
    assert nanjing_page.status_code == 200
    assert "旧南京员工" in nanjing_page.text
    assert "旧南昌员工" not in nanjing_page.text


def test_config_files_drive_options_validation_images_and_export_name(tmp_path, monkeypatch):
    client, _ = build_client(
        tmp_path,
        monkeypatch,
        {
            "stores.json": ["测试门店"],
            "request_types.json": ["陈列需求"],
            "urgency_levels.json": ["立刻处理"],
            "statuses.json": ["新建", "跟进中"],
            "brands.json": ["自有品牌"],
            "system.json": {
                "app_name": "配置化工单",
                "port": 8701,
                "max_image_mb": 1,
                "allowed_image_extensions": ["png"],
                "default_status": "新建",
                "excel_filename_prefix": "配置导出",
                "page_size": 25,
                "max_image_count": 2,
                "max_total_upload_mb": 3,
            },
        },
    )

    submit_page = client.get("/submit")
    assert submit_page.status_code == 200
    assert "测试门店" in submit_page.text
    assert "陈列需求" in submit_page.text
    assert "立刻处理" in submit_page.text
    assert "南京门东店" not in submit_page.text
    assert "建单需求" not in submit_page.text

    bad_request_type = client.post(
        "/submit",
        data={
            "store_name": "测试门店",
            "submitter": "配置测试",
            "request_type": "建单需求",
            "urgency": "立刻处理",
            "description": "配置文件校验应拒绝未配置的需求类型",
        },
    )
    assert bad_request_type.status_code == 400
    assert "请选择有效需求类型" in bad_request_type.text

    oversized_image = client.post(
        "/submit",
        data={
            "store_name": "测试门店",
            "submitter": "配置测试",
            "request_type": "陈列需求",
            "urgency": "立刻处理",
            "description": "图片大小限制来自 system.json",
        },
        files=[("images", ("large.png", VALID_PNG + b"0" * (1024 * 1024 + 1), "image/png"))],
    )
    assert oversized_image.status_code == 400
    assert "单张图片不能超过 1MB" in oversized_image.text

    response = client.post(
        "/submit",
        data={
            "store_name": "测试门店",
            "submitter": "配置测试",
            "request_type": "陈列需求",
            "urgency": "立刻处理",
            "brand": "自有品牌",
            "description": "配置化提交成功",
        },
    )
    assert response.status_code == 200

    logged_in_client(client)
    admin_page = client.get("/admin?status=新建")
    assert admin_page.status_code == 200
    assert "新建" in admin_page.text
    assert "跟进中" in admin_page.text
    assert "配置化提交成功" in admin_page.text

    detail_page = client.get("/admin/ticket/1")
    assert detail_page.status_code == 200
    assert "跟进中" in detail_page.text

    export_response = client.get("/admin/export")
    assert export_response.status_code == 200
    assert "%E9%85%8D%E7%BD%AE%E5%AF%BC%E5%87%BA_" in export_response.headers["content-disposition"]


def test_missing_and_invalid_config_files_fall_back_to_defaults(tmp_path, monkeypatch):
    client, main = build_client(tmp_path, monkeypatch, write_configs=False)
    submit_page = client.get("/submit")
    assert submit_page.status_code == 200
    assert "南京门东店" in submit_page.text
    assert "建单需求" in submit_page.text
    assert main.load_app_config().page_size == 50
    assert main.load_app_config().max_image_count == 5
    assert main.load_app_config().max_total_upload_mb == 30
    assert main.load_app_config().allowed_file_extensions == ["pdf", "doc", "docx", "xls", "xlsx", "csv", "txt", "zip", "rar"]
    assert main.load_app_config().max_file_mb == 20
    assert main.load_app_config().max_file_count == 5
    assert main.load_app_config().max_total_file_upload_mb == 50


def test_store_query_filters_by_store_paginates_and_hides_protected_links(tmp_path, monkeypatch):
    client, _ = build_client(
        tmp_path,
        monkeypatch,
        config_overrides={
            "system.json": dict(DEFAULT_TEST_CONFIG["system.json"], store_query_default_days=30, store_query_page_size=2)
        },
    )
    for index in range(3):
        submit_ticket(client, store_name="南京门东店", description=f"南京查询工单 {index}", product_name=f"南京商品{index}")
    submit_ticket_with_image_and_file(client, store_name="南昌万寿宫店", description="南昌不应出现", product_name="南昌商品")

    empty_store = client.get("/query?keyword=南京")
    assert empty_store.status_code == 400
    assert "请选择门店" in empty_store.text

    page = client.get("/query?store_name=南京门东店")
    assert page.status_code == 200
    assert "门店工单查询" in page.text
    assert "南京查询工单" in page.text
    assert "南昌不应出现" not in page.text
    assert "当前第 1 页" in page.text
    assert "page=2" in page.text
    assert "/admin/files" not in page.text
    assert "/admin/uploads" not in page.text
    assert "查看详情" in page.text
    assert "补充资料" in page.text

    keyword_page = client.get("/query?store_name=南京门东店&keyword=南京商品1")
    assert "南京商品1" in keyword_page.text
    assert "南京商品0" not in keyword_page.text

    ticket_no = rows_for(tmp_path, "tickets")[0]["ticket_no"]
    ticket_page = client.get(f"/query?store_name=南京门东店&ticket_no={ticket_no}")
    assert ticket_no in ticket_page.text

    empty_page = client.get("/query?store_name=山城巷店")
    assert "当前门店暂无符合条件的工单。" in empty_page.text


def test_store_ticket_detail_flow_safety_and_return_url(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    response = submit_ticket_with_image_and_file(
        client,
        store_name="南京门东店",
        description="这是一条门店完整详情问题说明",
        product_name="详情页测试商品",
        expected_finish_date=(date.today() + timedelta(days=1)).isoformat(),
    )
    assert response.status_code == 200

    logged_in_client(client)
    admin_post(
        client,
        "/admin/ticket/1",
        data={"status": "待门店补充", "assigned_to": "总部商品", "handler_note": "请补充门店详情证明"},
        follow_redirects=False,
    )

    query_path = "/query?store_name=南京门东店&keyword=详情页测试商品"
    query_page = client.get(query_path)
    assert query_page.status_code == 200
    assert "查看详情" in query_page.text
    assert "/query/ticket/1" in query_page.text
    assert "return_url=" in query_page.text

    detail_url = (
        "/query/ticket/1?"
        + f"store_name={quote('南京门东店')}&return_url={quote(query_path, safe='')}"
    )
    detail_page = client.get(detail_url)
    assert detail_page.status_code == 200
    assert "门店工单详情" in detail_page.text
    assert "REQ-" in detail_page.text
    assert "待门店补充" in detail_page.text
    assert "请补充门店详情证明" in detail_page.text
    assert "这是一条门店完整详情问题说明" in detail_page.text
    assert "时效状态" in detail_page.text
    assert "补充资料" in detail_page.text
    assert "图片数量" in detail_page.text
    assert "文件数量" in detail_page.text
    assert "暂无补充资料记录。" in detail_page.text
    assert "/admin/uploads" not in detail_page.text
    assert "/admin/files" not in detail_page.text
    assert "删除" not in detail_page.text
    assert "保存处理结果" not in detail_page.text
    assert 'name="status"' not in detail_page.text

    missing_store = client.get("/query/ticket/1")
    assert missing_store.status_code == 400
    assert "请选择门店后查看工单详情" in missing_store.text

    mismatch = client.get("/query/ticket/1?store_name=南昌万寿宫店")
    assert mismatch.status_code == 404
    assert "未找到该门店对应工单" in mismatch.text

    unsafe_return = client.get("/query/ticket/1?store_name=南京门东店&return_url=https%3A%2F%2Fevil.example%2Fquery")
    assert unsafe_return.status_code == 200
    assert "evil.example" not in unsafe_return.text

    supplement = client.post(
        "/query/ticket/1/supplement",
        data={
            "store_name": "南京门东店",
            "submitter": "小李",
            "note": "补充了现场陈列照片说明",
            "return_url": "/query/ticket/1?store_name=南京门东店",
        },
    )
    assert supplement.status_code == 200
    assert "返回工单详情" in supplement.text
    assert "返回查询结果" in supplement.text

    detail_after_supplement = client.get("/query/ticket/1?store_name=南京门东店")
    assert detail_after_supplement.status_code == 200
    assert "补充了现场陈列照片说明" in detail_after_supplement.text
    assert "小李" in detail_after_supplement.text

    admin_detail = client.get("/admin/ticket/1")
    assert admin_detail.status_code == 200
    assert "请补充门店详情证明" in admin_detail.text
    export_response = client.get("/admin/export")
    assert export_response.status_code == 200
    assert export_response.content.startswith(b"PK")


def test_store_supplement_records_sources_logs_and_status_transition(tmp_path, monkeypatch):
    client, _ = build_client(
        tmp_path,
        monkeypatch,
        config_overrides={
            "system.json": dict(
                DEFAULT_TEST_CONFIG["system.json"],
                supplement_status_after_store_update="待处理",
            )
        },
    )
    submit_ticket(client, store_name="南京门东店", description="需要门店补充")
    logged_in_client(client)
    admin_post(
        client,
        "/admin/ticket/1",
        data={"status": "待门店补充", "assigned_to": "总部商品", "handler_note": "请补图"},
        follow_redirects=False,
    )

    supplement_page = client.get("/query/ticket/1/supplement?store_name=南京门东店")
    assert supplement_page.status_code == 200
    assert "门店补充资料" in supplement_page.text
    assert "/admin/files" not in supplement_page.text

    mismatch = client.get("/query/ticket/1/supplement?store_name=南昌万寿宫店")
    assert mismatch.status_code == 403

    empty = client.post(
        "/query/ticket/1/supplement",
        data={"store_name": "南京门东店", "submitter": "小李", "note": ""},
    )
    assert empty.status_code == 400
    assert "请填写补充说明或上传附件" in empty.text

    response = client.post(
        "/query/ticket/1/supplement",
        data={"store_name": "南京门东店", "submitter": "小李", "note": "已补充现场照片和表格"},
        files=[
            ("images", ("supplement.png", VALID_PNG, "image/png")),
            ("files", ("supplement.xlsx", b"sku,qty\nA001,2\n", "application/octet-stream")),
        ],
    )
    assert response.status_code == 200
    assert "补充资料已提交，总部会继续处理。" in response.text
    assert "返回查询结果" in response.text

    supplements = rows_for(tmp_path, "ticket_supplements")
    images = rows_for(tmp_path, "ticket_images")
    files = rows_for(tmp_path, "ticket_files")
    tickets = rows_for(tmp_path, "tickets")
    logs = rows_for(tmp_path, "ticket_logs")
    assert supplements[-1]["submitter"] == "小李"
    assert supplements[-1]["image_count"] == 1
    assert supplements[-1]["file_count"] == 1
    assert images[-1]["source"] == "store_supplement"
    assert images[-1]["uploaded_by"] == "小李"
    assert files[-1]["source"] == "store_supplement"
    assert files[-1]["uploaded_by"] == "小李"
    assert tickets[0]["status"] == "待处理"
    assert logs[-1]["action"] == "门店补充资料"
    assert logs[-1]["operator"].startswith("门店:")
    assert logs[-1]["old_status"] == "待门店补充"
    assert logs[-1]["new_status"] == "待处理"


def test_request_type_rules_validate_fields_attachments_and_missing_config(tmp_path, monkeypatch):
    rules = {
        "建单需求": {
            "required_fields": ["brand", "product_name", "quantity"],
            "require_image": False,
            "require_file": False,
            "require_any_attachment": True,
            "description_hint": "请说明到货情况、采购单需求或建单原因",
        },
        "商品异常": {
            "required_fields": ["product_name", "description"],
            "require_image": True,
            "require_file": False,
            "require_any_attachment": False,
            "description_hint": "请说明异常现象、影响范围和处理诉求",
        },
    }
    client, _ = build_client(tmp_path, monkeypatch, config_overrides={"request_type_rules.json": rules})

    submit_page = client.get("/submit")
    assert "选择需求类型后会显示填写提示、建议填写和必填信息。" in submit_page.text
    assert "工单类型模板" in submit_page.text
    assert "建议填写" in submit_page.text
    assert "data-request-type-rules" in submit_page.text

    missing_brand = submit_ticket_with_file(client, request_type="建单需求", brand="", product_name="商品", quantity="1")
    assert missing_brand.status_code == 200
    assert rows_for(tmp_path, "tickets")[-1]["brand"] == ""

    missing_attachment = submit_ticket(client, request_type="建单需求", brand="品牌", product_name="商品", quantity="1")
    assert missing_attachment.status_code == 400
    assert "建单需求必须上传图片或文件附件。" in missing_attachment.text

    missing_image = submit_ticket(client, request_type="商品异常", product_name="商品异常商品")
    assert missing_image.status_code == 400
    assert "商品异常必须上传至少一张图片。" in missing_image.text

    unconfigured = submit_ticket(client, request_type="系统问题", description="系统配置未覆盖的需求")
    assert unconfigured.status_code == 200

    missing_rules_client, _ = build_client(tmp_path / "missing-rules", monkeypatch)
    assert submit_ticket(missing_rules_client, request_type="建单需求", brand="", product_name="", quantity="").status_code == 200


def test_admin_dashboard_due_status_and_attachment_statistics(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    yesterday = (date.today() - timedelta(days=1)).isoformat()
    today = date.today().isoformat()
    tomorrow = (date.today() + timedelta(days=1)).isoformat()

    submit_ticket_with_image(client, store_name="南京门东店", request_type="商品异常", expected_finish_date=yesterday, description="已超时工单")
    submit_ticket_with_file(client, store_name="南京门东店", request_type="缺货需求", expected_finish_date=today, description="今日到期工单")
    submit_ticket(client, store_name="南昌万寿宫店", request_type="新品需求", expected_finish_date=tomorrow, description="未到期工单")
    submit_ticket(client, store_name="山城巷店", request_type="系统问题", expected_finish_date="", description="未设置工单")
    logged_in_client(client)
    admin_post(
        client,
        "/admin/ticket/3",
        data={"status": "已完成", "assigned_to": "采购", "handler_note": "按时完成"},
        follow_redirects=False,
    )

    dashboard_redirect = TestClient(__import__("main").app).get("/admin/dashboard", follow_redirects=False)
    assert dashboard_redirect.status_code == 303

    dashboard = client.get("/admin/dashboard")
    assert dashboard.status_code == 200
    assert "统计看板" in dashboard.text
    assert "总工单数" in dashboard.text
    assert "商品异常" in dashboard.text
    assert "南京门东店" in dashboard.text
    assert "采购" in dashboard.text
    assert "超时工单" in dashboard.text
    assert "有图片工单数" in dashboard.text
    assert "有文件工单数" in dashboard.text
    assert "无附件工单数" in dashboard.text

    admin_page = client.get("/admin")
    assert "已超时" in admin_page.text
    assert "今日到期" in admin_page.text
    assert "未设置" in admin_page.text
    assert "按时完成" in admin_page.text

    overdue_filter = client.get("/admin?due_status=已超时")
    assert "已超时工单" in overdue_filter.text
    assert "今日到期工单" not in overdue_filter.text

    detail_page = client.get("/admin/ticket/1")
    assert "该工单已超过期望完成时间，请优先处理。" in detail_page.text


def test_lightweight_erp_admin_layout_navigation_and_placeholder_pages(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    submit_ticket_with_image_and_file(client, description="ERP 布局测试工单")
    logged_in_client(client)

    admin_page = client.get("/admin")
    assert admin_page.status_code == 200
    assert 'class="admin-layout"' in admin_page.text
    assert "/static/img/zhiyang-logo.png" in admin_page.text
    assert "止痒 ERP" in admin_page.text
    assert "工单管理" in admin_page.text
    assert "data-current-menu=\"tickets\"" in admin_page.text
    assert "数据导出" not in admin_page.text
    assert "导出 Excel" in admin_page.text
    assert "权限：系统管理员" in admin_page.text

    dashboard = client.get("/admin/dashboard")
    assert dashboard.status_code == 200
    assert "/static/img/zhiyang-logo.png" in dashboard.text
    assert "止痒运营协同总览" in dashboard.text
    assert "dashboard-hero" in dashboard.text
    assert "最近工单动态" in dashboard.text

    for label in ("业务总览", "工单管理", "门店查询", "配置管理", "账号设置", "系统设置"):
        assert label in dashboard.text
    assert "数据导出" not in dashboard.text

    for path, title in (
        ("/admin/settings", "配置管理"),
        ("/admin/account", "账号设置"),
        ("/admin/system", "系统设置"),
    ):
        unauthenticated = TestClient(__import__("main").app).get(path, follow_redirects=False)
        assert unauthenticated.status_code == 303
        assert unauthenticated.headers["location"].startswith("/admin/login")
        page = client.get(path)
        assert page.status_code == 200
        assert 'class="admin-layout"' in page.text
        assert title in page.text

    assert client.get("/submit").status_code == 200
    assert client.get("/query").status_code == 200
    assert client.get("/admin/ticket/1").status_code == 200
    export_response = client.get("/admin/export")
    assert export_response.status_code == 200
    assert export_response.content.startswith(b"PK")


def test_navigation_branding_export_links_and_public_topbar(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    submit_ticket(client, store_name="南京门东店", request_type="缺货需求", status="待处理")
    logged_in_client(client)

    filtered_admin = client.get("/admin?store_name=南京门东店&status=待处理&due_status=未设置&keyword=自动化")
    assert filtered_admin.status_code == 200
    assert "导出 Excel" in filtered_admin.text
    assert "store_name=%E5%8D%97%E4%BA%AC%E9%97%A8%E4%B8%9C%E5%BA%97" in filtered_admin.text
    assert "status=%E5%BE%85%E5%A4%84%E7%90%86" in filtered_admin.text
    assert "due_status=%E6%9C%AA%E8%AE%BE%E7%BD%AE" in filtered_admin.text
    assert "keyword=%E8%87%AA%E5%8A%A8%E5%8C%96" in filtered_admin.text

    query_page = client.get("/query")
    assert query_page.status_code == 200
    assert "public-topbar" in query_page.text
    assert "/static/img/zhiyang-logo.png" in query_page.text
    assert "提交新工单" in query_page.text
    assert "返回业务总览" in query_page.text
    assert "返回工单管理" in query_page.text
    assert "门店工单查询" in query_page.text

    submit_page = client.get("/submit")
    assert submit_page.status_code == 200
    assert "public-topbar" in submit_page.text
    assert "/static/img/zhiyang-logo.png" in submit_page.text
    assert "查询工单" in submit_page.text
    assert "返回业务总览" in submit_page.text
    assert "返回工单管理" in submit_page.text
    assert "门店需求提报" in submit_page.text

    detail_page = client.get("/query/ticket/1?store_name=南京门东店")
    assert detail_page.status_code == 200
    assert "public-topbar" in detail_page.text
    assert "/static/img/zhiyang-logo.png" in detail_page.text
    assert "返回查询结果" in detail_page.text
    assert "提交新工单" in detail_page.text
    assert "返回门店查询" in detail_page.text


def test_zhiyang_logo_asset_and_styles_are_present():
    style = (PROJECT_DIR / "static" / "style.css").read_text(encoding="utf-8")
    logo_path = PROJECT_DIR / "static" / "img" / "zhiyang-logo.png"

    assert logo_path.is_file()
    assert logo_path.stat().st_size > 0
    assert ".sidebar-logo-img" in style
    assert ".public-logo-img" in style
    assert "object-fit: contain" in style


def test_runtime_embedded_page_upload_replace_navigation_and_content(tmp_path, monkeypatch):
    client, main = build_client(tmp_path, monkeypatch)

    unauthenticated = client.get("/admin/embedded-pages", follow_redirects=False)
    assert unauthenticated.status_code == 303
    assert unauthenticated.headers["location"].startswith("/admin/login")

    logged_in_client(client)
    monkeypatch.setattr(main, "now_text", lambda: "2026-07-04 10:00:00")
    page = client.get("/admin/embedded-pages")
    assert page.status_code == 200
    assert "嵌入页面管理" in page.text
    assert 'name="page_key"' in page.text
    assert 'name="html_file"' in page.text
    assert 'accept=".html,.zip"' in page.text
    assert "ZIP" in page.text
    assert "100MB" in page.text

    csrf_token = csrf_token_for(client, "/admin/embedded-pages")
    html_content = b"<h1>V1 report</h1>"
    create = client.post(
        "/admin/embedded-pages",
        data={
            "csrf_token": csrf_token,
            "page_key": "daily-report",
            "title": "每日经营日报",
            "nav_label": "经营日报",
            "enabled": "1",
        },
        files={"html_file": ("daily.html", html_content, "text/html")},
        follow_redirects=False,
    )
    assert create.status_code == 303
    assert create.headers["location"].startswith("/admin/embedded-pages")

    pages = rows_for(tmp_path, "embedded_pages")
    assert len(pages) == 1
    assert pages[0]["page_key"] == "daily-report"
    assert pages[0]["title"] == "每日经营日报"
    assert pages[0]["nav_label"] == "经营日报"
    assert pages[0]["filename"] == "daily-report/index.html"
    assert pages[0]["storage_type"] == "html"
    assert pages[0]["entry_file"] == "index.html"
    assert pages[0]["file_size"] == len(html_content)
    assert pages[0]["enabled"] == 1
    assert pages[0]["updated_by"] == ADMIN_AUTH[0]
    embedded_file = tmp_path / "embedded_pages" / "daily-report" / "index.html"
    assert embedded_file.is_file()
    assert embedded_file.read_text(encoding="utf-8") == "<h1>V1 report</h1>"

    list_page = client.get("/admin/embedded-pages")
    assert "HTML" in list_page.text
    assert "index.html" in list_page.text
    assert "存在" in list_page.text

    dashboard = client.get("/admin/dashboard")
    assert dashboard.status_code == 200
    assert "扩展页面" in dashboard.text
    assert "经营日报" in dashboard.text
    assert "/admin/embed/daily-report" in dashboard.text

    embed_page = client.get("/admin/embed/daily-report")
    assert embed_page.status_code == 200
    assert "每日经营日报" in embed_page.text
    assert "/admin/embed-content/daily-report/index.html?v=20260704100000" in embed_page.text
    assert "embedded-frame" in embed_page.text

    content = client.get("/admin/embed-content/daily-report/index.html")
    assert content.status_code == 200
    assert content.text == "<h1>V1 report</h1>"
    assert content.headers["x-frame-options"] == "SAMEORIGIN"
    assert content.headers["x-content-type-options"] == "nosniff"
    assert content.headers["content-type"].startswith("text/html")

    fresh_client = TestClient(main.app)
    denied_content = fresh_client.get("/admin/embed-content/daily-report/index.html", follow_redirects=False)
    assert denied_content.status_code != 200

    monkeypatch.setattr(main, "now_text", lambda: "2026-07-04 10:05:00")
    zip_content = make_zip_bytes(
        {
            "index.html": '<link rel="stylesheet" href="assets/style.css"><h1>V2 package</h1>',
            "assets/style.css": "body { color: #c00; }",
            "assets/chart.js": "window.embeddedReportReady = true;",
        }
    )
    replace = client.post(
        "/admin/embedded-pages/daily-report/replace",
        data={"csrf_token": csrf_token},
        files={"html_file": ("daily.zip", zip_content, "application/zip")},
        follow_redirects=False,
    )
    assert replace.status_code == 303
    updated_pages = rows_for(tmp_path, "embedded_pages")
    assert updated_pages[0]["updated_at"] == "2026-07-04 10:05:00"
    assert updated_pages[0]["storage_type"] == "zip"
    assert updated_pages[0]["entry_file"] == "index.html"
    assert updated_pages[0]["file_size"] == len(zip_content)
    assert "V2 package" in embedded_file.read_text(encoding="utf-8")
    assert (tmp_path / "embedded_pages" / "daily-report" / "assets" / "style.css").is_file()
    updated_embed = client.get("/admin/embed/daily-report")
    assert "/admin/embed-content/daily-report/index.html?v=20260704100500" in updated_embed.text
    assert "V2 package" in client.get("/admin/embed-content/daily-report/index.html").text
    css = client.get("/admin/embed-content/daily-report/assets/style.css")
    assert css.status_code == 200
    assert css.text == "body { color: #c00; }"
    assert css.headers["content-type"].startswith("text/css")
    script = client.get("/admin/embed-content/daily-report/assets/chart.js")
    assert script.status_code == 200
    assert script.headers["content-type"].startswith("application/javascript")
    replaced_list = client.get("/admin/embedded-pages")
    assert "ZIP" in replaced_list.text
    assert "index.html" in replaced_list.text
    assert "存在" in replaced_list.text

    disable = client.post(
        "/admin/embedded-pages/daily-report/toggle",
        data={"csrf_token": csrf_token, "enabled": "0"},
        follow_redirects=False,
    )
    assert disable.status_code == 303
    disabled_dashboard = client.get("/admin/dashboard")
    assert "经营日报" not in disabled_dashboard.text
    assert client.get("/admin/embed/daily-report").status_code == 404


def test_embedded_page_upload_validation_and_backup_rules(tmp_path, monkeypatch):
    client, _ = build_client(
        tmp_path,
        monkeypatch,
        config_overrides={"system.json": dict(DEFAULT_TEST_CONFIG["system.json"], max_embedded_html_mb=1)},
    )
    logged_in_client(client)
    csrf_token = csrf_token_for(client, "/admin/embedded-pages")

    for bad_key in ("../test", "中文", "bad key"):
        response = client.post(
            "/admin/embedded-pages",
            data={
                "csrf_token": csrf_token,
                "page_key": bad_key,
                "title": "非法页面",
                "nav_label": "非法",
                "enabled": "1",
            },
            files={"html_file": ("bad.html", b"<h1>bad</h1>", "text/html")},
        )
        assert response.status_code == 400
        assert "page_key 只能使用小写字母、数字和短横线。" in response.text

    non_html = client.post(
        "/admin/embedded-pages",
        data={
            "csrf_token": csrf_token,
            "page_key": "not-html",
            "title": "非 HTML",
            "nav_label": "非 HTML",
            "enabled": "1",
        },
        files={"html_file": ("report.txt", b"plain", "text/plain")},
    )
    assert non_html.status_code == 400
    assert "只允许上传 .html 或 .zip 文件。" in non_html.text

    empty_html = client.post(
        "/admin/embedded-pages",
        data={
            "csrf_token": csrf_token,
            "page_key": "empty-html",
            "title": "空 HTML",
            "nav_label": "空 HTML",
            "enabled": "1",
        },
        files={"html_file": ("empty.html", b"", "text/html")},
    )
    assert empty_html.status_code == 400
    assert "HTML 文件内容不能为空。" in empty_html.text

    too_large = client.post(
        "/admin/embedded-pages",
        data={
            "csrf_token": csrf_token,
            "page_key": "large-html",
            "title": "超大 HTML",
            "nav_label": "超大",
            "enabled": "1",
        },
        files={"html_file": ("large.html", b"x" * (1024 * 1024 + 1), "text/html")},
    )
    assert too_large.status_code == 400
    assert "HTML 文件不能超过 1MB。" in too_large.text
    assert rows_for(tmp_path, "embedded_pages") == []
    assert not (tmp_path / "embedded_pages").exists()

    backup_script = (PROJECT_DIR / "backup.sh").read_text(encoding="utf-8")
    gitignore = (PROJECT_DIR / ".gitignore").read_text(encoding="utf-8")
    assert "data/embedded_pages" in backup_script
    assert "data/embedded_pages/" in gitignore


def test_embedded_zip_upload_security_validation(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    logged_in_client(client)
    csrf_token = csrf_token_for(client, "/admin/embedded-pages")

    def upload_zip(page_key, entries):
        return client.post(
            "/admin/embedded-pages",
            data={
                "csrf_token": csrf_token,
                "page_key": page_key,
                "title": "ZIP package",
                "nav_label": "ZIP",
                "enabled": "1",
            },
            files={"html_file": (f"{page_key}.zip", make_zip_bytes(entries), "application/zip")},
            follow_redirects=False,
        )

    single_html = upload_zip(
        "single-html",
        {
            "线下门店每日运营日报_2026-06-30.html": "<h1>单文件日报</h1>",
            "assets/style.css": "body {}",
        },
    )
    assert single_html.status_code == 303
    index_file = tmp_path / "embedded_pages" / "single-html" / "index.html"
    original_file = tmp_path / "embedded_pages" / "single-html" / "线下门店每日运营日报_2026-06-30.html"
    assert index_file.read_text(encoding="utf-8") == "<h1>单文件日报</h1>"
    assert not original_file.exists()
    assert client.get("/admin/embed-content/single-html/index.html").text == "<h1>单文件日报</h1>"

    multiple_html = upload_zip("multiple-html", {"daily.html": "<h1>A</h1>", "summary.html": "<h1>B</h1>"})
    assert multiple_html.status_code == 400
    assert "ZIP 根目录包含多个 HTML 文件，请将入口文件命名为 index.html。" in multiple_html.text

    no_html = upload_zip("missing-html", {"assets/style.css": "body {}"})
    assert no_html.status_code == 400
    assert "ZIP 根目录必须包含 index.html 或一个 HTML 文件。" in no_html.text

    traversal = upload_zip("bad-path", {"index.html": "<h1>ok</h1>", "../evil.txt": "bad"})
    assert traversal.status_code == 400
    assert "ZIP 文件路径不安全" in traversal.text

    blocked_extension = upload_zip("bad-ext", {"index.html": "<h1>ok</h1>", "assets/tool.exe": "bad"})
    assert blocked_extension.status_code == 400
    assert "exe" in blocked_extension.text

    rows = rows_for(tmp_path, "embedded_pages")
    assert [row["page_key"] for row in rows] == ["single-html"]


def test_session_security_secure_cookie_expiry_csrf_and_env_example(tmp_path, monkeypatch):
    monkeypatch.setenv("SESSION_MAX_AGE_HOURS", "1")
    monkeypatch.setenv("SESSION_COOKIE_SECURE", "true")
    secure_client, _secure_main = build_client(tmp_path / "secure-cookie", monkeypatch)

    login = login_admin(secure_client)
    assert_login_success(login)
    set_cookie = login.headers.get("set-cookie", "")
    assert "Max-Age=3600" in set_cookie
    assert "Secure" in set_cookie

    monkeypatch.setenv("SESSION_COOKIE_SECURE", "false")
    client, main = build_client(tmp_path / "csrf", monkeypatch)
    assert_login_success(login_admin(client))

    no_csrf = client.post(
        "/admin/ticket/1",
        data={"status": "处理中", "assigned_to": "采购", "handler_note": "缺少 csrf"},
        follow_redirects=False,
    )
    assert no_csrf.status_code == 403

    submit_ticket(client)
    ok = admin_post(
        client,
        "/admin/ticket/1",
        data={"status": "处理中", "assigned_to": "采购", "handler_note": "带 csrf"},
        follow_redirects=False,
    )
    assert ok.status_code == 303

    expired_token = main.create_admin_session(ADMIN_AUTH[0], issued_at=0, max_age_seconds=1)
    expired_client = TestClient(main.app)
    expired_client.cookies.set("admin_session", expired_token)
    expired = expired_client.get("/admin", follow_redirects=False)
    assert expired.status_code == 303

    env_example = (PROJECT_DIR / ".env.example").read_text(encoding="utf-8")
    assert "APP_ENV=development" in env_example
    assert "SESSION_MAX_AGE_HOURS=12" in env_example
    assert "SESSION_COOKIE_SECURE=false" in env_example


def test_notifications_for_new_tickets_api_reads_and_per_user_state(tmp_path, monkeypatch):
    client, main = build_client(
        tmp_path,
        monkeypatch,
        admin_users="admin:123456,caigou:123456",
    )
    normal = submit_ticket(client, urgency="普通", request_type="建单需求", description="普通提醒工单")
    warning = submit_ticket(client, urgency="加急", request_type="审单需求", description="加急提醒工单")
    urgent = submit_ticket(client, urgency="当天必须处理", request_type="商品异常", description="紧急提醒工单")
    assert normal.status_code == 200
    assert warning.status_code == 200
    assert urgent.status_code == 200

    events = rows_for(tmp_path, "notification_events")
    assert [event["event_type"] for event in events] == ["new_ticket", "new_ticket", "new_ticket"]
    assert [event["severity"] for event in events] == ["info", "warning", "urgent"]
    assert events[-1]["title"] == "新工单"
    assert "商品异常" in events[-1]["content"]

    unauthenticated = TestClient(main.app).get("/admin/api/notifications", follow_redirects=False)
    assert unauthenticated.status_code == 401

    assert_login_success(login_admin(client, "admin", "123456"))
    api_page = client.get("/admin/api/notifications?limit=2")
    assert api_page.status_code == 200
    payload = api_page.json()
    assert payload["unread_count"] == 3
    assert payload["latest_id"] == events[-1]["id"]
    assert len(payload["notifications"]) == 2
    assert payload["notifications"][0]["id"] == events[-1]["id"]
    assert payload["notifications"][0]["is_read"] is False
    assert payload["notifications"][0]["detail_url"] == "/admin/ticket/3"

    missing_csrf = client.post(f"/admin/api/notifications/{events[-1]['id']}/read")
    assert missing_csrf.status_code == 403

    csrf_token = csrf_token_for(client)
    mark_one = client.post(
        f"/admin/api/notifications/{events[-1]['id']}/read",
        data={"csrf_token": csrf_token},
    )
    assert mark_one.status_code == 200
    assert mark_one.json()["unread_count"] == 2
    assert client.get("/admin/api/notifications?unread_only=true").json()["unread_count"] == 2

    caigou_client = TestClient(main.app)
    assert_login_success(login_admin(caigou_client, "caigou", "123456"))
    caigou_payload = caigou_client.get("/admin/api/notifications").json()
    assert caigou_payload["unread_count"] == 3

    read_all = client.post("/admin/api/notifications/read-all", data={"csrf_token": csrf_token})
    assert read_all.status_code == 200
    assert read_all.json()["unread_count"] == 0
    assert caigou_client.get("/admin/api/notifications").json()["unread_count"] == 3


def test_notifications_for_status_transition_and_store_supplement(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    submit_ticket(client, store_name="南京门东店", description="状态提醒工单")
    logged_in_client(client)
    admin_post(
        client,
        "/admin/ticket/1",
        data={"status": "待门店补充", "assigned_to": "总部商品", "handler_note": "请补充"},
        follow_redirects=False,
    )
    after_status = rows_for(tmp_path, "notification_events")
    assert [event["event_type"] for event in after_status] == ["new_ticket", "need_store_supplement"]
    assert after_status[-1]["title"] == "待门店补充"
    assert after_status[-1]["severity"] == "warning"

    admin_post(
        client,
        "/admin/ticket/1",
        data={"status": "待门店补充", "assigned_to": "总部商品", "handler_note": "备注变化但状态不变"},
        follow_redirects=False,
    )
    assert len(rows_for(tmp_path, "notification_events")) == 2

    supplement = client.post(
        "/query/ticket/1/supplement",
        data={"store_name": "南京门东店", "submitter": "小李", "note": "补充了资料"},
    )
    assert supplement.status_code == 200
    events = rows_for(tmp_path, "notification_events")
    assert [event["event_type"] for event in events] == [
        "new_ticket",
        "need_store_supplement",
        "store_supplement",
    ]
    assert events[-1]["title"] == "门店补充资料"
    assert events[-1]["severity"] == "warning"
    assert events[-1]["created_by"] == "门店:小李"


def test_notification_ui_hooks_are_rendered_on_admin_pages(tmp_path, monkeypatch):
    client, _ = build_client(tmp_path, monkeypatch)
    submit_ticket(client, description="页面消息入口工单")
    logged_in_client(client)

    for path in ("/admin", "/admin/ticket/1", "/admin/dashboard"):
        page = client.get(path)
        assert page.status_code == 200
        assert "data-notification-root" in page.text
        assert "data-notification-count" in page.text
        assert "data-notification-list" in page.text
        assert "开启桌面提醒" in page.text
        assert "消息" in page.text
        assert re.search(r'/static/app\.js\?v=[A-Za-z0-9._-]+', page.text)

    script = (PROJECT_DIR / "static" / "app.js").read_text(encoding="utf-8")
    assert "/admin/api/notifications" in script
    assert "setInterval" in script
    assert "Notification.requestPermission" in script
    assert "showNotificationToast" in script
    assert "data-notification-root" in script
    style = (PROJECT_DIR / "static" / "style.css").read_text(encoding="utf-8")
    assert ".notification-badge[hidden]" in style


def test_nginx_https_example_and_navigation_files_are_present(tmp_path, monkeypatch):
    nginx = PROJECT_DIR / "deploy" / "nginx-store-request-tool.conf.example"
    assert nginx.exists()
    content = nginx.read_text(encoding="utf-8")
    assert "listen 80" in content
    assert "listen 443 ssl" in content
    assert "server_name request.example.com" in content
    assert "proxy_pass http://127.0.0.1:8701" in content
    assert "client_max_body_size 120M" in content
    assert "X-Frame-Options" in content
    assert "certbot" in content

    for template_name in ("query.html", "supplement.html", "dashboard.html"):
        assert (PROJECT_DIR / "templates" / template_name).exists()
    config_dir = tmp_path / "config"
    config_dir.mkdir(parents=True, exist_ok=True)
    (config_dir / "stores.json").write_text("{bad json", encoding="utf-8")
    (config_dir / "system.json").write_text(
        json.dumps(
            {
                "page_size": -1,
                "max_image_count": "bad",
                "max_total_upload_mb": 0,
                "allowed_file_extensions": "bad",
                "max_file_mb": -1,
                "max_file_count": "bad",
                "max_total_file_upload_mb": 0,
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    monkeypatch.setenv("STORE_REQUEST_DB_PATH", str(tmp_path / "tickets.db"))
    monkeypatch.setenv("STORE_REQUEST_UPLOAD_DIR", str(tmp_path / "uploads"))
    monkeypatch.setenv("STORE_REQUEST_CONFIG_DIR", str(config_dir))
    monkeypatch.setenv("ADMIN_USERNAME", ADMIN_AUTH[0])
    monkeypatch.setenv("ADMIN_PASSWORD", ADMIN_AUTH[1])
    monkeypatch.setenv("SESSION_SECRET", "test-session-secret")
    monkeypatch.syspath_prepend(str(PROJECT_DIR))
    sys.modules.pop("main", None)
    main = importlib.import_module("main")
    invalid_config_client = TestClient(main.app)

    invalid_submit_page = invalid_config_client.get("/submit")
    assert invalid_submit_page.status_code == 200
    assert "南京门东店" in invalid_submit_page.text
    assert main.load_app_config().page_size == 50
    assert main.load_app_config().max_image_count == 5
    assert main.load_app_config().max_total_upload_mb == 30
    assert main.load_app_config().allowed_file_extensions == ["pdf", "doc", "docx", "xls", "xlsx", "csv", "txt", "zip", "rar"]
    assert main.load_app_config().max_file_mb == 20
    assert main.load_app_config().max_file_count == 5
    assert main.load_app_config().max_total_file_upload_mb == 50
